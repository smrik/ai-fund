"""CIQ workbook parser for the committed cleandata workbook contract."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PARSER_VERSION = "ibm_standard_v4"
REQUIRED_SHEETS = [
    "Financial Statements",
    "Common Size",
    "Detailed Comps",
]
OPTIONAL_SHEETS = ["Summary Comps"]


class CIQTemplateContractError(ValueError):
    """Raised when workbook layout does not match the expected CIQ workbook contract."""


@dataclass(slots=True)
class CIQWorkbookPayload:
    ticker: str
    source_file: str
    file_hash: str
    parser_version: str
    template_fingerprint: str
    long_form_records: list[dict[str, Any]]
    valuation_snapshot: dict[str, Any]
    comps_snapshot: list[dict[str, Any]]
    rows_parsed: int


_METRIC_MAP = {
    "Revenues": "revenue",
    "Total Revenues": "revenue",
    "Operating Income": "operating_income",
    "Capital Expenditure": "capex",
    "Capital Expenditures": "capex",
    "Depreciation & Amort.": "da",
    "Depreciation And Amortization": "da",
    "D&A For EBITDA": "da_for_ebitda",
    "Depreciation & Amort., Total": "da_total",
    "Income Tax Expense": "tax",
    "EBT Excl Unusual Items": "ebt_excl_unusual",
    "Total Debt": "debt",
    "Cash and Equivalents": "cash",
    "Cash And Equivalents": "cash",
    "Cash & Equivalents": "cash",
    "Weighted Avg. Diluted Shares Out.": "shares",
    "Weighted Avg. Diluted Shares Out": "shares",
    "ROIC": "roic",
    "FCF Yield": "fcf_yield",
    "Total Debt/": "debt_to_ebitda",
}


_EXCEL_ERROR_TOKENS = (
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#N/A",
    "#NUM!",
    "#NULL!",
    "#SPILL!",
    "#CALC!",
    "#FIELD!",
    "#BLOCKED!",
    "#UNKNOWN!",
    "#GETTING_DATA",
)


def _json_safe_cell_value(value: Any) -> Any:
    """Preserve cached values without introducing non-JSON scalar types."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _excel_error_token(value: Any, *, embedded: bool) -> str | None:
    if value is None:
        return None
    text = str(value).upper()
    for token in _EXCEL_ERROR_TOKENS:
        if (embedded and token in text) or (not embedded and text == token):
            return token
    return None


def _cell_provenance(
    ws: Worksheet,
    formula_ws: Worksheet | None,
    row_index: int,
    column_index: int,
) -> dict[str, Any]:
    """Return additive, cell-exact provenance for a cached-value worksheet cell."""
    cached_cell = ws.cell(row_index, column_index)
    formula_cell = formula_ws.cell(row_index, column_index) if formula_ws is not None else None
    formula_value = formula_cell.value if formula_cell is not None else None
    formula_text = (
        str(formula_value)
        if isinstance(formula_value, str) and formula_value.startswith("=")
        else None
    )
    cached_value = _json_safe_cell_value(cached_cell.value)
    has_cached_value = cached_cell.value not in (None, "")
    formula_error = _excel_error_token(formula_text, embedded=True)
    cached_error = _excel_error_token(cached_cell.value, embedded=False)

    if formula_error is not None:
        formula_status = "formula_error"
    elif cached_error is not None:
        formula_status = "cached_error"
    elif formula_text is not None and has_cached_value:
        formula_status = "formula_cached"
    elif formula_text is not None:
        formula_status = "formula_cache_missing"
    else:
        formula_status = "literal"

    coordinate = cached_cell.coordinate
    return {
        "row_index": row_index,
        "source_row_id": f"{ws.title}!{row_index}",
        "a1_locator": coordinate,
        "cell_locator": f"{ws.title}!{coordinate}",
        "formula_text": formula_text,
        "cached_value": cached_value,
        "has_formula": formula_text is not None,
        "has_cached_value": has_cached_value,
        "formula_status": formula_status,
        "formula_error": formula_error,
        "cached_error": cached_error,
    }


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"NM", "N/A", "--", "-", "na", "n/m"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_ticker(raw: str) -> str:
    if not raw:
        return ""
    text = raw.strip()
    if ":" in text:
        return text.split(":", 1)[1].strip().upper()
    return text.upper()


def _metric_key(label: str) -> str | None:
    if not label:
        return None
    if label in _METRIC_MAP:
        return _METRIC_MAP[label]
    slug = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_")
    return slug or None


def _percent_to_decimal(value: float | None) -> float | None:
    if value is None:
        return None
    if abs(value) > 1.5:
        return value / 100.0
    return value


def _default_unit_scale(sheet: str, section: str) -> tuple[str | None, float]:
    combined = f"{sheet} {section}".lower()
    if "usd in millions" in combined:
        return "USD", 1_000_000.0
    if "percentage" in combined or "margin" in combined or "cagr" in combined:
        return "%", 0.01
    if "multiple" in combined or "/" in combined:
        return "x", 1.0
    return None, 1.0


def _to_iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        if value.year <= 1900:
            return None
        return value.date().isoformat()
    if isinstance(value, date):
        if value.year <= 1900:
            return None
        return value.isoformat()
    text = _safe_str(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        try:
            if int(text[:4]) <= 1900:
                return None
        except ValueError:
            return None
        return text
    return None


def _row_contains(ws: Worksheet, row: int, needle: str) -> tuple[int, str] | None:
    for col in range(1, min(ws.max_column, 120) + 1):
        value = _safe_str(ws.cell(row, col).value)
        if needle.lower() in value.lower():
            return col, value
    return None


def _contiguous_date_columns(
    ws: Worksheet,
    row: int,
    start_col: int,
    max_col: int,
    header_row: int | None = None,
) -> list[tuple[int, str]]:
    period_cols: list[tuple[int, str]] = []
    for c in range(start_col, max_col + 1):
        period_date = _to_iso_date(ws.cell(row, c).value)
        if period_date is None:
            if period_cols:
                break
            continue

        if header_row is not None:
            header_val = _safe_str(ws.cell(header_row, c).value)
            if not header_val:
                if period_cols:
                    break
                continue

        period_cols.append((c, period_date))

    return period_cols


def _period_columns(ws: Worksheet) -> tuple[int, list[tuple[int, str]], dict[int, str | None]]:
    max_col = min(ws.max_column, 250)

    clean_period_header = _safe_str(ws.cell(1, 1).value).lower() == "period"
    clean_period_date = _safe_str(ws.cell(2, 1).value).lower().startswith("period date")
    if clean_period_header and clean_period_date:
        period_cols = _contiguous_date_columns(ws, row=2, start_col=2, max_col=max_col, header_row=1)
        if not period_cols:
            raise CIQTemplateContractError(f"{ws.title}: no clean period columns detected")

        calc_types = {c: (_safe_str(ws.cell(1, c).value) or None) for c, _ in period_cols}
        return 3, period_cols, calc_types

    period_row = None
    for r in range(1, min(ws.max_row, 250) + 1):
        if _safe_str(ws.cell(r, 1).value) == "Period Date":
            period_row = r
            break

    if period_row is None:
        raise CIQTemplateContractError(f"{ws.title}: missing 'Period Date' header row")

    period_cols = _contiguous_date_columns(ws, row=period_row, start_col=2, max_col=max_col)
    if not period_cols:
        raise CIQTemplateContractError(f"{ws.title}: no period columns detected")

    calc_types: dict[int, str | None] = {}
    calc_row = period_row + 1
    if _safe_str(ws.cell(calc_row, 1).value) == "Calculation Type":
        data_start_row = calc_row + 1
        for c, _ in period_cols:
            calc_val = _safe_str(ws.cell(calc_row, c).value)
            calc_types[c] = calc_val or None
    else:
        data_start_row = period_row + 1
        for c, _ in period_cols:
            calc_types[c] = None

    return data_start_row, period_cols, calc_types


def _parse_time_series_sheet(
    ws: Worksheet,
    ticker: str,
    source_file: str,
    *,
    formula_ws: Worksheet | None = None,
) -> list[dict[str, Any]]:
    data_start_row, period_cols, calc_types = _period_columns(ws)
    records: list[dict[str, Any]] = []

    section = "Uncategorized"
    max_scan_row = min(ws.max_row, 5000)
    for r in range(data_start_row, max_scan_row + 1):
        label = _safe_str(ws.cell(r, 1).value)
        if not label:
            continue
        if label.startswith("Source:") or label.startswith("Copyright"):
            continue

        values = [ws.cell(r, c).value for c, _ in period_cols]
        formula_values = (
            [formula_ws.cell(r, c).value for c, _ in period_cols]
            if formula_ws is not None
            else []
        )
        has_values = any(v not in (None, "") for v in values) or any(
            isinstance(v, str) and v.startswith("=") for v in formula_values
        )

        if not has_values:
            section = label
            continue

        unit, scale = _default_unit_scale(ws.title, section)
        row_metric_key = _metric_key(label)
        for c, period_date in period_cols:
            raw_value = ws.cell(r, c).value
            provenance = _cell_provenance(ws, formula_ws, r, c)
            if raw_value in (None, "") and not provenance["has_formula"]:
                continue
            value_num = _to_num(raw_value)
            records.append(
                {
                    "ticker": ticker,
                    "sheet_name": ws.title,
                    "section_name": section,
                    "row_label": label,
                    "metric_key": row_metric_key,
                    "period_date": period_date,
                    "calc_type": calc_types.get(c),
                    "column_label": period_date,
                    "column_index": c,
                    "value_raw": "" if raw_value is None else str(raw_value),
                    "value_num": value_num,
                    "unit": unit,
                    "scale_factor": scale,
                    "source_file": source_file,
                    **provenance,
                }
            )

    return records


def _find_header_row(
    ws: Worksheet,
    header: str = "Ticker",
    max_rows: int = 200,
    max_cols: int = 120,
) -> tuple[int, int] | None:
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            if _safe_str(ws.cell(r, c).value).lower() == header.lower():
                return r, c
    return None


def _header_columns(
    ws: Worksheet,
    header_row: int,
    start_col: int,
    *,
    contiguous: bool,
) -> dict[int, str]:
    headers: dict[int, str] = {}
    max_col = min(ws.max_column, 250)

    for c in range(start_col, max_col + 1):
        value = _safe_str(ws.cell(header_row, c).value)
        if not value:
            if contiguous and headers:
                break
            continue
        headers[c] = value

    return headers


def _parse_comps_sheet(
    ws: Worksheet,
    target_ticker: str,
    source_file: str,
    *,
    contiguous_headers: bool,
    formula_ws: Worksheet | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    comps_rows: list[dict[str, Any]] = []

    header_pos = _find_header_row(ws, "Ticker")
    if header_pos is None:
        raise CIQTemplateContractError(f"{ws.title}: missing 'Ticker' header row")

    header_row, ticker_col = header_pos
    headers = _header_columns(ws, header_row, ticker_col, contiguous=contiguous_headers)
    headers.setdefault(ticker_col, "Ticker")
    if len(headers) <= 1 and contiguous_headers:
        headers = _header_columns(ws, header_row, ticker_col, contiguous=False)
        headers.setdefault(ticker_col, "Ticker")
    if len(headers) <= 1:
        raise CIQTemplateContractError(f"{ws.title}: no comps metric columns detected")

    metric_keys: dict[int, str | None] = {}
    metric_seen: dict[str, int] = {}
    for c, metric_label in headers.items():
        if c == ticker_col:
            continue
        base_key = _metric_key(metric_label)
        if base_key is None:
            metric_keys[c] = None
            continue
        count = metric_seen.get(base_key, 0) + 1
        metric_seen[base_key] = count
        metric_keys[c] = base_key if count == 1 else f"{base_key}__{count}"

    section = "Comps"
    name_col = next((c for c, h in headers.items() if h.lower() == "name"), None)

    max_scan_row = min(ws.max_row, 1200)
    for r in range(header_row + 1, max_scan_row + 1):
        raw_row_ticker = _safe_str(ws.cell(r, ticker_col).value)
        if not raw_row_ticker:
            continue
        if raw_row_ticker.startswith("Source:") or raw_row_ticker.startswith("Copyright"):
            continue

        row_ticker = _normalize_ticker(raw_row_ticker)
        row_name = _safe_str(ws.cell(r, name_col).value) if name_col is not None else ""

        for c, metric_label in headers.items():
            if c == ticker_col:
                continue
            metric_key = metric_keys.get(c)
            if metric_key is None:
                continue

            raw_value = ws.cell(r, c).value
            provenance = _cell_provenance(ws, formula_ws, r, c)
            if raw_value in (None, "") and not provenance["has_formula"]:
                continue

            value_num = _to_num(raw_value)

            records.append(
                {
                    "ticker": target_ticker,
                    "sheet_name": ws.title,
                    "section_name": section,
                    "row_label": raw_row_ticker,
                    "metric_key": metric_key,
                    "period_date": None,
                    "calc_type": None,
                    "column_label": metric_label,
                    "column_index": c,
                    "value_raw": "" if raw_value is None else str(raw_value),
                    "value_num": value_num,
                    "unit": None,
                    "scale_factor": 1.0,
                    "source_file": source_file,
                    **provenance,
                }
            )

            if row_ticker and value_num is not None:
                comps_rows.append(
                    {
                        "target_ticker": target_ticker,
                        "peer_ticker": row_ticker,
                        "source_sheet": ws.title,
                        "peer_name": row_name,
                        "section_name": section,
                        "metric_key": metric_key,
                        "metric_label": metric_label,
                        "value_raw": str(raw_value),
                        "value_num": value_num,
                        "unit": None,
                        "is_target": 1 if row_ticker == target_ticker else 0,
                        "source_file": source_file,
                    }
                )

    return records, comps_rows


def _series(
    records: list[dict[str, Any]],
    keys: set[str],
    *,
    sheet_names: set[str] | None = None,
    exclude_calc_types: set[str] | None = None,
) -> list[tuple[str, float]]:
    out_with_cols: list[tuple[str, int, float]] = []
    excluded = {item.upper() for item in (exclude_calc_types or set())}
    for row in records:
        mk = row.get("metric_key")
        pd = row.get("period_date")
        val = row.get("value_num")
        if sheet_names is not None and row.get("sheet_name") not in sheet_names:
            continue
        calc_type = _safe_str(row.get("calc_type")).upper()
        if calc_type in excluded:
            continue
        if mk in keys and pd and val is not None:
            out_with_cols.append((pd, int(row.get("column_index") or 0), float(val)))

    out_with_cols.sort(key=lambda x: (x[0], x[1]), reverse=True)
    deduped: list[tuple[str, float]] = []
    seen_periods: set[str] = set()
    for period_date, _column_index, value in out_with_cols:
        if period_date in seen_periods:
            continue
        seen_periods.add(period_date)
        deduped.append((period_date, value))
    return deduped


def _preferred_series(
    records: list[dict[str, Any]],
    keys_by_priority: tuple[str, ...],
    *,
    sheet_names: set[str] | None = None,
    exclude_calc_types: set[str] | None = None,
) -> list[tuple[str, float]]:
    """Return the first available, unambiguous metric series by source priority."""
    for key in keys_by_priority:
        values = _series(
            records,
            {key},
            sheet_names=sheet_names,
            exclude_calc_types=exclude_calc_types,
        )
        if values:
            return values
    return []


def _latest_value(
    records: list[dict[str, Any]],
    keys: set[str],
    *,
    sheet_names: set[str] | None = None,
    exclude_calc_types: set[str] | None = None,
) -> float | None:
    series = _series(records, keys, sheet_names=sheet_names, exclude_calc_types=exclude_calc_types)
    return series[0][1] if series else None


def _avg(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def _build_valuation_snapshot(
    ticker: str,
    source_file: str,
    long_records: list[dict[str, Any]],
) -> dict[str, Any]:
    fs_sheets = {"Financial Statements"}
    annual_exclusions = {"LTM", "NTM"}

    revenue = _series(long_records, {"revenue"}, sheet_names=fs_sheets)
    operating_income = _series(long_records, {"operating_income"}, sheet_names=fs_sheets)
    capex = _series(long_records, {"capex"}, sheet_names=fs_sheets)
    da = _preferred_series(
        long_records,
        ("da_total", "da_for_ebitda", "da"),
        sheet_names=fs_sheets,
    )

    annual_revenue = _series(
        long_records,
        {"revenue"},
        sheet_names=fs_sheets,
        exclude_calc_types=annual_exclusions,
    )
    annual_operating_income = _series(
        long_records,
        {"operating_income"},
        sheet_names=fs_sheets,
        exclude_calc_types=annual_exclusions,
    )
    annual_capex = _series(long_records, {"capex"}, sheet_names=fs_sheets, exclude_calc_types=annual_exclusions)
    annual_da = _preferred_series(
        long_records,
        ("da_total", "da_for_ebitda", "da"),
        sheet_names=fs_sheets,
        exclude_calc_types=annual_exclusions,
    )
    annual_tax_exp = _series(
        long_records,
        {"tax", "income_tax_expense"},
        sheet_names=fs_sheets,
        exclude_calc_types=annual_exclusions,
    )
    annual_ebt = _series(
        long_records,
        {"ebt_excl_unusual"},
        sheet_names=fs_sheets,
        exclude_calc_types=annual_exclusions,
    )

    revenue_vals = [v for _, v in revenue[:4]]
    op_vals = [v for _, v in operating_income[:4]]
    capex_vals = [abs(v) for _, v in capex[:4]]
    da_vals = [abs(v) for _, v in da[:4]]

    annual_revenue_by_period = dict(annual_revenue)
    annual_op_by_period = dict(annual_operating_income)
    annual_capex_by_period = {period: abs(value) for period, value in annual_capex}
    annual_da_by_period = {period: abs(value) for period, value in annual_da}
    annual_tax_by_period = {period: abs(value) for period, value in annual_tax_exp}
    annual_ebt_by_period = dict(annual_ebt)

    latest_margin = None
    if revenue_vals and op_vals and revenue_vals[0] not in (0, None):
        latest_margin = op_vals[0] / revenue_vals[0]

    annual_periods = [period for period, revenue_value in annual_revenue if revenue_value not in (0, None)]
    op_margins = [
        annual_op_by_period[period] / annual_revenue_by_period[period]
        for period in annual_periods
        if period in annual_op_by_period
    ]
    capex_pct = [
        annual_capex_by_period[period] / annual_revenue_by_period[period]
        for period in annual_periods
        if period in annual_capex_by_period
    ]
    da_pct = [
        annual_da_by_period[period] / annual_revenue_by_period[period]
        for period in annual_periods
        if period in annual_da_by_period
    ]

    tax_rates = []
    for period, tax_value in annual_tax_by_period.items():
        pretax_value = annual_ebt_by_period.get(period)
        if pretax_value is not None and pretax_value > 0:
            tax_rates.append(tax_value / pretax_value)

    revenue_cagr = None
    annual_revenue_vals = [v for _, v in annual_revenue[:4]]
    if len(annual_revenue_vals) >= 2 and annual_revenue_vals[-1] > 0:
        years = len(annual_revenue_vals) - 1
        revenue_cagr = (annual_revenue_vals[0] / annual_revenue_vals[-1]) ** (1 / years) - 1

    roic = _percent_to_decimal(_latest_value(long_records, {"roic"}, sheet_names=fs_sheets))
    fcf_yield = _percent_to_decimal(_latest_value(long_records, {"fcf_yield"}, sheet_names=fs_sheets))
    debt_to_ebitda = _latest_value(long_records, {"debt_to_ebitda"}, sheet_names=fs_sheets)

    as_of_date = revenue[0][0] if revenue else None

    return {
        "ticker": ticker,
        "as_of_date": as_of_date,
        "source_file": source_file,
        "revenue_mm": revenue_vals[0] if revenue_vals else None,
        "operating_income_mm": op_vals[0] if op_vals else None,
        "capex_mm": capex_vals[0] if capex_vals else None,
        "da_mm": da_vals[0] if da_vals else None,
        "total_debt_mm": _latest_value(long_records, {"debt", "total_debt"}, sheet_names=fs_sheets),
        "cash_mm": _latest_value(long_records, {"cash", "cash_and_equivalents"}, sheet_names=fs_sheets),
        "shares_out_mm": _latest_value(long_records, {"shares", "shares_diluted"}, sheet_names=fs_sheets),
        "ebit_margin": latest_margin,
        "op_margin_avg_3yr": _avg(op_margins[:3]),
        "capex_pct_avg_3yr": _avg(capex_pct[:3]),
        "da_pct_avg_3yr": _avg(da_pct[:3]),
        "effective_tax_rate": _avg(tax_rates),
        "effective_tax_rate_avg": _avg(tax_rates),
        "revenue_cagr_3yr": revenue_cagr,
        "roic": roic,
        "fcf_yield": fcf_yield,
        "debt_to_ebitda": debt_to_ebitda,
    }


def _workbook_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _validate_contract(wb) -> dict[str, Any]:
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]
    if missing:
        raise CIQTemplateContractError(f"Missing required sheets: {missing}")

    found_anchors: dict[str, list[dict[str, Any]]] = {}

    fs = wb["Financial Statements"]
    fs_clean = _safe_str(fs.cell(1, 1).value).lower() == "period" and _safe_str(fs.cell(2, 1).value).lower().startswith(
        "period date"
    )
    if fs_clean:
        start_row, period_cols, _ = _period_columns(fs)
        found_anchors["Financial Statements"] = [
            {
                "layout": "clean",
                "data_start_row": start_row,
                "first_period_col": period_cols[0][0],
                "period_count": len(period_cols),
            }
        ]
    else:
        legacy_checks = [(1, "Select Language"), (2, "Ticker"), (9, "INCOME STATEMENT")]
        sheet_found: list[dict[str, Any]] = []
        for row, needle in legacy_checks:
            match = _row_contains(fs, row, needle)
            if match is None:
                raise CIQTemplateContractError(
                    f"Financial Statements!row {row} expected to contain '{needle}' in at least one column"
                )
            col, value = match
            sheet_found.append({"row": row, "col": col, "value": value})
        found_anchors["Financial Statements"] = sheet_found

    cs = wb["Common Size"]
    cs_start_row, cs_period_cols, _ = _period_columns(cs)
    found_anchors["Common Size"] = [
        {
            "data_start_row": cs_start_row,
            "first_period_col": cs_period_cols[0][0],
            "period_count": len(cs_period_cols),
        }
    ]

    detailed = wb["Detailed Comps"]
    detailed_header = _find_header_row(detailed, "Ticker")
    if detailed_header is None:
        raise CIQTemplateContractError("Detailed Comps: missing 'Ticker' header row")

    header_row, ticker_col = detailed_header
    detailed_headers = _header_columns(detailed, header_row, ticker_col, contiguous=True)
    header_mode = "contiguous"
    if len(detailed_headers) <= 1:
        detailed_headers = _header_columns(detailed, header_row, ticker_col, contiguous=False)
        header_mode = "sparse_legacy"
    if len(detailed_headers) <= 1:
        raise CIQTemplateContractError("Detailed Comps: missing metric headers after 'Ticker'")

    found_anchors["Detailed Comps"] = [
        {
            "header_row": header_row,
            "ticker_col": ticker_col,
            "header_count": len(detailed_headers),
            "header_mode": header_mode,
        }
    ]

    if "Summary Comps" in wb.sheetnames:
        summary = wb["Summary Comps"]
        summary_header = _find_header_row(summary, "Ticker")
        if summary_header is not None:
            found_anchors["Summary Comps"] = [
                {
                    "header_row": summary_header[0],
                    "ticker_col": summary_header[1],
                }
            ]

    fp = {
        "sheets": wb.sheetnames,
        "anchors": found_anchors,
    }
    return fp


def _resolve_ticker(wb) -> str:
    if "Input" in wb.sheetnames:
        input_ticker = _normalize_ticker(_safe_str(wb["Input"].cell(2, 2).value))
        if input_ticker:
            return input_ticker

    fs_ticker = _normalize_ticker(_safe_str(wb["Financial Statements"].cell(2, 3).value))
    if fs_ticker:
        return fs_ticker

    raise CIQTemplateContractError("Could not resolve ticker from Input!B2 or Financial Statements!C2")


def parse_ciq_workbook(path: str | Path) -> CIQWorkbookPayload:
    """Parse a CIQ workbook into long-form records + deterministic snapshots."""
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    file_hash = _workbook_hash(workbook_path)
    cached_wb = load_workbook(workbook_path, data_only=True, read_only=False)
    formula_wb = None
    try:
        formula_wb = load_workbook(workbook_path, data_only=False, read_only=False)
        fingerprint = _validate_contract(cached_wb)
        ticker = _resolve_ticker(cached_wb)

        long_records: list[dict[str, Any]] = []
        long_records.extend(
            _parse_time_series_sheet(
                cached_wb["Financial Statements"],
                ticker,
                workbook_path.name,
                formula_ws=formula_wb["Financial Statements"],
            )
        )
        long_records.extend(
            _parse_time_series_sheet(
                cached_wb["Common Size"],
                ticker,
                workbook_path.name,
                formula_ws=formula_wb["Common Size"],
            )
        )

        detailed_records, detailed_comps = _parse_comps_sheet(
            cached_wb["Detailed Comps"],
            ticker,
            workbook_path.name,
            contiguous_headers=True,
            formula_ws=formula_wb["Detailed Comps"],
        )
        long_records.extend(detailed_records)

        summary_comps: list[dict[str, Any]] = []
        if "Summary Comps" in cached_wb.sheetnames:
            try:
                summary_records, summary_comps = _parse_comps_sheet(
                    cached_wb["Summary Comps"],
                    ticker,
                    workbook_path.name,
                    contiguous_headers=False,
                    formula_ws=formula_wb["Summary Comps"],
                )
                long_records.extend(summary_records)
            except CIQTemplateContractError:
                # Summary Comps is optional in newer templates.
                pass

        valuation_snapshot = _build_valuation_snapshot(ticker, workbook_path.name, long_records)
        comps_snapshot = detailed_comps + summary_comps

        return CIQWorkbookPayload(
            ticker=ticker,
            source_file=workbook_path.name,
            file_hash=file_hash,
            parser_version=PARSER_VERSION,
            template_fingerprint=json.dumps(fingerprint, sort_keys=True),
            long_form_records=long_records,
            valuation_snapshot=valuation_snapshot,
            comps_snapshot=comps_snapshot,
            rows_parsed=len(long_records),
        )
    finally:
        cached_wb.close()
        if formula_wb is not None:
            formula_wb.close()
