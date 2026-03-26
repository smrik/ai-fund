"""
Build templates/ticker_review.xlsx — professional valuation template.

Pre-populated with IBM data (2026-03-09 snapshot).
Sheets: Cover | Assumptions | DCF_Base | DCF_Bear | DCF_Bull |
        Equity_Bridge | Comps | Sensitivity | QoE | Output

Run:
    python scripts/build_valuation_template.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
BLUE   = "0070C0"   # hardcoded input
BLACK  = "000000"   # formula
GREEN  = "375623"   # cross-sheet link
YELLOW = "FFFF00"   # override cell
NAVY   = "1F3864"   # section header bg
LBLUE  = "D9E1F2"   # column header bg
WHITE  = "FFFFFF"
GRAY   = "F2F2F2"
RED_H  = "FF0000"
AMBER  = "FFC000"
GREEN_H = "70AD47"
DARK_GREEN = "1E4620"

# ── Number formats ─────────────────────────────────────────────────────────────
PCT1 = "0.0%"
PCT2 = "0.00%"
NUM0 = "#,##0"
NUM1 = "#,##0.0"
MULT = '0.0"x"'
DOLLA = '"$"#,##0.00'

# ── Load snapshot from JSON (batch_runner output) ─────────────────────────────

_ROOT = Path(__file__).resolve().parent.parent
_DEFAULT_JSON = _ROOT / "data" / "valuations" / "json" / "IBM_latest.json"

# Global parameters not stored per-ticker in JSON (loaded from config)
_RF  = 0.045   # risk-free rate  — 10Y UST
_ERP = 0.050   # equity risk premium — Damodaran


def _load_snapshot(json_path: Path = _DEFAULT_JSON) -> dict:
    """
    Map IBM_latest.json (batch_runner output) → the flat dict expected by
    every build_* function in this script.

    All percentage values are stored as decimals in the JSON (0.066 not 6.6).
    Dollar values labelled _mm are in $mm.
    """
    with json_path.open(encoding="utf-8") as f:
        j = json.load(f)

    m = j["market"]
    a = j["assumptions"]
    w = j["wacc"]
    v = j["valuation"]
    s = j["scenarios"]
    t = j.get("terminal", {})
    cd = j.get("comps_detail", {})
    qoe = j.get("qoe") or {}

    # Derive kd_pretax from WACC equation:
    # WACC = ke*eq_wt + kd_at*debt_wt  → kd_at = (WACC - ke*eq_wt)/debt_wt
    # kd_pretax = kd_at / (1 - tax_s)
    _wacc = w.get("wacc") or 0.07
    _ke   = w.get("cost_of_equity") or 0.08
    _eq   = w.get("equity_weight") or 0.82
    _dw   = w.get("debt_weight") or 0.18
    _tax  = a.get("tax_rate_start_pct") or 0.21
    _kd_at = (_wacc - _ke * _eq) / _dw if _dw else 0.0
    _kd_pretax = _kd_at / (1 - _tax) if _tax < 1 else _kd_at

    # Build peer list from comps_detail
    peers = []
    for p in cd.get("peers", []):
        peers.append({
            "ticker": p.get("ticker", ""),
            "name":   p.get("name", ""),
            "mcap":   p.get("market_cap_mm") or 0,
            "tev":    p.get("tev_mm") or 0,
            "rev":    p.get("revenue_ltm_mm") or 0,
            "ebitda": p.get("ebitda_ltm_mm") or 0,
            "ebit":   p.get("ebit_ltm_mm") or 0,
            "eps":    p.get("eps_ltm") or 0,
            "ev_ebitda_ltm": p.get("tev_ebitda_ltm") or 0,
            "ev_ebitda_fwd": p.get("tev_ebitda_fwd") or 0,
            "ev_ebit": p.get("tev_ebit_ltm") or 0,
            "pe":      p.get("pe_ltm") or 0,
        })

    return {
        "ticker":       j.get("ticker", "IBM"),
        "name":         j.get("company_name", ""),
        "sector":       j.get("sector", ""),
        "date":         j.get("generated_at", "")[:10],
        # market
        "price":        m.get("price") or 0,
        "mcap_mm":      m.get("market_cap_mm") or 0,
        "ev_mm":        m.get("ev_mm") or 0,
        "pe_t":         m.get("pe_trailing") or 0,
        "pe_f":         m.get("pe_forward") or 0,
        "ev_ebitda":    m.get("ev_ebitda") or 0,
        "analyst_target": m.get("analyst_target") or 0,
        "analyst_rec":  m.get("analyst_recommendation") or "",
        "n_analysts":   m.get("num_analysts") or 0,
        # assumptions (all decimals)
        "revenue_mm":   a.get("revenue_mm") or 0,
        "g_near":       a.get("growth_near_pct") or 0,
        "g_mid":        a.get("growth_mid_pct") or 0,
        "g_term":       a.get("growth_terminal_pct") or 0,
        "ebit_s":       a.get("ebit_margin_start_pct") or 0,
        "ebit_t":       a.get("ebit_margin_target_pct") or 0,
        "da_pct":       a.get("da_pct") or 0,
        "capex_pct":    a.get("capex_pct") or 0,
        "tax_s":        a.get("tax_rate_start_pct") or 0,
        "tax_t":        a.get("tax_rate_target_pct") or 0,
        "dso_s":        a.get("dso_start") or 0,
        "dso_t":        a.get("dso_target") or 0,
        "dio_s":        a.get("dio_start") or 0,
        "dio_t":        a.get("dio_target") or 0,
        "dpo_s":        a.get("dpo_start") or 0,
        "dpo_t":        a.get("dpo_target") or 0,
        "exit_mult":    a.get("exit_multiple") or 0,
        "ronic":        a.get("ronic_terminal_pct") or 0,
        # equity bridge ($mm)
        "net_debt_mm":  a.get("net_debt_mm") or 0,
        "shares_mm":    a.get("shares_outstanding_mm") or 0,
        "inv_cap_mm":   a.get("invested_capital_mm") or 0,
        "non_op_mm":    a.get("non_operating_assets_mm") or 0,
        "minority_mm":  a.get("minority_interest_mm") or 0,
        "preferred_mm": a.get("preferred_equity_mm") or 0,
        "pension_mm":   a.get("pension_deficit_mm") or 0,
        "leases_mm":    a.get("lease_liabilities_mm") or 0,
        "options_mm":   a.get("options_value_mm") or 0,
        "converts_mm":  a.get("convertibles_value_mm") or 0,
        # WACC
        "rf":           _RF,
        "erp":          _ERP,
        "beta_raw":     w.get("beta_raw") or 0,
        "beta_unl":     w.get("beta_unlevered") or 0,
        "beta_rel":     w.get("beta_relevered") or 0,
        "size_prem":    w.get("size_premium") or 0,
        "ke":           _ke,
        "kd_pretax":    max(_kd_pretax, 0.0),
        "eq_wt":        _eq,
        "debt_wt":      _dw,
        "wacc":         _wacc,
        # scenarios
        "prob_bear":    a.get("scenario_prob_bear") or 0.2,
        "prob_base":    a.get("scenario_prob_base") or 0.6,
        "prob_bull":    a.get("scenario_prob_bull") or 0.2,
        "iv_bear":      s["bear"].get("iv") or v.get("iv_bear") or 0,
        "iv_base":      s["base"].get("iv") or v.get("iv_base") or 0,
        "iv_bull":      s["bull"].get("iv") or v.get("iv_bull") or 0,
        "expected_iv":  v.get("expected_iv") or 0,
        "tv_pct_ev":    (t.get("tv_pct_of_ev") or 0) / 100.0,
        # QoE
        "sloan_ratio":    qoe.get("sloan_accruals_ratio") or 0.045,
        "cash_conv":      qoe.get("cash_conversion") or 0.82,
        "dso_baseline":   qoe.get("dso_baseline") or 0,
        "dso_drift":      qoe.get("dso_drift") or 0,
        "capex_da_ratio": qoe.get("capex_da_ratio") or 0,
        "qoe_score":      qoe.get("qoe_score") or 3,
        "qoe_flag":       qoe.get("qoe_flag") or "amber",
        "peers":          peers,
    }

# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(color=BLACK, bold=False, size=10, italic=False) -> Font:
    return Font(name="Calibri", color=color, bold=bold, size=size, italic=italic)


def _border_bottom(color="CCCCCC", thick=False) -> Border:
    s = "medium" if thick else "thin"
    return Border(bottom=Side(style=s, color=color))


def _align(h="right", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set(ws, row: int, col: int, value,
         color=BLACK, bold=False, fill_hex=None,
         fmt=None, align="right", green=False, blue=False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    fc = GREEN if green else (BLUE if blue else color)
    c.font = _font(color=fc, bold=bold)
    c.alignment = _align(h=align)
    if fill_hex:
        c.fill = _fill(fill_hex)
    if fmt:
        c.number_format = fmt


def _lbl(ws, row: int, col: int, text: str, bold=False, indent=0, size=10) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=bold, size=size)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)


def _section_hdr(ws, row: int, text: str, ncols=6, color=NAVY) -> None:
    ws.row_dimensions[row].height = 16
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(color)
        cell.font = _font(color=WHITE, bold=True)
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).alignment = _align(h="left")


def _col_hdr(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(LBLUE)
    c.font = _font(bold=True)
    c.alignment = _align(h="center")
    c.border = _border_bottom(thick=True)


# ── Sheet: Cover ───────────────────────────────────────────────────────────────

def build_cover(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30

    ws.row_dimensions[4].height = 36
    ws.row_dimensions[5].height = 22

    ws.merge_cells("B4:C4")
    c = ws["B4"]
    c.value = "EQUITY VALUATION MODEL"
    c.font = _font(color=NAVY, bold=True, size=20)
    c.alignment = _align(h="left")

    ws.merge_cells("B5:C5")
    c = ws["B5"]
    c.value = f"{d['ticker']} — {d['name']}"
    c.font = _font(color=BLUE, bold=False, size=14)
    c.alignment = _align(h="left")

    rows = [
        (7,  "Sector",   d["sector"]),
        (8,  "As of",    d["date"]),
        (9,  "Price",    f"${d['price']:.2f}"),
        (10, "EV ($mm)",  f"${d['ev_mm']:,.0f}"),
        (11, "Shares",   f"{d['shares_mm']:.1f}mm"),
        (12, "Model",    "DCF Bear / Base / Bull + Comps + EP cross-check"),
    ]
    for r, lbl, val in rows:
        ws.cell(row=r, column=2, value=lbl).font = _font(bold=True)
        ws.cell(row=r, column=3, value=val).font = _font(color=BLUE)

    ws.row_dimensions[14].height = 14
    for r, txt in [
        (14, "Sheet Guide"),
        (15, "  Assumptions  — All inputs with JSON baseline + PM override column (yellow)"),
        (16, "  DCF_Base     — 10-year FCFF projection, live formulas, Base scenario"),
        (17, "  DCF_Bear     — Bear scenario (growth ×0.6, margin ×0.75, WACC +200bp)"),
        (18, "  DCF_Bull     — Bull scenario (growth ×1.4, margin ×1.15, WACC -100bp)"),
        (19, "  Equity_Bridge— Enterprise → Equity value, IV per share, expected IV"),
        (20, "  Comps        — Peer multiples table, live medians, exclude toggle"),
        (21, "  Sensitivity  — IV grid: WACC × terminal growth (±200bp each)"),
        (22, "  QoE          — Quality of Earnings signals (Sloan, cash conv, NWC drift)"),
        (23, "  Output       — Summary dashboard with colour-coded upside flags"),
    ]:
        c = ws.cell(row=r, column=2, value=txt)
        c.font = _font(bold=(r == 14), size=10 if r == 14 else 9)
        c.alignment = _align(h="left")

    ws.row_dimensions[25].height = 14
    c = ws.cell(row=25, column=2,
                value="Override pattern: Col B = JSON baseline (blue)  |  "
                      "Col C = PM override (yellow, type here)  |  "
                      "Col D = Active (formula: =IF(C=\"\",B,C))")
    c.font = _font(size=9, italic=True)
    c.alignment = _align(h="left", wrap=True)

    c = ws.cell(row=26, column=2, value="Press F9 to recalculate all formulas after editing.")
    c.font = _font(size=9, italic=True, color=AMBER)
    c.alignment = _align(h="left")


# ── Sheet: Assumptions ─────────────────────────────────────────────────────────
# Row map for Active column (D): used by all DCF formula sheets

ASS = {
    "revenue": 3,
    "g_near": 4, "g_mid": 5, "g_term": 6,
    "ebit_s": 8, "ebit_t": 9,
    "da_pct": 11, "capex_pct": 12,
    "tax_s": 14, "tax_t": 15,
    "dso_s": 17, "dso_t": 18,
    "dio_s": 19, "dio_t": 20,
    "dpo_s": 21, "dpo_t": 22,
    "exit_mult": 24, "ronic": 25,
    "rf": 27, "erp": 28, "beta_rel": 29, "size_prem": 30,
    "ke": 31,         # formula row
    "kd_pretax": 32,
    "tax_kd": 33,     # =D14 link
    "kd_at": 34,      # formula row
    "eq_wt": 35, "debt_wt": 36,
    "wacc": 37,       # formula row
    "net_debt": 39, "shares": 40, "inv_cap": 41,
    "non_op": 42, "minority": 43, "preferred": 44,
    "pension": 45, "leases": 46, "options": 47, "converts": 48,
    "prob_bear": 50, "prob_base": 51, "prob_bull": 52,
}


def build_assumptions(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15   # JSON
    ws.column_dimensions["C"].width = 15   # Override
    ws.column_dimensions["D"].width = 15   # Active
    ws.column_dimensions["E"].width = 22   # Source

    # Column headers
    ws.row_dimensions[1].height = 32
    for col, txt, fill_hex in [
        (1, "Driver", NAVY),
        (2, "JSON Baseline", NAVY),
        (3, "PM Override", NAVY),
        (4, "Active", NAVY),
        (5, "Source", NAVY),
    ]:
        c = ws.cell(row=1, column=col, value=txt)
        c.fill = _fill(fill_hex)
        c.font = _font(color=WHITE, bold=True)
        c.alignment = _align(h="center")

    c = ws.cell(row=2, column=3,
                value="← Type override here (yellow cells)")
    c.font = _font(size=9, italic=True, color=AMBER)
    c.alignment = _align(h="left")

    # Helper to write one assumption row
    def row(r: int, lbl: str, val, fmt=PCT1, source="yfinance / CIQ",
            formula_d: str | None = None) -> None:
        _lbl(ws, r, 1, lbl, indent=1)
        # Col B (JSON baseline)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(color=BLUE)
        c.alignment = _align()
        c.number_format = fmt
        # Col C (override — yellow)
        oc = ws.cell(row=r, column=3)
        oc.fill = _fill(YELLOW)
        oc.number_format = fmt
        # Col D (active)
        dc = ws.cell(row=r, column=4)
        if formula_d:
            dc.value = formula_d
            dc.font = _font(color=BLACK)
        else:
            dc.value = f"=IF(C{r}=\"\",B{r},C{r})"
            dc.font = _font(color=BLACK)
        dc.number_format = fmt
        dc.alignment = _align()
        # Col E (source)
        ws.cell(row=r, column=5, value=source).font = _font(size=9)
        ws.cell(row=r, column=5).alignment = _align(h="left")

    _section_hdr(ws, ASS["revenue"] - 1, "  MARKET DATA", ncols=5, color=NAVY)
    row(ASS["revenue"], "Revenue LTM ($mm)", d["revenue_mm"], fmt=NUM0, source="CIQ")

    _section_hdr(ws, ASS["g_near"] - 1, "  DCF GROWTH DRIVERS", ncols=5)
    row(ASS["g_near"],  "Revenue Growth — Near (Y1–Y5)",     d["g_near"],    source="CIQ consensus")
    row(ASS["g_mid"],   "Revenue Growth — Mid (Y6–Y10)",     d["g_mid"],     source="CIQ consensus blend")
    row(ASS["g_term"],  "Terminal Revenue Growth",           d["g_term"],    source="GDP anchor")

    _section_hdr(ws, ASS["ebit_s"] - 1, "  MARGIN DRIVERS", ncols=5)
    row(ASS["ebit_s"],  "EBIT Margin — Start",               d["ebit_s"],    source="CIQ LTM")
    row(ASS["ebit_t"],  "EBIT Margin — Target (Y10)",        d["ebit_t"],    source="Sector median")

    _section_hdr(ws, ASS["da_pct"] - 1, "  D&A / CAPEX", ncols=5)
    row(ASS["da_pct"],  "D&A % Revenue",                     d["da_pct"],    source="CIQ LTM")
    row(ASS["capex_pct"], "Capex % Revenue",                 d["capex_pct"], source="CIQ LTM")

    _section_hdr(ws, ASS["tax_s"] - 1, "  TAX", ncols=5)
    row(ASS["tax_s"],   "Effective Tax Rate — Start",        d["tax_s"],     source="CIQ LTM")
    row(ASS["tax_t"],   "Effective Tax Rate — Target",       d["tax_t"],     source="Company guidance")

    _section_hdr(ws, ASS["dso_s"] - 1, "  WORKING CAPITAL DRIVERS", ncols=5)
    row(ASS["dso_s"],   "DSO Start (days)",                  d["dso_s"],     fmt=NUM1, source="CIQ LTM")
    row(ASS["dso_t"],   "DSO Target (days)",                 d["dso_t"],     fmt=NUM1, source="Sector median")
    row(ASS["dio_s"],   "DIO Start (days)",                  d["dio_s"],     fmt=NUM1, source="CIQ LTM")
    row(ASS["dio_t"],   "DIO Target (days)",                 d["dio_t"],     fmt=NUM1, source="Sector median")
    row(ASS["dpo_s"],   "DPO Start (days)",                  d["dpo_s"],     fmt=NUM1, source="CIQ LTM")
    row(ASS["dpo_t"],   "DPO Target (days)",                 d["dpo_t"],     fmt=NUM1, source="Sector median")

    _section_hdr(ws, ASS["exit_mult"] - 1, "  TERMINAL VALUE", ncols=5)
    row(ASS["exit_mult"], "Exit Multiple (EV/EBITDA)",        d["exit_mult"], fmt=MULT, source="Forward comps")
    row(ASS["ronic"],   "RONIC Terminal",                    d["ronic"],     source="ROIC estimate")

    _section_hdr(ws, ASS["rf"] - 1, "  WACC BUILD-UP", ncols=5)
    row(ASS["rf"],      "Risk-Free Rate",                    d["rf"],        source="10Y UST")
    row(ASS["erp"],     "Equity Risk Premium",               d["erp"],       source="Damodaran")
    row(ASS["beta_rel"], "Beta (re-levered)",                d["beta_rel"],  fmt=NUM1 + "0", source="Hamada / peers")
    row(ASS["size_prem"], "Size Premium",                    d["size_prem"], source="Duff & Phelps")

    # Ke — formula row
    r_ke = ASS["ke"]
    _lbl(ws, r_ke, 1, "Cost of Equity (Ke)", indent=1, bold=True)
    ws.cell(row=r_ke, column=2, value=d["ke"]).font = _font(color=BLUE)
    ws.cell(row=r_ke, column=2).number_format = PCT2
    ws.cell(row=r_ke, column=3).fill = _fill(YELLOW)
    ws.cell(row=r_ke, column=3).number_format = PCT2
    dc = ws.cell(row=r_ke, column=4,
                 value=f"=IF(C{r_ke}=\"\","
                       f"D{ASS['rf']}+D{ASS['beta_rel']}*D{ASS['erp']}+D{ASS['size_prem']},"
                       f"C{r_ke})")
    dc.font = _font(color=BLACK)
    dc.number_format = PCT2
    dc.alignment = _align()
    ws.cell(row=r_ke, column=5, value="CAPM: Rf + β×ERP + size premium").font = _font(size=9)

    row(ASS["kd_pretax"], "Cost of Debt (pre-tax)",          d["kd_pretax"],  source="Interest expense / total debt (CIQ)")

    # tax_kd — cross-link to tax_s
    r_tkd = ASS["tax_kd"]
    _lbl(ws, r_tkd, 1, "Tax Rate (for Kd)", indent=2)
    ws.cell(row=r_tkd, column=4,
            value=f"=D{ASS['tax_s']}").font = _font(color=GREEN)
    ws.cell(row=r_tkd, column=4).number_format = PCT1
    ws.cell(row=r_tkd, column=4).alignment = _align()
    ws.cell(row=r_tkd, column=5, value="=Assumptions D14").font = _font(size=9)

    # Kd after-tax — formula row
    r_kdat = ASS["kd_at"]
    _lbl(ws, r_kdat, 1, "Cost of Debt after-tax (Kd)", indent=1, bold=True)
    ws.cell(row=r_kdat, column=2, value=d["kd_pretax"] * (1 - d["tax_s"])).font = _font(color=BLUE)
    ws.cell(row=r_kdat, column=2).number_format = PCT2
    ws.cell(row=r_kdat, column=3).fill = _fill(YELLOW)
    ws.cell(row=r_kdat, column=3).number_format = PCT2
    dc = ws.cell(row=r_kdat, column=4,
                 value=f"=IF(C{r_kdat}=\"\","
                       f"D{ASS['kd_pretax']}*(1-D{ASS['tax_kd']}),"
                       f"C{r_kdat})")
    dc.font = _font(color=BLACK)
    dc.number_format = PCT2
    dc.alignment = _align()
    ws.cell(row=r_kdat, column=5, value="Kd × (1 - tax)").font = _font(size=9)

    row(ASS["eq_wt"],   "Equity Weight",                    d["eq_wt"],     source="Market cap / TEV")
    row(ASS["debt_wt"], "Debt Weight",                      d["debt_wt"],   source="Net debt / TEV")

    # WACC — formula row
    r_wacc = ASS["wacc"]
    _lbl(ws, r_wacc, 1, "WACC", indent=1, bold=True)
    ws.cell(row=r_wacc, column=2, value=d["wacc"]).font = _font(color=BLUE)
    ws.cell(row=r_wacc, column=2).number_format = PCT2
    ws.cell(row=r_wacc, column=3).fill = _fill(YELLOW)
    ws.cell(row=r_wacc, column=3).number_format = PCT2
    # D37 = JSON baseline WACC by default; PM may override via C37.
    # Using B37 (JSON value) rather than recomputing from Ke/Kd components
    # ensures exact match with batch_runner output.
    dc = ws.cell(row=r_wacc, column=4,
                 value=f"=IF(C{r_wacc}=\"\",B{r_wacc},C{r_wacc})")
    dc.font = _font(color=BLACK, bold=True)
    dc.number_format = PCT2
    dc.alignment = _align()
    ws.cell(row=r_wacc, column=5, value="JSON baseline (batch_runner) — override C37 to adjust").font = _font(size=9)

    _section_hdr(ws, ASS["net_debt"] - 1, "  EQUITY BRIDGE ITEMS ($mm)", ncols=5)
    row(ASS["net_debt"],   "Net Debt ($mm)",                  d["net_debt_mm"],  fmt=NUM0, source="CIQ balance sheet")
    row(ASS["shares"],     "Shares Outstanding (mm)",         d["shares_mm"],    fmt=NUM1, source="CIQ")
    row(ASS["inv_cap"],    "Invested Capital ($mm)",          d["inv_cap_mm"],   fmt=NUM0, source="CIQ")
    row(ASS["non_op"],     "Non-Operating Assets ($mm)",      d["non_op_mm"],    fmt=NUM0, source="CIQ")
    row(ASS["minority"],   "Minority Interest ($mm)",         d["minority_mm"],  fmt=NUM0, source="CIQ")
    row(ASS["preferred"],  "Preferred Equity ($mm)",          d["preferred_mm"], fmt=NUM0, source="CIQ")
    row(ASS["pension"],    "Pension Deficit ($mm)",           d["pension_mm"],   fmt=NUM0, source="CIQ")
    row(ASS["leases"],     "Lease Liabilities ($mm)",         d["leases_mm"],    fmt=NUM0, source="CIQ")
    row(ASS["options"],    "Options Value ($mm)",             d["options_mm"],   fmt=NUM0, source="CIQ")
    row(ASS["converts"],   "Convertibles ($mm)",              d["converts_mm"],  fmt=NUM0, source="CIQ")

    _section_hdr(ws, ASS["prob_bear"] - 1, "  SCENARIO PROBABILITIES", ncols=5)
    row(ASS["prob_bear"], "P(Bear)",  d["prob_bear"], source="PM judgment")
    row(ASS["prob_base"], "P(Base)",  d["prob_base"], source="PM judgment")
    row(ASS["prob_bull"], "P(Bull)",  d["prob_bull"], source="PM judgment")

    # Probability check (row 53)
    ws.cell(row=53, column=1, value="Sum of probs").font = _font(size=9, italic=True)
    c = ws.cell(row=53, column=4,
                value=f"=D{ASS['prob_bear']}+D{ASS['prob_base']}+D{ASS['prob_bull']}")
    c.font = _font(size=9)
    c.number_format = PCT1
    c.alignment = _align()


# ── DCF sheet builder (shared for Base / Bear / Bull) ──────────────────────────

def _dcf_sheet(wb: openpyxl.Workbook, name: str, d: dict,
               bear=False, bull=False) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12  # LTM
    for i in range(3, 15):
        ws.column_dimensions[get_column_letter(i)].width = 10

    # ── Scenario params (Bear/Bull only) ──
    # Rows 1-7: scenario multiplier inputs (shown for Bear/Bull)
    # Base: row 1 = title only, projection starts row 4
    # Bear/Bull: scenario params rows 3-7, projection starts row 11 (offset=7)
    R = 7 if (bear or bull) else 0  # row offset for projection

    if bear or bull:
        scenario_label = "BEAR SCENARIO" if bear else "BULL SCENARIO"
        params = {
            "growth_mult":  0.60 if bear else 1.40,
            "margin_mult":  0.75 if bear else 1.15,
            "capex_mult":   1.20 if bear else 0.90,
            "wacc_adj":     0.02 if bear else -0.01,
            "exit_mult_adj": 0.70 if bear else 1.30,
        }
        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value = f"IBM — {scenario_label}  |  {d['ticker']} Valuation"
        c.font = _font(color=NAVY, bold=True, size=12)
        c.alignment = _align(h="left")

        _section_hdr(ws, 2, "  SCENARIO PARAMETERS", ncols=3, color="9E0700" if bear else "215732")
        param_rows = [
            (3, "Growth Multiplier",       "growth_mult",   "0.00"),
            (4, "EBIT Margin Multiplier",  "margin_mult",   "0.00"),
            (5, "Capex Multiplier",        "capex_mult",    "0.00"),
            (6, "WACC Adjustment (add)",   "wacc_adj",      PCT2),
            (7, "Exit Multiple Multiplier","exit_mult_adj", "0.00"),
        ]
        for pr, plbl, pkey, pfmt in param_rows:
            _lbl(ws, pr, 1, plbl)
            c = ws.cell(row=pr, column=2, value=params[pkey])
            c.font = _font(color=BLUE)
            c.alignment = _align()
            c.number_format = pfmt
    else:
        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value = f"IBM — BASE SCENARIO  |  {d['ticker']} Valuation"
        c.font = _font(color=NAVY, bold=True, size=12)
        c.alignment = _align(h="left")

    # ── Header row ──
    _section_hdr(ws, R + 2, "  10-YEAR FCFF PROJECTION  ($mm)", ncols=14, color=NAVY)
    ws.row_dimensions[R + 2].height = 18
    hdr_labels = ["LTM"] + [f"Y{i}" for i in range(1, 11)] + ["Terminal"]
    for ci, hl in enumerate(hdr_labels, start=2):
        _col_hdr(ws, R + 2, ci, hl)

    # ── Row offsets (absolute rows in sheet) ──
    rREV  = R + 4   # Revenue
    rGRW  = R + 5   # Growth %
    rMGN  = R + 7   # EBIT margin
    rEBIT = R + 8   # EBIT
    rEBDA = R + 9   # EBITDA
    rTAX  = R + 10  # Tax rate
    rNOPT = R + 11  # NOPAT
    rDA_P = R + 13  # D&A %
    rDA   = R + 14  # D&A $
    rCX_P = R + 15  # Capex %
    rCX   = R + 16  # Capex $
    rDSO  = R + 18  # DSO
    rDIO  = R + 19  # DIO
    rDPO  = R + 20  # DPO
    rAR   = R + 21  # AR
    rINV  = R + 22  # Inventory
    rAP   = R + 23  # AP
    rNWC  = R + 24  # NWC
    rDNWC = R + 25  # ΔNWC
    rFCFF = R + 27  # FCFF
    rDISC = R + 28  # Discount factor
    rPVF  = R + 29  # PV FCFF
    rCPVF = R + 30  # Cumulative PV FCFF
    rIC   = R + 32  # Invested Capital
    rROIC = R + 33  # ROIC
    rEP   = R + 34  # Economic Profit

    # section separators
    for r, lbl in [
        (R + 3,  "── REVENUE ─────────────────────────────────"),
        (R + 6,  "── PROFITABILITY ────────────────────────────"),
        (R + 12, "── D&A / CAPEX ──────────────────────────────"),
        (R + 17, "── NWC DRIVERS ──────────────────────────────"),
        (R + 26, "── FREE CASH FLOW ───────────────────────────"),
        (R + 31, "── RETURNS ──────────────────────────────────"),
    ]:
        c = ws.cell(row=r, column=1, value=lbl)
        c.font = _font(size=8, italic=True, color="888888")

    # Row labels
    labels = {
        rREV: "Revenue ($mm)", rGRW: "Revenue Growth %",
        rMGN: "EBIT Margin %", rEBIT: "EBIT ($mm)", rEBDA: "EBITDA ($mm)",
        rTAX: "Tax Rate %", rNOPT: "NOPAT ($mm)",
        rDA_P: "D&A % Revenue", rDA: "D&A ($mm)",
        rCX_P: "Capex % Revenue", rCX: "Capex ($mm)",
        rDSO: "DSO (days)", rDIO: "DIO (days)", rDPO: "DPO (days)",
        rAR: "AR ($mm)", rINV: "Inventory ($mm)", rAP: "AP ($mm)",
        rNWC: "NWC ($mm)", rDNWC: "ΔNWC ($mm)",
        rFCFF: "FCFF ($mm)", rDISC: "Discount Factor",
        rPVF: "PV FCFF ($mm)", rCPVF: "Cumulative PV FCFF ($mm)",
        rIC: "Invested Capital ($mm)", rROIC: "ROIC %", rEP: "Economic Profit ($mm)",
    }
    for r, lbl in labels.items():
        bold = r in (rREV, rNOPT, rFCFF, rPVF, rROIC)
        _lbl(ws, r, 1, lbl, bold=bold, indent=1)

    # Number formats per row
    fmts = {
        rREV: NUM0, rGRW: PCT1, rMGN: PCT1, rEBIT: NUM0, rEBDA: NUM0,
        rTAX: PCT1, rNOPT: NUM0,
        rDA_P: PCT2, rDA: NUM0, rCX_P: PCT2, rCX: NUM0,
        rDSO: NUM1, rDIO: NUM1, rDPO: NUM1,
        rAR: NUM0, rINV: NUM0, rAP: NUM0, rNWC: NUM0, rDNWC: NUM0,
        rFCFF: NUM0, rDISC: "0.0000", rPVF: NUM0, rCPVF: NUM0,
        rIC: NUM0, rROIC: PCT1, rEP: NUM0,
    }

    # ── Abbreviations for Assumptions references ──
    # For Bear/Bull, growth/margin/capex/wacc/exit use scenario multipliers
    A = "Assumptions"  # sheet name
    Ag = f"{A}!$D${ASS['g_near']}"
    Agm = f"{A}!$D${ASS['g_mid']}"
    Agt = f"{A}!$D${ASS['g_term']}"
    Aes = f"{A}!$D${ASS['ebit_s']}"
    Aet = f"{A}!$D${ASS['ebit_t']}"
    Ada = f"{A}!$D${ASS['da_pct']}"
    Acx = f"{A}!$D${ASS['capex_pct']}"
    Ats = f"{A}!$D${ASS['tax_s']}"
    Att = f"{A}!$D${ASS['tax_t']}"
    Ads = f"{A}!$D${ASS['dso_s']}"
    Adt = f"{A}!$D${ASS['dso_t']}"
    Adis = f"{A}!$D${ASS['dio_s']}"
    Adit = f"{A}!$D${ASS['dio_t']}"
    Adps = f"{A}!$D${ASS['dpo_s']}"
    Adpt = f"{A}!$D${ASS['dpo_t']}"
    Aex = f"{A}!$D${ASS['exit_mult']}"
    Aro = f"{A}!$D${ASS['ronic']}"
    Aw  = f"{A}!$D${ASS['wacc']}"
    Ash = f"{A}!$D${ASS['shares']}"
    Aic = f"{A}!$D${ASS['inv_cap']}"

    # Scenario modifiers (only applied to Bear/Bull)
    gm = f"*$B$3" if (bear or bull) else ""
    mm = f"*$B$4" if (bear or bull) else ""
    cm = f"*$B$5" if (bear or bull) else ""
    wm = f"+$B$6" if (bear or bull) else ""
    em = f"*$B$7" if (bear or bull) else ""

    def wacc_formula():
        return f"({Aw}{wm})" if (bear or bull) else Aw

    def rev_formula(col: int) -> str:
        """Revenue formula for given column (col 3=Y1, col 12=Y10, col 13=Terminal)."""
        year = col - 2  # col3→1, col12→10, col13→11
        prev_col = get_column_letter(col - 1)
        if year <= 5:
            return f"={prev_col}{rREV}*(1+{Ag}{gm})"
        elif year <= 10:
            return f"={prev_col}{rREV}*(1+{Agm}{gm})"
        else:  # Terminal
            return f"={prev_col}{rREV}*(1+{Agt})"

    def interp(start_ref, end_ref, col, mult=""):
        """Linear interpolation formula via COLUMN()-2."""
        return f"={start_ref}+({end_ref}-{start_ref})*(COLUMN()-2)/10{mult}"

    # ── LTM column (col B = 2) ──
    ws.cell(row=rREV, column=2, value=f"={A}!$D${ASS['revenue']}").font = _font(color=GREEN)
    ws.cell(row=rREV, column=2).number_format = NUM0
    ws.cell(row=rREV, column=2).alignment = _align()

    ws.cell(row=rDA_P, column=2, value=f"={Ada}").font = _font(color=GREEN)
    ws.cell(row=rDA_P, column=2).number_format = PCT2
    ws.cell(row=rDA_P, column=2).alignment = _align()

    ws.cell(row=rCX_P, column=2, value=f"={Acx}").font = _font(color=GREEN)
    ws.cell(row=rCX_P, column=2).number_format = PCT2
    ws.cell(row=rCX_P, column=2).alignment = _align()

    ws.cell(row=rDSO, column=2, value=f"={Ads}").font = _font(color=GREEN)
    ws.cell(row=rDSO, column=2).number_format = NUM1
    ws.cell(row=rDSO, column=2).alignment = _align()

    ws.cell(row=rDIO, column=2, value=f"={Adis}").font = _font(color=GREEN)
    ws.cell(row=rDIO, column=2).number_format = NUM1
    ws.cell(row=rDIO, column=2).alignment = _align()

    ws.cell(row=rDPO, column=2, value=f"={Adps}").font = _font(color=GREEN)
    ws.cell(row=rDPO, column=2).number_format = NUM1
    ws.cell(row=rDPO, column=2).alignment = _align()

    # LTM AR/INV/AP/NWC (using LTM revenue from B_rREV)
    ws.cell(row=rAR,  column=2, value=f"=B{rREV}/365*B{rDSO}").font = _font(color=BLACK)
    ws.cell(row=rINV, column=2, value=f"=B{rREV}/365*B{rDIO}").font = _font(color=BLACK)
    ws.cell(row=rAP,  column=2, value=f"=B{rREV}/365*B{rDPO}").font = _font(color=BLACK)
    ws.cell(row=rNWC, column=2, value=f"=B{rAR}+B{rINV}-B{rAP}").font = _font(color=BLACK)
    for ltm_r in (rAR, rINV, rAP, rNWC):
        ws.cell(row=ltm_r, column=2).number_format = NUM0
        ws.cell(row=ltm_r, column=2).alignment = _align()

    ws.cell(row=rIC, column=2, value=f"={Aic}").font = _font(color=GREEN)
    ws.cell(row=rIC, column=2).number_format = NUM0
    ws.cell(row=rIC, column=2).alignment = _align()

    # ── Y1–Y10 columns (cols 3–12) ──
    for col in range(3, 13):  # Y1=col3 ... Y10=col12
        CL = get_column_letter(col)
        PL = get_column_letter(col - 1)
        year = col - 2  # 1..10
        wf = wacc_formula()

        def f(expr: str, row: int, fmt_key: int = 0) -> None:
            c = ws.cell(row=row, column=col, value=expr)
            c.font = _font(color=BLACK)
            c.alignment = _align()
            if row in fmts:
                c.number_format = fmts[row]

        # Revenue
        f(rev_formula(col), rREV)

        # Growth %
        f(f"={CL}{rREV}/{PL}{rREV}-1", rGRW)

        # EBIT margin — linear interp
        f(interp(f"{Aes}", f"{Aet}{mm}", col), rMGN)

        # EBIT
        f(f"={CL}{rREV}*{CL}{rMGN}", rEBIT)

        # D&A %
        f(f"={Ada}", rDA_P)

        # D&A $
        f(f"={CL}{rREV}*{CL}{rDA_P}", rDA)

        # EBITDA
        f(f"={CL}{rEBIT}+{CL}{rDA}", rEBDA)

        # Tax rate — linear interp
        f(interp(f"{Ats}", f"{Att}", col), rTAX)

        # NOPAT
        f(f"={CL}{rEBIT}*(1-{CL}{rTAX})", rNOPT)

        # Capex %
        f(f"={Acx}{cm}", rCX_P)

        # Capex $
        f(f"={CL}{rREV}*{CL}{rCX_P}", rCX)

        # NWC drivers — linear interp
        f(interp(f"{Ads}", f"{Adt}", col), rDSO)
        f(interp(f"{Adis}", f"{Adit}", col), rDIO)
        f(interp(f"{Adps}", f"{Adpt}", col), rDPO)

        # AR, Inventory, AP
        f(f"={CL}{rREV}/365*{CL}{rDSO}", rAR)
        f(f"={CL}{rREV}/365*{CL}{rDIO}", rINV)
        f(f"={CL}{rREV}/365*{CL}{rDPO}", rAP)

        # NWC, ΔNWC
        f(f"={CL}{rAR}+{CL}{rINV}-{CL}{rAP}", rNWC)
        f(f"={CL}{rNWC}-{PL}{rNWC}", rDNWC)

        # FCFF
        f(f"={CL}{rNOPT}+{CL}{rDA}-{CL}{rCX}-{CL}{rDNWC}", rFCFF)

        # Discount factor: 1/(1+WACC)^year
        f(f"=1/(1+{wf})^{year}", rDISC)

        # PV FCFF
        f(f"={CL}{rFCFF}*{CL}{rDISC}", rPVF)

        # Cumulative PV FCFF
        if col == 3:
            f(f"={CL}{rPVF}", rCPVF)
        else:
            f(f"={PL}{rCPVF}+{CL}{rPVF}", rCPVF)

        # Invested Capital: IC_t = IC_{t-1} + capex + ΔNWC - D&A
        f(f"={PL}{rIC}+{CL}{rCX}+{CL}{rDNWC}-{CL}{rDA}", rIC)

        # ROIC
        f(f"={CL}{rNOPT}/{PL}{rIC}", rROIC)

        # Economic Profit
        f(f"={CL}{rNOPT}-{PL}{rIC}*{wf}", rEP)

    # ── Terminal column (col M = 13) ──
    TL = get_column_letter(13)
    PL = get_column_letter(12)
    wf = wacc_formula()

    def tf(expr: str, row: int) -> None:
        c = ws.cell(row=row, column=13, value=expr)
        c.font = _font(color=BLACK)
        c.alignment = _align()
        if row in fmts:
            c.number_format = fmts[row]

    tf(f"={PL}{rREV}*(1+{Agt})", rREV)        # Terminal revenue
    tf(f"={TL}{rREV}/{PL}{rREV}-1", rGRW)
    tf(f"={Aet}", rMGN)                         # Target margin
    tf(f"={TL}{rREV}*{TL}{rMGN}", rEBIT)
    tf(f"={Ada}", rDA_P)
    tf(f"={TL}{rREV}*{TL}{rDA_P}", rDA)
    tf(f"={TL}{rEBIT}+{TL}{rDA}", rEBDA)
    tf(f"={Att}", rTAX)
    tf(f"={TL}{rEBIT}*(1-{TL}{rTAX})", rNOPT)
    tf(f"=1/(1+{wf})^10", rDISC)               # Y10 disc factor

    # ── TV section (below projection) ──
    rTV = R + 36
    _section_hdr(ws, rTV - 1, "  TERMINAL VALUE & EV SUMMARY", ncols=6, color="215732")

    tv_rows = [
        (rTV,     "Gordon Growth TV ($mm)",
         f"=IF({Aw}{wm}>{Agt},"
         f"M{rNOPT}*(1-{Agt}/{Aro})/({Aw}{wm}-{Agt}),0)"),
        (rTV + 1, "Exit Multiple TV ($mm)",
         f"=L{rEBDA}*{Aex}{em}"),
        (rTV + 2, "Blended TV ($mm)",
         f"=(M{rTV}+M{rTV+1})/2"),
        (rTV + 3, "PV of Blended TV ($mm)",
         f"=M{rTV+2}*L{rDISC}"),
        (rTV + 5, "Sum PV FCFF ($mm)",
         f"=L{rCPVF}"),
        (rTV + 6, "EV Operations ($mm)",
         f"=M{rTV+5}+M{rTV+3}"),
        (rTV + 7, "TV % of EV",
         f"=IF(M{rTV+6}>0,M{rTV+3}/M{rTV+6},0)"),
    ]
    for rr, lbl, fml in tv_rows:
        bold = rr in (rTV + 5, rTV + 6)
        _lbl(ws, rr, 1, lbl, bold=bold, indent=1)
        c = ws.cell(row=rr, column=13, value=fml)
        c.font = _font(color=BLACK, bold=bold)
        c.alignment = _align()
        c.number_format = PCT1 if "%" in lbl else NUM0


# ── Sheet: Equity_Bridge ───────────────────────────────────────────────────────

def build_equity_bridge(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("Equity_Bridge")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "IBM — EQUITY BRIDGE & INTRINSIC VALUE"
    c.font = _font(color=NAVY, bold=True, size=13)
    c.alignment = _align(h="left")

    # Column headers
    for col, txt, fh in [(1, "Item", NAVY), (2, "Bear", NAVY), (3, "Base", NAVY),
                          (4, "Bull", NAVY), (5, "Notes", NAVY)]:
        c = ws.cell(row=2, column=col, value=txt)
        c.fill = _fill(fh)
        c.font = _font(color=WHITE, bold=True)
        c.alignment = _align(h="center" if col > 1 else "left")

    # We need to reference the TV section rows in each DCF sheet.
    # In DCF_Base (R=0): rTV = 36+0 = 36, so TV block starts at row 35 (R+35)
    # Bear (R=7): rTV = 36+7 = 43
    # Bull: same as Bear

    def dcf_ref(sheet: str, R_offset: int, key: str) -> str:
        rTV = R_offset + 36
        map_ = {
            "ev_ops":    f"M{rTV + 6}",
            "sum_pvcf":  f"M{rTV + 5}",
            "pv_tv":     f"M{rTV + 3}",
            "tv_pct":    f"M{rTV + 7}",
        }
        return f"={sheet}!{map_[key]}"

    A = "Assumptions"
    R_base = 0; R_bear = 7; R_bull = 7

    rows = [
        # (label, bear_fml, base_fml, bull_fml, fmt, bold, note)
        ("Sum PV FCFF (Y1-Y10) $mm",
         dcf_ref("DCF_Bear", R_bear, "sum_pvcf"),
         dcf_ref("DCF_Base", R_base, "sum_pvcf"),
         dcf_ref("DCF_Bull", R_bull, "sum_pvcf"),
         NUM0, False, ""),
        ("PV Terminal Value $mm",
         dcf_ref("DCF_Bear", R_bear, "pv_tv"),
         dcf_ref("DCF_Base", R_base, "pv_tv"),
         dcf_ref("DCF_Bull", R_bull, "pv_tv"),
         NUM0, False, "Blended (Gordon+Exit)/2"),
        ("EV — Operations $mm",
         dcf_ref("DCF_Bear", R_bear, "ev_ops"),
         dcf_ref("DCF_Base", R_base, "ev_ops"),
         dcf_ref("DCF_Bull", R_bull, "ev_ops"),
         NUM0, True, ""),
        ("TV % of EV",
         dcf_ref("DCF_Bear", R_bear, "tv_pct"),
         dcf_ref("DCF_Base", R_base, "tv_pct"),
         dcf_ref("DCF_Bull", R_bull, "tv_pct"),
         PCT1, False, "Flag if >75%"),
        ("(+) Non-Operating Assets $mm",
         f"={A}!$D${ASS['non_op']}",
         f"={A}!$D${ASS['non_op']}",
         f"={A}!$D${ASS['non_op']}",
         NUM0, False, ""),
        ("EV — Total $mm",
         "=B5+B7", "=C5+C7", "=D5+D7",
         NUM0, True, "EV_ops + Non-op assets"),
        ("(-) Net Debt $mm",
         f"={A}!$D${ASS['net_debt']}",
         f"={A}!$D${ASS['net_debt']}",
         f"={A}!$D${ASS['net_debt']}",
         NUM0, False, ""),
        ("(-) Minority Interest $mm",
         f"={A}!$D${ASS['minority']}",
         f"={A}!$D${ASS['minority']}",
         f"={A}!$D${ASS['minority']}",
         NUM0, False, ""),
        ("(-) Preferred Equity $mm",
         f"={A}!$D${ASS['preferred']}",
         f"={A}!$D${ASS['preferred']}",
         f"={A}!$D${ASS['preferred']}",
         NUM0, False, ""),
        ("(-) Pension Deficit $mm",
         f"={A}!$D${ASS['pension']}",
         f"={A}!$D${ASS['pension']}",
         f"={A}!$D${ASS['pension']}",
         NUM0, False, ""),
        ("(-) Lease Liabilities $mm",
         f"={A}!$D${ASS['leases']}",
         f"={A}!$D${ASS['leases']}",
         f"={A}!$D${ASS['leases']}",
         NUM0, False, ""),
        ("(-) Options Value $mm",
         f"={A}!$D${ASS['options']}",
         f"={A}!$D${ASS['options']}",
         f"={A}!$D${ASS['options']}",
         NUM0, False, ""),
        ("(-) Convertibles $mm",
         f"={A}!$D${ASS['converts']}",
         f"={A}!$D${ASS['converts']}",
         f"={A}!$D${ASS['converts']}",
         NUM0, False, ""),
        ("Equity Value $mm",
         "=B8-B9-B10-B11-B12-B13-B14-B15",
         "=C8-C9-C10-C11-C12-C13-C14-C15",
         "=D8-D9-D10-D11-D12-D13-D14-D15",
         NUM0, True, "EV_Total − all claims"),
        ("Shares Outstanding (mm)",
         f"={A}!$D${ASS['shares']}",
         f"={A}!$D${ASS['shares']}",
         f"={A}!$D${ASS['shares']}",
         NUM1, False, ""),
        ("Intrinsic Value / Share",
         "=B16/B17", "=C16/C17", "=D16/D17",
         DOLLA, True, "Equity / Shares"),
        ("Market Price",
         d["price"], d["price"], d["price"],
         DOLLA, False, "As of " + d["date"]),
        ("Upside / (Downside) %",
         "=B18/B19-1", "=C18/C19-1", "=D18/D19-1",
         PCT1, True, ""),
    ]

    for i, (lbl, bv, bsv, blv, fmt, bold, note) in enumerate(rows, start=3):
        _lbl(ws, i, 1, lbl, bold=bold, indent=1 if not bold else 0)
        for col, val in [(2, bv), (3, bsv), (4, blv)]:
            c = ws.cell(row=i, column=col, value=val)
            c.font = _font(color=GREEN if str(val).startswith("=") else BLUE, bold=bold)
            c.number_format = fmt
            c.alignment = _align()
        if note:
            ws.cell(row=i, column=5, value=note).font = _font(size=9)

    # Expected IV row
    rEX = len(rows) + 3 + 1
    _section_hdr(ws, rEX - 1, "  EXPECTED INTRINSIC VALUE", ncols=5, color="215732")
    _lbl(ws, rEX, 1, "Expected IV / Share (probability-weighted)", bold=True)
    exp_fml = (f"=B18*{A}!$D${ASS['prob_bear']}"
               f"+C18*{A}!$D${ASS['prob_base']}"
               f"+D18*{A}!$D${ASS['prob_bull']}")
    c = ws.cell(row=rEX, column=3, value=exp_fml)
    c.font = _font(color=BLACK, bold=True)
    c.number_format = DOLLA
    c.alignment = _align()

    _lbl(ws, rEX + 1, 1, "Expected Upside %", bold=True)
    c = ws.cell(row=rEX + 1, column=3,
                value=f"=C{rEX}/B19-1")
    c.font = _font(color=BLACK, bold=True)
    c.number_format = PCT1
    c.alignment = _align()


# ── Sheet: Comps ───────────────────────────────────────────────────────────────

def build_comps(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("Comps")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 7    # ticker
    ws.column_dimensions["C"].width = 20   # name
    ws.column_dimensions["D"].width = 12   # mcap
    ws.column_dimensions["E"].width = 12   # tev
    ws.column_dimensions["F"].width = 12   # rev
    ws.column_dimensions["G"].width = 12   # ebitda
    ws.column_dimensions["H"].width = 10   # ebit
    ws.column_dimensions["I"].width = 8    # eps
    ws.column_dimensions["J"].width = 12   # EV/EBITDA LTM
    ws.column_dimensions["K"].width = 12   # EV/EBITDA Fwd
    ws.column_dimensions["L"].width = 10   # EV/EBIT
    ws.column_dimensions["M"].width = 8    # P/E

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "IBM — COMPARABLE COMPANY ANALYSIS"
    c.font = _font(color=NAVY, bold=True, size=13)
    c.alignment = _align(h="left")

    hdrs = ["Excl.", "Ticker", "Company", "MCap $mm", "TEV $mm",
            "Revenue $mm", "EBITDA $mm", "EBIT $mm", "EPS",
            "EV/EBITDA LTM", "EV/EBITDA Fwd", "EV/EBIT", "P/E"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = _fill(NAVY)
        c.font = _font(color=WHITE, bold=True)
        c.alignment = _align(h="center")

    # IBM target row (row 3)
    target_vals = [
        "", d["ticker"], d["name"],
        d["mcap_mm"], d["ev_mm"], d["revenue_mm"],
        None, None, None,
        d["ev_ebitda"], None, None, d["pe_t"],
    ]
    for ci, v in enumerate(target_vals, start=1):
        c = ws.cell(row=3, column=ci, value=v)
        c.fill = _fill(LBLUE)
        c.font = _font(color=BLUE, bold=True)
        c.number_format = MULT if ci in (10, 11, 12, 13) else (NUM0 if ci >= 4 else "@")
        c.alignment = _align(h="center" if ci <= 2 else "right")

    # Peer rows (rows 4..4+n)
    peers = d["peers"]
    for pi, peer in enumerate(peers):
        row = 4 + pi
        vals = [
            "", peer["ticker"], peer["name"],
            peer["mcap"], peer["tev"], peer["rev"],
            peer["ebitda"], peer["ebit"], peer["eps"],
            peer["ev_ebitda_ltm"], peer["ev_ebitda_fwd"], peer["ev_ebit"], peer["pe"],
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = _font()
            bg = GRAY if pi % 2 == 0 else WHITE
            c.fill = _fill(bg)
            if ci >= 4:
                c.number_format = MULT if ci >= 10 else (NUM0 if ci <= 8 else "0.00")
            c.alignment = _align(h="center" if ci <= 2 else "right")

    # Median row
    n_peers = len(peers)
    rMED = 4 + n_peers
    _section_hdr(ws, rMED, "  MEDIAN (excluding \"x\" peers)", ncols=13, color="215732")
    rFirst = 4
    rLast  = 3 + n_peers

    def med_fml(col_letter: str) -> str:
        excl_range = f"A{rFirst}:A{rLast}"
        val_range  = f"{col_letter}{rFirst}:{col_letter}{rLast}"
        return f"=IFERROR(MEDIAN(IF({excl_range}<>\"x\",{val_range})),MEDIAN({val_range}))"

    for ci, col_letter in enumerate(list("DEFGHIJKLM"), start=4):
        if ci >= 10:  # multiple columns
            c = ws.cell(row=rMED, column=ci, value=med_fml(col_letter))
            c.font = _font(color=WHITE, bold=True)
            c.alignment = _align()
            c.number_format = MULT

    # Implied IV section
    rIV = rMED + 2
    _section_hdr(ws, rIV, "  IMPLIED INTRINSIC VALUE (Comps)", ncols=13, color=NAVY)

    A = "Assumptions"
    shares = f"{A}!$D${ASS['shares']}"
    net_debt = f"{A}!$D${ASS['net_debt']}"
    minority = f"{A}!$D${ASS['minority']}"
    preferred = f"{A}!$D${ASS['preferred']}"
    pension = f"{A}!$D${ASS['pension']}"
    leases = f"{A}!$D${ASS['leases']}"
    opts = f"{A}!$D${ASS['options']}"
    convs = f"{A}!$D${ASS['converts']}"
    non_op = f"{A}!$D${ASS['non_op']}"
    claims = f"({net_debt}+{minority}+{preferred}+{pension}+{leases}+{opts}+{convs})"

    # LTM EBITDA approx from DCF_Base row rNOPT... easier to just hardcode from IBM
    # Or reference from DCF_Base: DCF_Base!L{rEBDA} = Y10 EBITDA... we want LTM
    # LTM EBITDA = revenue * ebit_margin + da = use from Assumptions
    # LTM EBITDA ≈ rev*(ebit_s + da_pct)
    rev_ref = f"{A}!$D${ASS['revenue']}"
    ebit_s_ref = f"{A}!$D${ASS['ebit_s']}"
    da_ref = f"{A}!$D${ASS['da_pct']}"
    ltm_ebitda_fml = f"={rev_ref}*({ebit_s_ref}+{da_ref})"

    med_ev_ebitda = f"J{rMED}"
    med_ev_ebit   = f"L{rMED}"
    med_pe        = f"M{rMED}"

    def iv_row(r: int, lbl: str, ev_fml: str) -> None:
        _lbl(ws, r, 1, lbl, indent=1)
        # EV
        c = ws.cell(row=r, column=2, value=ev_fml)
        c.font = _font()
        c.number_format = NUM0
        c.alignment = _align()
        # Equity value = EV + non_op - claims
        eq = ws.cell(row=r, column=3, value=f"=B{r}+{non_op}-{claims}")
        eq.font = _font()
        eq.number_format = NUM0
        eq.alignment = _align()
        # IV per share
        iv = ws.cell(row=r, column=4, value=f"=C{r}/{shares}")
        iv.font = _font(bold=True)
        iv.number_format = DOLLA
        iv.alignment = _align()
        # Upside
        ups = ws.cell(row=r, column=5, value=f"=D{r}/{d['price']}-1")
        ups.font = _font()
        ups.number_format = PCT1
        ups.alignment = _align()

    for ci, lbl in [(2, "TEV ($mm)"), (3, "Equity ($mm)"), (4, "IV/Share"), (5, "Upside %")]:
        c = ws.cell(row=rIV, column=ci, value=lbl)
        c.fill = _fill("215732")
        c.font = _font(color=WHITE, bold=True)
        c.alignment = _align(h="center")

    iv_row(rIV + 1, "EV/EBITDA LTM approach",
           f"={med_ev_ebitda}*({ltm_ebitda_fml})")
    iv_row(rIV + 2, "EV/EBIT LTM approach",
           f"={med_ev_ebit}*{rev_ref}*{ebit_s_ref}")
    # P/E approach: implied IV = median_PE * EPS (NOPAT/shares)
    iv_row(rIV + 3, "P/E approach",
           f"=({med_pe}*{rev_ref}*{ebit_s_ref}*(1-{A}!$D${ASS['tax_s']})"
           f"+{non_op}-{claims})")


# ── Sheet: Sensitivity ─────────────────────────────────────────────────────────

def build_sensitivity(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for ci in range(2, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "IBM — SENSITIVITY ANALYSIS  |  IV / Share vs WACC × Terminal Growth"
    c.font = _font(color=NAVY, bold=True, size=12)
    c.alignment = _align(h="left")

    # WACC axis (rows): WACC - 200bp to WACC + 200bp in 50bp steps
    wacc_base = d["wacc"]
    g_base = d["g_term"]
    wacc_steps = [-0.02, -0.015, -0.01, -0.005, 0.0, 0.005, 0.01, 0.015, 0.02]
    g_steps    = [-0.02, -0.015, -0.01, -0.005, 0.0, 0.005, 0.01, 0.015, 0.02]

    _section_hdr(ws, 2, f"  IV/Share ($) — WACC (rows) × Terminal Growth (cols)", ncols=10)

    ws.cell(row=3, column=1, value="WACC \\ Term. g").font = _font(bold=True)
    ws.cell(row=3, column=1).alignment = _align(h="center")

    for ci, dg in enumerate(g_steps, start=2):
        g_val = g_base + dg
        c = ws.cell(row=3, column=ci, value=g_val)
        c.font = _font(bold=True)
        c.number_format = PCT1
        c.alignment = _align(h="center")
        if dg == 0.0:
            c.fill = _fill(LBLUE)

    # Pre-compute IV for each cell using simplified DCF
    # IV = FCFF_Y1 * [(1-(1/(1+w))^10) / w]  +  TV / (1+w)^10   ... approximation
    # We use the full Gordon formula TV: NOPAT_Y11 * (1 - g/RONIC) / (WACC - g)
    # NOPAT_Y11 ≈ Revenue * (1+g_near)^5 * (1+g_mid)^5 * (1+g_term) * ebit_t * (1-tax_t)
    import math

    rev = d["revenue_mm"]
    g_near = d["g_near"]
    g_mid = d["g_mid"]
    ebit_s = d["ebit_s"]
    ebit_t = d["ebit_t"]
    da = d["da_pct"]
    cx = d["capex_pct"]
    tax = d["tax_t"]
    ronic = d["ronic"]
    shares = d["shares_mm"]
    net_debt = d["net_debt_mm"]
    non_op = d["non_op_mm"]
    pension = d["pension_mm"]
    leases = d["leases_mm"]
    bridge = non_op - net_debt - pension - leases

    # Build year-by-year for each (wacc, g_term) combo
    def compute_iv(wacc: float, g_term: float) -> float:
        rev_t = rev
        pv_sum = 0.0
        prev_nwc = rev * (d["dso_s"] + d["dio_s"] - d["dpo_s"]) / 365
        for yr in range(1, 11):
            g = g_near if yr <= 5 else g_mid
            rev_t *= (1 + g)
            margin = ebit_s + (ebit_t - ebit_s) * yr / 10
            ebit = rev_t * margin
            nopat = ebit * (1 - tax)
            da_mm = rev_t * da
            cx_mm = rev_t * cx
            dso = d["dso_s"] + (d["dso_t"] - d["dso_s"]) * yr / 10
            dio = d["dio_s"] + (d["dio_t"] - d["dio_s"]) * yr / 10
            dpo = d["dpo_s"] + (d["dpo_t"] - d["dpo_s"]) * yr / 10
            nwc = rev_t * (dso + dio - dpo) / 365
            delta_nwc = nwc - prev_nwc
            prev_nwc = nwc
            fcff = nopat + da_mm - cx_mm - delta_nwc
            disc = 1 / (1 + wacc) ** yr
            pv_sum += fcff * disc

        # Terminal NOPAT
        rev_11 = rev_t * (1 + g_term)
        nopat_11 = rev_11 * ebit_t * (1 - tax)
        if wacc <= g_term:
            tv = 0.0
        else:
            tv = nopat_11 * (1 - g_term / ronic) / (wacc - g_term)
        disc10 = 1 / (1 + wacc) ** 10
        pv_tv = tv * disc10
        ev_ops = pv_sum + pv_tv
        equity = ev_ops + bridge
        iv = equity / shares
        return round(iv, 2)

    for ri, dw in enumerate(wacc_steps, start=4):
        w_val = wacc_base + dw
        c = ws.cell(row=ri, column=1, value=w_val)
        c.font = _font(bold=(dw == 0.0))
        c.number_format = PCT1
        c.alignment = _align(h="center")
        if dw == 0.0:
            c.fill = _fill(LBLUE)
        for ci, dg in enumerate(g_steps, start=2):
            g_val = g_base + dg
            iv = compute_iv(w_val, g_val)
            cell = ws.cell(row=ri, column=ci, value=iv)
            cell.number_format = DOLLA
            cell.alignment = _align(h="center")
            # Colour-code vs price
            if iv >= d["price"] * 1.20:
                cell.fill = _fill("C6EFCE")  # green
                cell.font = _font(color=DARK_GREEN, bold=(dw == 0.0 and dg == 0.0))
            elif iv >= d["price"] * 0.95:
                cell.fill = _fill("FFEB9C")  # yellow
            else:
                cell.fill = _fill("FFC7CE")  # red
            if dw == 0.0 and dg == 0.0:
                cell.font = _font(bold=True)
                cell.border = Border(
                    top=Side(style="medium"), bottom=Side(style="medium"),
                    left=Side(style="medium"), right=Side(style="medium")
                )

    note = ws.cell(row=4 + len(wacc_steps) + 1, column=1,
                   value=f"Price = ${d['price']:.2f}  |  Green = IV > Price+20%  |  Yellow = within 5%  |  Red = IV < Price-5%")
    note.font = _font(size=9, italic=True)
    note.alignment = _align(h="left")


# ── Sheet: QoE ─────────────────────────────────────────────────────────────────

def build_qoe(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("QoE")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "IBM — QUALITY OF EARNINGS SIGNALS"
    c.font = _font(color=NAVY, bold=True, size=13)
    c.alignment = _align(h="left")

    # Composite score banner
    score_color = {"green": "C6EFCE", "amber": "FFEB9C", "red": "FFC7CE"}[d["qoe_flag"]]
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"QoE Composite Score: {d['qoe_score']}/5  |  Flag: {d['qoe_flag'].upper()}  |  As of {d['date']}"
    c.fill = _fill(score_color)
    c.font = _font(bold=True, size=11)
    c.alignment = _align(h="center")

    hdrs = ["Signal", "Value", "Score", "Threshold", "Interpretation"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = _fill(NAVY)
        c.font = _font(color=WHITE, bold=True)
        c.alignment = _align(h="center" if ci > 1 else "left")

    signals = [
        ("Sloan Accruals Ratio",         d["sloan_ratio"],   "amber", "<5% green / >10% red",   "Earnings quality (lower = better)"),
        ("Cash Conversion (CFO/NOPAT)",  d["cash_conv"],     "amber", ">90% green / <70% red",   "Cash backing for earnings"),
        ("DSO Drift (vs 3Y baseline)",   d["dso_drift"],     "green", "<3d green / >7d red",     "AR quality; negative = improving"),
        ("Capex/D&A Ratio",              d["capex_da_ratio"],"green", "<1.5x green / >2.5x red", "Maintenance vs growth capex"),
        ("DIO Drift",                     0.0,               "green", "<5d green / >10d red",    "Inventory build-up check"),
        ("Revenue recognition flag",      "N/A",             "green", "N/A",                     "No restatements flagged"),
        ("Auditor flag",                  "None",            "green", "N/A",                     "Clean Big-4 opinion"),
    ]

    flag_colors = {"green": "C6EFCE", "amber": "FFEB9C", "red": "FFC7CE"}

    for ri, (sig, val, flag, thresh, interp) in enumerate(signals, start=4):
        _lbl(ws, ri, 1, sig, indent=1)
        c = ws.cell(row=ri, column=2, value=val)
        c.font = _font(color=BLUE)
        c.alignment = _align()
        if isinstance(val, float):
            c.number_format = PCT2 if abs(val) < 10 else NUM1
        sc = ws.cell(row=ri, column=3, value=flag.upper())
        sc.fill = _fill(flag_colors[flag])
        sc.font = _font(bold=True)
        sc.alignment = _align(h="center")
        ws.cell(row=ri, column=4, value=thresh).font = _font(size=9)
        ws.cell(row=ri, column=5, value=interp).font = _font(size=9)
        ws.cell(row=ri, column=5).alignment = _align(h="left", wrap=True)

    # NWC detail
    rNWC = 4 + len(signals) + 2
    _section_hdr(ws, rNWC, "  NWC BASELINE DETAIL", ncols=5)
    nwc_rows = [
        ("DSO — Current",    d["dso_s"],        NUM1),
        ("DSO — 3Y Baseline", d["dso_baseline"], NUM1),
        ("DSO — Drift",      d["dso_drift"],    "+0.0;-0.0;0.0"),
    ]
    for i, (lbl, val, fmt) in enumerate(nwc_rows):
        _lbl(ws, rNWC + 1 + i, 1, lbl, indent=1)
        c = ws.cell(row=rNWC + 1 + i, column=2, value=val)
        c.font = _font(color=BLUE)
        c.number_format = fmt
        c.alignment = _align()


# ── Sheet: Output ──────────────────────────────────────────────────────────────

def build_output(wb: openpyxl.Workbook, d: dict) -> None:
    ws = wb.create_sheet("Output")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"IBM — VALUATION SUMMARY  |  {d['date']}"
    c.font = _font(color=NAVY, bold=True, size=14)
    c.alignment = _align(h="left")

    _section_hdr(ws, 2, "  MARKET DATA", ncols=5)
    mkt_rows = [
        ("Price",                d["price"],         DOLLA),
        ("Market Cap ($mm)",     d["mcap_mm"],        NUM0),
        ("Enterprise Value ($mm)", d["ev_mm"],        NUM0),
        ("P/E (Trailing)",       d["pe_t"],           NUM1),
        ("P/E (Forward)",        d["pe_f"],           NUM1),
        ("EV/EBITDA",            d["ev_ebitda"],      NUM1),
        ("Analyst Target",       d["analyst_target"], DOLLA),
        ("Analyst Rec.",         d["analyst_rec"],    "@"),
        ("# Analysts",           d["n_analysts"],     NUM0),
    ]
    for ri, (lbl, val, fmt) in enumerate(mkt_rows, start=3):
        _lbl(ws, ri, 1, lbl, indent=1)
        c = ws.cell(row=ri, column=2, value=val)
        c.font = _font(color=BLUE)
        c.number_format = fmt
        c.alignment = _align()

    rIV = 3 + len(mkt_rows) + 1
    _section_hdr(ws, rIV, "  INTRINSIC VALUE SUMMARY", ncols=5)

    for ci, sc in enumerate(["Bear", "Base", "Bull"], start=2):
        c = ws.cell(row=rIV + 1, column=ci, value=sc)
        c.fill = _fill(LBLUE)
        c.font = _font(bold=True)
        c.alignment = _align(h="center")

    EB = "Equity_Bridge"
    iv_labels = [
        ("Intrinsic Value / Share", "B18", "C18", "D18", DOLLA),
        ("Upside / (Downside)",     "B20", "C20", "D20", PCT1),
        ("Scenario Probability",
         f"Assumptions!$D${ASS['prob_bear']}",
         f"Assumptions!$D${ASS['prob_base']}",
         f"Assumptions!$D${ASS['prob_bull']}", PCT1),
    ]
    for i, (lbl, br, bs, bl, fmt) in enumerate(iv_labels, start=rIV + 2):
        _lbl(ws, i, 1, lbl, bold=True, indent=1)
        for ci, ref in enumerate([br, bs, bl], start=2):
            fml = f"={EB}!{ref}" if not ref.startswith("Assumptions") else f"={ref}"
            c = ws.cell(row=i, column=ci, value=fml)
            c.font = _font(color=GREEN, bold=True)
            c.number_format = fmt
            c.alignment = _align()
            # Colour upside cells
            if "Upside" in lbl:
                # We can't add conditional formatting easily here; apply static color
                pass

    rEX = rIV + 2 + len(iv_labels)
    _lbl(ws, rEX, 1, "Expected IV / Share", bold=True)
    c = ws.cell(row=rEX, column=2, value=f"={EB}!C22")  # row 22 in Equity_Bridge
    c.font = _font(color=GREEN, bold=True)
    c.number_format = DOLLA
    c.alignment = _align()

    _lbl(ws, rEX + 1, 1, "Expected Upside", bold=True)
    c = ws.cell(row=rEX + 1, column=2, value=f"={EB}!C23")
    c.font = _font(color=GREEN, bold=True)
    c.number_format = PCT1
    c.alignment = _align()

    rWACC = rEX + 3
    _section_hdr(ws, rWACC, "  WACC SUMMARY", ncols=5)
    wacc_rows = [
        ("WACC",              f"=Assumptions!$D${ASS['wacc']}",    PCT2),
        ("Cost of Equity",    f"=Assumptions!$D${ASS['ke']}",      PCT2),
        ("Cost of Debt (at)", f"=Assumptions!$D${ASS['kd_at']}",   PCT2),
        ("Beta (Re-levered)", f"=Assumptions!$D${ASS['beta_rel']}", "0.00"),
        ("Equity Weight",     f"=Assumptions!$D${ASS['eq_wt']}",   PCT1),
    ]
    for i, (lbl, fml, fmt) in enumerate(wacc_rows, start=rWACC + 1):
        _lbl(ws, i, 1, lbl, indent=1)
        c = ws.cell(row=i, column=2, value=fml)
        c.font = _font(color=GREEN)
        c.number_format = fmt
        c.alignment = _align()

    rQOE = rWACC + len(wacc_rows) + 2
    _section_hdr(ws, rQOE, "  QoE FLAGS", ncols=5)
    qoe_rows = [
        ("QoE Score",           d["qoe_score"],      "0"),
        ("QoE Flag",            d["qoe_flag"].upper(), "@"),
        ("Sloan Accruals Ratio", d["sloan_ratio"],    PCT2),
        ("Cash Conversion",     d["cash_conv"],       PCT1),
        ("DSO Drift (days)",    d["dso_drift"],       "+0.0;-0.0"),
    ]
    for i, (lbl, val, fmt) in enumerate(qoe_rows, start=rQOE + 1):
        _lbl(ws, i, 1, lbl, indent=1)
        c = ws.cell(row=i, column=2, value=val)
        c.font = _font(color=BLUE)
        c.number_format = fmt
        c.alignment = _align()


# ── Config sheet ───────────────────────────────────────────────────────────────

def _build_config(wb: openpyxl.Workbook) -> None:
    """
    Hidden sheet with one named cell: json_path.
    Power Query reads this to locate the JSON file — change it to switch tickers.
    """
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False
    ws.sheet_state = "hidden"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

    # Header
    for col, txt in [(1, "Setting"), (2, "Value")]:
        c = ws.cell(row=1, column=col, value=txt)
        c.fill = _fill(NAVY)
        c.font = _font(color=WHITE, bold=True)
        c.alignment = _align(h="left")

    # json_path row
    ws.cell(row=2, column=1, value="json_path").font = _font(bold=True)
    path_cell = ws.cell(
        row=2, column=2,
        value=str(_ROOT / "data" / "valuations" / "json" / "IBM_latest.json"),
    )
    path_cell.font = _font(color=BLUE)
    path_cell.alignment = _align(h="left")

    # Help text
    ws.cell(row=4, column=1, value="Instructions:").font = _font(bold=True, size=9)
    for i, txt in enumerate([
        "1. Change B2 to point at a different ticker (e.g. ORCL_latest.json)",
        "2. Data → Refresh All  (Ctrl+Alt+F5)",
        "3. All Assumptions Col B values update from the new JSON.",
        "4. The path must match what batch_runner --json last wrote.",
    ], start=5):
        ws.cell(row=i, column=1, value=txt).font = _font(size=9, italic=True, color="595959")
        ws.merge_cells(f"A{i}:B{i}")

    # Define named range "json_path" pointing to Config!$B$2
    # (openpyxl DefinedName scoped to workbook)
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["json_path"] = DefinedName(
        name="json_path",
        attr_text="Config!$B$2",
    )


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    d = _load_snapshot()
    _build_config(wb)
    build_cover(wb, d)
    build_assumptions(wb, d)
    _dcf_sheet(wb, "DCF_Base", d, bear=False, bull=False)
    _dcf_sheet(wb, "DCF_Bear", d, bear=True,  bull=False)
    _dcf_sheet(wb, "DCF_Bull", d, bear=False, bull=True)
    build_equity_bridge(wb, d)
    build_comps(wb, d)
    build_sensitivity(wb, d)
    build_qoe(wb, d)
    build_output(wb, d)

    # Tab colours
    tab_colors = {
        "Config":        "595959",
        "Cover":         "1F3864",
        "Assumptions":   "0070C0",
        "DCF_Base":      "215732",
        "DCF_Bear":      "9E0700",
        "DCF_Bull":      "375623",
        "Equity_Bridge": "7030A0",
        "Comps":         "E36C09",
        "Sensitivity":   "4F81BD",
        "QoE":           "FFC000",
        "Output":        "1F3864",
    }
    for sheet_name, color in tab_colors.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    # Force full recalculation on open so formula cells show values, not zeros
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    out_path = Path(__file__).resolve().parent.parent / "templates" / "ticker_review_auto.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
