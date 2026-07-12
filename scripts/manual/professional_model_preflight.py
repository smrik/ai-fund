"""Read-only source preflight for the professional integrated financial model."""
from __future__ import annotations

import argparse
import hashlib
import importlib.metadata
import json
import platform
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from ciq.workbook_parser import CIQTemplateContractError, parse_ciq_workbook  # noqa: E402
from config import DB_PATH  # noqa: E402


EXCEL_ERROR_VALUES = frozenset(
    {
        "#NULL!",
        "#DIV/0!",
        "#VALUE!",
        "#REF!",
        "#NAME?",
        "#NUM!",
        "#N/A",
        "#GETTING_DATA",
        "#SPILL!",
        "#CALC!",
    }
)


class PreflightError(RuntimeError):
    """Raised when a source cannot safely enter professional-model assembly."""


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _mtime_iso(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat().replace("+00:00", "Z")


def _sqlite_ro_uri(path: Path) -> str:
    return f"file:{path.resolve().as_posix()}?mode=ro"


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_workbook_path(
    *,
    ticker: str,
    workbook_path: str | Path | None,
    search_roots: Iterable[str | Path] | None = None,
) -> Path:
    """Resolve one ticker workbook without silently choosing among duplicates."""
    expected_ticker = ticker.strip().upper()
    if not expected_ticker:
        raise PreflightError("ticker is required")

    if workbook_path is not None:
        explicit = Path(workbook_path).expanduser().resolve()
        if not explicit.exists() or not explicit.is_file():
            raise PreflightError(f"workbook does not exist: {explicit}")
        return explicit

    roots = tuple(Path(root).expanduser().resolve() for root in (search_roots or (REPO_ROOT / "data" / "exports",)))
    expected_name = f"{expected_ticker}_Standard.xlsx".lower()
    candidates: dict[str, Path] = {}
    for root in roots:
        if not root.exists():
            continue
        for candidate in root.glob("*_Standard.xlsx"):
            if candidate.is_file() and candidate.name.lower() == expected_name and not candidate.name.startswith("~$"):
                resolved = candidate.resolve()
                candidates[str(resolved).lower()] = resolved

    matches = sorted(candidates.values(), key=lambda item: str(item).lower())
    if not matches:
        searched = ", ".join(str(root) for root in roots)
        raise PreflightError(f"no {expected_ticker}_Standard.xlsx found under: {searched}")
    if len(matches) > 1:
        listed = ", ".join(str(path) for path in matches)
        raise PreflightError(f"multiple workbook candidates found; pass --workbook-path explicitly: {listed}")
    return matches[0]


def inspect_workbook_cells(workbook_path: str | Path, *, max_error_records: int = 100) -> dict:
    """Inspect workbook structure, formulas, and both formula and cached errors."""
    path = Path(workbook_path)
    formula_book = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    cached_book = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        formula_count = 0
        formula_errors: list[dict] = []
        sheets: list[dict] = []
        for sheet in formula_book.worksheets:
            sheet_formula_count = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        formula_count += 1
                        sheet_formula_count += 1
                        formula_text = str(cell.value).upper()
                        if any(token in formula_text for token in EXCEL_ERROR_VALUES):
                            formula_errors.append(
                                {
                                    "sheet": sheet.title,
                                    "cell": cell.coordinate,
                                    "value": cell.value,
                                    "kind": "formula_reference_error",
                                }
                            )
            sheets.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "formula_count": sheet_formula_count,
                }
            )

        cached_errors: list[dict] = []
        for sheet in cached_book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    normalized = str(cell.value).strip().upper() if cell.value is not None else ""
                    if cell.data_type == "e" or normalized in EXCEL_ERROR_VALUES:
                        cached_errors.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "value": cell.value,
                                "kind": "cached_value_error",
                            }
                        )

        all_errors = [*formula_errors, *cached_errors]
        errors = all_errors[:max_error_records]
        return {
            "sheet_count": len(sheets),
            "sheets": sheets,
            "formula_count": formula_count,
            "formula_error_count": len(formula_errors),
            "cached_error_count": len(cached_errors),
            "error_count": len(all_errors),
            "errors": errors,
            "errors_truncated": len(all_errors) > len(errors),
        }
    finally:
        formula_book.close()
        cached_book.close()


def _matching_ingest(db_path: Path, *, source_file: str, file_hash: str, ticker: str, parser_version: str) -> dict:
    if not db_path.exists():
        return {"ingest_status": "not_ingested", "run_id": None, "db_path": str(db_path.resolve())}
    try:
        with sqlite3.connect(_sqlite_ro_uri(db_path), uri=True) as conn:
            conn.row_factory = sqlite3.Row
            table = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'ciq_ingest_runs'"
            ).fetchone()
            if table is None:
                return {"ingest_status": "not_ingested", "run_id": None, "db_path": str(db_path.resolve())}
            row = conn.execute(
                """
                SELECT id AS run_id, source_file, file_hash, ticker, parser_version,
                       ingest_ts, status, as_of_date
                FROM ciq_ingest_runs
                WHERE source_file = ? AND file_hash = ? AND ticker = ? AND parser_version = ?
                ORDER BY id DESC
                LIMIT 1
                """,
                (source_file, file_hash, ticker, parser_version),
            ).fetchone()
    except sqlite3.Error as exc:
        raise PreflightError(f"cannot inspect CIQ ingest database {db_path}: {exc}") from exc

    if row is None:
        return {"ingest_status": "not_ingested", "run_id": None, "db_path": str(db_path.resolve())}
    result = dict(row)
    result["ingest_status"] = "matched"
    result["db_path"] = str(db_path.resolve())
    return result


def _git_identity(repo_root: Path = REPO_ROOT) -> dict:
    def run(*args: str) -> str | None:
        completed = subprocess.run(
            ["git", *args],
            cwd=repo_root,
            text=True,
            capture_output=True,
            timeout=10,
            check=False,
        )
        return completed.stdout.strip() if completed.returncode == 0 else None

    status = run("status", "--porcelain") or ""
    return {
        "branch": run("rev-parse", "--abbrev-ref", "HEAD"),
        "head": run("rev-parse", "HEAD"),
        "dirty_path_count": len([line for line in status.splitlines() if line.strip()]),
    }


def _package_versions(names: Iterable[str]) -> dict[str, str | None]:
    versions: dict[str, str | None] = {}
    for name in names:
        try:
            versions[name] = importlib.metadata.version(name)
        except importlib.metadata.PackageNotFoundError:
            versions[name] = None
    return versions


def _package_available(name: str) -> bool:
    try:
        importlib.metadata.version(name)
    except importlib.metadata.PackageNotFoundError:
        return False
    return True


def _read_workbook_metadata(workbook_path: Path, expected_ticker: str) -> dict:
    book = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    try:
        metadata = {
            "workbook_refresh_date": None,
            "currency": None,
            "conversion": None,
            "fiscal_year_end": None,
            "target_ltm_period_end": None,
        }
        if "Input" in book.sheetnames:
            sheet = book["Input"]
            inputs = {
                str(sheet.cell(row, 1).value or "").strip().lower(): sheet.cell(row, 2).value
                for row in range(1, min(sheet.max_row, 100) + 1)
            }
            metadata["workbook_refresh_date"] = _to_iso(inputs.get("date"))
            metadata["currency"] = str(inputs.get("currency") or "").strip() or None
            metadata["conversion"] = str(inputs.get("conversion") or "").strip() or None

        if "Detailed Comps" in book.sheetnames:
            sheet = book["Detailed Comps"]
            headers = {
                str(sheet.cell(1, column).value or "").strip().lower(): column
                for column in range(1, min(sheet.max_column, 250) + 1)
            }
            ticker_col = headers.get("ticker")
            if ticker_col:
                for row in range(2, min(sheet.max_row, 1200) + 1):
                    raw_ticker = str(sheet.cell(row, ticker_col).value or "").strip().upper()
                    if raw_ticker.split(":")[-1] != expected_ticker:
                        continue
                    metadata["fiscal_year_end"] = _to_iso(sheet.cell(row, headers.get("fye", 0)).value) if headers.get("fye") else None
                    metadata["target_ltm_period_end"] = _to_iso(sheet.cell(row, headers.get("ltm", 0)).value) if headers.get("ltm") else None
                    break
        return metadata
    finally:
        book.close()


def _to_iso(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    return text[:10] if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-" else None


def build_preflight_manifest(
    *,
    ticker: str,
    workbook_path: str | Path,
    db_path: str | Path = DB_PATH,
    require_ingested: bool = False,
) -> dict:
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise PreflightError(f"workbook does not exist: {path}")

    try:
        payload = parse_ciq_workbook(path)
    except CIQTemplateContractError as exc:
        raise PreflightError(f"CIQ workbook contract failed: {exc}") from exc

    expected_ticker = ticker.strip().upper()
    if payload.ticker.upper() != expected_ticker:
        raise PreflightError(f"ticker mismatch: expected {expected_ticker}, workbook has {payload.ticker}")

    actual_hash = sha256_file(path)
    if actual_hash != payload.file_hash:
        raise PreflightError(f"parser hash mismatch: direct={actual_hash}, parser={payload.file_hash}")

    workbook = inspect_workbook_cells(path)
    ingest = _matching_ingest(
        Path(db_path),
        source_file=payload.source_file,
        file_hash=payload.file_hash,
        ticker=payload.ticker,
        parser_version=payload.parser_version,
    )
    try:
        template_fingerprint = json.loads(payload.template_fingerprint)
    except (TypeError, json.JSONDecodeError):
        template_fingerprint = payload.template_fingerprint

    blockers: list[str] = []
    warnings: list[str] = []
    as_of_date = payload.valuation_snapshot.get("as_of_date")
    if not as_of_date:
        blockers.append("workbook_as_of_date_missing")
    if workbook["formula_error_count"]:
        blockers.append(f"formula_reference_errors:{workbook['formula_error_count']}")
    if workbook["cached_error_count"]:
        blockers.append(f"cached_excel_errors:{workbook['cached_error_count']}")
    if ingest["ingest_status"] != "matched":
        message = "matching_ciq_ingest_run_missing"
        (blockers if require_ingested else warnings).append(message)
    elif ingest.get("status") != "completed":
        blockers.append(f"ciq_ingest_status:{ingest.get('status')}")

    workbook_metadata = _read_workbook_metadata(path, expected_ticker)
    if not workbook_metadata["workbook_refresh_date"]:
        warnings.append("workbook_refresh_date_missing")
    if not workbook_metadata["currency"]:
        warnings.append("workbook_currency_missing")

    manifest = {
        "schema_version": "professional_model_preflight_v1",
        "generated_at": _utc_now_iso(),
        "ticker": expected_ticker,
        "status": "blocked" if blockers else "ready",
        "blockers": blockers,
        "warnings": warnings,
        "source": {
            "path": str(path),
            "source_file": payload.source_file,
            "sha256": actual_hash,
            "size_bytes": path.stat().st_size,
            "file_modified_at": _mtime_iso(path),
            "workbook_as_of_date": as_of_date,
            **workbook_metadata,
            **ingest,
        },
        "parser": {
            "parser_version": payload.parser_version,
            "template_fingerprint": template_fingerprint,
            "rows_parsed": payload.rows_parsed,
            "long_form_count": len(payload.long_form_records),
            "comps_fact_count": len(payload.comps_snapshot),
            "valuation_snapshot": payload.valuation_snapshot,
        },
        "workbook": workbook,
        "environment": {
            "python_executable": sys.executable,
            "python_version": platform.python_version(),
            "platform": platform.platform(),
            "packages": _package_versions(("pytest", "pandas", "openpyxl", "pydantic", "xlwings")),
            "recalculation_tools": {
                "libreoffice": shutil.which("soffice"),
                "excel_com_candidate": platform.system() == "Windows" and _package_available("xlwings"),
            },
        },
        "git": _git_identity(),
    }
    return manifest


def write_manifest(manifest: dict, *, output_root: str | Path = REPO_ROOT / "output" / "professional_models") -> Path:
    ticker = str(manifest["ticker"]).upper()
    fingerprint = str(manifest["source"]["sha256"])[:12]
    path = Path(output_root) / ticker / fingerprint / "preflight_fingerprint.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Fingerprint and validate a CIQ Standard workbook for model assembly.")
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--workbook-path")
    parser.add_argument("--db-path", default=str(DB_PATH))
    parser.add_argument("--output-root", default=str(REPO_ROOT / "output" / "professional_models"))
    parser.add_argument("--require-ingested", action="store_true")
    args = parser.parse_args(argv)

    try:
        workbook_path = resolve_workbook_path(ticker=args.ticker, workbook_path=args.workbook_path)
        manifest = build_preflight_manifest(
            ticker=args.ticker,
            workbook_path=workbook_path,
            db_path=args.db_path,
            require_ingested=args.require_ingested,
        )
        output_path = write_manifest(manifest, output_root=args.output_root)
    except PreflightError as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1

    print(
        json.dumps(
            {
                "status": manifest["status"],
                "ticker": manifest["ticker"],
                "workbook": manifest["source"]["path"],
                "sha256": manifest["source"]["sha256"],
                "run_id": manifest["source"].get("run_id"),
                "as_of_date": manifest["source"].get("workbook_as_of_date"),
                "rows_parsed": manifest["parser"]["rows_parsed"],
                "formula_reference_errors": manifest["workbook"]["formula_error_count"],
                "cached_excel_errors": manifest["workbook"]["cached_error_count"],
                "manifest": str(output_path),
                "blockers": manifest["blockers"],
                "warnings": manifest["warnings"],
            },
            indent=2,
        )
    )
    return 1 if manifest["blockers"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
