"""
Excel Template Generator — creates the PM review workbook.

Creates a formatted .xlsx template with:
1. Instructions sheet — how to connect Power Query
2. Overrides sheet — where the PM enters assumption adjustments
3. Model sheet —  columns that merge Python data with overrides (formulas)
4. Dashboard sheet — summary view with conditional formatting

The template is designed to be connected to data/valuations/latest.csv
via Power Query. The PM sets this up once, then just clicks Refresh All.

Usage:
    python -m src.stage_02_valuation.create_template
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


ROOT_DIR = Path(__file__).resolve().parent.parent.parent
TEMPLATE_DIR = ROOT_DIR / "templates"


# ── Styles ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
OVERRIDE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NUMBER_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
PCT_FMT = "0.0%"
MONEY_FMT = "$#,##0.00"
MONEY_MM_FMT = "$#,##0"
MULT_FMT = "0.0x"


def _style_header_row(ws, num_cols):
    """Apply header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def create_instructions_sheet(wb: Workbook):
    """Create the Instructions tab with setup guide."""
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "2F5496"

    content = [
        ("ALPHA POD — Valuation Review Template", "A1", Font(size=16, bold=True, color="2F5496")),
        ("", "A2", None),
        ("SETUP (one-time)", "A3", Font(size=13, bold=True)),
        ("1. In Excel: Data → Get Data → From Text/CSV", "A4", None),
        ("2. Navigate to: ai-fund/data/valuations/latest.csv", "A5", None),
        ("3. Click 'Load To...' → select 'Table' and 'Existing worksheet'", "A6", None),
        ("4. Place at cell A1 on the 'Data' sheet", "A7", None),
        ("5. Done! The Data sheet now auto-refreshes from Python output", "A8", None),
        ("", "A9", None),
        ("DAILY WORKFLOW", "A10", Font(size=13, bold=True)),
        ("1. Run:  python -m src.stage_02_valuation.batch_runner", "A11", None),
        ("2. Open this workbook → Data → Refresh All (Ctrl+Alt+F5)", "A12", None),
        ("3. Review the Dashboard tab — sorted by upside", "A13", None),
        ("4. To adjust assumptions: go to Overrides tab, enter your values", "A14", None),
        ("5. Model tab auto-merges Python defaults with your overrides", "A15", None),
        ("", "A16", None),
        ("OVERRIDE RULES", "A17", Font(size=13, bold=True)),
        ("• Leave a cell blank in Overrides = use Python default", "A18", None),
        ("• Enter a value in Overrides = your value takes priority", "A19", None),
        ("• Add a reason so you remember why you changed it", "A20", None),
        ("• Overrides persist across refreshes — Python never touches this file", "A21", None),
    ]

    for text, cell_ref, font in content:
        cell = ws[cell_ref]
        cell.value = text
        if font:
            cell.font = font

    ws.column_dimensions["A"].width = 70


def create_data_sheet(wb: Workbook):
    """Create empty Data sheet where Power Query loads latest.csv."""
    ws = wb.create_sheet("Data")
    ws.sheet_properties.tabColor = "808080"

    # Placeholder headers matching latest.csv columns
    headers = [
        "ticker", "company_name", "sector", "industry",
        "price", "market_cap_mm", "ev_mm",
        "pe_trailing", "pe_forward", "ev_ebitda",
        "revenue_mm", "op_margin", "profit_margin", "rev_growth",
        "fcf_mm", "net_debt_mm", "beta_raw",
        "wacc", "cost_of_equity", "beta_relevered", "beta_unlevered",
        "size_premium", "equity_weight", "peers_used",
        "iv_bear", "iv_base", "iv_bull",
        "upside_base_pct", "upside_bear_pct", "upside_bull_pct",
        "margin_of_safety",
        "growth_near", "growth_mid", "ebit_margin_used",
        "exit_multiple_used", "tv_pct_of_ev",
        "analyst_target", "analyst_recommendation", "num_analysts",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    ws.cell(row=2, column=1, value="← Power Query loads data here. Data → Refresh All to update.")
    ws.cell(row=2, column=1).font = Font(italic=True, color="999999")


def create_overrides_sheet(wb: Workbook):
    """Create the Overrides tab where the PM enters adjustments."""
    ws = wb.create_sheet("Overrides")
    ws.sheet_properties.tabColor = "FFC000"

    headers = [
        "Ticker", "growth_near (%)", "growth_mid (%)",
        "ebit_margin (%)", "wacc (%)", "exit_multiple (x)",
        "Reason",
    ]
    col_widths = [10, 15, 15, 15, 12, 16, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    # Example rows
    examples = [
        ("HALO", 12.0, 8.0, None, None, None, "Cap growth — drug patent cliff in 2030"),
        ("LYFT", None, None, None, None, 25.0, "SaaS-like unit economics justify higher multiple"),
        ("", None, None, None, None, None, ""),
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = OVERRIDE_FILL
            cell.font = Font(name="Calibri", size=10)

    # Fill the override area with yellow background
    for row in range(5, 50):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = OVERRIDE_FILL

    ws.row_dimensions[1].height = 30


def create_model_sheet(wb: Workbook):
    """
    Create the Model sheet with formulas that merge Data + Overrides.
    Uses VLOOKUP: =IF(override exists, override, data default).
    """
    ws = wb.create_sheet("Model")
    ws.sheet_properties.tabColor = "00B050"

    headers = [
        "Ticker", "Company", "Sector", "Price",
        "Growth Near (%)", "Growth Mid (%)", "EBIT Margin (%)",
        "WACC (%)", "Exit Multiple (x)",
        "IV Bear", "IV Base", "IV Bull",
        "Upside Base (%)", "Margin of Safety (%)",
        "Source",  # "Default" or "Override"
    ]
    col_widths = [10, 25, 18, 10, 14, 14, 14, 10, 14, 10, 10, 10, 14, 16, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30

    # Add formula rows — the user populates these by referencing Data + Overrides
    # Row 2: formula template (copy down)
    ws.cell(row=3, column=1).value = "← Fill formulas referencing Data + Overrides sheets"
    ws.cell(row=3, column=1).font = Font(italic=True, color="999999")

    # Add formula guide in row 2
    formula_guide = [
        "=Data!A2",                    # Ticker
        "=Data!B2",                    # Company
        "=Data!C2",                    # Sector
        "=Data!E2",                    # Price
        '=IFERROR(VLOOKUP(A2,Overrides!A:G,2,0), Data!AF2)',   # Growth Near
        '=IFERROR(VLOOKUP(A2,Overrides!A:G,3,0), Data!AG2)',   # Growth Mid
        '=IFERROR(VLOOKUP(A2,Overrides!A:G,4,0), Data!AH2)',   # EBIT Margin
        '=IFERROR(VLOOKUP(A2,Overrides!A:G,5,0), Data!R2)',    # WACC
        '=IFERROR(VLOOKUP(A2,Overrides!A:G,6,0), Data!AI2)',   # Exit Multiple
        "=Data!Y2",                    # IV Bear
        "=Data!Z2",                    # IV Base
        "=Data!AA2",                   # IV Bull
        "=Data!AB2",                   # Upside Base
        "=Data!AE2",                   # Margin of Safety
        '=IF(ISNA(VLOOKUP(A2,Overrides!A:G,2,0)),"Default","Override")',
    ]

    for col, formula in enumerate(formula_guide, 1):
        cell = ws.cell(row=2, column=col, value=formula)
        cell.font = Font(name="Calibri", size=10, italic=True, color="4472C4")

    # Note about the formulas
    ws.cell(row=4, column=1).value = (
        "FORMULA GUIDE: Row 2 has template formulas. "
        "After Power Query loads data into the Data sheet, adjust column references "
        "to match, then copy row 2 down for all tickers."
    )
    ws.cell(row=4, column=1).font = Font(italic=True, color="999999", size=9)
    ws.merge_cells("A4:O4")

    # Conditional formatting: highlight overridden rows
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA")
    ws.conditional_formatting.add(
        "O2:O500",
        CellIsRule(operator="equal", formula=['"Override"'], fill=green_fill)
    )

    # Conditional formatting: upside > 20% = green, < -20% = red
    ws.conditional_formatting.add(
        "M2:M500",
        CellIsRule(operator="greaterThan", formula=["20"], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE"))
    )
    ws.conditional_formatting.add(
        "M2:M500",
        CellIsRule(operator="lessThan", formula=["-20"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE"))
    )


def create_dashboard_sheet(wb: Workbook):
    """Create a dashboard summary sheet with formatted sections."""
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "4472C4"

    # Move to first position
    wb.move_sheet(ws, offset=-3)

    # Title
    ws["A1"] = "ALPHA POD — Valuation Dashboard"
    ws["A1"].font = Font(size=16, bold=True, color="2F5496")
    ws["A2"] = '=TEXT(NOW(), "YYYY-MM-DD HH:MM")'
    ws["A2"].font = Font(size=10, color="808080")

    # Quick stats section
    ws["A4"] = "QUICK STATS"
    ws["A4"].font = Font(size=12, bold=True)
    ws["A4"].fill = SECTION_FILL

    stats = [
        ("Total Valued:", '=COUNTA(Data!A:A)-1'),
        ("Avg Upside (base):", '=AVERAGE(Data!AB:AB)'),
        ("Median WACC:", '=MEDIAN(Data!R:R)'),
        ("Names > 20% upside:", '=COUNTIF(Data!AB:AB,">20")'),
        ("Names > 50% upside:", '=COUNTIF(Data!AB:AB,">50")'),
    ]
    for i, (label, formula) in enumerate(stats):
        ws.cell(row=5 + i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=5 + i, column=2, value=formula)

    # Sector breakdown
    ws["A12"] = "SECTOR BREAKDOWN"
    ws["A12"].font = Font(size=12, bold=True)
    ws["A12"].fill = SECTION_FILL

    ws["A13"] = "Sector"
    ws["B13"] = "Count"
    ws["C13"] = "Avg Upside"
    for cell in [ws["A13"], ws["B13"], ws["C13"]]:
        cell.font = Font(bold=True)

    sectors = [
        "Technology", "Healthcare", "Industrials", "Consumer Cyclical",
        "Consumer Defensive", "Energy", "Basic Materials", "Communication Services"
    ]
    for i, sector in enumerate(sectors):
        row = 14 + i
        ws.cell(row=row, column=1, value=sector)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Data!C:C,"{sector}")')
        ws.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIF(Data!C:C,"{sector}",Data!AB:AB),"-")')

    # Instructions reference
    ws["A24"] = "HOW TO USE"
    ws["A24"].font = Font(size=12, bold=True)
    ws["A24"].fill = SECTION_FILL
    ws["A25"] = "1. Run batch_runner → Ctrl+Alt+F5 to refresh"
    ws["A26"] = "2. Review this Dashboard for overview"
    ws["A27"] = "3. Overrides tab to adjust assumptions for specific tickers"
    ws["A28"] = "4. Model tab shows merged output (Python default + your overrides)"

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


def create_template():
    """Generate the full review template."""
    wb = Workbook()

    create_instructions_sheet(wb)
    create_data_sheet(wb)
    create_overrides_sheet(wb)
    create_model_sheet(wb)
    create_dashboard_sheet(wb)

    # Save
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    output_path = TEMPLATE_DIR / "valuation_review.xlsx"
    wb.save(output_path)

    print("=" * 60)
    print("ALPHA POD — Template Created")
    print("=" * 60)
    print(f"  Output: {output_path}")
    print()
    print("  Next steps:")
    print("  1. Open the template in Excel")
    print("  2. Go to the 'Data' sheet")
    print("  3. Data → Get Data → From Text/CSV")
    print("  4. Select: data/valuations/latest.csv")
    print("  5. Load To → Existing worksheet, cell A1 on Data sheet")
    print("  6. Done! Dashboard + Model will auto-populate")
    print()

    return output_path


if __name__ == "__main__":
    create_template()
