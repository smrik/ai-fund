"""Deterministic 26-sheet renderer for the professional financial model.

This module is deliberately additive.  It does not retrieve, infer, or mutate
source data.  Callers must supply a normalized payload with:

* a single, immutable source-workbook run identity;
* source-located historical observations for every supplied canonical line;
* exactly five forecast periods for base, upside, and downside;
* frozen valuation assumptions and explicitly typed optional-data states.

The workbook is a presentation/calculation artifact, not a second source of
truth.  Source observations and forecast assumptions are hardcoded only on the
``Historical_Data`` and ``Scenarios``/``Assumptions`` sheets.  Model sheets,
WACC, DCF, valuation bridges, sensitivities, and integrity checks use formulas.
The renderer never opens the source workbook and never creates external links.

Limitations of this first Task-11 slice are intentionally visible: segment,
consensus, and SOTP analyses are only calculated when normalized evidence is
supplied; otherwise their sheets carry a typed ``UNAVAILABLE`` state.  Native
Excel recalculation/parity is also outside this renderer, so its returned
manifest remains blocked on ``recalculation_not_run`` until a later stage.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import date
from enum import Enum
import hashlib
import json
import math
from pathlib import Path
import re
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from src.contracts.professional_financial_model import (
    AvailabilityState,
    AvailabilityStatus,
    CellClassification,
    CellKind,
    CheckResult,
    CheckStatus,
    DriverApprovalRecord,
    ConsensusSnapshot,
    DefinedNameMapping,
    LineCellMapping,
    PROFESSIONAL_WORKBOOK_SHEETS,
    SourcePresentationRecord,
    WorkbookCheckCell,
    WorkbookManifest,
)


RENDERER_VERSION = "professional-model-openpyxl-0.2.0"
REQUIRED_SHEET_ORDER = PROFESSIONAL_WORKBOOK_SHEETS
REQUIRED_SCENARIOS = ("base", "upside", "downside")
MODEL_LINE_SHEETS = frozenset(
    {
        "Segment_Build",
        "Income_Statement",
        "Balance_Sheet",
        "Cash_Flow",
        "Working_Capital",
        "PP&E_Intangibles",
        "Debt_Cash_Interest",
        "Capital_Allocation",
        "Taxes",
        "Shares_EPS",
    }
)
REQUIRED_DCF_LINES = frozenset(
    {
        "revenue",
        "ebit",
        "depreciation_amortization",
        "capex",
        "change_in_net_working_capital",
        "cf.unlevered_fcf",
    }
)
BRIDGE_ASSET_KEYS = (
    "bridge_cash",
    "bridge_short_term_investments",
    "bridge_long_term_investments",
)
BRIDGE_BORROWING_KEYS = (
    "bridge_short_term_borrowings",
    "bridge_current_long_term_debt",
    "bridge_long_term_debt",
)
BRIDGE_LEASE_KEYS = (
    "bridge_current_lease_liabilities",
    "bridge_long_term_lease_liabilities",
)
BRIDGE_OTHER_CLAIM_KEYS = (
    "bridge_minority_interest",
    "bridge_pension_liability",
)
BRIDGE_REFERENCE_KEYS = (
    "bridge_gross_debt",
    "bridge_total_borrowings",
    "bridge_lease_liabilities",
)
DIAGNOSTIC_METHODS = (
    ("diagnostic_comps_ev_ebitda_per_share", "Diagnostic comps: EV / EBITDA", "per_share"),
    ("diagnostic_comps_ev_ebit_per_share", "Diagnostic comps: EV / EBIT", "per_share"),
    ("diagnostic_comps_pe_per_share", "Diagnostic comps: P / E", "per_share"),
    ("diagnostic_economic_profit_per_share", "Diagnostic economic profit", "per_share"),
    ("diagnostic_v1_gordon_per_share", "Diagnostic v1 Gordon growth", "per_share"),
    ("diagnostic_v1_exit_per_share", "Diagnostic v1 exit multiple", "per_share"),
    ("diagnostic_reverse_dcf_implied_growth_pct_points", "Diagnostic reverse-DCF implied growth", "percentage_points"),
)
WACC_REQUIRED_INPUTS = frozenset(
    {
        "risk_free_rate",
        "beta",
        "equity_risk_premium",
        "pre_tax_cost_of_debt",
        "tax_rate",
        "debt_value",
        "equity_value",
        "size_premium",
        "backend_selected_wacc",
    }
)
DCF_TIMING_INPUTS = frozenset(
    {
        *(f"dcf_discount_exponent_{index}" for index in range(1, 6)),
        "dcf_terminal_discount_exponent",
        *(f"dcf_annual_fy26_fcff_{scenario}" for scenario in ("base", "upside", "downside")),
        *(f"dcf_ytd_fcff_{scenario}" for scenario in ("base", "upside", "downside")),
        *(f"dcf_stub_fcff_{scenario}" for scenario in ("base", "upside", "downside")),
    }
)
DCF_SCENARIO_INPUTS = frozenset(
    {
        *(f"dcf_wacc_{scenario}" for scenario in ("base", "upside", "downside")),
        *(f"dcf_nopat_tax_rate_{scenario}" for scenario in ("base", "upside", "downside")),
        *(f"dcf_terminal_growth_{scenario}" for scenario in ("base", "upside", "downside")),
    }
)
DCF_REQUIRED_VALUATION_INPUTS = frozenset(
    {
        *WACC_REQUIRED_INPUTS,
        "current_fully_diluted_shares",
        *DCF_TIMING_INPUTS,
        *DCF_SCENARIO_INPUTS,
        *BRIDGE_ASSET_KEYS,
        *BRIDGE_BORROWING_KEYS,
        *BRIDGE_LEASE_KEYS,
        *BRIDGE_OTHER_CLAIM_KEYS,
    }
)
BRIDGE_RECONCILIATION_INPUTS = frozenset(BRIDGE_REFERENCE_KEYS)
PRICE_COMPARISON_INPUTS = frozenset({"current_price"})
REQUIRED_VALUATION_INPUTS = frozenset(
    DCF_REQUIRED_VALUATION_INPUTS
    | BRIDGE_RECONCILIATION_INPUTS
    | PRICE_COMPARISON_INPUTS
)


_SHA256_RE = re.compile(r"^[0-9a-f]{64}$")
_A1_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")


# Banker-style conventions: assumptions blue, formulas black, sourced facts
# green, checks red/green, and unavailable states gray.  Fills are restrained.
NAVY = "17365D"
MEDIUM_BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF2F8"
INPUT_BLUE = "0000FF"
SOURCE_GREEN = "008000"
CHECK_GREEN = "008000"
CHECK_RED = "C00000"
FORMULA_BLACK = "000000"
GRAY = "7F7F7F"
LIGHT_GRAY = "E7E6E6"
PALE_YELLOW = "FFF2CC"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE4D6"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="B7B7B7")


def _clean_text(value: str, field_name: str) -> str:
    cleaned = str(value).strip()
    if not cleaned:
        raise ValueError(f"{field_name} is required")
    return cleaned


def _canonical_hash(value: Any) -> str:
    def default(item: Any) -> Any:
        if isinstance(item, date):
            return item.isoformat()
        if isinstance(item, Enum):
            return item.value
        if hasattr(item, "model_dump"):
            return item.model_dump(mode="json")
        raise TypeError(f"cannot serialize {type(item)!r}")

    encoded = json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        allow_nan=False,
        default=default,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _formula_text_hash(workbook: Workbook) -> str:
    rows: list[dict[str, str]] = []
    identities: set[tuple[str, str]] = set()
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                sheet = str(worksheet.title).strip()
                coordinate = str(cell.coordinate).replace("$", "").strip().upper()
                formula_text = str(cell.value)
                identity = (sheet, coordinate)
                if not sheet or not coordinate or not formula_text or identity in identities:
                    raise ValueError("formula evidence must be unique and complete")
                identities.add(identity)
                rows.append(
                    {"sheet": sheet, "cell": coordinate, "formula_text": formula_text}
                )
    encoded = json.dumps(
        sorted(rows, key=lambda item: (item["sheet"], item["cell"])),
        sort_keys=True, separators=(",", ":"), ensure_ascii=False, allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()

def _quote_sheet(sheet: str) -> str:
    return "'" + sheet.replace("'", "''") + "'"


def _absolute_ref(sheet: str, coordinate: str) -> str:
    match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", coordinate.upper())
    if match is None:
        raise ValueError(f"invalid A1 coordinate: {coordinate}")
    return f"{_quote_sheet(sheet)}!${match.group(1)}${match.group(2)}"


@dataclass(frozen=True)
class TypedAvailability:
    """Explicit availability for evidence-dependent workbook sections."""

    status: str
    reason_code: str | None = None
    message: str | None = None

    def __post_init__(self) -> None:
        status = str(self.status).lower().strip()
        if status not in {"available", "unavailable", "blocking", "pm_required"}:
            raise ValueError(f"unsupported availability status: {self.status}")
        object.__setattr__(self, "status", status)
        if status == "available":
            if self.reason_code or self.message:
                raise ValueError("available state cannot carry a reason")
        elif not self.reason_code or not self.message:
            raise ValueError("non-available state requires reason_code and message")


@dataclass(frozen=True)
class SourceWorkbookRun:
    """Frozen identity and preflight outcome for one ingested workbook run."""

    source_file: str
    source_path: str
    source_hash: str
    run_id: int
    parser_version: str
    status: str
    fact_count: int
    formula_error_count: int

    def __post_init__(self) -> None:
        for field_name in ("source_file", "source_path", "parser_version"):
            object.__setattr__(self, field_name, _clean_text(getattr(self, field_name), field_name))
        source_hash = str(self.source_hash).lower().strip()
        if not _SHA256_RE.fullmatch(source_hash):
            raise ValueError("source_hash must be a lowercase SHA-256 digest")
        object.__setattr__(self, "source_hash", source_hash)
        if self.run_id <= 0 or self.fact_count < 0 or self.formula_error_count < 0:
            raise ValueError("source run/count fields must be non-negative and run_id positive")
        status = str(self.status).lower().strip()
        if status not in {"ready", "ok", "blocked", "unavailable", "partial"}:
            raise ValueError(f"unsupported source status: {self.status}")
        object.__setattr__(self, "status", status)


@dataclass(frozen=True)
class HistoricalSourceCell:
    """One historical value with exact source-workbook lineage."""

    period_key: str
    value: float | int | None
    source_sheet: str | None
    source_cell: str | None
    source_row_id: str
    source_formula: str | None = None
    formula_status: str = "calculated"

    raw_value: Any = None
    normalized_value: float | int | None = None
    derived_value: float | int | None = None
    transformation_rule: str | None = None
    unit: str | None = None
    unit_kind: str | None = None
    scale: float | None = None
    currency: str | None = None
    period_type: str | None = None
    period_end: date | None = None
    formula_error: str | None = None
    upstream_dependencies: tuple[str, ...] = ()
    downstream_dependencies: tuple[str, ...] = ()

    def __post_init__(self) -> None:
        object.__setattr__(self, "period_key", _clean_text(self.period_key, "period_key"))
        if self.source_sheet is None or self.source_cell is None:
            if self.source_sheet is not None or self.source_cell is not None:
                raise ValueError("historical source sheet/cell must both be supplied or both absent")
            if self.formula_status != "typed_unavailable":
                raise ValueError("only typed-unavailable history may omit a source locator")
        else:
            object.__setattr__(self, "source_sheet", _clean_text(self.source_sheet, "source_sheet"))
            cell = str(self.source_cell).upper().replace("$", "").strip()
            if not _A1_RE.fullmatch(cell):
                raise ValueError(f"invalid historical source cell: {self.source_cell}")
            object.__setattr__(self, "source_cell", cell)
        object.__setattr__(self, "source_row_id", _clean_text(self.source_row_id, "source_row_id"))

        object.__setattr__(
            self,
            "upstream_dependencies",
            tuple(sorted({_clean_text(item, "upstream dependency") for item in self.upstream_dependencies})),
        )
        object.__setattr__(
            self,
            "downstream_dependencies",
            tuple(sorted({_clean_text(item, "downstream dependency") for item in self.downstream_dependencies})),
        )

@dataclass(frozen=True)
class ScenarioForecast:
    """One scenario's ordered, frozen forecast results for one model line."""

    scenario_key: str
    values: tuple[tuple[str, float | int | None], ...]

    def __post_init__(self) -> None:
        object.__setattr__(self, "scenario_key", _clean_text(self.scenario_key, "scenario_key").lower())
        normalized = tuple((_clean_text(period, "forecast period"), value) for period, value in self.values)
        if len({period for period, _ in normalized}) != len(normalized):
            raise ValueError("scenario forecast contains duplicate periods")
        object.__setattr__(self, "values", normalized)


@dataclass(frozen=True)
class ModelLine:
    """Canonical historical and scenario series assigned to one model sheet."""

    canonical_key: str
    label: str
    sheet: str
    unit: str
    historical: tuple[HistoricalSourceCell, ...]
    scenario_forecasts: tuple[ScenarioForecast, ...]

    def __post_init__(self) -> None:
        for field_name in ("canonical_key", "label", "sheet", "unit"):
            object.__setattr__(self, field_name, _clean_text(getattr(self, field_name), field_name))
        if self.sheet not in MODEL_LINE_SHEETS:
            raise ValueError(f"unsupported model-line sheet: {self.sheet}")


@dataclass(frozen=True)
class ComparableCompany:
    """Optional source-normalized public-company comparable observation."""

    ticker: str
    company_name: str
    enterprise_value: float | None = None
    equity_value: float | None = None
    revenue: float | None = None
    ebitda: float | None = None
    net_income: float | None = None
    share_price: float | None = None

    def __post_init__(self) -> None:
        object.__setattr__(self, "ticker", _clean_text(self.ticker, "ticker").upper())
        object.__setattr__(self, "company_name", _clean_text(self.company_name, "company_name"))


@dataclass(frozen=True)
class SOTPComponent:
    """Optional source-normalized SOTP component; absent evidence stays typed unavailable."""

    component_key: str
    label: str
    metric: float
    multiple: float
    net_debt_allocation: float = 0.0


@dataclass(frozen=True)
class NormalizedProfessionalWorkbookPayload:
    """Complete deterministic input boundary for :func:`render_professional_model_workbook`.

    ``valuation_inputs`` accepts frozen numeric values for risk-free rate, beta,
    equity risk premium, pre-tax cost of debt, tax rate, debt value, equity
    value, terminal growth, net debt, diluted shares, current price, and an
    optional method-specific valuation diagnostics.  The renderer validates required scenario and
    period coverage but does not invent missing finance data.
    """

    ticker: str
    company_name: str
    as_of_date: date
    currency: str
    unit_convention: str
    source: SourceWorkbookRun
    historical_periods: tuple[str, ...]
    forecast_periods: tuple[str, ...]
    lines: tuple[ModelLine, ...]
    valuation_inputs: Mapping[str, float | int | None]
    availability: Mapping[str, TypedAvailability]
    comparables: tuple[ComparableCompany, ...] = ()
    consensus_snapshot: ConsensusSnapshot | None = None
    sotp_components: tuple[SOTPComponent, ...] = ()
    backend_checks: Mapping[str, float | int | str | None] = field(default_factory=dict)
    source_presentations: tuple[SourcePresentationRecord, ...] = ()
    driver_approvals: tuple[DriverApprovalRecord, ...] = ()
    warnings: tuple[str, ...] = ()
    current_price_source: str | None = None
    current_price_as_of: date | None = None
    valuation_date: date | None = None
    decision_context: Mapping[str, str] = field(default_factory=dict)
    blockers: tuple[str, ...] = ()

    def __post_init__(self) -> None:
        object.__setattr__(self, "ticker", _clean_text(self.ticker, "ticker").upper())
        object.__setattr__(self, "company_name", _clean_text(self.company_name, "company_name"))
        currency = _clean_text(self.currency, "currency").upper()
        if not re.fullmatch(r"[A-Z]{3}", currency):
            raise ValueError("currency must be a three-letter ISO code")
        object.__setattr__(self, "currency", currency)
        object.__setattr__(self, "unit_convention", _clean_text(self.unit_convention, "unit_convention"))
        historical = tuple(_clean_text(period, "historical period") for period in self.historical_periods)
        forecast = tuple(_clean_text(period, "forecast period") for period in self.forecast_periods)
        if not historical or len(set(historical)) != len(historical):
            raise ValueError("historical_periods must be non-empty and unique")
        if len(forecast) != 5 or len(set(forecast)) != 5:
            raise ValueError("forecast_periods must contain exactly five unique periods")
        object.__setattr__(self, "historical_periods", historical)
        object.__setattr__(self, "forecast_periods", forecast)
        if not self.lines:
            raise ValueError("at least one model line is required")
        keys = [line.canonical_key for line in self.lines]
        if len(set(keys)) != len(keys):
            raise ValueError("model line canonical keys must be unique")
        required_availability = {"segments", "consensus", "sotp"}
        if not required_availability <= set(self.availability):
            raise ValueError("availability must explicitly type segments, consensus, and sotp")
        consensus_state = self.availability["consensus"].status
        if consensus_state == "available" and self.consensus_snapshot is None:
            raise ValueError("available consensus state requires a typed ConsensusSnapshot")
        if self.consensus_snapshot is not None:
            if self.consensus_snapshot.ticker != self.ticker:
                raise ValueError("consensus snapshot ticker must match payload ticker")
            if consensus_state != "available":
                raise ValueError("a supplied consensus snapshot must be marked available")
        object.__setattr__(self, "valuation_inputs", dict(sorted(self.valuation_inputs.items())))
        object.__setattr__(self, "availability", dict(sorted(self.availability.items())))
        object.__setattr__(self, "backend_checks", dict(sorted(self.backend_checks.items())))
        source_ids = [record.source_id for record in self.source_presentations]
        if len(set(source_ids)) != len(source_ids):
            raise ValueError("source presentation IDs must be unique")
        object.__setattr__(
            self,
            "source_presentations",
            tuple(sorted(self.source_presentations, key=lambda record: record.source_id)),
        )
        approval_ids = [
            (record.scenario_key, record.driver_key)
            for record in self.driver_approvals
        ]
        if len(set(approval_ids)) != len(approval_ids):
            raise ValueError("driver approval scenario/key identities must be unique")
        object.__setattr__(
            self,
            "driver_approvals",
            tuple(
                sorted(
                    self.driver_approvals,
                    key=lambda record: (record.scenario_key or "", record.driver_key),
                )
            ),
        )
        if self.valuation_date is not None and self.valuation_date < self.as_of_date:
            raise ValueError("valuation_date cannot precede the financial cutoff date")
        if self.current_price_source is not None:
            object.__setattr__(
                self,
                "current_price_source",
                _clean_text(self.current_price_source, "current_price_source"),
            )
        object.__setattr__(
            self,
            "decision_context",
            {
                _clean_text(str(key), "decision_context key"): _clean_text(
                    str(value), "decision_context value"
                )
                for key, value in sorted(self.decision_context.items())
            },
        )
        object.__setattr__(self, "warnings", tuple(sorted(set(self.warnings))))
        object.__setattr__(self, "blockers", tuple(sorted(set(self.blockers))))


class _Renderer:
    def __init__(self, payload: NormalizedProfessionalWorkbookPayload) -> None:
        self.payload = payload
        self.model_input_hash = _canonical_hash(asdict(payload))
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        for sheet_name in REQUIRED_SHEET_ORDER:
            self.workbook.create_sheet(sheet_name)
        self.line_mappings: list[LineCellMapping] = []
        self.classifications: dict[tuple[str, str], CellClassification] = {}
        self.defined_names: dict[str, DefinedNameMapping] = {}
        self.check_cells: dict[str, WorkbookCheckCell] = {}
        self.historical_cells: dict[tuple[str, str], tuple[str, str]] = {}
        self.scenario_cells: dict[tuple[str, str, str], tuple[str, str]] = {}
        self.output_cells: dict[tuple[str, str, str], tuple[str, str]] = {}
        self.assumption_cells: dict[str, tuple[str, str]] = {}
        self.dcf_result_cells: dict[str, tuple[str, str]] = {}
        self.wacc_state = "BLOCKED"
        self.dcf_state = "BLOCKED"
        self.comps_state = "UNAVAILABLE"
        self.dcf_decision_state = "UNVERIFIED"
        self.blockers = set(payload.blockers)
        self.warnings = set(payload.warnings)
        self.lines = tuple(
            sorted(
                payload.lines,
                key=lambda line: (REQUIRED_SHEET_ORDER.index(line.sheet), line.canonical_key),
            )
        )

    def validate(self) -> None:
        expected_history = self.payload.historical_periods
        expected_forecast = self.payload.forecast_periods
        for line in self.lines:
            history_keys = tuple(item.period_key for item in line.historical)
            if history_keys != expected_history:
                raise ValueError(
                    f"{line.canonical_key} historical periods must exactly match payload historical_periods"
                )
            scenario_map = {item.scenario_key: item for item in line.scenario_forecasts}
            if set(scenario_map) != set(REQUIRED_SCENARIOS):
                raise ValueError(
                    f"{line.canonical_key} must provide base, upside, and downside forecasts"
                )
            for scenario in REQUIRED_SCENARIOS:
                forecast_keys = tuple(period for period, _ in scenario_map[scenario].values)
                if forecast_keys != expected_forecast:
                    raise ValueError(
                        f"{line.canonical_key}/{scenario} forecast periods must exactly match payload forecast_periods"
                    )
        if self.payload.source.status not in {"ready", "ok"} or self.payload.source.formula_error_count:
            self.blockers.add("source_preflight_blocked")
        supplied_line_keys = {line.canonical_key for line in self.lines}
        missing_dcf = REQUIRED_DCF_LINES - supplied_line_keys
        if missing_dcf:
            self.blockers.add("missing_dcf_inputs:" + ",".join(sorted(missing_dcf)))
        missing_checks = {
            "total_assets",
            "total_liabilities_and_equity",
            "cash",
            "cf.net_change_cash",
        } - supplied_line_keys
        if missing_checks:
            self.blockers.add("missing_check_inputs:" + ",".join(sorted(missing_checks)))
        valuation_values = self.payload.valuation_inputs
        if self.payload.valuation_date is None:
            self.blockers.add("valuation_date_explicit_absent")
        for blocker_prefix, required_inputs in (
            ("missing_wacc_inputs", WACC_REQUIRED_INPUTS),
            ("missing_dcf_valuation_inputs", DCF_REQUIRED_VALUATION_INPUTS),
            ("missing_bridge_reconciliation_inputs", BRIDGE_RECONCILIATION_INPUTS),
            ("missing_price_comparison_inputs", PRICE_COMPARISON_INPUTS),
        ):
            missing = {
                key
                for key in required_inputs
                if key not in valuation_values or valuation_values[key] is None
            }
            if missing:
                self.blockers.add(f"{blocker_prefix}:" + ",".join(sorted(missing)))
        if self.payload.availability["segments"].status != "available":
            self.warnings.add("segments_unavailable")
        if self.payload.availability["consensus"].status != "available":
            self.warnings.add("consensus_unavailable")
        if self.payload.availability["sotp"].status != "available":
            self.warnings.add("sotp_unavailable")

    def set_cell(
        self,
        sheet: str,
        coordinate: str,
        value: Any,
        kind: CellKind,
        *,
        number_format: str | None = None,
        fill: PatternFill | None = None,
        font: Font | None = None,
        alignment: Alignment | None = None,
        border: Border | None = None,
        comment: str | None = None,
    ) -> None:
        cell = self.workbook[sheet][coordinate]
        cell.value = value
        if number_format:
            cell.number_format = number_format
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        else:
            color = {
                CellKind.EDITABLE: INPUT_BLUE,
                CellKind.SOURCE: SOURCE_GREEN,
                CellKind.FORMULA: FORMULA_BLACK,
                CellKind.CHECK: FORMULA_BLACK,
                CellKind.UNAVAILABLE: GRAY,
            }.get(kind, FORMULA_BLACK)
            cell.font = Font(name="Arial", size=10, color=color)
        cell.alignment = alignment or Alignment(vertical="center")
        if border:
            cell.border = border
        if comment:
            cell.comment = Comment(comment, "ai-fund")
        self.classifications[(sheet, coordinate)] = CellClassification(
            sheet=sheet,
            cell=coordinate,
            kind=kind,
        )

    def set_formula(
        self,
        sheet: str,
        coordinate: str,
        formula: str,
        *,
        kind: CellKind = CellKind.FORMULA,
        number_format: str | None = None,
        comment: str | None = None,
    ) -> None:
        if not formula.startswith("="):
            raise ValueError("Excel formula must start with =")
        if any(token in formula for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!")):
            raise ValueError(f"formula contains an error literal: {formula}")
        if "[" in formula or "]" in formula:
            raise ValueError(f"external-link syntax is prohibited: {formula}")
        self.set_cell(
            sheet,
            coordinate,
            formula,
            kind,
            number_format=number_format,
            font=Font(name="Arial", size=10, color=FORMULA_BLACK),
            comment=comment,
        )

    def title(self, sheet: str, title: str, subtitle: str | None = None) -> None:
        worksheet = self.workbook[sheet]
        worksheet.merge_cells("A1:L1")
        self.set_cell(
            sheet,
            "A1",
            title,
            CellKind.STATIC,
            fill=PatternFill("solid", fgColor=NAVY),
            font=Font(name="Arial", size=15, bold=True, color=WHITE),
            alignment=Alignment(vertical="center"),
        )
        worksheet.row_dimensions[1].height = 24
        if subtitle:
            worksheet.merge_cells("A2:L2")
            self.set_cell(
                sheet,
                "A2",
                subtitle,
                CellKind.STATIC,
                font=Font(name="Arial", size=9, italic=True, color=GRAY),
            )

    def header_row(self, sheet: str, row: int, values: Sequence[str], start_col: int = 1) -> None:
        for offset, value in enumerate(values):
            coordinate = f"{get_column_letter(start_col + offset)}{row}"
            self.set_cell(
                sheet,
                coordinate,
                value,
                CellKind.STATIC,
                fill=PatternFill("solid", fgColor=MEDIUM_BLUE),
                font=Font(name="Arial", size=9, bold=True, color=WHITE),
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                border=Border(bottom=THIN_GRAY),
            )

    def add_name(self, name: str, sheet: str, coordinate: str) -> None:
        definition = DefinedName(name, attr_text=_absolute_ref(sheet, coordinate))
        self.workbook.defined_names.add(definition)
        self.defined_names[name] = DefinedNameMapping(name=name, sheet=sheet, cell=coordinate)

    def add_check(self, check_id: str, sheet: str, coordinate: str) -> None:
        if check_id in self.check_cells:
            raise ValueError(f"duplicate workbook check ID: {check_id}")
        self.check_cells[check_id] = WorkbookCheckCell(check_id=check_id, sheet=sheet, cell=coordinate)

    def finish_sheet(self, sheet: str, freeze: str = "C5") -> None:
        worksheet = self.workbook[sheet]
        worksheet.freeze_panes = freeze
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5
        worksheet.print_title_rows = "1:4"
        worksheet.column_dimensions["A"].width = max(18, worksheet.column_dimensions["A"].width or 0)
        worksheet.column_dimensions["B"].width = max(30, worksheet.column_dimensions["B"].width or 0)


    def render_cover(self) -> None:
        sheet = "Cover"
        self.title(
            sheet,
            f"{self.payload.company_name} ({self.payload.ticker})",
            "Professional integrated financial model | evidence and decision gates",
        )
        labels = (
            (3, "Ticker", self.payload.ticker),
            (4, "Company", self.payload.company_name),
            (5, "Model as of", self.payload.as_of_date),
            (6, "Currency / units", f"{self.payload.currency} | {self.payload.unit_convention}"),
            (7, "Source workbook", self.payload.source.source_file),
            (8, "Source preflight", self.payload.source.status.upper()),
            (9, "Source SHA-256", self.payload.source.source_hash),
            (10, "Selected CIQ run", self.payload.source.run_id),
        )
        for row, label, value in labels:
            self.set_cell(
                sheet,
                f"A{row}",
                label,
                CellKind.STATIC,
                font=Font(name="Arial", size=10, bold=True, color=NAVY),
            )
            self.set_cell(sheet, f"B{row}", value, CellKind.SOURCE if row >= 7 else CellKind.STATIC)

        status_rows = (
            (11, "Package workflow state", "=Checks!$C$5", True),
            (12, "Calculation verification", "UNVERIFIED", False),
            (13, "Verification authority", "External SHA-bound verification sidecar; never inferred from nonblank cells", False),
            (14, "Model input hash", self.model_input_hash, False),
            (15, "Current price", self.payload.valuation_inputs.get("current_price"), False),
            (16, "Current price source", self.payload.current_price_source or "UNVERIFIED", False),
            (17, "Current price as of", self.payload.current_price_as_of or "UNVERIFIED", False),
        )
        for row, label, value, is_formula in status_rows:
            self.set_cell(
                sheet,
                f"A{row}",
                label,
                CellKind.STATIC,
                font=Font(name="Arial", size=10, bold=True, color=NAVY),
            )
            if is_formula:
                self.set_formula(sheet, f"B{row}", str(value), kind=CellKind.CHECK)
            else:
                kind = CellKind.SOURCE if row in {15, 16, 17} else CellKind.CHECK
                self.set_cell(
                    sheet,
                    f"B{row}",
                    value,
                    kind,
                    number_format="$0.00" if row == 15 else None,
                    comment=(
                        "UNVERIFIED is fail-closed. Only the isolated recalculation verifier may issue a SHA-bound verification record."
                        if row == 12
                        else None
                    ),
                )
        self.set_cell(sheet, "A18", "Hard blockers", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, "B18", '=COUNTIF(Checks!$C$6:$C$17,"BLOCKED")+COUNTIF(Checks!$C$6:$C$17,"FAIL")', kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "A19", "PM review gates", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, "B19", '=COUNTIF(Checks!$C$6:$C$17,"NEEDS_PM_REVIEW")', kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "A20", "Partial modules", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, "B20", '=COUNTIF(Checks!$C$6:$C$17,"PARTIAL")', kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "A22", "Interpretation", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "B22",
            "FULL requires explicit positive evidence for every required gate. Blank, unknown, degraded, stale, and unverified states never pass. Method availability is separate from decision eligibility.",
            CellKind.STATIC,
            alignment=Alignment(wrap_text=True, vertical="top"),
        )
        self.header_row(sheet, 3, ("Navigation", "Open"), start_col=4)
        for row, (label, target) in enumerate(
            (("Executive summary", "Summary"), ("Integrity checks", "Checks"), ("PM review queue", "PM_Review_Queue"), ("DCF", "DCF"), ("Valuation", "Valuation"), ("Sources", "Sources")),
            start=4,

        ):
            self.set_cell(sheet, f"D{row}", label, CellKind.STATIC)
            self.set_cell(sheet, f"E{row}", "OPEN", CellKind.STATIC, font=Font(name="Arial", size=10, color=MEDIUM_BLUE, underline="single"))
            self.workbook[sheet][f"E{row}"].hyperlink = f"#'{target}'!A1"
        self.add_name("Calculation_Status", sheet, "B12")
        self.add_name("Source_Hash", sheet, "B9")
        self.add_name("Source_Run_ID", sheet, "B10")
        self.workbook[sheet].column_dimensions["A"].width = 25
        self.workbook[sheet].column_dimensions["B"].width = 88
        self.workbook[sheet].column_dimensions["D"].width = 24
        self.workbook[sheet].column_dimensions["E"].width = 14
        self.finish_sheet(sheet, "A3")

    def render_sources(self) -> None:
        sheet = "Sources"
        self.title(sheet, "Source register", "Exact workbook/run/cell lineage; formula text is preserved as text, never executed here")
        metadata = (
            ("Source file", self.payload.source.source_file),
            ("Source path", self.payload.source.source_path),
            ("Source hash", self.payload.source.source_hash),
            ("Run ID", self.payload.source.run_id),
            ("Parser", self.payload.source.parser_version),
            ("Preflight status", self.payload.source.status.upper()),
            ("Facts", self.payload.source.fact_count),
            ("Formula errors", self.payload.source.formula_error_count),
        )
        for index, (label, value) in enumerate(metadata, start=3):
            self.set_cell(sheet, f"A{index}", label, CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
            self.set_cell(sheet, f"B{index}", value, CellKind.SOURCE)

        table_row = 13
        headers = (
            "Canonical key", "Period", "Raw source value", "Normalized value", "Derived value",
            "Final cached value", "Transformation / normalization", "Unit", "Unit kind", "Scale",
            "Currency", "Period type", "Period end", "Source sheet", "Source cell", "Source row ID",
            "Source formula (text)", "Formula status", "Formula error", "Upstream dependencies",
            "Downstream dependencies", "Run ID", "Source hash",
        )
        self.header_row(sheet, table_row, headers)
        row = table_row + 1
        for line in self.lines:
            for fact in line.historical:
                values: tuple[Any, ...] = (
                    line.canonical_key,
                    fact.period_key,
                    fact.raw_value,
                    fact.normalized_value,
                    fact.derived_value,
                    fact.value,
                    fact.transformation_rule,
                    fact.unit,
                    fact.unit_kind,
                    fact.scale,
                    fact.currency,
                    fact.period_type,
                    fact.period_end,
                    fact.source_sheet,
                    fact.source_cell,
                    fact.source_row_id,
                    ("'" + fact.source_formula) if fact.source_formula else None,
                    fact.formula_status,
                    fact.formula_error,
                    ", ".join(fact.upstream_dependencies),
                    ", ".join(fact.downstream_dependencies),
                    self.payload.source.run_id,
                    self.payload.source.source_hash,
                )
                for column, value in enumerate(values, start=1):
                    kind = CellKind.CHECK if column == 19 and value else CellKind.SOURCE
                    self.set_cell(
                        sheet,
                        f"{get_column_letter(column)}{row}",
                        value,
                        kind,
                        number_format="#,##0.0;[Red](#,##0.0);-" if 3 <= column <= 6 else None,
                    )
                row += 1
        for record in self.payload.source_presentations:
            presented_value = (
                record.derived_value
                if record.derived_value is not None
                else record.normalized_value
                if record.normalized_value is not None
                else record.raw_value
            )
            record_period = record.as_of_date or record.period_end
            values = (
                record.canonical_key,
                record_period.isoformat() if record_period else "NONE",
                record.raw_value,
                record.normalized_value,
                record.derived_value,
                presented_value,
                record.transform,
                record.unit,
                record.unit_kind.value,
                record.scale,
                record.currency,
                record.period_type.value,
                record_period,
                "; ".join(record.source_refs),
                None,
                record.source_id,
                None,
                record.formula_status.value,
                (
                    f"{record.error_code}: {record.error_message}"
                    if record.error_code
                    else None
                ),
                None,
                ", ".join(record.downstream_dependencies),
                self.payload.source.run_id,
                self.payload.source.source_hash,
            )
            for column, value in enumerate(values, start=1):
                kind = CellKind.CHECK if column == 19 and value else CellKind.SOURCE
                self.set_cell(sheet, f"{get_column_letter(column)}{row}", value, kind)
            row += 1
        worksheet = self.workbook[sheet]
        worksheet.auto_filter.ref = f"A{table_row}:W{max(table_row, row - 1)}"
        widths = {
            "A": 34, "B": 12, "C": 18, "D": 18, "E": 18, "F": 18, "G": 42,
            "H": 16, "I": 14, "J": 12, "K": 12, "L": 15, "M": 14, "N": 24,
            "O": 14, "P": 28, "Q": 44, "R": 18, "S": 28, "T": 38, "U": 38,
            "V": 12, "W": 68,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, f"A{table_row + 1}")

    def render_historical_data(self) -> None:
        sheet = "Historical_Data"
        self.title(sheet, "Historical data", "Green-font values are immutable source observations; comments carry exact locators")
        headers = ("Canonical key", "Display label", *self.payload.historical_periods, "Model sheet", "Unit")
        self.header_row(sheet, 4, headers)
        for row, line in enumerate(self.lines, start=5):
            self.set_cell(sheet, f"A{row}", line.canonical_key, CellKind.STATIC)
            self.set_cell(sheet, f"B{row}", line.label, CellKind.STATIC)
            for period_index, fact in enumerate(line.historical, start=3):
                coordinate = f"{get_column_letter(period_index)}{row}"
                locator = (
                    f"{fact.source_sheet}!{fact.source_cell}"
                    if fact.source_sheet and fact.source_cell
                    else "NO SOURCE LOCATOR: typed unavailable"
                )
                lineage_bits = [
                    f"Source: {self.payload.source.source_file}",
                    locator,
                    f"Run {self.payload.source.run_id}",
                    f"SHA-256 {self.payload.source.source_hash}",
                    f"Row ID {fact.source_row_id}",
                    f"Period {fact.period_key} | type {fact.period_type or 'UNVERIFIED'} | end {fact.period_end or 'UNVERIFIED'}",
                    f"Unit {fact.unit or 'UNVERIFIED'} | kind {fact.unit_kind or 'UNVERIFIED'} | scale {fact.scale if fact.scale is not None else 'UNVERIFIED'} | currency {fact.currency or 'UNVERIFIED'}",
                    f"Raw {fact.raw_value!r}",
                    f"Normalized {fact.normalized_value!r}",
                    f"Derived {fact.derived_value!r}",
                    f"Transform {fact.transformation_rule or 'UNVERIFIED'}",
                    f"Formula status {fact.formula_status}",
                    f"Formula error {fact.formula_error or 'None'}",
                    f"Upstream {', '.join(fact.upstream_dependencies) or 'None'}",
                    f"Downstream {', '.join(fact.downstream_dependencies) or 'None'}",
                ]
                if fact.source_formula:
                    lineage_bits.append(f"Source formula {fact.source_formula}")
                comment = " | ".join(lineage_bits)
                self.set_cell(
                    sheet, coordinate, fact.value, CellKind.SOURCE,
                    number_format="#,##0.0;[Red](#,##0.0);-", fill=PatternFill("solid", fgColor=PALE_GREEN), comment=comment,
                )
                self.historical_cells[(line.canonical_key, fact.period_key)] = (sheet, coordinate)
            trailing = 3 + len(self.payload.historical_periods)
            self.set_cell(sheet, f"{get_column_letter(trailing)}{row}", line.sheet, CellKind.STATIC)
            self.set_cell(sheet, f"{get_column_letter(trailing + 1)}{row}", line.unit, CellKind.STATIC)
        worksheet = self.workbook[sheet]
        worksheet.auto_filter.ref = f"A4:{get_column_letter(4 + len(self.payload.historical_periods))}{4 + len(self.lines)}"
        worksheet.column_dimensions["A"].width = 38
        worksheet.column_dimensions["B"].width = 34
        for index in range(3, 3 + len(self.payload.historical_periods)):
            worksheet.column_dimensions[get_column_letter(index)].width = 14
        self.finish_sheet(sheet, "C5")
    def render_assumptions(self) -> None:
        sheet = "Assumptions"
        self.title(sheet, "Assumptions and valuation evidence", "Blue cells are review inputs; green cells are frozen bridge/diagnostic evidence and are never blended")
        self.set_cell(sheet, "A4", "Selected scenario", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B4", "base", CellKind.EDITABLE, fill=PatternFill("solid", fgColor=PALE_YELLOW), font=Font(name="Arial", size=10, bold=True, color=INPUT_BLUE))
        validation = DataValidation(type="list", formula1='"base,upside,downside"', allow_blank=False)
        self.workbook[sheet].add_data_validation(validation)
        validation.add(self.workbook[sheet]["B4"])
        self.add_name("Selected_Scenario", sheet, "B4")
        self.set_cell(sheet, "A5", "Source status", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B5", self.payload.source.status.upper(), CellKind.SOURCE)

        self.header_row(sheet, 7, ("Input / evidence key", "Value", "Units / interpretation", "Evidence state"))
        pct_keys = {"risk_free_rate", "equity_risk_premium", "pre_tax_cost_of_debt", "tax_rate", "terminal_growth", "size_premium", "backend_selected_wacc", "backend_cost_of_equity", "backend_debt_weight", "backend_equity_weight"}
        for row, key in enumerate(sorted(self.payload.valuation_inputs), start=8):
            value = self.payload.valuation_inputs[key]
            self.set_cell(sheet, f"A{row}", key, CellKind.STATIC)
            is_diagnostic = key.startswith("diagnostic_")
            is_bridge = key.startswith("bridge_")
            is_frozen_evidence = is_diagnostic or is_bridge or key.startswith("dcf_") or key in {"net_debt", "current_price", "current_basic_shares", "current_fully_diluted_shares"}
            if key in pct_keys or key.startswith("dcf_wacc_") or key.startswith("dcf_terminal_growth_"):
                number_format = "0.0%"
                units = "%"
            elif key == "diagnostic_reverse_dcf_implied_growth_pct_points":
                number_format = '0.0"%"'
                units = "percentage points (raw, not decimal)"
            elif key.endswith("_per_share") or key == "current_price":
                number_format = "$0.00"
                units = f"{self.payload.currency}/share"
            elif key == "beta":
                number_format = "0.00x"
                units = "x"
            elif key in {"current_basic_shares", "current_fully_diluted_shares"}:
                number_format = "#,##0.0"
                units = "mm shares"
            else:
                number_format = "#,##0.0;[Red](#,##0.0);-"
                units = self.payload.unit_convention
            if value is None:
                kind = CellKind.UNAVAILABLE
                fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                evidence_state = "UNAVAILABLE"
            elif is_frozen_evidence:
                kind = CellKind.SOURCE
                fill = PatternFill("solid", fgColor=PALE_GREEN)
                evidence_state = "DIAGNOSTIC / NON-APPROVED" if is_diagnostic else "FROZEN SOURCE EVIDENCE"
            else:
                kind = CellKind.EDITABLE
                fill = PatternFill("solid", fgColor=PALE_YELLOW)
                evidence_state = "AVAILABLE"
            evidence_comment = (
                f"Source: {self.payload.current_price_source or 'UNVERIFIED'} | "
                f"As of: {self.payload.current_price_as_of or 'UNVERIFIED'}"
                if key == "current_price"
                else None
            )
            self.set_cell(sheet, f"B{row}", value, kind, number_format=number_format, fill=fill, comment=evidence_comment)
            interpretation = units
            if key == "net_debt":
                interpretation += " | legacy reference only; DCF uses component bridge"
            self.set_cell(sheet, f"C{row}", interpretation, CellKind.STATIC)
            self.set_cell(sheet, f"D{row}", evidence_state, CellKind.STATIC if value is not None else CellKind.UNAVAILABLE)
            self.assumption_cells[key] = (sheet, f"B{row}")
        self.workbook[sheet].column_dimensions["A"].width = 50
        self.workbook[sheet].column_dimensions["B"].width = 20
        self.workbook[sheet].column_dimensions["C"].width = 54
        self.workbook[sheet].column_dimensions["D"].width = 28
        self.finish_sheet(sheet, "A8")
    def render_scenarios(self) -> None:
        sheet = "Scenarios"
        self.title(
            sheet,
            "Scenario forecast outputs",
            "Deterministic engine outputs are calculated diagnostics, not source facts; driver fingerprints and PM gates control decision use.",
        )
        pm_required = any(
            item.startswith(("pm_approval_required:", "source_or_pm_required:"))
            for item in self.blockers
        )
        policy_pass = self.payload.backend_checks.get("scenario.policy_gate") == "PASS"
        # This sheet currently stores immutable backend-output evidence. It is not
        # a formula-first driver matrix, so metadata may not assert structural PASS.
        self.blockers.add("scenario_formula_first_architecture_not_implemented")
        scenario_control_gate = "BLOCKED"
        scenario_state = "NEEDS_PM_REVIEW" if pm_required or not policy_pass else "BLOCKED"
        self.set_cell(sheet, "A3", "Output state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "B3",
            scenario_state,
            CellKind.CHECK,
            fill=PatternFill("solid", fgColor=PALE_YELLOW),
        )
        self.set_cell(sheet, "D3", "Formula-first control gate", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "E3", scenario_control_gate, CellKind.CHECK, fill=PatternFill("solid", fgColor=PALE_GREEN if scenario_control_gate == "PASS" else PALE_RED))
        self.add_name("Scenario_Model_Gate", sheet, "E3")
        headers = ("Scenario", "Canonical key", "Display label", "Model sheet", "Unit", *self.payload.forecast_periods)
        self.header_row(sheet, 4, headers)
        row = 5
        for scenario in REQUIRED_SCENARIOS:
            for line in self.lines:
                scenario_data = next(item for item in line.scenario_forecasts if item.scenario_key == scenario)
                self.set_cell(sheet, f"A{row}", scenario.upper(), CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
                self.set_cell(sheet, f"B{row}", line.canonical_key, CellKind.STATIC)
                self.set_cell(sheet, f"C{row}", line.label, CellKind.STATIC)
                self.set_cell(sheet, f"D{row}", line.sheet, CellKind.STATIC)
                self.set_cell(sheet, f"E{row}", line.unit, CellKind.STATIC)
                for period_index, (period, value) in enumerate(scenario_data.values, start=6):
                    coordinate = f"{get_column_letter(period_index)}{row}"
                    self.set_cell(
                        sheet,
                        coordinate,
                        value,
                        CellKind.CHECK,
                        number_format="#,##0.0;[Red](#,##0.0);-",
                        fill=PatternFill("solid", fgColor=PALE_YELLOW),
                        comment=(
                            "Calculated backend output; not a source fact or approved decision input."
                            if period_index == 6
                            else None
                        ),
                    )
                    self.scenario_cells[(scenario, line.canonical_key, period)] = (sheet, coordinate)
                row += 1
        worksheet = self.workbook[sheet]
        worksheet.auto_filter.ref = f"A4:{get_column_letter(5 + len(self.payload.forecast_periods))}{row - 1}"
        for column, width in {"A": 13, "B": 38, "C": 32, "D": 24, "E": 18}.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, "F5")
    def availability_banner(self, sheet: str, state: TypedAvailability) -> None:
        self.set_cell(sheet, "A4", "Evidence state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        kind = CellKind.STATIC if state.status == "available" else CellKind.UNAVAILABLE
        self.set_cell(
            sheet,
            "B4",
            state.status.upper(),
            kind,
            fill=PatternFill("solid", fgColor=PALE_GREEN if state.status == "available" else LIGHT_GRAY),
            font=Font(name="Arial", size=10, bold=True, color=CHECK_GREEN if state.status == "available" else GRAY),
        )
        if state.status != "available":
            self.set_cell(sheet, "A5", "Reason", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
            self.set_cell(sheet, "B5", f"{state.reason_code}: {state.message}", CellKind.UNAVAILABLE, alignment=Alignment(wrap_text=True))

    def render_model_sheets(self) -> None:
        for sheet in MODEL_LINE_SHEETS:
            lines = tuple(line for line in self.lines if line.sheet == sheet)
            label = sheet.replace("_", " ")
            self.title(sheet, label, "Historical source links and scenario forecast formulas; change rows support analytical review")
            if sheet == "Segment_Build":
                state = self.payload.availability["segments"]
                # Source-backed segment lines may coexist with a typed state;
                # no lines are synthesized merely to make the sheet look full.
                self.availability_banner(sheet, state)
            else:
                state = TypedAvailability("available") if lines else TypedAvailability(
                    "unavailable", "normalized_lines_absent", f"No normalized line series were supplied for {label}."
                )
                self.availability_banner(sheet, state)
            if not lines:
                self.finish_sheet(sheet, "A4")
                continue

            periods = (*self.payload.historical_periods, *self.payload.forecast_periods)
            fiscal_actuals = tuple(
                period
                for period in self.payload.historical_periods
                if re.fullmatch(r"FY[0-9]{2,4}", period.upper())
            )
            annual_anchor_period = fiscal_actuals[-1] if fiscal_actuals else self.payload.historical_periods[-1]
            first_forecast_period = self.payload.forecast_periods[0]
            row = 7
            for scenario in REQUIRED_SCENARIOS:
                self.workbook[sheet].merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(periods))
                self.set_cell(
                    sheet,
                    f"A{row}",
                    f"{scenario.upper()} SCENARIO",
                    CellKind.STATIC,
                    fill=PatternFill("solid", fgColor=LIGHT_BLUE),
                    font=Font(name="Arial", size=10, bold=True, color=NAVY),
                )
                row += 1
                self.header_row(sheet, row, ("Canonical key", "Line item", *periods))
                row += 1
                for line in lines:
                    value_row = row
                    change_row = row + 1
                    self.set_cell(sheet, f"A{value_row}", line.canonical_key, CellKind.STATIC)
                    self.set_cell(sheet, f"B{value_row}", line.label, CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=FORMULA_BLACK))
                    self.set_cell(sheet, f"A{change_row}", "", CellKind.STATIC)
                    self.set_cell(sheet, f"B{change_row}", "YoY / change", CellKind.STATIC, font=Font(name="Arial", size=8, italic=True, color=GRAY))
                    for period_index, period in enumerate(periods, start=3):
                        coordinate = f"{get_column_letter(period_index)}{value_row}"
                        if period in self.payload.historical_periods:
                            source_sheet, source_cell = self.historical_cells[(line.canonical_key, period)]
                            reference = _absolute_ref(source_sheet, source_cell)
                        else:
                            input_sheet, input_cell = self.scenario_cells[(scenario, line.canonical_key, period)]
                            reference = _absolute_ref(input_sheet, input_cell)
                        formula = f'=IF(AND(Scenario_Model_Gate="PASS",ISNUMBER({reference})),{reference},"")'
                        # Exact source lineage lives on the immutable Historical_Data
                        # hardcodes and in the manifest's line-cell mappings. Repeating
                        # the same comment on every linked statement output materially
                        # increases native Excel load time without adding audit evidence.
                        self.set_formula(sheet, coordinate, formula, number_format="#,##0.0;[Red](#,##0.0);-")
                        self.output_cells[(scenario, line.canonical_key, period)] = (sheet, coordinate)
                        self.line_mappings.append(
                            LineCellMapping(
                                canonical_key=line.canonical_key,
                                scenario_key=scenario,
                                period_key=period,
                                sheet=sheet,
                                cell=coordinate,
                            )
                        )
                        if period_index > 3:
                            if period == first_forecast_period:
                                anchor_index = periods.index(annual_anchor_period) + 3
                                previous = f"{get_column_letter(anchor_index)}{value_row}"
                            else:
                                previous = f"{get_column_letter(period_index - 1)}{value_row}"
                            change_coordinate = f"{get_column_letter(period_index)}{change_row}"
                            self.set_formula(
                                sheet,
                                change_coordinate,
                                f'=IF(OR(NOT(ISNUMBER({coordinate})),NOT(ISNUMBER({previous})),{previous}=0),"",{coordinate}/{previous}-1)',
                                number_format="0.0%;[Red](0.0%);-",
                            )
                    row += 2
                row += 2
            worksheet = self.workbook[sheet]
            worksheet.column_dimensions["A"].width = 38
            worksheet.column_dimensions["B"].width = 32
            for index in range(3, 3 + len(periods)):
                worksheet.column_dimensions[get_column_letter(index)].width = 13
            self.finish_sheet(sheet, "C9")
    def assumption_ref(self, key: str) -> str | None:
        location = self.assumption_cells.get(key)
        return _absolute_ref(*location) if location else None

    def component_net_claims_expression(self) -> str | None:
        asset_refs = tuple(self.assumption_ref(key) for key in BRIDGE_ASSET_KEYS)
        claim_keys = (*BRIDGE_BORROWING_KEYS, *BRIDGE_LEASE_KEYS, *BRIDGE_OTHER_CLAIM_KEYS)
        claim_refs = tuple(self.assumption_ref(key) for key in claim_keys)
        if any(reference is None for reference in (*asset_refs, *claim_refs)):
            return None
        return f"({'+'.join(claim_refs)})-({'+'.join(asset_refs)})"

    def diagnostic_ref(self, key: str) -> str | None:
        if key not in {item[0] for item in DIAGNOSTIC_METHODS}:
            raise KeyError(key)
        if self.payload.valuation_inputs.get(key) is None:
            return None
        return self.assumption_ref(key)

    def output_ref(self, scenario: str, line_key: str, period: str) -> str | None:
        location = self.output_cells.get((scenario, line_key, period))
        return _absolute_ref(*location) if location else None

    def render_wacc(self) -> None:
        sheet = "WACC"
        self.title(sheet, "Weighted average cost of capital", "CAPM cost of equity and after-tax cost of debt")
        rows = {
            "risk_free_rate": 5,
            "beta": 6,
            "equity_risk_premium": 7,
            "debt_value": 9,
            "equity_value": 10,
            "pre_tax_cost_of_debt": 13,
            "tax_rate": 14,
        }
        source_ready = self.payload.source.status in {"ready", "ok"} and not self.payload.source.formula_error_count
        valuation_missing = tuple(sorted(item for item in self.blockers if item.startswith("missing_wacc_inputs")))
        degraded = tuple(sorted(item for item in self.blockers if item.startswith("wacc_degraded")))
        if valuation_missing:
            self.wacc_state = "BLOCKED"
        elif degraded:
            self.wacc_state = "NEEDS_PM_REVIEW"
        else:
            self.wacc_state = "UNVERIFIED"
        if valuation_missing or not source_ready:
            self.wacc_decision_state = "BLOCKED"
        elif degraded:
            self.wacc_decision_state = "NEEDS_PM_REVIEW"
        else:
            self.wacc_decision_state = "UNVERIFIED"
        availability_reasons = valuation_missing + degraded
        decision_reasons = (
            (() if source_ready else ("source_preflight_blocked_unproven_global",))
            + degraded
            + (("calculation_verification_unverified",) if self.wacc_decision_state == "UNVERIFIED" else ())
        )
        self.set_cell(sheet, "A3", "Input availability", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "B3",
            self.wacc_state,
            CellKind.CHECK,
            fill=PatternFill(
                "solid",
                fgColor=PALE_YELLOW if self.wacc_state in {"UNVERIFIED", "NEEDS_PM_REVIEW"} else PALE_RED,
            ),
        )
        self.set_cell(sheet, "D3", "Decision eligibility", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "E3",
            self.wacc_decision_state,
            CellKind.CHECK,
            fill=PatternFill(
                "solid",
                fgColor=PALE_YELLOW if self.wacc_decision_state in {"NEEDS_PM_REVIEW", "UNVERIFIED"} else PALE_RED,
            ),
        )
        self.set_cell(sheet, "A4", "Availability evidence", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        self.set_cell(sheet, "B4", "; ".join(availability_reasons) if availability_reasons else "Numeric inputs present", CellKind.CHECK if availability_reasons else CellKind.STATIC)
        self.set_cell(sheet, "D4", "Decision gate", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        self.set_cell(sheet, "E4", "; ".join(decision_reasons) if decision_reasons else "None", CellKind.CHECK if decision_reasons else CellKind.STATIC)
        labels = {
            "risk_free_rate": "Risk-free rate",
            "beta": "Levered beta",
            "equity_risk_premium": "Equity risk premium",
            "debt_value": "Book total debt incl. leases (market value unavailable)",
            "equity_value": "Market value of equity",
            "pre_tax_cost_of_debt": "Pre-tax cost of debt",
            "tax_rate": "Marginal tax rate",
        }
        for key, row in rows.items():
            self.set_cell(sheet, f"A{row}", labels[key], CellKind.STATIC)
            reference = self.assumption_ref(key)
            self.set_formula(sheet, f"B{row}", f"={reference}" if reference else '=""', number_format="0.0%" if key in {"risk_free_rate", "equity_risk_premium", "pre_tax_cost_of_debt", "tax_rate"} else "#,##0.0")
        self.set_cell(sheet, "A8", "Cost of equity", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        size_premium_ref = self.assumption_ref("size_premium")
        self.set_formula(sheet, "B8", f'=IF(COUNT(B5:B7,{size_premium_ref})=4,B5+B6*B7+{size_premium_ref},"")' if size_premium_ref else '=""', number_format="0.0%")
        self.set_cell(sheet, "A11", "Debt weight", CellKind.STATIC)
        self.set_formula(sheet, "B11", '=IFERROR(IF(AND(ISNUMBER(B9),ISNUMBER(B10),B9+B10<>0),B9/(B9+B10),""),"")', number_format="0.0%")
        self.set_cell(sheet, "A12", "Equity weight", CellKind.STATIC)
        self.set_formula(sheet, "B12", '=IF(B11="","",1-B11)', number_format="0.0%")
        self.set_cell(sheet, "A15", "After-tax cost of debt", CellKind.STATIC)
        self.set_formula(sheet, "B15", '=IF(COUNT(B13:B14)=2,B13*(1-B14),"")', number_format="0.0%")
        self.set_cell(sheet, "A16", "WACC", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY), fill=PatternFill("solid", fgColor=LIGHT_BLUE))
        self.set_formula(sheet, "B16", '=IF(COUNT(B8,B11,B12,B15)=4,B8*B12+B15*B11,"")', number_format="0.0%")
        self.workbook[sheet]["B16"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        self.workbook[sheet]["B16"].font = Font(name="Arial", size=11, bold=True, color=FORMULA_BLACK)
        self.add_name("WACC_Base", sheet, "B16")
        backend_wacc_ref = self.assumption_ref("backend_selected_wacc")
        self.set_cell(sheet, "D5", "Backend selected WACC", CellKind.STATIC)
        self.set_formula(sheet, "E5", f"={backend_wacc_ref}" if backend_wacc_ref else '=""', number_format="0.000%")
        self.set_cell(sheet, "D6", "Workbook minus backend (bps)", CellKind.STATIC)
        self.set_formula(sheet, "E6", '=IF(COUNT(B16,E5)=2,(B16-E5)*10000,"")', number_format="0.0")
        self.set_cell(sheet, "D7", "WACC parity <= 1 bp", CellKind.STATIC)
        self.set_formula(sheet, "E7", '=IF(NOT(ISNUMBER(E6)),"BLOCKED",IF(ABS(E6)<=1,"PASS","BLOCKED"))', kind=CellKind.CHECK)
        self.set_cell(sheet, "D8", "Method selection / debt basis / beta evidence require explicit PM approval", CellKind.STATIC, alignment=Alignment(wrap_text=True))
        self.workbook[sheet].column_dimensions["A"].width = 34
        self.workbook[sheet].column_dimensions["B"].width = 54
        self.finish_sheet(sheet, "A5")
    def render_dcf(self) -> None:
        sheet = "DCF"
        self.title(sheet, "Discounted cash flow valuation", "Integrated FCFF, terminal value, EV-to-equity bridge, and per-share value")
        required_available = REQUIRED_DCF_LINES <= {line.canonical_key for line in self.lines}
        line_index = {line.canonical_key: line for line in self.lines}
        required_values_available = required_available
        for line_key in REQUIRED_DCF_LINES:
            line = line_index.get(line_key)
            if line is None:
                required_values_available = False
                continue
            forecasts = {item.scenario_key: dict(item.values) for item in line.scenario_forecasts}
            for scenario in REQUIRED_SCENARIOS:
                values = forecasts.get(scenario, {})
                for period in self.payload.forecast_periods[1:]:
                    value = values.get(period)
                    if (
                        not isinstance(value, (int, float))
                        or isinstance(value, bool)
                        or not math.isfinite(float(value))
                    ):
                        required_values_available = False
        if not required_values_available:
            self.blockers.add("dcf_required_forecast_values_incomplete")
        shares_value = self.payload.valuation_inputs.get("current_fully_diluted_shares")
        if not isinstance(shares_value, (int, float)) or shares_value <= 0:
            self.blockers.add("current_fully_diluted_shares_invalid")
            required_values_available = False
        discount_exponents = tuple(
            self.payload.valuation_inputs.get(f"dcf_discount_exponent_{index}")
            for index in range(1, 6)
        )
        terminal_exponent = self.payload.valuation_inputs.get("dcf_terminal_discount_exponent")
        discount_timing_valid = (
            self.payload.valuation_date is not None
            and all(
                isinstance(value, (int, float))
                and not isinstance(value, bool)
                and math.isfinite(float(value))
                and float(value) > 0.0
                for value in discount_exponents
            )
            and tuple(float(value) for value in discount_exponents)
            == tuple(sorted(float(value) for value in discount_exponents))
            and isinstance(terminal_exponent, (int, float))
            and float(terminal_exponent) >= float(discount_exponents[-1])
        )
        if not discount_timing_valid:
            self.blockers.add("dcf_discount_timing_invalid_or_unbound")
            required_values_available = False
        nopat_tax_valid = all(
            isinstance(self.payload.valuation_inputs.get(f"dcf_nopat_tax_rate_{scenario}"), (int, float))
            and 0.0 <= float(self.payload.valuation_inputs[f"dcf_nopat_tax_rate_{scenario}"]) <= 1.0
            for scenario in REQUIRED_SCENARIOS
        )
        if not nopat_tax_valid:
            self.blockers.add("dcf_nopat_tax_governance_invalid")
            required_values_available = False
        valuation_available = not any(item.startswith("missing_dcf_valuation_inputs") for item in self.blockers)
        source_ready = self.payload.source.status in {"ready", "ok"} and not self.payload.source.formula_error_count
        scenario_control_pass = (
            self.payload.backend_checks.get("scenario.formula_first_gate") == "PASS"
            and self.payload.backend_checks.get("scenario.policy_gate") == "PASS"
        )
        wacc_gt_g_pass = all(
            isinstance(self.payload.valuation_inputs.get(f"dcf_wacc_{scenario}"), (int, float))
            and isinstance(self.payload.valuation_inputs.get(f"dcf_terminal_growth_{scenario}"), (int, float))
            and self.payload.valuation_inputs[f"dcf_wacc_{scenario}"]
            > self.payload.valuation_inputs[f"dcf_terminal_growth_{scenario}"]
            for scenario in REQUIRED_SCENARIOS
        )
        availability_blockers = tuple(
            sorted(
                tuple(item for item in self.blockers if item.startswith("missing_dcf_inputs"))
                + tuple(item for item in self.blockers if item.startswith("missing_dcf_valuation_inputs"))
                + tuple(
                    item
                    for item in self.blockers
                    if item in {
                        "dcf_required_forecast_values_incomplete",
                        "current_fully_diluted_shares_invalid",
                        "dcf_discount_timing_invalid_or_unbound",
                        "dcf_nopat_tax_governance_invalid",
                        "valuation_date_explicit_absent",
                    }
                )
                + (("wacc_inputs_blocked",) if self.wacc_state == "BLOCKED" else ())
            )
        )
        policy_blockers = tuple(
            sorted(
                item
                for item in self.blockers
                if item.startswith(("wacc_degraded", "pm_approval_required:", "source_or_pm_required:"))
            )
        )
        mathematical_ready = (
            required_values_available
            and valuation_available
            and self.wacc_state != "BLOCKED"
        )
        if not mathematical_ready:
            self.dcf_state = "BLOCKED"
        elif policy_blockers or self.wacc_state in {"DEGRADED", "NEEDS_PM_REVIEW"}:
            self.dcf_state = "NEEDS_PM_REVIEW"
        else:
            self.dcf_state = "UNVERIFIED"
        calculation_preconditions_pass = (
            mathematical_ready
            and source_ready
            and scenario_control_pass
            and wacc_gt_g_pass
            and not policy_blockers
        )
        self.dcf_calculation_gate = "UNVERIFIED" if calculation_preconditions_pass else "BLOCKED"
        if not mathematical_ready or not source_ready:
            self.dcf_decision_state = "BLOCKED"
        elif policy_blockers or self.wacc_state in {"DEGRADED", "NEEDS_PM_REVIEW"}:
            self.dcf_decision_state = "NEEDS_PM_REVIEW"
        else:
            self.dcf_decision_state = "UNVERIFIED"
        eligibility_reasons = tuple(
            sorted(
                (() if source_ready else ("source_preflight_blocked_unproven_global",))
                + policy_blockers
                + (
                    ("calculation_verification_unverified",)
                    if self.dcf_decision_state == "UNVERIFIED"
                    else ()
                )
            )
        )
        self.set_cell(sheet, "A3", "Method availability", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "B3",
            self.dcf_state,
            CellKind.CHECK,
            fill=PatternFill(
                "solid",
                fgColor=PALE_YELLOW if self.dcf_state in {"UNVERIFIED", "NEEDS_PM_REVIEW"} else PALE_RED,
            ),
        )
        self.set_cell(sheet, "D3", "Decision eligibility", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "E3",
            self.dcf_decision_state,
            CellKind.CHECK,
            fill=PatternFill(
                "solid",
                fgColor=PALE_YELLOW if self.dcf_decision_state in {"NEEDS_PM_REVIEW", "UNVERIFIED"} else PALE_RED,
            ),
        )
        self.set_cell(sheet, "A4", "Availability gate", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        self.set_cell(sheet, "B4", "; ".join(availability_blockers) if availability_blockers else "All mathematical inputs present", CellKind.CHECK if availability_blockers else CellKind.STATIC)
        self.set_cell(sheet, "G3", "Calculation gate", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        if calculation_preconditions_pass:
            self.set_formula(
                sheet,
                "H3",
                '=IF(AND(WACC!$E$7="PASS",Scenario_Model_Gate="PASS"),"PASS","BLOCKED")',
                kind=CellKind.CHECK,
            )
        else:
            self.set_cell(sheet, "H3", "BLOCKED", CellKind.CHECK, fill=PatternFill("solid", fgColor=PALE_RED))
        self.add_name("DCF_Calculation_Gate", sheet, "H3")
        self.set_cell(sheet, "G4", "Calculation evidence", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        calculation_reasons = (
            availability_blockers
            + policy_blockers
            + (() if source_ready else ("source_preflight",))
            + (() if scenario_control_pass else ("scenario_formula_or_policy_gate",))
            + (() if wacc_gt_g_pass else ("wacc_must_exceed_terminal_growth",))
            + (("wacc_parity_formula_pending",) if calculation_preconditions_pass else ())
        )
        self.set_cell(sheet, "H4", "; ".join(calculation_reasons) if calculation_reasons else "All required calculation gates PASS", CellKind.CHECK if calculation_reasons else CellKind.STATIC)
        self.set_cell(sheet, "D4", "Eligibility gate", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        self.set_cell(sheet, "E4", "; ".join(eligibility_reasons) if eligibility_reasons else "None", CellKind.CHECK if eligibility_reasons else CellKind.STATIC)
        starts = {"base": 5, "upside": 35, "downside": 65}
        forecast = self.payload.forecast_periods
        dcf_period_labels = ("Q4 FY26 Stub", *forecast[1:])
        for scenario in REQUIRED_SCENARIOS:
            start = starts[scenario]
            self.workbook[sheet].merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
            self.set_cell(sheet, f"A{start}", f"{scenario.upper()} SCENARIO", CellKind.STATIC,
                          fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="Arial", size=10, bold=True, color=NAVY))
            self.header_row(sheet, start + 1, ("Line item", "Units", *dcf_period_labels))
            labels = (
                (start + 2, "Revenue", self.payload.unit_convention),
                (start + 3, "EBIT", self.payload.unit_convention),
                (start + 4, "Tax rate", "%"),
                (start + 5, "NOPAT", self.payload.unit_convention),
                (start + 6, "D&A", self.payload.unit_convention),
                (start + 7, "Capital expenditures", self.payload.unit_convention),
                (start + 8, "Change in net working capital", self.payload.unit_convention),
                (start + 9, "Integrated unlevered free cash flow", self.payload.unit_convention),
                (start + 10, "ACT/365 midpoint exponent", "years"),
                (start + 11, "Discount factor", "x"),
                (start + 12, "Present value of integrated FCFF", self.payload.unit_convention),
                (start + 13, "Terminal growth", "%"),
                (start + 14, "Terminal value", self.payload.unit_convention),
                (start + 15, "Present value of terminal value", self.payload.unit_convention),
                (start + 16, "Enterprise value", self.payload.unit_convention),
                (start + 17, "Less: component-based net claims", self.payload.unit_convention),
                (start + 18, "Equity value", self.payload.unit_convention),
                (start + 19, f"Current fully diluted shares (as of {self.payload.as_of_date.isoformat()})", "mm"),
                (start + 20, "Implied value per share", f"{self.payload.currency}/share"),
                (start + 21, "Current price", f"{self.payload.currency}/share"),
                (start + 22, "Upside / (downside)", "%"),
                (start + 23, "Narrow-form FCFF reference", self.payload.unit_convention),
                (start + 24, "Integrated less narrow-form reference", self.payload.unit_convention),
                (start + 25, "Discount timing convention", "text"),
                (start + 26, "PV of terminal value / enterprise value", "%"),
                (start + 27, "Full FY26 FCFF reference", self.payload.unit_convention),
                (start + 28, "Nine-month FY26 FCFF actual", self.payload.unit_convention),
                (start + 29, "FY26 annual less YTD less Q4 stub", self.payload.unit_convention),
            )
            for row, label, unit in labels:
                bold = row in {start + 9, start + 16, start + 18, start + 20, start + 22}
                self.set_cell(sheet, f"A{row}", label, CellKind.STATIC, font=Font(name="Arial", size=9, bold=bold, color=NAVY if bold else FORMULA_BLACK))
                self.set_cell(sheet, f"B{row}", unit, CellKind.STATIC, font=Font(name="Arial", size=8, color=GRAY))
            wacc_ref = self.assumption_ref(f"dcf_wacc_{scenario}")
            terminal_ref = self.assumption_ref(f"dcf_terminal_growth_{scenario}")
            terminal_exponent_ref = self.assumption_ref("dcf_terminal_discount_exponent")
            nopat_tax_ref = self.assumption_ref(f"dcf_nopat_tax_rate_{scenario}")
            for index, period in enumerate(forecast, start=3):
                column = get_column_letter(index)
                revenue_ref = self.output_ref(scenario, "revenue", period)
                ebit_ref = self.output_ref(scenario, "ebit", period)
                da_ref = self.output_ref(scenario, "depreciation_amortization", period)
                capex_ref = self.output_ref(scenario, "capex", period)
                nwc_ref = self.output_ref(scenario, "change_in_net_working_capital", period)
                annual_fcff_ref = self.output_ref(scenario, "cf.unlevered_fcf", period)
                component_refs = (
                    (2, revenue_ref),
                    (3, ebit_ref),
                    (6, da_ref),
                    (7, capex_ref),
                    (8, nwc_ref),
                )
                for row_offset, reference in component_refs:
                    formula = (
                        f'=IF(DCF_Calculation_Gate="PASS",{reference},"")'
                        if index > 3 and reference
                        else '=""'
                    )
                    self.set_formula(sheet, f"{column}{start + row_offset}", formula, number_format="#,##0.0;[Red](#,##0.0);-")
                self.set_formula(
                    sheet,
                    f"{column}{start + 4}",
                    f'=IF(DCF_Calculation_Gate="PASS",{nopat_tax_ref},"")' if nopat_tax_ref else '=""',
                    number_format="0.0%",
                )
                nopat_formula = (
                    f'=IF(AND(DCF_Calculation_Gate="PASS",COUNT({column}{start + 3}:{column}{start + 4})=2),'
                    f'{column}{start + 3}*(1-{column}{start + 4}),"")'
                    if index > 3
                    else '=""'
                )
                self.set_formula(sheet, f"{column}{start + 5}", nopat_formula, number_format="#,##0.0;[Red](#,##0.0);-")
                discounted_fcff_ref = self.assumption_ref(f"dcf_stub_fcff_{scenario}") if index == 3 else annual_fcff_ref
                self.set_formula(sheet, f"{column}{start + 9}", f'=IF(DCF_Calculation_Gate="PASS",{discounted_fcff_ref},"")' if discounted_fcff_ref else '=""', number_format="#,##0.0;[Red](#,##0.0);-")
                exponent_ref = self.assumption_ref(f"dcf_discount_exponent_{index - 2}")
                self.set_formula(sheet, f"{column}{start + 10}", f'=IF(DCF_Calculation_Gate="PASS",{exponent_ref},"")' if exponent_ref else '=""', number_format="0.000")
                self.set_formula(sheet, f"{column}{start + 11}", f'=IF(DCF_Calculation_Gate="PASS",1/(1+{wacc_ref})^{column}{start + 10},"")' if wacc_ref else '=""', number_format="0.000x")
                self.set_formula(sheet, f"{column}{start + 12}", f'=IF(DCF_Calculation_Gate="PASS",{column}{start + 9}*{column}{start + 11},"")', number_format="#,##0.0;[Red](#,##0.0);-")
                self.set_formula(sheet, f"{column}{start + 13}", f'=IF(DCF_Calculation_Gate="PASS",{terminal_ref},"")' if terminal_ref else '=""', number_format="0.0%")
                if index == 3:
                    self.set_formula(sheet, f"{column}{start + 23}", '=""', number_format="#,##0.0")
                    self.set_formula(sheet, f"{column}{start + 24}", '=""', number_format="#,##0.0")
                else:
                    self.set_formula(sheet, f"{column}{start + 23}", f'=IF(DCF_Calculation_Gate="PASS",{column}{start + 5}+{column}{start + 6}-ABS({column}{start + 7})-{column}{start + 8},"")', number_format="#,##0.0;[Red](#,##0.0);-")
                    self.set_formula(sheet, f"{column}{start + 24}", f'=IF(DCF_Calculation_Gate="PASS",{column}{start + 9}-{column}{start + 23},"")', number_format="#,##0.0;[Red](#,##0.0);-")
            last_column = get_column_letter(2 + len(forecast))
            if wacc_ref and terminal_ref and terminal_exponent_ref:
                self.set_formula(sheet, f"{last_column}{start + 14}", f'=IF(AND(DCF_Calculation_Gate="PASS",{wacc_ref}>{terminal_ref}),{last_column}{start + 9}*(1+{terminal_ref})/({wacc_ref}-{terminal_ref}),"")', number_format="#,##0.0;[Red](#,##0.0);-")
                self.set_formula(sheet, f"{last_column}{start + 15}", f'=IF(DCF_Calculation_Gate="PASS",{last_column}{start + 14}/(1+{wacc_ref})^{terminal_exponent_ref},"")', number_format="#,##0.0;[Red](#,##0.0);-")
            else:
                self.set_formula(sheet, f"{last_column}{start + 14}", '=""', number_format="#,##0.0")
                self.set_formula(sheet, f"{last_column}{start + 15}", '=""', number_format="#,##0.0")
            self.set_formula(sheet, f"C{start + 16}", f'=IF(DCF_Calculation_Gate="PASS",SUM(C{start + 12}:{last_column}{start + 12})+{last_column}{start + 15},"")', number_format="#,##0.0;[Red](#,##0.0);-")
            net_claims = self.component_net_claims_expression()
            shares = self.assumption_ref("current_fully_diluted_shares")
            price = self.assumption_ref("current_price")
            self.set_formula(sheet, f"C{start + 17}", f'=IF(DCF_Calculation_Gate="PASS",{net_claims},"")' if net_claims else '=""', number_format="#,##0.0;[Red](#,##0.0);-")
            self.set_formula(sheet, f"C{start + 18}", f'=IF(DCF_Calculation_Gate="PASS",C{start + 16}-C{start + 17},"")', number_format="#,##0.0;[Red](#,##0.0);-")
            self.set_formula(sheet, f"C{start + 19}", f'=IF(DCF_Calculation_Gate="PASS",{shares},"")' if shares else '=""', number_format="#,##0.0")
            self.set_formula(sheet, f"C{start + 20}", f'=IF(AND(DCF_Calculation_Gate="PASS",ISNUMBER(C{start + 18}),ISNUMBER(C{start + 19}),C{start + 19}>0),C{start + 18}/C{start + 19},"")', number_format="$0.00")
            self.set_formula(sheet, f"C{start + 21}", f'=IF(AND(DCF_Calculation_Gate="PASS",ISNUMBER({price}),{price}>0),{price},"")' if price else '=""', number_format="$0.00")
            self.set_formula(sheet, f"C{start + 22}", f'=IF(AND(DCF_Calculation_Gate="PASS",ISNUMBER(C{start + 20}),ISNUMBER(C{start + 21}),C{start + 21}>0),C{start + 20}/C{start + 21}-1,"")', number_format="0.0%;[Red](0.0%)")
            self.set_formula(sheet, f"C{start + 25}", '=IF(DCF_Calculation_Gate="PASS","ACT/365 MID_YEAR; Q4 FY26 STUB","")')
            self.set_formula(sheet, f"C{start + 26}", f'=IF(AND(DCF_Calculation_Gate="PASS",ISNUMBER(C{start + 16}),C{start + 16}<>0),{last_column}{start + 15}/C{start + 16},"")', number_format="0.0%")
            annual_ref = self.assumption_ref(f"dcf_annual_fy26_fcff_{scenario}")
            ytd_ref = self.assumption_ref(f"dcf_ytd_fcff_{scenario}")
            stub_ref = self.assumption_ref(f"dcf_stub_fcff_{scenario}")
            self.set_formula(sheet, f"C{start + 27}", f"={annual_ref}" if annual_ref else '=""', number_format="#,##0.0")
            self.set_formula(sheet, f"C{start + 28}", f"={ytd_ref}" if ytd_ref else '=""', number_format="#,##0.0")
            self.set_formula(sheet, f"C{start + 29}", f'=IF(COUNT(C{start + 27}:C{start + 28},{stub_ref})=3,C{start + 27}-C{start + 28}-{stub_ref},"")' if stub_ref else '=""', number_format="#,##0.0")
            self.dcf_result_cells[scenario] = (sheet, f"C{start + 20}")
        self.add_name("DCF_Base_Per_Share", sheet, "C25")
        self.workbook[sheet].column_dimensions["A"].width = 40
        self.workbook[sheet].column_dimensions["B"].width = 22
        for column in "CDEFG":
            self.workbook[sheet].column_dimensions[column].width = 14
        self.finish_sheet(sheet, "C7")
    def render_comps(self) -> None:
        sheet = "Comps"
        self.title(sheet, "Comparable company valuation", "Frozen exact-run/cached peer evidence and separate diagnostics; no blend or target")
        source_ready = self.payload.source.status in {"ready", "ok"} and not self.payload.source.formula_error_count

        def positive_number(value: float | None) -> bool:
            return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value)) and float(value) > 0.0

        valid_peer_evidence = any(
            (positive_number(item.enterprise_value) and positive_number(item.revenue))
            or (positive_number(item.enterprise_value) and positive_number(item.ebitda))
            or (positive_number(item.equity_value) and positive_number(item.net_income))
            for item in self.payload.comparables
        )
        diagnostic_keys = tuple(item[0] for item in DIAGNOSTIC_METHODS[:3])
        diagnostic_present = any(self.diagnostic_ref(key) for key in diagnostic_keys)
        if not source_ready:
            self.comps_state = "BLOCKED"
        elif valid_peer_evidence:
            self.comps_state = "AVAILABLE"
        elif self.payload.comparables or diagnostic_present:
            self.comps_state = "PARTIAL"
        else:
            self.comps_state = "UNAVAILABLE"
        self.set_cell(sheet, "A3", "Evidence state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(
            sheet,
            "B3",
            self.comps_state,
            CellKind.STATIC if self.comps_state == "AVAILABLE" else CellKind.UNAVAILABLE,
            fill=PatternFill("solid", fgColor=PALE_GREEN if self.comps_state == "AVAILABLE" else PALE_YELLOW if self.comps_state == "PARTIAL" else PALE_RED if self.comps_state == "BLOCKED" else LIGHT_GRAY),
        )
        evidence_note = {
            "AVAILABLE": "Frozen exact-run/cached peer evidence supports at least one trading multiple; LTM/NTM basis and as-of comparability are not independently verified.",
            "PARTIAL": "Frozen exact-run/cached peer evidence is incomplete; LTM/NTM basis and as-of comparability are not independently verified; separate diagnostics remain non-approved.",
            "BLOCKED": "Frozen exact-run/cached peer evidence remains diagnostic only; source preflight is blocked and LTM/NTM basis/as-of comparability is not independently verified.",
            "UNAVAILABLE": "No usable peer-denominator evidence is present; LTM/NTM basis and as-of comparability are not independently verified.",
        }[self.comps_state]
        self.set_cell(sheet, "C3", evidence_note, CellKind.STATIC, alignment=Alignment(wrap_text=True))
        headers = ("Ticker", "Company", "Enterprise value", "Equity value", "Revenue", "EBITDA", "Net income", "Share price", "EV / Revenue", "EV / EBITDA", "P / E")
        self.header_row(sheet, 5, headers)
        row = 6
        for comparable in sorted(self.payload.comparables, key=lambda item: item.ticker):
            raw = (
                comparable.ticker, comparable.company_name, comparable.enterprise_value, comparable.equity_value,
                comparable.revenue, comparable.ebitda, comparable.net_income, comparable.share_price,
            )
            for column, value in enumerate(raw, start=1):
                self.set_cell(sheet, f"{get_column_letter(column)}{row}", value, CellKind.SOURCE,
                              number_format="#,##0.0;[Red](#,##0.0);-" if column >= 3 else None)
            self.set_formula(sheet, f"I{row}", f'=IFERROR(IF(AND(ISNUMBER(C{row}),C{row}>0,ISNUMBER(E{row}),E{row}>0),C{row}/E{row},"NM"),"NM")', number_format="0.0x")
            self.set_formula(sheet, f"J{row}", f'=IFERROR(IF(AND(ISNUMBER(C{row}),C{row}>0,ISNUMBER(F{row}),F{row}>0),C{row}/F{row},"NM"),"NM")', number_format="0.0x")
            self.set_formula(sheet, f"K{row}", f'=IFERROR(IF(AND(ISNUMBER(D{row}),D{row}>0,ISNUMBER(G{row}),G{row}>0),D{row}/G{row},"NM"),"NM")', number_format="0.0x")
            row += 1
        if self.payload.comparables:
            self.set_cell(sheet, f"A{row}", "Median", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
            for column in "IJK":
                self.set_formula(sheet, f"{column}{row}", f'=IF(COUNT({column}6:{column}{row - 1})>0,MEDIAN({column}6:{column}{row - 1}),"")', number_format="0.0x")
            row += 2
        self.header_row(sheet, row, ("Separate diagnostic method", "Per-share value", "Evidence state", "Decision use"))
        row += 1
        for key, label, _unit in DIAGNOSTIC_METHODS[:3]:
            reference = self.diagnostic_ref(key)
            self.set_cell(sheet, f"A{row}", label, CellKind.STATIC)
            self.set_formula(sheet, f"B{row}", f"={reference}" if reference else '=""', number_format="$0.00")
            state = "BLOCKED" if self.comps_state == "BLOCKED" else "DIAGNOSTIC" if reference else "UNAVAILABLE"
            self.set_cell(sheet, f"C{row}", state, CellKind.CHECK if state == "BLOCKED" else CellKind.UNAVAILABLE)
            self.set_cell(sheet, f"D{row}", "NON-APPROVED; never blended or used as a target", CellKind.STATIC)
            row += 1
        worksheet = self.workbook[sheet]
        for column, width in {"A": 34, "B": 30, "C": 20, "D": 46, "E": 16, "F": 16, "G": 16, "H": 14, "I": 14, "J": 14, "K": 12}.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, "C6")
    def render_sotp(self) -> None:
        sheet = "SOTP"
        self.title(sheet, "Sum-of-the-parts valuation", "Only source-normalized segment evidence is permitted")
        state = self.payload.availability["sotp"]
        source_ready = self.payload.source.status in {"ready", "ok"} and not self.payload.source.formula_error_count
        shares = self.assumption_ref("current_fully_diluted_shares")
        if not source_ready:
            state = TypedAvailability("blocking", "source_preflight_blocked", "SOTP is blocked until the exact source run passes preflight.")
            self.blockers.add("sotp_source_preflight_blocked")
        elif state.status == "available" and not self.payload.sotp_components:
            state = TypedAvailability("blocking", "components_absent", "SOTP was marked available but no normalized components were supplied.")
            self.blockers.add("sotp_components_absent")
        elif state.status == "available" and not shares:
            state = TypedAvailability("blocking", "current_fdso_absent", "SOTP requires exact-dated current fully diluted shares.")
            self.blockers.add("sotp_current_fdso_absent")
        self.availability_banner(sheet, state)
        if state.status != "available":
            self.finish_sheet(sheet, "A4")
            return
        self.header_row(sheet, 7, ("Component", "Source-bound metric", "Approved multiple", "Segment enterprise value", "Evidence state", "Approval"))
        row = 8
        for component in sorted(self.payload.sotp_components, key=lambda item: item.component_key):
            self.set_cell(sheet, f"A{row}", component.label, CellKind.SOURCE)
            self.set_cell(sheet, f"B{row}", component.metric, CellKind.SOURCE, number_format="#,##0.0")
            self.set_cell(sheet, f"C{row}", component.multiple, CellKind.CHECK, number_format="0.0x")
            self.set_formula(sheet, f"D{row}", f"=B{row}*C{row}", number_format="#,##0.0")
            self.set_cell(sheet, f"E{row}", "SOURCE / PERIOD / PEER METHOD REQUIRED", CellKind.CHECK)
            self.set_cell(sheet, f"F{row}", "PM APPROVAL REQUIRED", CellKind.CHECK)
            row += 1
        self.set_cell(sheet, f"A{row}", "Total segment enterprise value", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, f"D{row}", f"=SUM(D8:D{row - 1})", number_format="#,##0.0")
        net_claims = self.component_net_claims_expression()
        self.set_cell(sheet, f"A{row + 1}", "Less: consolidated component net claims", CellKind.STATIC)
        self.set_formula(sheet, f"D{row + 1}", f"={net_claims}" if net_claims else '=""', number_format="#,##0.0")
        self.set_cell(sheet, f"A{row + 2}", "Consolidated equity value", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, f"D{row + 2}", f'=IF(COUNT(D{row}:D{row + 1})=2,D{row}-D{row + 1},"")', number_format="#,##0.0")
        self.set_cell(sheet, f"A{row + 3}", f"Current FDSO ({self.payload.as_of_date.isoformat()})", CellKind.STATIC)
        self.set_formula(sheet, f"D{row + 3}", f"={shares}", number_format="#,##0.0")
        self.set_cell(sheet, f"A{row + 4}", "SOTP diagnostic per share", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, f"D{row + 4}", f'=IF(AND(ISNUMBER(D{row + 2}),ISNUMBER(D{row + 3}),D{row + 3}>0),D{row + 2}/D{row + 3},"")', number_format="$0.00")
        self.add_name("SOTP_Per_Share", sheet, f"D{row + 4}")
        self.finish_sheet(sheet, "B8")
    def render_consensus_bridge(self) -> None:
        sheet = "Consensus_Bridge"
        self.title(
            sheet,
            "Consensus bridge",
            "Immutable snapshot, exact period types/end dates, guarded model deltas, guidance, and revisions",
        )
        snapshot = self.payload.consensus_snapshot
        if snapshot is None:
            evidence_state = "BLOCKED"
            decision_state = "BLOCKED"
            reason = self.payload.availability["consensus"].reason_code or "consensus_snapshot_unavailable"
            detail = self.payload.availability["consensus"].message or "No qualified consensus snapshot was supplied."
            self.blockers.add("consensus_snapshot_unavailable")
        else:
            observations = snapshot.observations
            has_gaps = any(
                item.value_state.status.value != "available"
                or item.mapping_method.value != "EXACT_PERIOD_END"
                for item in observations
            )
            pm_review = any(
                item.method_status.decision_eligibility.value == "NEEDS_PM_REVIEW"
                for item in observations
            )
            eligible = any(
                item.method_status.decision_eligibility.value == "ELIGIBLE"
                for item in observations
            )
            evidence_state = "PARTIAL" if has_gaps else "UNVERIFIED"
            decision_state = "NEEDS_PM_REVIEW" if pm_review else "UNVERIFIED" if eligible else "BLOCKED"
            reason = "typed_snapshot_supplied"
            detail = (
                f"{len(observations)} typed observations; values remain non-controlling until "
                "period, lineage, coverage, and decision gates pass."
            )
        self.set_cell(sheet, "A3", "Evidence workflow state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B3", evidence_state, CellKind.CHECK)
        self.set_cell(sheet, "D3", "Decision workflow state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "E3", decision_state, CellKind.CHECK)
        self.set_cell(sheet, "A4", "Reason", CellKind.STATIC)
        self.set_cell(sheet, "B4", reason, CellKind.CHECK)
        self.set_cell(sheet, "D4", "Snapshot detail", CellKind.STATIC)
        self.set_cell(sheet, "E4", detail, CellKind.STATIC, alignment=Alignment(wrap_text=True))
        if snapshot is None:
            self.finish_sheet(sheet, "B3")
            return

        snapshot_locator = snapshot.source_snapshot_locator
        if re.match(r"^[A-Za-z]:[\\/]", snapshot_locator):
            snapshot_locator = Path(snapshot_locator).name
        self.set_cell(sheet, "A5", "Snapshot as-of / source", CellKind.STATIC)
        self.set_cell(
            sheet,
            "B5",
            f"{snapshot.as_of_date.isoformat()} | {snapshot.source_name} | {snapshot_locator}",
            CellKind.SOURCE,
        )
        self.set_cell(sheet, "D5", "Snapshot hash", CellKind.STATIC)
        self.set_cell(sheet, "E5", snapshot.snapshot_hash, CellKind.SOURCE)
        headers = (
            "Observation ID",
            "Metric",
            "Source metric",
            "Statistic",
            "Period type",
            "Period end",
            "Mapping",
            "Model period",
            "Consensus",
            "Base model",
            "Delta",
            "Unit",
            "Analysts",
            "Value state",
            "Decision eligibility",
            "Source / transform",
        )
        self.header_row(sheet, 7, headers)
        metric_to_line = {
            "REVENUE": "revenue",
            "EBIT": "ebit",
            "D_AND_A": "depreciation_amortization",
            "DILUTED_EPS": "shares.diluted_eps",
        }
        row = 8
        for observation in snapshot.observations:
            mapped_period = observation.mapped_model_period_key
            line_key = metric_to_line.get(observation.metric)
            model_reference = (
                self.output_ref("base", line_key, mapped_period)
                if line_key and mapped_period in self.payload.forecast_periods
                else None
            )
            source_locator = observation.source_locator
            if re.match(r"^[A-Za-z]:[\\/]", source_locator):
                source_locator = Path(source_locator).name
            raw_values = (
                observation.observation_id,
                observation.metric,
                observation.source_metric,
                observation.statistic.value,
                observation.period_type.value,
                observation.period_end.isoformat(),
                observation.mapping_method.value,
                mapped_period or "UNMAPPED",
            )
            for column, value in enumerate(raw_values, start=1):
                self.set_cell(sheet, f"{get_column_letter(column)}{row}", value, CellKind.SOURCE)
            self.set_cell(
                sheet,
                f"I{row}",
                observation.value,
                CellKind.SOURCE if observation.value is not None else CellKind.UNAVAILABLE,
                number_format="#,##0.0",
            )
            self.set_formula(sheet, f"J{row}", f"={model_reference}" if model_reference else '=""', number_format="#,##0.0")
            exact_mapping = observation.mapping_method.value == "EXACT_PERIOD_END"
            delta_formula = f'=IF(COUNT(I{row}:J{row})=2,J{row}-I{row},"")' if exact_mapping else '=""'
            self.set_formula(sheet, f"K{row}", delta_formula, number_format="#,##0.0;[Red](#,##0.0);-")
            self.set_cell(sheet, f"L{row}", observation.unit, CellKind.SOURCE)
            analyst_display = observation.analyst_count if observation.analyst_count is not None else "UNAVAILABLE"
            self.set_cell(sheet, f"M{row}", analyst_display, CellKind.SOURCE)
            self.set_cell(sheet, f"N{row}", observation.value_state.status.value.upper(), CellKind.CHECK)
            self.set_cell(sheet, f"O{row}", observation.method_status.decision_eligibility.value, CellKind.CHECK)
            self.set_cell(
                sheet,
                f"P{row}",
                f"{observation.source_as_of_date.isoformat()} | {source_locator} | {observation.transformation}",
                CellKind.SOURCE,
                alignment=Alignment(wrap_text=True),
            )
            row += 1

        row += 1
        self.header_row(sheet, row, ("Section", "Workflow state", "Evidence requirement"))
        self.set_cell(sheet, f"A{row + 1}", "Guidance", CellKind.STATIC)
        self.set_cell(sheet, f"B{row + 1}", "BLOCKED", CellKind.CHECK)
        self.set_cell(sheet, f"C{row + 1}", "No typed, source-located guidance packet supplied.", CellKind.UNAVAILABLE)
        self.set_cell(sheet, f"A{row + 2}", "Estimate revisions", CellKind.STATIC)
        self.set_cell(sheet, f"B{row + 2}", "BLOCKED", CellKind.CHECK)
        self.set_cell(sheet, f"C{row + 2}", "No prior snapshot or revision-series contract supplied.", CellKind.UNAVAILABLE)
        worksheet = self.workbook[sheet]
        for column, width in {
            "A": 24, "B": 18, "C": 18, "D": 14, "E": 14, "F": 13, "G": 22, "H": 14,
            "I": 14, "J": 14, "K": 14, "L": 14, "M": 12, "N": 15, "O": 20, "P": 58,
        }.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, "I8")

    def render_sensitivities(self) -> None:
        sheet = "Sensitivities"
        self.title(sheet, "DCF sensitivity analysis", "Base-scenario implied value per share across WACC and terminal-growth cases")
        self.set_cell(sheet, "A3", "DCF calculation gate", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, "B3", "=DCF_Calculation_Gate", kind=CellKind.CHECK)
        self.set_cell(sheet, "A4", "Terminal growth / WACC", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        base_wacc_ref = self.assumption_ref("dcf_wacc_base")
        for index, delta in enumerate((-0.02, -0.01, 0.0, 0.01, 0.02), start=3):
            column = get_column_letter(index)
            self.set_formula(sheet, f"{column}6", f'=IF($B$3="PASS",{base_wacc_ref}+{delta},"")' if base_wacc_ref else '=""', number_format="0.0%")
        terminal_ref = self.assumption_ref("dcf_terminal_growth_base")
        terminal_exponent_ref = self.assumption_ref("dcf_terminal_discount_exponent")
        shares_ref = self.assumption_ref("current_fully_diluted_shares")
        net_claims = self.component_net_claims_expression()
        shares_expr = shares_ref
        debt_expr = net_claims
        for row, delta in enumerate((-0.01, -0.005, 0.0, 0.005, 0.01), start=7):
            self.set_formula(sheet, f"B{row}", f'=IF($B$3="PASS",{terminal_ref}+{delta},"")' if terminal_ref else '=""', number_format="0.0%")
            for index in range(3, 8):
                column = get_column_letter(index)
                pv_terms = "+".join(f"DCF!{period_column}14/(1+{column}$6)^DCF!{period_column}15" for period_column in "CDEFG")
                if not shares_expr or not debt_expr or not terminal_exponent_ref:
                    self.set_formula(sheet, f"{column}{row}", '=""', number_format="$0.00")
                    continue
                formula = (
                    f'=IF($B$3<>"PASS","",IF(AND({column}$6>$B{row},ISNUMBER({shares_expr}),{shares_expr}>0),('
                    f'{pv_terms}+DCF!G14*(1+$B{row})/({column}$6-$B{row})/(1+{column}$6)^{terminal_exponent_ref}-({debt_expr}))/'
                    f'{shares_expr},""))'
                )
                self.set_formula(sheet, f"{column}{row}", formula, number_format="$0.00")
        worksheet = self.workbook[sheet]
        worksheet.conditional_formatting.add("C7:G11", CellIsRule(operator="greaterThan", formula=["DCF_Base_Per_Share"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
        worksheet.conditional_formatting.add("C7:G11", CellIsRule(operator="lessThan", formula=["DCF_Base_Per_Share"], fill=PatternFill("solid", fgColor=PALE_RED)))
        for column in "BCDEFG":
            worksheet.column_dimensions[column].width = 15
        self.finish_sheet(sheet, "C7")
    def render_valuation(self) -> None:
        sheet = "Valuation"
        self.title(sheet, "Valuation summary", "Separate approved and diagnostic methods; no blend, target, probability weight, or recommendation")
        self.header_row(sheet, 4, ("Method", "Base / diagnostic value", "Upside", "Downside", "Evidence state", "Policy / evidence note"))
        self.set_cell(sheet, "A5", "Integrated FCFF DCF", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        for column, scenario in zip("BCD", REQUIRED_SCENARIOS, strict=True):
            dcf_sheet, dcf_cell = self.dcf_result_cells[scenario]
            self.set_formula(sheet, f"{column}5", f'=IF(DCF_Calculation_Gate="PASS",{_absolute_ref(dcf_sheet, dcf_cell)},"")', number_format="$0.00")
        self.set_cell(sheet, "E5", f"{self.dcf_state} / {self.dcf_decision_state}", CellKind.CHECK)
        self.set_cell(sheet, "F5", "Mathematically available values are calculated diagnostics only; the separate eligibility state controls decision use and no value is a target.", CellKind.STATIC)

        source_ready = self.payload.source.status in {"ready", "ok"} and not self.payload.source.formula_error_count
        for row, (key, label, unit_kind) in enumerate(DIAGNOSTIC_METHODS, start=6):
            reference = self.diagnostic_ref(key)
            self.set_cell(sheet, f"A{row}", label, CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
            if reference:
                self.set_formula(sheet, f"B{row}", f"={reference}", number_format="$0.00" if unit_kind == "per_share" else '0.0"%"')
            else:
                self.set_formula(sheet, f"B{row}", '=""', number_format="$0.00" if unit_kind == "per_share" else '0.0"%"')
            self.set_formula(sheet, f"C{row}", '=""')
            self.set_formula(sheet, f"D{row}", '=""')
            state = "BLOCKED" if not source_ready else "DIAGNOSTIC" if reference else "UNAVAILABLE"
            self.set_cell(sheet, f"E{row}", state, CellKind.CHECK if state == "BLOCKED" else CellKind.UNAVAILABLE)
            note = "Percentage points; raw 16.9 means 16.9%, not a 16.9 decimal rate." if unit_kind == "percentage_points" else "Frozen method-specific legacy diagnostic."
            self.set_cell(sheet, f"F{row}", note + " NON-APPROVED; never blended or used as a target.", CellKind.STATIC)

        fcfe_state = str(self.payload.backend_checks.get("fcfe.state") or "UNAVAILABLE").upper()
        fcfe_reason = str(self.payload.backend_checks.get("fcfe.reason_code") or "fcfe_method_evidence_not_provided")
        fcfe_detail = str(self.payload.backend_checks.get("fcfe.detail") or "No source-backed FCFE detail was supplied.")
        self.set_cell(sheet, "A13", "FCFE cross-check", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        for column in "BCD":
            self.set_formula(sheet, f"{column}13", '=""')
        self.set_cell(sheet, "E13", fcfe_state, CellKind.CHECK if fcfe_state in {"BLOCKED", "FAIL"} else CellKind.UNAVAILABLE)
        self.set_cell(sheet, "F13", f"{fcfe_reason}: {fcfe_detail}", CellKind.STATIC, alignment=Alignment(wrap_text=True))

        self.set_cell(sheet, "A14", "SOTP", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        sotp_available = "SOTP_Per_Share" in self.defined_names
        for column in "BCD":
            if sotp_available:
                self.set_formula(sheet, f"{column}14", "=SOTP_Per_Share", number_format="$0.00")
            else:
                self.set_formula(sheet, f"{column}14", '=""', number_format="$0.00")
        self.set_cell(sheet, "E14", "AVAILABLE" if sotp_available else "UNAVAILABLE", CellKind.STATIC if sotp_available else CellKind.UNAVAILABLE)
        self.set_cell(sheet, "F14", "No proxy SOTP without normalized segment evidence.", CellKind.STATIC)

        self.set_cell(sheet, "A15", "Historical trading range", CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
        for column in "BCD":
            self.set_formula(sheet, f"{column}15", '=""')
        self.set_cell(sheet, "E15", "UNAVAILABLE", CellKind.UNAVAILABLE)
        self.set_cell(sheet, "F15", "No source-backed historical valuation-range payload was supplied.", CellKind.STATIC)

        self.set_cell(sheet, "A18", "Single-count DCF EV-to-equity bridge", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY))
        self.header_row(sheet, 19, ("Bridge line", "Base", "Upside", "Downside", "Count policy", "Evidence note"))
        bridge_rows = (
            (20, "Enterprise value", "dcf_ev"),
            (21, "Add: cash", "bridge_cash"),
            (22, "Add: short-term investments", "bridge_short_term_investments"),
            (23, "Add: long-term investments", "bridge_long_term_investments"),
            (24, "Less: short-term borrowings", "bridge_short_term_borrowings"),
            (25, "Less: current long-term debt", "bridge_current_long_term_debt"),
            (26, "Less: long-term debt", "bridge_long_term_debt"),
            (27, "Less: current lease liabilities", "bridge_current_lease_liabilities"),
            (28, "Less: long-term lease liabilities", "bridge_long_term_lease_liabilities"),
            (29, "Less: minority interest", "bridge_minority_interest"),
            (30, "Less: pension liability", "bridge_pension_liability"),
            (31, "Component-based net claims", "net_claims"),
            (32, "Equity value", "dcf_equity"),
            (33, "Reported gross debt (reference only)", "bridge_gross_debt"),
            (34, "Reported gross-debt reconciliation", "gross_debt_check"),
            (35, "Total-borrowings reconciliation", "borrowings_check"),
            (36, "Lease-liabilities reconciliation", "leases_check"),
            (37, "EV-to-equity bridge check", "equity_check"),
        )
        dcf_starts = {"base": 5, "upside": 35, "downside": 65}
        refs = {key: self.assumption_ref(key) for key in (*BRIDGE_ASSET_KEYS, *BRIDGE_BORROWING_KEYS, *BRIDGE_LEASE_KEYS, *BRIDGE_OTHER_CLAIM_KEYS, *BRIDGE_REFERENCE_KEYS)}
        def sum_refs(keys: tuple[str, ...]) -> str | None:
            values = tuple(refs[key] for key in keys)
            return "+".join(values) if all(values) else None

        assets_expression = sum_refs(BRIDGE_ASSET_KEYS)
        borrowings_expression = sum_refs(BRIDGE_BORROWING_KEYS)
        leases_expression = sum_refs(BRIDGE_LEASE_KEYS)
        other_claims_expression = sum_refs(BRIDGE_OTHER_CLAIM_KEYS)
        net_claims_expression = (
            f"({borrowings_expression}+{leases_expression}+{other_claims_expression})-({assets_expression})"
            if all((assets_expression, borrowings_expression, leases_expression, other_claims_expression))
            else None
        )
        for row, label, key in bridge_rows:
            self.set_cell(sheet, f"A{row}", label, CellKind.STATIC, font=Font(name="Arial", size=9, bold=row in {31, 32, 34, 35, 36, 37}, color=NAVY if row in {31, 32, 34, 35, 36, 37} else FORMULA_BLACK))
            for column, scenario in zip("BCD", REQUIRED_SCENARIOS, strict=True):
                start = dcf_starts[scenario]
                if key == "dcf_ev":
                    expression = f"DCF!$C${start + 16}"
                elif key == "dcf_equity":
                    expression = f"DCF!$C${start + 18}"
                elif key == "net_claims":
                    expression = net_claims_expression
                elif key == "gross_debt_check":
                    expression = (
                        f"{refs['bridge_gross_debt']}-({borrowings_expression}+{leases_expression})"
                        if refs["bridge_gross_debt"] and borrowings_expression and leases_expression
                        else None
                    )
                elif key == "borrowings_check":
                    expression = (
                        f"{refs['bridge_total_borrowings']}-({borrowings_expression})"
                        if refs["bridge_total_borrowings"] and borrowings_expression
                        else None
                    )
                elif key == "leases_check":
                    expression = (
                        f"{refs['bridge_lease_liabilities']}-({leases_expression})"
                        if refs["bridge_lease_liabilities"] and leases_expression
                        else None
                    )
                elif key == "equity_check":
                    expression = f"{column}32-({column}20-{column}31)"
                else:
                    expression = refs[key]
                if expression is None:
                    formula = '=""'
                elif key in {"dcf_ev", "dcf_equity", "equity_check"}:
                    formula = f'=IF(DCF_Calculation_Gate="PASS",{expression},"")'
                else:
                    formula = f"={expression}"
                self.set_formula(sheet, f"{column}{row}", formula, number_format="#,##0.0;[Red](#,##0.0);-")
            count_policy = "REFERENCE ONLY; not counted" if key in BRIDGE_REFERENCE_KEYS or key.endswith("_check") else "COUNT ONCE"
            self.set_cell(sheet, f"E{row}", count_policy, CellKind.STATIC)
            evidence_note = "Reported debt reconciles borrowings plus leases; aggregate is excluded from net claims." if key == "bridge_gross_debt" else "Exact frozen bridge component or formula."
            self.set_cell(sheet, f"F{row}", evidence_note, CellKind.STATIC)
        self.workbook[sheet].column_dimensions["A"].width = 40
        for column in "BCD":
            self.workbook[sheet].column_dimensions[column].width = 18
        self.workbook[sheet].column_dimensions["E"].width = 24
        self.workbook[sheet].column_dimensions["F"].width = 72
        self.finish_sheet(sheet, "B5")
    def render_accounting_qoe(self) -> None:
        sheet = "Accounting_QoE"
        self.title(
            sheet,
            "Accounting quality of earnings",
            "Concise reported-to-normalized findings; generic model integrity checks remain on Checks",
        )
        topics = (
            ("Reported to normalized EBIT / EPS / FCF", ("normalized", "unusual", "fcff_definition"), "Separate reported and normalized measures.", "Prevent duplicate normalization in DCF/comps.", "Which adjustments, if any, are recurring?"),
            ("Revenue recognition / deferred revenue", ("revenue", "deferred_revenue"), "Assess recognition timing and contract liabilities.", "Normalize only source-supported timing effects.", "What evidence changes the durability view?"),
            ("Stock-based compensation", ("stock_compensation", "apic"), "Keep SBC visible and reconcile equity/cash effects.", "Do not add back without an explicit valuation policy.", "What dilution and cash-equivalent treatment is approved?"),
            ("Capitalized software / R&D", ("software", "research", "development"), "Identify capitalization and amortization policy.", "Align normalized EBIT and invested capital.", "Are capitalization periods comparable and supportable?"),
            ("Leases", ("lease",), "Use one GAAP or lease-adjusted convention end to end.", "Align EBIT, FCFF, WACC, and EV bridge.", "Which lease convention is approved?"),
            ("Unusual / restructuring items", ("unusual", "restructur"), "Reconcile reported to normalized earnings by period.", "Apply each adjustment once.", "Which items are nonrecurring and why?"),
            ("M&A / PPA / intangibles", ("acquisition", "ppa", "intangible", "amortization", "impair"), "Separate goodwill, finite-lived intangibles, and amortization.", "Align D&A, capex, FCFF, and multiples.", "What PPA and amortization evidence is available?"),
            ("Income taxes", ("tax_", "deferred_tax", "cash_tax"), "Reconcile book, current, deferred, payable, and cash tax.", "Use distinct NOPAT, marginal, effective, and cash rates.", "Which marginal/NOPAT tax policy is approved?"),
            ("Working capital / cash conversion", ("working_capital", "receivables", "inventory", "payables", "cash_flow"), "Reconcile operating working-capital definitions and conversion.", "Avoid financing/nonoperating double counts.", "Which balances are operating versus nonoperating?"),
            ("Reserves / credit losses", ("reserve", "credit_loss", "allowance"), "Assess reserve roll-forwards and loss coverage.", "Normalize only evidence-backed reserve changes.", "What reserve evidence changes the view?"),
            ("Non-GAAP reconciliation", ("non_gaap", "adjusted"), "Reconcile every non-GAAP measure to reported results.", "Do not reuse vendor-adjusted figures without definitions.", "Which definitions are approved for valuation?"),
        )
        topic_rows: list[tuple[str, str, str, str, str, str]] = []
        any_evidence = False
        any_blocked = False
        for label, tokens, accounting_treatment, valuation_treatment, pm_question in topics:
            evidence_keys = tuple(
                key
                for key in sorted(self.payload.backend_checks)
                if any(token in key.casefold() for token in tokens)
            )
            status_values = {
                str(self.payload.backend_checks[key]).strip().upper()
                for key in evidence_keys
                if key.endswith(".status")
            }
            if status_values & {"FAIL", "BLOCKED", "UNAVAILABLE"}:
                state = "BLOCKED"
            elif "PASS" in status_values:
                state = "PARTIAL"
            else:
                state = "BLOCKED"
            any_evidence = any_evidence or bool(evidence_keys)
            any_blocked = any_blocked or state == "BLOCKED"
            anchors = ", ".join(evidence_keys[:6]) if evidence_keys else "No source-located QoE evidence supplied"
            topic_rows.append((label, state, anchors, accounting_treatment, valuation_treatment, pm_question))

        evidence_state = "BLOCKED" if not any_evidence or any_blocked else "PARTIAL"
        self.set_cell(sheet, "A3", "Evidence workflow state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B3", evidence_state, CellKind.CHECK)
        self.set_cell(sheet, "D3", "Decision workflow state", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "E3", "BLOCKED", CellKind.CHECK)
        self.set_cell(sheet, "A4", "Interpretation", CellKind.STATIC)
        self.set_cell(
            sheet,
            "B4",
            "Backend arithmetic checks are evidence anchors only; they do not make QoE findings decision-ready.",
            CellKind.STATIC,
            alignment=Alignment(wrap_text=True),
        )
        self.header_row(
            sheet,
            7,
            (
                "Finding area",
                "Workflow state",
                "Period / units",
                "Evidence anchors",
                "Accounting treatment",
                "Valuation treatment",
                "Materiality",
                "Confidence",
                "PM question",
                "No-double-count rule",
                "Approval / result",
                "What changes the view",
            ),
        )
        for row, (label, state, anchors, accounting_treatment, valuation_treatment, pm_question) in enumerate(topic_rows, start=8):
            values = (
                label,
                state,
                "Exact period and units required",
                anchors,
                accounting_treatment,
                valuation_treatment,
                "UNVERIFIED",
                "LOW",
                pm_question,
                "One adjustment, one schedule, one valuation bridge.",
                "BLOCKED",
                "Source-located evidence plus fingerprinted approval.",
            )
            for column, value in enumerate(values, start=1):
                self.set_cell(
                    sheet,
                    f"{get_column_letter(column)}{row}",
                    value,
                    CellKind.CHECK if column in {2, 7, 8, 11} else CellKind.STATIC,
                    alignment=Alignment(wrap_text=True),
                )
        worksheet = self.workbook[sheet]
        for column, width in {
            "A": 34, "B": 16, "C": 24, "D": 54, "E": 40, "F": 40,
            "G": 15, "H": 12, "I": 42, "J": 36, "K": 18, "L": 42,
        }.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, "D8")

    def render_pm_review_queue(self) -> None:
        sheet = "PM_Review_Queue"
        self.title(
            sheet,
            "PM review queue",
            "Normalized root causes, stable approval keys, annual paths, provenance, impact, and consumed evidence",
        )
        headers = (
            "Priority",
            "Category",
            "Module",
            "Scenario",
            "Stable approval key",
            "Current FY26E-FY30E path",
            "Proposed FY26E-FY30E path",
            "Units",
            "Source / method / as-of",
            "Materiality / impact",
            "PM question / required action",
            "Workflow state",
            "Current fingerprint",
            "Reviewer / rationale / timestamp",
            "Downstream dependencies",
            "Evidence locator",
        )
        self.header_row(sheet, 4, headers)

        def classify(code: str) -> tuple[str, str, str, str]:
            lowered = code.casefold()
            if "formula" in lowered or "recalc" in lowered or "artifact" in lowered:
                category = "SYSTEM REPAIR"
            elif "source" in lowered or "consensus" in lowered or "segment" in lowered:
                category = "SOURCE ACQUISITION"
            elif "pm_approval" in lowered or "source_or_pm" in lowered:
                category = "PM JUDGMENT"
            else:
                category = "FINANCE POLICY"
            module = next(
                (
                    label
                    for token, label in (
                        ("wacc", "WACC"),
                        ("dcf", "DCF"),
                        ("share", "Shares_EPS"),
                        ("tax", "Taxes"),
                        ("ppe", "PP&E_Intangibles"),
                        ("lease", "Debt_Cash_Interest"),
                        ("debt", "Debt_Cash_Interest"),
                        ("consensus", "Consensus_Bridge"),
                        ("segment", "Segment_Build"),
                        ("comps", "Comps"),
                        ("sotp", "SOTP"),
                        ("scenario", "Scenarios"),
                    )
                    if token in lowered
                ),
                "Package",
            )
            priority = "P0" if any(token in lowered for token in ("source_formula", "dcf", "wacc", "share", "tax", "stub", "scenario")) else "P1"
            state = "NEEDS_PM_REVIEW" if category in {"PM JUDGMENT", "FINANCE POLICY"} else "BLOCKED"
            return priority, category, module, state

        unavailable_path = " | ".join(f"{period}:UNAVAILABLE" for period in self.payload.forecast_periods)
        provenance = (
            f"run {self.payload.source.run_id} | {self.payload.source.source_hash} | "
            f"financial cutoff {self.payload.as_of_date.isoformat()}"
        )
        queue_rows: list[tuple[Any, ...]] = []
        normalized_codes: set[str] = set()
        for blocker in sorted(self.blockers):
            code = "source_formula_errors" if blocker.startswith("source_formula_errors:") else blocker
            if code in normalized_codes:
                continue
            normalized_codes.add(code)
            priority, category, module, state = classify(code)
            scenario = next((item for item in ("Base", "Upside", "Downside") if f":{item}:" in blocker), "ALL")
            stable_key = _canonical_hash({"category": category, "module": module, "scenario": scenario, "code": code})[:16]
            evidence_locator = (
                str(self.payload.backend_checks.get("source.formula_error_cells") or "See Sources / Checks")
                if code == "source_formula_errors"
                else "See Checks and Sources"
            )
            question = (
                "Repair/re-source this root cause and rerun exact-run validation."
                if category in {"SYSTEM REPAIR", "SOURCE ACQUISITION"}
                else "Approve an explicit policy/path with rationale; do not approve derived outputs."
            )
            queue_rows.append(
                (
                    priority,
                    category,
                    module,
                    scenario,
                    stable_key,
                    unavailable_path,
                    unavailable_path,
                    "UNVERIFIED",
                    provenance,
                    "Decision or valuation impact requires quantified sensitivity",
                    question,
                    state,
                    "UNAVAILABLE",
                    "UNASSIGNED | rationale required | timestamp required",
                    f"{module} outputs; Valuation; Summary; Checks",
                    evidence_locator,
                )
            )

        for approval in self.payload.driver_approvals:
            scenario = approval.scenario_key or "ALL"
            category = "PM JUDGMENT" if approval.driver_group.value == "finance_semantic" else "SYSTEM REPAIR"
            state = "PARTIAL" if approval.approval_state.value == "APPROVED" else "NEEDS_PM_REVIEW"
            stable_key = _canonical_hash(
                {
                    "scenario": scenario,
                    "driver": approval.driver_key,
                    "record_hash": approval.record_hash,
                }
            )[:16]
            reviewer = (
                f"{approval.approved_by or 'UNASSIGNED'} | {approval.approval_ref or 'rationale required'} | "
                f"{approval.approved_at.isoformat() if approval.approved_at else 'timestamp required'}"
            )
            queue_rows.append(
                (
                    "P2" if state == "PARTIAL" else "P0",
                    category,
                    "Scenarios",
                    scenario,
                    stable_key,
                    unavailable_path,
                    unavailable_path,
                    "Driver contract",
                    provenance,
                    "Scenario path and downstream valuation sensitivity",
                    "Confirm the five-year path, source/method, and approval fingerprint.",
                    state,
                    approval.current_driver_fingerprint,
                    reviewer,
                    "Statements; schedules; DCF; Valuation",
                    approval.record_hash or "UNAVAILABLE",
                )
            )

        queue_rows.append(
            (
                "P0",
                "SYSTEM REPAIR",
                "Package",
                "ALL",
                _canonical_hash({"gate": "calculation_verification", "model": self.model_input_hash})[:16],
                unavailable_path,
                unavailable_path,
                "N/A",
                provenance,
                "All calculated outputs remain unverified",
                "Run isolated native recalculation and bind the authoritative sidecar to workbook/input/formula hashes.",
                "UNVERIFIED",
                self.model_input_hash,
                "UNASSIGNED | automated verification | timestamp pending",
                "All workbook outputs",
                "Cover!B14; Checks!calculation_verification",
            )
        )
        for row, values in enumerate(queue_rows, start=5):
            for column, value in enumerate(values, start=1):
                self.set_cell(
                    sheet,
                    f"{get_column_letter(column)}{row}",
                    value,
                    CellKind.CHECK if column in {1, 2, 12} else CellKind.STATIC,
                    alignment=Alignment(wrap_text=True, vertical="top"),
                )
        worksheet = self.workbook[sheet]
        for column, width in {
            "A": 10, "B": 20, "C": 22, "D": 12, "E": 20, "F": 58, "G": 58, "H": 16,
            "I": 52, "J": 38, "K": 54, "L": 20, "M": 34, "N": 52, "O": 42, "P": 48,
        }.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, "E5")

    def render_checks(self) -> None:
        sheet = "Checks"
        self.title(sheet, "Model integrity checks", "FULL requires explicit PASS on every required gate; unknown and degraded values fail closed")
        self.header_row(sheet, 4, ("Check ID", "Description", "Status", "Difference / count", "Tolerance"))
        self.set_cell(sheet, "A5", "model_readiness", CellKind.STATIC)
        self.set_cell(sheet, "B5", "Fail-closed package workflow state", CellKind.STATIC)
        self.set_formula(sheet, "C5", '=IF(COUNTIF(C6:C17,"FAIL")+COUNTIF(C6:C17,"BLOCKED")>0,"BLOCKED",IF(COUNTIF(C6:C17,"NEEDS_PM_REVIEW")>0,"NEEDS_PM_REVIEW",IF(COUNTIF(C6:C17,"PASS")=ROWS(C6:C17),"FULL",IF(COUNTIF(C6:C17,"PASS")+COUNTIF(C6:C17,"PARTIAL")>0,"PARTIAL","UNVERIFIED"))))', kind=CellKind.CHECK)
        self.set_formula(sheet, "D5", '=ROWS(C6:C17)-COUNTIF(C6:C17,"PASS")', kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "E5", 0, CellKind.STATIC)
        self.add_name("Model_Status", sheet, "C5")
        self.add_check("model_readiness", sheet, "C5")

        source_pass = self.payload.source.status in {"ready", "ok"} and self.payload.source.formula_error_count == 0
        self.set_cell(sheet, "A6", "source_preflight", CellKind.STATIC)
        self.set_cell(sheet, "B6", "Exact source run has no source-formula blockers", CellKind.STATIC)
        self.set_cell(sheet, "C6", "PASS" if source_pass else "BLOCKED", CellKind.CHECK)
        self.set_cell(sheet, "D6", self.payload.source.formula_error_count, CellKind.CHECK)
        self.set_cell(sheet, "E6", 0, CellKind.STATIC)
        self.add_check("source_preflight", sheet, "C6")

        scenario_last_row = 4 + len(self.lines) * len(REQUIRED_SCENARIOS)
        expected_scenario_rows = len(self.lines)
        scenario_range = f"Scenarios!$A$5:$A${scenario_last_row}"
        scenario_counts = [f'COUNTIF({scenario_range},"{scenario.upper()}")' for scenario in REQUIRED_SCENARIOS]
        self.set_cell(sheet, "A7", "scenario_completeness", CellKind.STATIC)
        self.set_cell(sheet, "B7", "Each scenario has exactly one frozen output row for every supplied model line", CellKind.STATIC)
        self.set_formula(sheet, "D7", "=MIN(" + ",".join(scenario_counts) + ")", kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "E7", expected_scenario_rows, CellKind.STATIC)
        self.set_formula(sheet, "C7", '=IF(AND(' + ",".join(f"{count}=$E$7" for count in scenario_counts) + '),"PASS","FAIL")', kind=CellKind.CHECK)
        self.add_check("scenario_completeness", sheet, "C7")

        first_forecast_column = get_column_letter(6)
        last_forecast_column = get_column_letter(5 + len(self.payload.forecast_periods))
        forecast_range = f"Scenarios!${first_forecast_column}$5:${last_forecast_column}${scenario_last_row}"
        expected_forecast_values = len(self.lines) * len(REQUIRED_SCENARIOS) * len(self.payload.forecast_periods)
        self.set_cell(sheet, "A8", "forecast_completeness", CellKind.STATIC)
        self.set_cell(sheet, "B8", "Every scenario/model-line row has a numeric value for every forecast year", CellKind.STATIC)
        self.set_formula(sheet, "D8", f"=COUNT({forecast_range})", kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "E8", expected_forecast_values, CellKind.STATIC)
        forecast_gap_state = (
            "NEEDS_PM_REVIEW"
            if any(item.startswith(("pm_approval_required:", "source_or_pm_required:")) for item in self.blockers)
            else "FAIL"
        )
        self.add_check("forecast_completeness", sheet, "C8")

        balance_terms: list[str] = []
        balance_inputs_complete = True
        balance_value_refs: list[str] = []
        for scenario in REQUIRED_SCENARIOS:
            for period in self.payload.forecast_periods:
                assets = self.output_ref(scenario, "total_assets", period)
                liabilities_equity = self.output_ref(scenario, "total_liabilities_and_equity", period)
                if not assets or not liabilities_equity:
                    balance_inputs_complete = False
                    continue
                balance_terms.append(f"ABS({assets}-{liabilities_equity})")
                balance_value_refs.extend((assets, liabilities_equity))
        self.set_cell(sheet, "A9", "balance_sheet", CellKind.STATIC)
        self.set_cell(sheet, "B9", "Maximum absolute assets-less-claims residual across all scenarios and forecast years", CellKind.STATIC)
        if balance_inputs_complete and len(balance_terms) == len(REQUIRED_SCENARIOS) * len(self.payload.forecast_periods):
            expected_balance_values = len(REQUIRED_SCENARIOS) * len(self.payload.forecast_periods) * 2
            balance_count = "COUNT(" + ",".join(balance_value_refs) + ")"
            self.set_formula(sheet, "D9", f'=IF({balance_count}={expected_balance_values},MAX(' + ",".join(balance_terms) + '),"")', kind=CellKind.CHECK, number_format="#,##0.0;[Red](#,##0.0);-")
            self.set_cell(sheet, "E9", 0.1, CellKind.STATIC)
            self.set_formula(sheet, "C9", f'=IF({balance_count}<>{expected_balance_values},"BLOCKED",IF(ABS(D9)<=E9,"PASS","FAIL"))', kind=CellKind.CHECK)
        else:
            self.set_cell(sheet, "C9", "BLOCKED", CellKind.CHECK)
            self.set_cell(sheet, "D9", "Required scenario/year lines absent", CellKind.UNAVAILABLE)
            self.set_cell(sheet, "E9", 0.1, CellKind.STATIC)
            self.blockers.add("balance_sheet_check_inputs_absent")
        self.add_check("balance_sheet", sheet, "C9")

        actual_anchor_period = self.payload.historical_periods[-1]
        historical_cash_location = self.historical_cells.get(("cash", actual_anchor_period))
        historical_cash_ref = _absolute_ref(*historical_cash_location) if historical_cash_location else None
        cash_terms: list[str] = []
        cash_inputs_complete = historical_cash_ref is not None
        cash_value_refs: list[str] = [historical_cash_ref] if historical_cash_ref else []
        for scenario in REQUIRED_SCENARIOS:
            prior_cash = historical_cash_ref
            for period in self.payload.forecast_periods:
                ending_cash = self.output_ref(scenario, "cash", period)
                net_change_cash = self.output_ref(scenario, "cf.net_change_cash", period)
                if not ending_cash or not net_change_cash or not prior_cash:
                    cash_inputs_complete = False
                else:
                    cash_terms.append(f"ABS({ending_cash}-({prior_cash}+{net_change_cash}))")
                prior_cash = ending_cash
                if ending_cash and net_change_cash:
                    cash_value_refs.extend((ending_cash, net_change_cash))
        self.set_cell(sheet, "A10", "cash_flow_tie", CellKind.STATIC)
        self.set_cell(sheet, "B10", "Maximum absolute ending-cash roll-forward residual across all scenarios and forecast years", CellKind.STATIC)
        if cash_inputs_complete and len(cash_terms) == len(REQUIRED_SCENARIOS) * len(self.payload.forecast_periods):
            expected_cash_values = 1 + len(REQUIRED_SCENARIOS) * len(self.payload.forecast_periods) * 2
            cash_count = "COUNT(" + ",".join(cash_value_refs) + ")"
            self.set_formula(sheet, "D10", f'=IF({cash_count}={expected_cash_values},MAX(' + ",".join(cash_terms) + '),"")', kind=CellKind.CHECK, number_format="#,##0.0;[Red](#,##0.0);-")
            self.set_cell(sheet, "E10", 0.1, CellKind.STATIC)
            self.set_formula(sheet, "C10", f'=IF({cash_count}<>{expected_cash_values},"BLOCKED",IF(ABS(D10)<=E10,"PASS","FAIL"))', kind=CellKind.CHECK)
        else:
            self.set_cell(sheet, "C10", "BLOCKED", CellKind.CHECK)
            self.set_cell(sheet, "D10", "Required scenario/year cash roll-forward lines absent", CellKind.UNAVAILABLE)
            self.set_cell(sheet, "E10", 0.1, CellKind.STATIC)
            self.blockers.add("cash_flow_check_inputs_absent")
        self.add_check("cash_flow_tie", sheet, "C10")

        self.set_cell(sheet, "A11", "valuation_bridge", CellKind.STATIC)
        self.set_cell(sheet, "B11", "Maximum component, reported-debt, and EV-to-equity bridge residual across scenarios", CellKind.STATIC)
        self.set_formula(sheet, "D11", '=IF(AND(DCF_Calculation_Gate="PASS",COUNT(Valuation!B34:D37)=12),MAX(ABS(Valuation!B34),ABS(Valuation!B35),ABS(Valuation!B36),ABS(Valuation!B37),ABS(Valuation!C34),ABS(Valuation!C35),ABS(Valuation!C36),ABS(Valuation!C37),ABS(Valuation!D34),ABS(Valuation!D35),ABS(Valuation!D36),ABS(Valuation!D37)),"")', kind=CellKind.CHECK, number_format="#,##0.0;[Red](#,##0.0);-")
        self.set_cell(sheet, "E11", 0.1, CellKind.STATIC)
        self.set_formula(sheet, "C11", '=IF(DCF_Calculation_Gate<>"PASS","BLOCKED",IF(COUNT(Valuation!B34:D37)<>12,"PARTIAL",IF(NOT(ISNUMBER(D11)),"BLOCKED",IF(ABS(D11)<=E11,"PASS","FAIL"))))', kind=CellKind.CHECK)
        self.add_check("valuation_bridge", sheet, "C11")

        dcf_gate_state = {
            "ELIGIBLE": "PASS",
            "NEEDS_PM_REVIEW": "NEEDS_PM_REVIEW",
            "INELIGIBLE": "BLOCKED",
            "UNVERIFIED": "UNVERIFIED",
        }.get(self.dcf_decision_state, "UNVERIFIED")
        self.set_cell(sheet, "A12", "dcf_decision_gate", CellKind.STATIC)
        self.set_cell(sheet, "B12", "DCF method availability and decision eligibility are evaluated separately", CellKind.STATIC)
        self.set_cell(sheet, "C12", dcf_gate_state, CellKind.CHECK)
        self.set_cell(sheet, "D12", 0 if dcf_gate_state == "PASS" else 1, CellKind.CHECK)
        self.set_cell(sheet, "E12", 0, CellKind.STATIC)
        self.add_check("dcf_decision_gate", sheet, "C12")

        valuation_inputs_available = not any(item.startswith("missing_dcf_valuation_inputs") for item in self.blockers)
        valuation_auxiliary_available = not any(
            item.startswith(("missing_bridge_reconciliation_inputs", "missing_price_comparison_inputs"))
            for item in self.blockers
        )
        valuation_gate_state = (
            "BLOCKED" if not valuation_inputs_available
            else "PASS" if valuation_auxiliary_available
            else "PARTIAL"
        )
        self.set_cell(sheet, "A13", "valuation_input_gate", CellKind.STATIC)
        self.set_cell(sheet, "B13", "Intrinsic DCF inputs are required; bridge-reference and exact-dated price evidence are separate comparison gates", CellKind.STATIC)
        self.set_cell(sheet, "C13", valuation_gate_state, CellKind.CHECK)
        self.set_cell(sheet, "D13", sum(item.startswith(("missing_dcf_valuation_inputs", "missing_bridge_reconciliation_inputs", "missing_price_comparison_inputs")) for item in self.blockers), CellKind.CHECK)
        self.set_cell(sheet, "E13", 0, CellKind.STATIC)
        self.add_check("valuation_input_gate", sheet, "C13")

        self.set_cell(sheet, "A14", "scenario_selector", CellKind.STATIC)
        self.set_cell(sheet, "B14", "Selected scenario must exactly match base, upside, or downside", CellKind.STATIC)
        self.set_formula(sheet, "C14", '=IF(OR(Assumptions!$B$4="base",Assumptions!$B$4="upside",Assumptions!$B$4="downside"),"PASS","FAIL")', kind=CellKind.CHECK)
        self.set_formula(sheet, "D14", '=COUNTIF(Assumptions!$B$4,"base")+COUNTIF(Assumptions!$B$4,"upside")+COUNTIF(Assumptions!$B$4,"downside")', kind=CellKind.CHECK, number_format="0")
        self.set_cell(sheet, "E14", 1, CellKind.STATIC)
        self.add_check("scenario_selector", sheet, "C14")

        self.set_cell(sheet, "A15", "calculation_verification", CellKind.STATIC)
        self.set_cell(sheet, "B15", "External record must bind workbook SHA, model input hash, formula parity, cache scan, errors, engine, and timestamp", CellKind.STATIC)
        self.set_cell(sheet, "C15", "UNVERIFIED", CellKind.CHECK)
        self.set_cell(sheet, "D15", 1, CellKind.CHECK)
        self.set_cell(sheet, "E15", 0, CellKind.STATIC)
        self.add_check("calculation_verification", sheet, "C15")

        pm_blockers = tuple(
            item
            for item in self.blockers
            if item.startswith(("pm_approval_required:", "source_or_pm_required:"))
        )
        finance_semantic_records = tuple(
            approval
            for approval in self.payload.driver_approvals
            if approval.driver_group.value == "finance_semantic"
        )
        nonapproved_records = tuple(
            approval
            for approval in finance_semantic_records
            if approval.approval_state.value != "APPROVED"
        )
        approval_issue_count = (
            len(pm_blockers)
            + len(nonapproved_records)
            + (0 if finance_semantic_records else 1)
        )
        self.set_cell(sheet, "A16", "pm_driver_approvals", CellKind.STATIC)
        self.set_cell(sheet, "B16", "Finance-semantic driver approvals must be current and fingerprint-matched", CellKind.STATIC)
        self.set_cell(sheet, "C16", "NEEDS_PM_REVIEW" if approval_issue_count else "PASS", CellKind.CHECK)
        self.set_cell(sheet, "D16", approval_issue_count, CellKind.CHECK)
        self.set_cell(sheet, "E16", 0, CellKind.STATIC)
        self.add_check("pm_driver_approvals", sheet, "C16")

        optional_states = tuple(self.payload.availability[key].status for key in ("segments", "consensus", "sotp"))
        consensus_usable = (
            self.payload.consensus_snapshot is not None
            and bool(self.payload.consensus_snapshot.observations)
            and all(
                item.mapping_method.value == "EXACT_PERIOD_END"
                for item in self.payload.consensus_snapshot.observations
            )
        )
        optional_complete = all(state == "available" for state in optional_states) and consensus_usable
        self.set_cell(sheet, "A17", "optional_modules", CellKind.STATIC)
        self.set_cell(sheet, "B17", "Segment, consensus, and SOTP modules are scoped separately from core calculation availability", CellKind.STATIC)
        self.set_cell(sheet, "C17", "PASS" if optional_complete else "PARTIAL", CellKind.CHECK)
        self.set_cell(sheet, "D17", sum(state != "available" for state in optional_states), CellKind.CHECK)
        self.set_cell(sheet, "E17", 0, CellKind.STATIC)
        self.add_check("optional_modules", sheet, "C17")


        def backend_group_state(tokens: tuple[str, ...]) -> tuple[str, int]:
            matching = tuple(
                str(value).strip().upper()
                for key, value in self.payload.backend_checks.items()
                if key.endswith(".status") and any(token in key for token in tokens)
            )
            if not matching:
                return "BLOCKED", 0
            if any(value in {"FAIL", "BLOCKED", "UNAVAILABLE", "ERROR"} for value in matching):
                return "BLOCKED", len(matching)
            if all(value == "PASS" for value in matching):
                return "PASS", len(matching)
            if any(value in {"NEEDS_PM_REVIEW", "PM_REQUIRED", "REVIEW"} for value in matching):
                return "NEEDS_PM_REVIEW", len(matching)
            return "UNVERIFIED", len(matching)

        def static_check(
            row: int,
            check_id: str,
            description: str,
            state: str,
            difference: Any,
            tolerance: Any = 0,
        ) -> None:
            self.set_cell(sheet, f"A{row}", check_id, CellKind.STATIC)
            self.set_cell(sheet, f"B{row}", description, CellKind.STATIC)
            self.set_cell(sheet, f"C{row}", state, CellKind.CHECK)
            self.set_cell(sheet, f"D{row}", difference, CellKind.CHECK)
            self.set_cell(sheet, f"E{row}", tolerance, CellKind.STATIC)
            self.add_check(check_id, sheet, f"C{row}")

        row = 18
        package_backend_state = str(
            self.payload.backend_checks.get("workflow.package.state") or "UNVERIFIED"
        ).strip().upper()
        package_gate_state = {
            "FULL": "PASS",
            "PARTIAL": "PARTIAL",
            "NEEDS_PM_REVIEW": "NEEDS_PM_REVIEW",
            "BLOCKED": "BLOCKED",
            "UNVERIFIED": "UNVERIFIED",
        }.get(package_backend_state, "UNVERIFIED")
        static_check(
            row,
            "backend_package_workflow",
            "Canonical backend package workflow state must reconcile to workbook gates",
            package_gate_state,
            package_backend_state,
        )
        row += 1

        static_check(
            row,
            "scenario_formula_first_architecture",
            "Visible driver matrix and schedule formulas must structurally control every scenario output",
            "BLOCKED",
            1,
        )
        self.blockers.add("scenario_formula_first_architecture_not_implemented")
        row += 1

        annual_refs = tuple(
            self.assumption_ref(f"dcf_annual_fy26_fcff_{scenario}")
            for scenario in REQUIRED_SCENARIOS
        )
        ytd_refs = tuple(
            self.assumption_ref(f"dcf_ytd_fcff_{scenario}")
            for scenario in REQUIRED_SCENARIOS
        )
        stub_refs = tuple(
            self.assumption_ref(f"dcf_stub_fcff_{scenario}")
            for scenario in REQUIRED_SCENARIOS
        )
        all_stub_refs = tuple(reference for reference in (*annual_refs, *ytd_refs, *stub_refs) if reference)
        stub_differences = tuple(
            f"ABS({annual_ref}-{ytd_ref}-{stub_ref})"
            for annual_ref, ytd_ref, stub_ref in zip(annual_refs, ytd_refs, stub_refs, strict=True)
            if annual_ref and ytd_ref and stub_ref
        )
        self.set_cell(sheet, f"A{row}", "first_year_stub", CellKind.STATIC)
        self.set_cell(sheet, f"B{row}", "FY26 annual FCFF must equal nine-month actual plus Q4 stub for every scenario", CellKind.STATIC)
        if len(all_stub_refs) == 9 and len(stub_differences) == 3:
            self.set_formula(
                sheet,
                f"D{row}",
                f'=IF(COUNT({",".join(all_stub_refs)})=9,MAX({",".join(stub_differences)}),"")',
                kind=CellKind.CHECK,
                number_format="#,##0.0;[Red](#,##0.0);-",
            )
            self.set_formula(
                sheet,
                f"C{row}",
                f'=IF(COUNT({",".join(all_stub_refs)})<>9,"BLOCKED",IF(D{row}<=E{row},"PASS","FAIL"))',
                kind=CellKind.CHECK,
            )
        else:
            self.set_cell(sheet, f"C{row}", "BLOCKED", CellKind.CHECK)
            self.set_cell(sheet, f"D{row}", "Typed stub inputs absent", CellKind.UNAVAILABLE)
        self.set_cell(sheet, f"E{row}", 0.1, CellKind.STATIC)
        self.add_check("first_year_stub", sheet, f"C{row}")
        row += 1

        grouped_checks = (
            ("statement_identities", ("balance_sheet:", "cash_flow:"), "Historical and forecast statement identities"),
            ("tax_current_deferred_cash", ("tax_payable_roll_forward", "tax_cash_conversion", "book_current_deferred_tax_identity", "deferred_tax_"), "Book/current/deferred/cash tax and DTA/DTL roll-forwards"),
            ("ppe_intangibles_rollforwards", ("ppe_gross_roll_forward", "accumulated_depreciation", "net_ppe", "da_split", "amortization", "net_intangibles"), "Gross/accumulated/net PP&E and intangible schedules"),
            ("debt_leases_interest", ("debt_roll_forward", "lease_liabilities", "interest_expense", "interest_income", "reported_debt"), "Debt, leases, and pretax interest identities"),
            ("shares_eps_dps", ("shares_tie", "basic_share", "diluted", "eps_tie", "dividend", "stock_compensation"), "Basic/FDSO, EPS, DPS, dividend, and stock-compensation identities"),
            ("retained_earnings_equity", ("retained_earnings", "apic_roll_forward", "treasury_stock", "aoci_roll_forward"), "Retained earnings and equity roll-forwards"),
            ("working_capital_definition", ("receivables_driver", "inventory_driver", "payables_driver", "deferred_revenue_driver"), "Operating working-capital definitions and driver ties"),
            ("fcff_definition_consistency", ("fcff_definition",), "Historical and forecast FCFF use one definition"),
        )
        for check_id, tokens, description in grouped_checks:
            state, count = backend_group_state(tokens)
            static_check(row, check_id, description, state, count)
            row += 1

        segment_state, segment_count = backend_group_state(("segment_kpi_evidence", "segment_reconciliation"))
        static_check(
            row,
            "segment_reconciliation",
            "Segment revenue and operating profit reconcile to consolidated results; assets/KPIs remain typed",
            segment_state,
            segment_count,
        )
        row += 1

        consensus_exact = (
            self.payload.consensus_snapshot is not None
            and bool(self.payload.consensus_snapshot.observations)
            and all(
                item.mapping_method.value == "EXACT_PERIOD_END"
                and item.mapped_model_period_end == item.period_end
                for item in self.payload.consensus_snapshot.observations
            )
        )
        static_check(
            row,
            "consensus_period_basis",
            "Consensus preserves CY/FY/NTM/quarter type and exact period end; no CY+1 to FY1 alias",
            "PARTIAL" if consensus_exact else "BLOCKED",
            len(self.payload.consensus_snapshot.observations) if self.payload.consensus_snapshot else 0,
        )
        row += 1

        static_check(
            row,
            "comps_comparability",
            "Peers require market as-of, financial period, metric definition, tier, statistic, locator, and bridge parity",
            "PARTIAL" if self.payload.comparables else "BLOCKED",
            len(self.payload.comparables),
        )
        row += 1

        sotp_ready = (
            self.payload.availability["sotp"].status == "available"
            and bool(self.payload.sotp_components)
            and bool(self.assumption_ref("current_fully_diluted_shares"))
        )
        static_check(
            row,
            "sotp_methodology",
            "Typed segment evidence, peer methodology, consolidated bridge, current FDSO, and approval",
            "NEEDS_PM_REVIEW" if sotp_ready else "BLOCKED",
            len(self.payload.sotp_components),
        )
        row += 1

        self.set_cell(sheet, f"A{row}", "wacc_parity", CellKind.STATIC)
        self.set_cell(sheet, f"B{row}", "Workbook/backend WACC methodology parity must be within one basis point", CellKind.STATIC)
        self.set_formula(sheet, f"C{row}", '=IF(WACC!$E$7="PASS","PASS","BLOCKED")', kind=CellKind.CHECK)
        self.set_formula(sheet, f"D{row}", '=WACC!$E$6', kind=CellKind.CHECK, number_format="0.0")
        self.set_cell(sheet, f"E{row}", 1.0, CellKind.STATIC)
        self.add_check("wacc_parity", sheet, f"C{row}")
        row += 1

        self.set_cell(sheet, f"A{row}", "dcf_sensitivity_central_tie", CellKind.STATIC)
        self.set_cell(sheet, f"B{row}", "Independently calculated central sensitivity cell must tie DCF base per-share value", CellKind.STATIC)
        self.set_formula(
            sheet,
            f"D{row}",
            '=IF(COUNT(Sensitivities!E9,DCF_Base_Per_Share)=2,Sensitivities!E9-DCF_Base_Per_Share,"")',
            kind=CellKind.CHECK,
            number_format="$0.00;[Red]($0.00);-",
        )
        self.set_formula(
            sheet,
            f"C{row}",
            f'=IF(NOT(ISNUMBER(D{row})),"BLOCKED",IF(ABS(D{row})<=E{row},"PASS","FAIL"))',
            kind=CellKind.CHECK,
        )
        self.set_cell(sheet, f"E{row}", 0.01, CellKind.STATIC)
        self.add_check("dcf_sensitivity_central_tie", sheet, f"C{row}")
        row += 1

        self.set_cell(sheet, f"A{row}", "ev_bridge_independent", CellKind.STATIC)
        self.set_cell(sheet, f"B{row}", "Duplicated enterprise-to-equity bridges must independently reconcile in every scenario", CellKind.STATIC)
        self.set_formula(
            sheet,
            f"D{row}",
            '=IF(COUNT(Valuation!B34:D37)=12,MAX(ABS(Valuation!B34),ABS(Valuation!B35),ABS(Valuation!B36),ABS(Valuation!B37),ABS(Valuation!C34),ABS(Valuation!C35),ABS(Valuation!C36),ABS(Valuation!C37),ABS(Valuation!D34),ABS(Valuation!D35),ABS(Valuation!D36),ABS(Valuation!D37)),"")',
            kind=CellKind.CHECK,
            number_format="#,##0.0;[Red](#,##0.0);-",
        )
        self.set_formula(
            sheet,
            f"C{row}",
            f'=IF(NOT(ISNUMBER(D{row})),"BLOCKED",IF(ABS(D{row})<=E{row},"PASS","FAIL"))',
            kind=CellKind.CHECK,
        )
        self.set_cell(sheet, f"E{row}", 0.1, CellKind.STATIC)
        self.add_check("ev_bridge_independent", sheet, f"C{row}")
        row += 1

        self.set_cell(sheet, f"A{row}", "terminal_value_concentration", CellKind.STATIC)
        self.set_cell(sheet, f"B{row}", "Terminal-value concentration is calculated; the decision threshold requires PM policy", CellKind.STATIC)
        self.set_formula(sheet, f"D{row}", '=DCF!$C$31', kind=CellKind.CHECK, number_format="0.0%")
        self.set_formula(sheet, f"C{row}", f'=IF(ISNUMBER(D{row}),"NEEDS_PM_REVIEW","BLOCKED")', kind=CellKind.CHECK)
        self.set_cell(sheet, f"E{row}", "PM threshold required", CellKind.STATIC)
        self.add_check("terminal_value_concentration", sheet, f"C{row}")
        row += 1

        fdso_value = self.payload.valuation_inputs.get("current_fully_diluted_shares")
        fdso_state = "BLOCKED"
        static_check(
            row,
            "current_fdso_evidence",
            "Current fully diluted shares require exact-dated options/RSU/PSU/convert evidence and approved treasury-stock/if-converted rules",
            fdso_state,
            fdso_value if fdso_value is not None else "UNAVAILABLE",
        )
        row += 1

        static_check(
            row,
            "formula_cache_circular_sign_unit",
            "Native formula/cache parity, circularity, formula errors, signs, and units require authoritative recalculation",
            "UNVERIFIED",
            1,
        )
        row += 1

        static_check(
            row,
            "source_lineage",
            "Raw, normalized, derived, transform, unit, period, source cell, run, and hash lineage",
            "PASS" if source_pass else "BLOCKED",
            self.payload.source.run_id,
        )
        row += 1

        static_check(
            row,
            "artifact_review_identity",
            "Workbook SHA, model-input hash, formula hash, source identity, repair ledger, and consumed approval evidence",
            "UNVERIFIED",
            self.model_input_hash,
        )
        row += 1

        hard_blockers = tuple(
            item
            for item in self.blockers
            if not item.startswith(("pm_approval_required:", "source_or_pm_required:"))
        )
        unresolved_state = "BLOCKED" if hard_blockers else "NEEDS_PM_REVIEW" if self.blockers else "PASS"
        static_check(
            row,
            "unresolved_blocker_inventory",
            "Every blocker is normalized to one root cause and appears in the PM queue",
            unresolved_state,
            len(self.blockers),
        )
        row += 1

        last_required_row = row - 1
        required_status_range = f"C6:C{last_required_row}"
        required_id_range = f"A6:A{last_required_row}"
        self.set_formula(
            sheet,
            "C5",
            (
                f'=IF(OR(COUNTIF({required_status_range},"FAIL")+COUNTIF({required_status_range},"BLOCKED")>0,'
                f'SUMPRODUCT(--(COUNTIF({required_id_range},{required_id_range})=1))<>ROWS({required_id_range})),'
                f'"BLOCKED",IF(COUNTIF({required_status_range},"NEEDS_PM_REVIEW")>0,"NEEDS_PM_REVIEW",'
                f'IF(COUNTIF({required_status_range},"PASS")=ROWS({required_status_range}),"FULL",'
                f'IF(COUNTIF({required_status_range},"PASS")+COUNTIF({required_status_range},"PARTIAL")='
                f'ROWS({required_status_range}),"PARTIAL","UNVERIFIED"))))'
            ),
            kind=CellKind.CHECK,
        )
        self.set_formula(
            sheet,
            "D5",
            f'=ROWS({required_status_range})-COUNTIF({required_status_range},"PASS")',
            kind=CellKind.CHECK,
            number_format="0",
        )

        worksheet = self.workbook[sheet]
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"FULL"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"BLOCKED"'], fill=PatternFill("solid", fgColor=PALE_RED)))
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=PALE_RED)))
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"NEEDS_PM_REVIEW"'], fill=PatternFill("solid", fgColor=PALE_YELLOW)))
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"UNVERIFIED"'], fill=PatternFill("solid", fgColor=PALE_YELLOW)))
        worksheet.conditional_formatting.add(f"C5:C{last_required_row}", CellIsRule(operator="equal", formula=['"PARTIAL"'], fill=PatternFill("solid", fgColor=LIGHT_GRAY)))
        for column, width in {"A": 30, "B": 72, "C": 16, "D": 22, "E": 14}.items():
            worksheet.column_dimensions[column].width = width
        self.finish_sheet(sheet, "C5")
    def render_summary(self) -> None:
        sheet = "Summary"
        self.title(sheet, "Executive decision summary", "Scenario operating paths, evidence gates, and separate valuation methods; no blended target")
        self.set_cell(sheet, "A4", "Model status", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY))
        self.set_formula(sheet, "B4", "=Checks!$C$5", kind=CellKind.CHECK)
        self.set_cell(sheet, "A5", "Selected scenario", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, "B5", "=Selected_Scenario")
        self.set_formula(sheet, "C5", '=IF(OR(B5="base",B5="upside",B5="downside"),"PASS","FAIL")', kind=CellKind.CHECK)
        self.set_cell(sheet, "A6", "Source status", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_formula(sheet, "B6", "=Cover!$B$8")
        self.set_cell(sheet, "A7", "Current price source", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B7", self.payload.current_price_source or "UNVERIFIED", CellKind.SOURCE)
        self.set_cell(sheet, "A8", "Current price as of", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B8", self.payload.current_price_as_of or "UNVERIFIED", CellKind.SOURCE)
        self.header_row(sheet, 9, ("Scenario", "DCF / share", "Current price", "Upside / (downside)", "TV / EV", "Timing", "Decision state"))
        dcf_starts = {"base": 5, "upside": 35, "downside": 65}
        current_price_ref = self.assumption_ref("current_price")
        for row, scenario in enumerate(REQUIRED_SCENARIOS, start=10):
            start = dcf_starts[scenario]
            self.set_cell(sheet, f"A{row}", scenario.upper(), CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
            self.set_formula(sheet, f"B{row}", f'=IF(DCF_Calculation_Gate="PASS",DCF!$C${start + 20},"")', number_format="$0.00")
            self.set_formula(sheet, f"C{row}", f'=IF(ISNUMBER({current_price_ref}),{current_price_ref},"")' if current_price_ref else '=""', number_format="$0.00")
            self.set_formula(sheet, f"D{row}", f'=IF(DCF_Calculation_Gate="PASS",DCF!$C${start + 22},"")', number_format="0.0%;[Red](0.0%)")
            self.set_formula(sheet, f"E{row}", f'=IF(DCF_Calculation_Gate="PASS",DCF!$C${start + 26},"")', number_format="0.0%")
            self.set_formula(sheet, f"F{row}", f'=IF(DCF_Calculation_Gate="PASS",DCF!$C${start + 25},"")')
            self.set_formula(sheet, f"G{row}", '=DCF!$E$3', kind=CellKind.CHECK)

        self.set_cell(sheet, "A15", "Separate comparative diagnostics", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY))
        self.header_row(sheet, 16, ("Method", "Diagnostic value", "Units", "Evidence state", "Decision use"))
        for row, (key, label, unit_kind) in enumerate(DIAGNOSTIC_METHODS, start=17):
            reference = self.diagnostic_ref(key)
            self.set_cell(sheet, f"A{row}", label, CellKind.STATIC)
            if reference:
                self.set_formula(sheet, f"B{row}", f"={reference}", number_format="$0.00" if unit_kind == "per_share" else '0.0"%"')
            else:
                self.set_formula(sheet, f"B{row}", '=""', number_format="$0.00" if unit_kind == "per_share" else '0.0"%"')
            self.set_cell(sheet, f"C{row}", f"{self.payload.currency}/share" if unit_kind == "per_share" else "percentage points", CellKind.STATIC)
            state = "BLOCKED" if self.payload.source.status not in {"ready", "ok"} or self.payload.source.formula_error_count else "DIAGNOSTIC" if reference else "UNAVAILABLE"
            self.set_cell(sheet, f"D{row}", state, CellKind.CHECK if state == "BLOCKED" else CellKind.UNAVAILABLE)
            self.set_cell(sheet, f"E{row}", "NON-APPROVED; never blended, weighted, or presented as a target", CellKind.STATIC)

        fcfe_row = 25
        fcfe_state = str(self.payload.backend_checks.get("fcfe.state") or "UNAVAILABLE").upper()
        fcfe_reason = str(self.payload.backend_checks.get("fcfe.reason_code") or "fcfe_method_evidence_not_provided")
        fcfe_detail = str(self.payload.backend_checks.get("fcfe.detail") or "No FCFE detail supplied.")
        self.set_cell(sheet, f"A{fcfe_row}", "FCFE cross-check", CellKind.STATIC)
        self.set_formula(sheet, f"B{fcfe_row}", '=""')
        self.set_cell(sheet, f"C{fcfe_row}", "per share", CellKind.STATIC)
        self.set_cell(sheet, f"D{fcfe_row}", fcfe_state, CellKind.UNAVAILABLE)
        self.set_cell(sheet, f"E{fcfe_row}", f"{fcfe_reason}: {fcfe_detail}", CellKind.STATIC, alignment=Alignment(wrap_text=True))

        sotp_row = 26
        sotp_available = "SOTP_Per_Share" in self.defined_names
        self.set_cell(sheet, f"A{sotp_row}", "SOTP", CellKind.STATIC)
        self.set_formula(sheet, f"B{sotp_row}", "=SOTP_Per_Share" if sotp_available else '=""', number_format="$0.00")
        self.set_cell(sheet, f"C{sotp_row}", f"{self.payload.currency}/share", CellKind.STATIC)
        self.set_cell(sheet, f"D{sotp_row}", "AVAILABLE" if sotp_available else "UNAVAILABLE", CellKind.STATIC if sotp_available else CellKind.UNAVAILABLE)
        self.set_cell(sheet, f"E{sotp_row}", "Requires normalized segment evidence.", CellKind.STATIC)

        historical_row = 27
        self.set_cell(sheet, f"A{historical_row}", "Historical trading range", CellKind.STATIC)
        self.set_formula(sheet, f"B{historical_row}", '=""')
        self.set_cell(sheet, f"C{historical_row}", f"{self.payload.currency}/share", CellKind.STATIC)
        self.set_cell(sheet, f"D{historical_row}", "UNAVAILABLE", CellKind.UNAVAILABLE)
        self.set_cell(sheet, f"E{historical_row}", "No source-backed historical valuation-range payload supplied.", CellKind.STATIC)

        self.set_cell(sheet, "A30", "Selected-scenario dashboard", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY))
        selected_rows = ((31, "DCF / share", "B", "$0.00"), (32, "Current price", "C", "$0.00"), (33, "Upside / (downside)", "D", "0.0%"), (34, "TV / EV", "E", "0.0%"), (35, "Timing convention", "F", None))
        for row, label, source_column, number_format in selected_rows:
            self.set_cell(sheet, f"A{row}", label, CellKind.STATIC)
            lookup = f"INDEX({source_column}$10:{source_column}$12,MATCH(UPPER($B$5),$A$10:$A$12,0))"
            formula = (
                f'=IFERROR(IF(AND($C$5="PASS",ISNUMBER({lookup})),{lookup},""),"")'
                if label == "Current price"
                else f'=IFERROR(IF(AND($C$5="PASS",DCF_Calculation_Gate="PASS"),{lookup},""),"")'
            )
            self.set_formula(sheet, f"B{row}", formula, number_format=number_format)
        self.set_cell(sheet, "A37", "Valuation policy", CellKind.STATIC, font=Font(name="Arial", size=10, bold=True, color=NAVY))
        self.set_cell(sheet, "B37", "DCF, each diagnostic, FCFE, SOTP, and historical range remain separate. No probability-weighted or blended value is presented.", CellKind.STATIC, alignment=Alignment(wrap_text=True))

        self.set_cell(sheet, "A40", "Scenario operating paths", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY))
        self.header_row(sheet, 41, ("Scenario", "Metric", *self.payload.forecast_periods, "Evidence / decision state"))
        scenario_path_state = (
            "NEEDS_PM_REVIEW"
            if any(item.startswith("pm_approval_required:") for item in self.blockers)
            else "UNVERIFIED"
        )
        row = 42
        path_rows: dict[tuple[str, str], int] = {}
        for scenario in REQUIRED_SCENARIOS:
            start = dcf_starts[scenario]
            metrics = (
                ("DCF / share (terminal output)", "dcf"),
                ("Revenue", "revenue"),
                ("EBIT margin", "ebit_margin"),
                ("Diluted EPS", "shares.diluted_eps"),
                ("Integrated FCFF", "cf.unlevered_fcf"),
            )
            for label, metric_key in metrics:
                path_rows[(scenario, metric_key)] = row
                self.set_cell(sheet, f"A{row}", scenario.upper(), CellKind.STATIC, font=Font(name="Arial", size=9, bold=True, color=NAVY))
                self.set_cell(sheet, f"B{row}", label, CellKind.STATIC)
                for period_index, period in enumerate(self.payload.forecast_periods, start=3):
                    coordinate = f"{get_column_letter(period_index)}{row}"
                    if metric_key == "dcf":
                        formula = (
                            f'=IF(AND(DCF_Calculation_Gate="PASS",{period_index}={2 + len(self.payload.forecast_periods)}),DCF!$C${start + 20},"")'
                        )
                        number_format = "$0.00"
                    elif metric_key == "ebit_margin":
                        revenue_ref = self.output_ref(scenario, "revenue", period)
                        ebit_ref = self.output_ref(scenario, "ebit", period)
                        formula = (
                            f'=IFERROR(IF(AND(ISNUMBER({ebit_ref}),ISNUMBER({revenue_ref}),{revenue_ref}<>0),{ebit_ref}/{revenue_ref},""),"")'
                            if revenue_ref and ebit_ref
                            else '=""'
                        )
                        number_format = "0.0%"
                    else:
                        output_ref = self.output_ref(scenario, metric_key, period)
                        formula = f"={output_ref}" if output_ref else '=""'
                        number_format = "$0.00" if metric_key == "shares.diluted_eps" else "#,##0.0;[Red](#,##0.0);-"
                    self.set_formula(sheet, coordinate, formula, number_format=number_format)
                state = self.dcf_decision_state if metric_key == "dcf" else scenario_path_state
                self.set_cell(sheet, f"H{row}", state, CellKind.CHECK)
                row += 1
            row += 1

        chart = LineChart()
        chart.title = "Base revenue and integrated FCFF"
        chart.y_axis.title = self.payload.unit_convention
        chart.x_axis.title = "Forecast period"
        chart.style = 13
        chart.height = 7
        chart.width = 15
        chart.set_categories(Reference(self.workbook[sheet], min_col=3, max_col=7, min_row=41))
        for metric_key in ("revenue", "cf.unlevered_fcf"):
            data_row = path_rows[("base", metric_key)]
            chart.add_data(
                Reference(self.workbook[sheet], min_col=2, max_col=7, min_row=data_row, max_row=data_row),
                titles_from_data=True,
                from_rows=True,
            )
        self.workbook[sheet].add_chart(chart, "J41")

        self.set_cell(sheet, "A62", "Decision context and falsification", CellKind.STATIC, font=Font(name="Arial", size=11, bold=True, color=NAVY))
        self.header_row(sheet, 63, ("Decision field", "State", "Source-backed statement", "Required next step"))
        context_rows = (
            ("variant_estimate_bridge", "Variant estimate / consensus bridge", "PARTIAL"),
            ("downside_mechanism", "Downside mechanism", "NEEDS_PM_REVIEW"),
            ("falsifier", "Thesis falsifier", "NEEDS_PM_REVIEW"),
            ("catalyst", "Catalyst", "NEEDS_PM_REVIEW"),
        )
        for context_row, (key, label, absent_state) in enumerate(context_rows, start=64):
            value = self.payload.decision_context.get(key)
            self.set_cell(sheet, f"A{context_row}", label, CellKind.STATIC)
            self.set_cell(sheet, f"B{context_row}", "NEEDS_PM_REVIEW" if value else absent_state, CellKind.CHECK)
            self.set_cell(sheet, f"C{context_row}", value or "Not supplied; no narrative has been invented.", CellKind.SOURCE if value else CellKind.UNAVAILABLE, alignment=Alignment(wrap_text=True))
            self.set_cell(sheet, f"D{context_row}", "PM must supply or approve explicit evidence before decision use.", CellKind.STATIC, alignment=Alignment(wrap_text=True))

        self.header_row(sheet, 4, ("Navigation", "Open"), start_col=10)
        for nav_row, (label, target) in enumerate((("Checks", "Checks"), ("Sources", "Sources"), ("PM queue", "PM_Review_Queue"), ("DCF", "DCF"), ("Valuation", "Valuation")), start=5):
            self.set_cell(sheet, f"J{nav_row}", label, CellKind.STATIC)
            self.set_cell(sheet, f"K{nav_row}", "OPEN", CellKind.STATIC, font=Font(name="Arial", size=10, color=MEDIUM_BLUE, underline="single"))
            self.workbook[sheet][f"K{nav_row}"].hyperlink = f"#'{target}'!A1"
        self.workbook[sheet].column_dimensions["A"].width = 36
        self.workbook[sheet].column_dimensions["B"].width = 30
        self.workbook[sheet].column_dimensions["C"].width = 22
        self.workbook[sheet].column_dimensions["D"].width = 24
        self.workbook[sheet].column_dimensions["E"].width = 30
        self.workbook[sheet].column_dimensions["F"].width = 18
        self.workbook[sheet].column_dimensions["G"].width = 18
        self.workbook[sheet].column_dimensions["H"].width = 24
        self.workbook[sheet].column_dimensions["J"].width = 18
        self.workbook[sheet].column_dimensions["K"].width = 12
        self.finish_sheet(sheet, "B10")
    def parity_results(self) -> tuple[CheckResult, ...]:
        source_pass = self.payload.source.status in {"ready", "ok"} and self.payload.source.formula_error_count == 0
        balance_inputs = all(
            self.output_ref(scenario, key, period)
            for scenario in REQUIRED_SCENARIOS
            for period in self.payload.forecast_periods
            for key in ("total_assets", "total_liabilities_and_equity")
        )
        fiscal_actuals = tuple(
            period
            for period in self.payload.historical_periods
            if re.fullmatch(r"FY[0-9]{2,4}", period.upper())
        )
        annual_anchor_period = fiscal_actuals[-1] if fiscal_actuals else self.payload.historical_periods[-1]
        cash_inputs = ("cash", annual_anchor_period) in self.historical_cells and all(
            self.output_ref(scenario, key, period)
            for scenario in REQUIRED_SCENARIOS
            for period in self.payload.forecast_periods
            for key in ("cash", "cf.net_change_cash")
        )
        dcf_inputs = self.dcf_state == "AVAILABLE"
        valuation_inputs = not any(item.startswith("missing_dcf_valuation_inputs") for item in self.blockers)
        return (
            CheckResult(
                check_id="source_preflight",
                status=CheckStatus.PASS if source_pass else CheckStatus.BLOCKED,
                difference=float(self.payload.source.formula_error_count),
                tolerance=0.0,
                message=None if source_pass else "Source preflight is blocked.",
            ),
            CheckResult(check_id="scenario_completeness", status=CheckStatus.PASS, difference=0.0, tolerance=0.0),
            CheckResult(check_id="forecast_completeness", status=CheckStatus.PASS, difference=0.0, tolerance=0.0),
            CheckResult(
                check_id="balance_sheet",
                status=CheckStatus.REVIEW if balance_inputs else CheckStatus.BLOCKED,
                message="All scenario/year residual formulas await native recalculation." if balance_inputs else "Required balance-sheet lines are absent.",
            ),
            CheckResult(
                check_id="cash_flow_tie",
                status=CheckStatus.REVIEW if cash_inputs else CheckStatus.BLOCKED,
                message="All scenario/year cash roll-forward formulas await native recalculation." if cash_inputs else "Required cash roll-forward lines are absent.",
            ),
            CheckResult(
                check_id="valuation_bridge",
                status=CheckStatus.REVIEW if dcf_inputs else CheckStatus.BLOCKED,
                message="Pending native recalculation." if dcf_inputs else "DCF calculation gate is not AVAILABLE.",
            ),
            CheckResult(
                check_id="dcf_input_gate",
                status=CheckStatus.PASS if dcf_inputs else CheckStatus.BLOCKED,
                difference=0.0 if dcf_inputs else 1.0,
                tolerance=0.0,
                message=None if dcf_inputs else "Source, WACC, integrated FCFF, valuation inputs, or PM policy is blocked/degraded.",
            ),
            CheckResult(
                check_id="valuation_input_gate",
                status=CheckStatus.PASS if valuation_inputs else CheckStatus.BLOCKED,
                difference=0.0 if valuation_inputs else 1.0,
                tolerance=0.0,
                message=None if valuation_inputs else "Required valuation inputs are absent.",
            ),
        )
    def build(self, output_path: Path) -> WorkbookManifest:
        self.validate()
        self.render_cover()
        self.render_sources()
        self.render_assumptions()
        self.render_historical_data()
        self.render_scenarios()
        self.render_model_sheets()
        self.render_wacc()
        self.render_dcf()
        self.render_comps()
        self.render_sotp()
        self.render_consensus_bridge()
        self.render_sensitivities()
        self.render_valuation()
        self.render_accounting_qoe()
        self.render_checks()
        self.render_pm_review_queue()
        self.render_summary()

        for sheet_name in REQUIRED_SHEET_ORDER:
            worksheet = self.workbook[sheet_name]
            worksheet.sheet_view.zoomScale = 90
            worksheet.oddFooter.center.text = f"{self.payload.ticker} | {self.payload.as_of_date.isoformat()} | Confidential"
            worksheet.oddFooter.right.text = "Page &P of &N"
            worksheet.print_area = worksheet.dimensions
        self.workbook.properties.creator = "ai-fund professional model renderer"
        self.workbook.properties.title = f"{self.payload.ticker} professional financial model"
        self.workbook.properties.subject = "Integrated historical, scenario, and valuation model"
        self.workbook.calculation.calcMode = "auto"
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True

        for worksheet in self.workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    formula = str(cell.value)
                    if "[" in formula or "]" in formula:
                        raise ValueError(f"external-link formula is prohibited at {worksheet.title}!{cell.coordinate}")
                    if any(token in formula for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!")):
                        raise ValueError(f"formula-error literal at {worksheet.title}!{cell.coordinate}")
                    if re.search(r"(?<![A-Za-z0-9_])None(?![A-Za-z0-9_])", formula):
                        raise ValueError(f"missing-reference literal at {worksheet.title}!{cell.coordinate}")
        expected_formula_text_hash = _formula_text_hash(self.workbook)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)

        payload_dict = asdict(self.payload)
        model_input_hash = _canonical_hash(payload_dict)
        result_hash = _canonical_hash(
            {
                "scenario_forecasts": [
                    {
                        "canonical_key": line.canonical_key,
                        "forecasts": [asdict(forecast) for forecast in line.scenario_forecasts],
                    }
                    for line in self.lines
                ],
                "valuation_inputs": dict(self.payload.valuation_inputs),
                "backend_checks": dict(self.payload.backend_checks),
            }
        )
        manifest_blockers = set(self.blockers)
        manifest_blockers.add("recalculation_not_run")
        return WorkbookManifest(
            ticker=self.payload.ticker,
            expected_formula_text_hash=expected_formula_text_hash,
            source_hash=self.payload.source.source_hash,
            model_input_hash=model_input_hash,
            result_hash=result_hash,
            sheet_order=REQUIRED_SHEET_ORDER,
            line_cell_mappings=tuple(self.line_mappings),
            cell_classifications=tuple(self.classifications.values()),
            defined_names=tuple(self.defined_names.values()),
            check_cells=tuple(self.check_cells.values()),
            renderer_version=RENDERER_VERSION,
            recalculation_state=AvailabilityState(
                status=AvailabilityStatus.BLOCKING,
                reason_code="recalculation_not_run",
                message="Native Excel full recalculation and parity verification have not been run.",
            ),
            parity_results=self.parity_results(),
            warnings=tuple(self.warnings),
            blockers=tuple(manifest_blockers),
        )


def render_professional_model_workbook(
    payload: NormalizedProfessionalWorkbookPayload,
    output_path: str | Path,
) -> WorkbookManifest:
    """Render a deterministic professional model and return its audit manifest.

    The returned manifest is intentionally ``blocking`` on native recalculation.
    A downstream artifact-QA stage should recalculate the workbook, verify cached
    values and formula parity, then produce the final distributable state.
    """

    if not isinstance(payload, NormalizedProfessionalWorkbookPayload):
        raise TypeError("payload must be a NormalizedProfessionalWorkbookPayload")
    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("output_path must use the .xlsx extension")
    return _Renderer(payload).build(path)


__all__ = [
    "ComparableCompany",
    "HistoricalSourceCell",
    "ModelLine",
    "NormalizedProfessionalWorkbookPayload",
    "RENDERER_VERSION",
    "REQUIRED_SCENARIOS",
    "REQUIRED_SHEET_ORDER",
    "SOTPComponent",
    "ScenarioForecast",
    "SourceWorkbookRun",
    "TypedAvailability",
    "render_professional_model_workbook",
]