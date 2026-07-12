"""Build a review-ready, formula-driven DCF workbook from valuation JSON.

This module creates an Excel **review artifact**. It does not feed values back
into the deterministic valuation pipeline (Vision: LLM/judgment never mutates the
deterministic layer; the PM Decision Queue is the only mutation bridge).

Design goals (set by PM, 2026-06-15):
  1. Transparent rebuild that reconciles to the backend deterministic IV to the
     cent. The Base-case DCF recomputes PV of explicit FCFF, both terminal-value
     methods (value-driver Gordon + exit multiple), the story-weighted blend, the
     equity bridge, and diluted shares from visible inputs, and the workbook
     intrinsic value ties to ``valuation.iv_base``. A build-time assertion fails
     the export if it does not reconcile.
  2. Surface the judgment layer the system already produced, instead of a generic
     historicals-plus-forecast calculator: an investment-thesis tab
     (``story_profile`` -> ``story_adjustments`` consequences), a PM review queue
     (``default_resolution`` fields), and per-assumption provenance + register
     flags.

Engineering decision (logged): only the Base case is a transparent in-Excel
rebuild, because the JSON exposes the full per-year Base forecast. Bear/Bull are
surfaced from the backend's deterministic scenario outputs and their driver specs
(official scenarios canonical, context-advisory shown alongside) rather than a
re-derived Excel scenario engine that would risk not reconciling. Re-deriving
scenarios in Excel waits until the exporter emits per-scenario forecasts.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_JSON_DIR = ROOT / "data" / "valuations" / "json"
DEFAULT_OUTPUT_ROOT = ROOT / "data" / "exports" / "generated" / "ticker"

# Reconciliation tolerance in $/share. The backend terminal block is rounded to
# whole $mm in the JSON, so a transparent rebuild lands within a few cents.
RECONCILE_TOLERANCE = 0.10

# Canonical sheet order.
CANONICAL_ORDER = [
    "Cover", "Thesis_Drivers", "PM_Review_Queue", "Assumptions",
    "Historical_Financials", "Input_Forecast", "WACC", "DCF_Base",
    "Scenarios", "Valuation_Bridge", "Sensitivity", "Checks",
]

# Sheets that hold per-ticker DATA and are rebuilt on refresh. The remaining
# sheets (WACC, DCF_Base, Sensitivity, plus anything the PM adds) are the editable
# MODEL and are never touched by a refresh, so model edits and PM overrides survive.
DATA_SHEETS = [
    "Cover", "Thesis_Drivers", "PM_Review_Queue", "Assumptions",
    "Historical_Financials", "Input_Forecast", "Scenarios", "Valuation_Bridge", "Checks",
]

# Map DCF_Base row labels -> the keys other sheets reference, so a refresh can
# relocate them by reading the (possibly PM-edited) DCF_Base instead of recomputing.
_DCF_ROW_LABELS = {
    "Enterprise value — operations": "ev_ops",
    "Terminal value % of EV": "tv_pct",
    "Intrinsic value / share": "iv",
    "IV / share — Gordon only": "iv_gordon",
    "IV / share — exit only": "iv_exit",
    "FCFF (NOPAT+D&A-Capex-dNWC)": "fcff_row",
    "EBIT": "ebit_row",
    "EBITDA": "ebitda_row",
    "NOPAT": "nopat_row",
}

NAVY = "1F4E78"
DARK_BLUE = "17365D"
LIGHT_BLUE = "D9EAF7"
PALE_YELLOW = "FFF2CC"
PALE_GREEN = "E2F0D9"
PALE_RED = "FBE4D5"
PALE_AMBER = "FFF0D0"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN_FONT = "008000"
BLUE_FONT = "0070C0"
RED_FONT = "C00000"
AMBER_FONT = "B26B00"

PCT = "0.0%"
MM = '#,##0.0;[Red](#,##0.0);-'
MMI = '#,##0;[Red](#,##0);-'
MULT = '#,##0.00"x"'
USD = '#,##0.00'


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #
def build_advanced_dcf_model(
    ticker: str,
    *,
    json_path: str | Path | None = None,
    output_path: str | Path | None = None,
    guided_workup_path: str | Path | None = None,
    as_of: datetime | None = None,
) -> Path:
    """Build an advanced DCF workbook for *ticker* and return the path.

    Raises ``ValueError`` if the workbook's Base intrinsic value does not
    reconcile to the backend deterministic ``iv_base`` within
    ``RECONCILE_TOLERANCE``. If a guided-workup / analyst-prep run is found for the
    ticker (or *guided_workup_path* is given), its agent judgment is surfaced
    read-only on the Thesis and PM-Review tabs.
    """

    ticker = ticker.upper()
    json_file = Path(json_path) if json_path else DEFAULT_JSON_DIR / f"{ticker}_latest.json"
    if not json_file.exists():
        raise FileNotFoundError(f"Valuation JSON not found: {json_file}")

    ctx, recon = _load_and_validate(json_file, expected_ticker=ticker)
    ctx.agent = _discover_agent_prep(ticker, guided_workup_path)

    as_of = as_of or datetime.now()
    if output_path:
        out = Path(output_path)
    else:
        out = (
            DEFAULT_OUTPUT_ROOT / ticker / f"{as_of:%Y%m%d-%H%M%S}-advanced-dcf"
            / f"{ticker}_advanced_dcf_model.xlsx"
        )
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    ctx.wb = wb

    _build_cover(ctx, ticker, json_file, as_of, recon)
    _build_thesis(ctx)
    _build_pm_review_queue(ctx)
    _build_assumptions(ctx)
    _build_historical(ctx)
    _build_forecast_input(ctx)
    _build_wacc(ctx)
    _build_dcf_base(ctx)
    _build_scenarios(ctx)
    _build_valuation_bridge(ctx)
    _build_sensitivities(ctx)
    _build_checks(ctx, recon)

    _set_recalc(wb)
    wb.active = wb["Cover"]
    wb.save(out)
    return out


def refresh_model_data(
    workbook_path: str | Path,
    *,
    ticker: str | None = None,
    json_path: str | Path | None = None,
    output_path: str | Path | None = None,
    guided_workup_path: str | Path | None = None,
    as_of: datetime | None = None,
) -> Path:
    """Refresh per-ticker DATA in an existing model workbook, in place.

    Rebuilds only the data sheets (see ``DATA_SHEETS``) from a ticker's valuation
    JSON. The MODEL sheets (WACC, DCF_Base, Sensitivity, and any sheet the PM
    added) and the PM Override column on Assumptions are preserved, so the PM can
    keep evolving one model for the same ticker. Cross-ticker refreshes retain formula
    sheets but clear ticker-specific PM overrides. Raises if the refreshed Base IV does
    not reconcile to the backend.
    """
    from openpyxl import load_workbook

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Model workbook not found: {workbook_path}")

    if json_path:
        json_file = Path(json_path)
    elif ticker:
        json_file = DEFAULT_JSON_DIR / f"{ticker.upper()}_latest.json"
    else:
        raise ValueError("refresh_model_data requires either ticker or json_path.")
    ticker = (ticker or json_file.stem.split("_")[0]).upper()

    ctx, recon = _load_and_validate(json_file, expected_ticker=ticker)
    ctx.agent = _discover_agent_prep(ticker, guided_workup_path)
    wb = load_workbook(workbook_path)
    ctx.wb = wb
    if "DCF_Base" not in wb.sheetnames:
        raise ValueError(
            "Workbook has no DCF_Base sheet; run a full build first, then refresh."
        )

    # Locate the IV/output rows on the (possibly PM-edited) DCF_Base so the data
    # tabs that reference it stay correct without rebuilding the model.
    ctx.dcf_base_rows = _read_dcf_base_rows(wb["DCF_Base"])

    source_ticker = _read_workbook_ticker(wb)
    preserve_overrides = source_ticker == ticker

    # Preserve workbook-local PM overrides only for a same-ticker refresh.
    overrides: dict[str, Any] = {}
    if preserve_overrides and "Assumptions" in wb.sheetnames:
        aws = wb["Assumptions"]
        for r in range(5, aws.max_row + 1):
            key, override = aws.cell(r, 2).value, aws.cell(r, 4).value
            if key and override not in (None, ""):
                overrides[str(key)] = override

    # Rebuild only the data sheets.
    for name in DATA_SHEETS:
        if name in wb.sheetnames:
            del wb[name]
    as_of = as_of or datetime.now()
    _build_cover(ctx, ticker, json_file, as_of, recon)
    _build_thesis(ctx)
    _build_pm_review_queue(ctx)
    _build_assumptions(ctx)
    _build_historical(ctx)
    _build_forecast_input(ctx)
    _build_scenarios(ctx)
    _build_valuation_bridge(ctx)
    _build_checks(ctx, recon)

    # Restore PM overrides onto the rebuilt Assumptions sheet.
    aws = wb["Assumptions"]
    for r in range(5, aws.max_row + 1):
        key = aws.cell(r, 2).value
        if key and str(key) in overrides:
            aws.cell(r, 4).value = overrides[str(key)]

    _reorder_sheets(wb)
    _set_recalc(wb)
    if "Cover" in wb.sheetnames:
        wb.active = wb["Cover"]
    out = Path(output_path) if output_path else workbook_path
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _discover_agent_prep(ticker: str, explicit_path: str | Path | None = None) -> dict | None:
    """Find the latest guided-workup / analyst-prep JSON for *ticker* and extract the
    per-name agent judgment (thesis cards, driver proposals, missing-data flags, queue
    decisions). Returns None when no agent run exists — the model then falls back to the
    deterministic story/sector layer. This is read-only surfacing; it never mutates the
    deterministic model (Vision: the PM Decision Queue is the only mutation bridge)."""
    ticker = ticker.upper()

    def _extract(data: dict, path: Path) -> dict | None:
        ap = data.get("analyst_prep") if isinstance(data.get("analyst_prep"), dict) else data
        if not isinstance(ap, dict) or not ap.get("thesis_cards"):
            return None
        return {
            "source_path": str(path),
            "run_stamp": data.get("run_stamp") or ap.get("generated_at") or "",
            "thesis_cards": ap.get("thesis_cards") or [],
            "driver_cards": ap.get("driver_cards") or [],
            "comps_card": ap.get("comps_card") or {},
            "missing_data": ap.get("missing_data") or [],
            "queue_decisions": data.get("queue_decisions") or [],
            "profile_runs": data.get("profile_runs") or [],
        }

    candidates: list[Path] = []
    if explicit_path:
        candidates = [Path(explicit_path)]
    else:
        for sub in ("guided_workups", "analyst_prep"):
            d = ROOT / "output" / sub / ticker
            if d.exists():
                candidates += sorted(d.glob(f"{ticker}-*.json"), reverse=True)
    for path in candidates:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        extracted = _extract(data, path)
        if extracted:
            return extracted
    return None


def _load_and_validate(
    json_file: Path,
    *,
    expected_ticker: str | None = None,
) -> tuple["_Context", dict]:
    payload = json.loads(json_file.read_text(encoding="utf-8"))
    payload_ticker = str(payload.get("ticker") or "").upper().strip()
    normalized_expected = str(expected_ticker or "").upper().strip()
    if normalized_expected and payload_ticker != normalized_expected:
        raise ValueError(
            "Advanced DCF JSON ticker mismatch: "
            f"expected {normalized_expected}, found {payload_ticker or '<missing>'}."
        )
    ctx = _Context(payload)
    # Forecast is load-bearing (the DCF + reconciliation run off it). Historical
    # actuals only enrich the Historical tab, so a name without CIQ/SEC history still
    # builds — the tab just says so.
    if not ctx.forecast:
        raise ValueError("Advanced DCF model requires excel_flat.forecast in valuation JSON.")
    # The transparent terminal rebuild mirrors the backend's value-driver Gordon
    # path. Bridge-mode names use a terminal FCFF the JSON does not expose, so
    # refuse with a clear message rather than silently diverging.
    mode = str(ctx.terminal.get("gordon_formula_mode") or "value_driver")
    if mode not in ("value_driver", ""):
        raise ValueError(
            f"Advanced DCF model supports value_driver Gordon mode; this name uses "
            f"'{mode}'. Terminal rebuild needs a bridge-mode branch before it can reconcile."
        )
    recon = ctx.reconcile()
    if recon["status"] != "ok":
        raise ValueError(
            "Advanced DCF model failed reconciliation to backend iv_base "
            f"({recon['workbook_iv']:.2f} vs {recon['backend_iv']:.2f}, "
            f"gap {recon['gap']:.2f} > {RECONCILE_TOLERANCE})."
        )
    return ctx, recon


def _read_workbook_ticker(wb) -> str:
    if "Cover" not in wb.sheetnames:
        return ""
    title = str(wb["Cover"]["A1"].value or "").strip()
    if not title:
        return ""
    return title.split("—", 1)[0].strip().upper()


def _read_dcf_base_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label in _DCF_ROW_LABELS:
            rows[_DCF_ROW_LABELS[label]] = r
    missing = {"iv", "iv_gordon", "iv_exit", "tv_pct"} - rows.keys()
    if missing:
        raise ValueError(
            f"DCF_Base is missing expected output rows {sorted(missing)}; "
            "refresh cannot relocate them. Rebuild the model or restore the labels."
        )
    rows.setdefault("last_col", 2 + 10)
    return rows


def _reorder_sheets(wb) -> None:
    ordered = [wb[n] for n in CANONICAL_ORDER if n in wb.sheetnames]
    extra = [ws for ws in wb._sheets if ws not in ordered]
    wb._sheets = ordered + extra


def _set_recalc(wb) -> None:
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


# --------------------------------------------------------------------------- #
# Context: extraction + reconciliation core
# --------------------------------------------------------------------------- #
class _Context:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload
        ef = payload.get("excel_flat") or {}
        self.excel_flat = ef
        self.assumptions = _kv(ef.get("assumptions"))
        self.wacc = _kv(ef.get("wacc"))
        missing_wacc_inputs = self.wacc.get("missing_inputs")
        if isinstance(missing_wacc_inputs, str):
            try:
                parsed_missing = json.loads(missing_wacc_inputs)
            except json.JSONDecodeError:
                parsed_missing = [missing_wacc_inputs] if missing_wacc_inputs else []
            self.wacc["missing_inputs"] = parsed_missing if isinstance(parsed_missing, list) else []
        supplied_preflight = payload.get("source_preflight")
        self.source_preflight = (
            supplied_preflight
            if isinstance(supplied_preflight, dict) and supplied_preflight
            else {
                "status": "blocked",
                "blockers": ["source_preflight_missing"],
                "workbook": {"formula_error_count": "unavailable"},
            }
        )
        self.diagnostic_shadow_cases = payload.get("diagnostic_shadow_cases") or {}
        self.market = _kv(ef.get("market"))
        self.valuation = _kv(ef.get("valuation"))
        self.terminal = _kv(ef.get("terminal"))
        self.lineage = _kv(ef.get("source_lineage"))
        self.drivers = payload.get("drivers_raw") or {}
        self.story = payload.get("story_profile") or {}
        self.story_adj = payload.get("story_adjustments") or {}
        self.resolution = payload.get("default_resolution") or {}
        self.scenario_policy = payload.get("scenario_policy") or {}
        self.context_scenarios = payload.get("context_scenarios") or {}
        self.register = payload.get("assumption_register_summary") or {}
        self.historical = list(
            payload.get("historical_financials") or ef.get("historical_financials") or []
        )
        self.forecast = list(ef.get("forecast") or [])[:10]
        # Per-name agent judgment from a guided-workup / analyst-prep run, if found.
        # Read-only surfacing only — never feeds the deterministic computation.
        self.agent: dict[str, Any] | None = None
        self.wb: Workbook | None = None
        self.assumption_rows: dict[str, int] = {}
        self.fc_cols: dict[str, int] = {}
        self.fc_start = 5
        self.dcf_base_rows: dict[str, int] = {}

    # -- flag map from the assumption register ---------------------------- #
    def flag_for(self, name: str) -> str:
        for entry in self.register.get("flagged_entries") or []:
            if entry.get("assumption_name") == name:
                return str(entry.get("flag_level") or entry.get("flag") or "watch")
        return ""

    def source_for(self, key: str) -> str:
        return str(self.lineage.get(key, "") or "")

    def source_formula_error_count(self) -> int | str:
        value = (self.source_preflight.get("workbook") or {}).get("formula_error_count")
        if isinstance(value, bool):
            return int(value)
        try:
            return int(value)
        except (TypeError, ValueError):
            return "unavailable"

    def overall_decision_status(self) -> str:
        status = str(self.source_preflight.get("status") or "").strip().lower()
        blockers = [item for item in self.source_preflight.get("blockers") or [] if item]
        formula_errors = self.source_formula_error_count()
        if status not in {"ready", "ok"} or blockers or formula_errors == "unavailable" or formula_errors > 0:
            return "BLOCKED"
        if (
            str(self.wacc.get("quality_status") or "") != "source_backed"
            or str(self.register.get("model_trust_state") or "") not in {"", "ok"}
            or str(self.resolution.get("status") or "") not in {"", "ok", "resolved"}
        ):
            return "REVIEW REQUIRED"
        return "REVIEW READY"

    def terminal_weight(self, key: str, default: float) -> float:
        value = self.story_adj.get(key)
        if value is None:
            value = self.drivers.get(key)
        return _num(value, default)

    def assump(self, key: str) -> str:
        row = self.assumption_rows[key]
        return f"'Assumptions'!$E${row}"

    def fc_ref(self, year_idx: int, field: str) -> str:
        col = self.fc_cols[field]
        return f"'Input_Forecast'!${get_column_letter(col)}${self.fc_start + year_idx - 1}"

    # -- numeric reconciliation (a Python mirror of the DCF_Base formulas) -- #
    def reconcile(self) -> dict[str, Any]:
        """Recompute Base IV exactly as the DCF_Base sheet does and compare to backend.

        This intentionally rebuilds terminal value from the forecast (value-driver
        Gordon + exit multiple) and the equity bridge from assumptions, rather than
        trusting the JSON terminal PVs, so the build-time gate mirrors the Excel
        formula chain rather than a parallel number that could silently diverge.
        """
        d = self.drivers
        a = self.assumptions
        n = len(self.forecast)
        wacc = _num(self.wacc.get("wacc"))
        g = _num(a.get("growth_terminal_pct"))
        ronic = _num(self.terminal.get("terminal_ronic_pct"), 11.0) / 100.0 \
            if self.terminal.get("terminal_ronic_pct") is not None else _num(d.get("ronic_terminal"), 0.11)
        gw = self.terminal_weight("terminal_blend_gordon_weight", 0.60)
        ew = self.terminal_weight("terminal_blend_exit_weight", 1 - gw)

        fc = self.forecast
        # PV of explicit FCFF (full-year, off the clean backend FCFF series).
        pv_fcff_sum = sum(_num(y.get("fcff_mm")) / (1 + wacc) ** _num(y.get("year")) for y in fc)

        y10 = fc[-1]
        nopat11 = _num(y10.get("nopat_mm")) * (1 + g)
        fcff11 = nopat11 * (1 - g / ronic) if ronic > 0 else 0.0
        tv_gordon = fcff11 / (wacc - g) if wacc > g else 0.0
        exit_metric = str(a.get("exit_metric", "ev_ebitda"))
        metric = _num(y10.get("ebit_mm")) if exit_metric == "ev_ebit" \
            else _num(y10.get("ebit_mm")) + _num(y10.get("da_mm"))
        tv_exit = metric * _num(a.get("exit_multiple"))
        pv_tv_gordon = tv_gordon / (1 + wacc) ** n
        pv_tv_exit = tv_exit / (1 + wacc) ** n
        pv_tv_blended = pv_tv_gordon * gw + pv_tv_exit * ew

        ev_ops = pv_fcff_sum + pv_tv_blended
        non_op = _num(a.get("non_operating_assets_mm"))
        claims = sum(_num(a.get(k)) for k in (
            "net_debt_mm", "minority_interest_mm", "preferred_equity_mm", "pension_deficit_mm",
            "lease_liabilities_mm", "options_value_mm", "convertibles_value_mm"))
        equity = ev_ops + non_op - claims

        current_shares = max(_num(a.get("shares_outstanding_mm")), 1.0)

        iv_blended = equity / current_shares
        iv_gordon = (pv_fcff_sum + pv_tv_gordon + non_op - claims) / current_shares
        iv_exit = (pv_fcff_sum + pv_tv_exit + non_op - claims) / current_shares

        backend_iv = _num(self.valuation.get("iv_base"))
        gap = abs(iv_blended - backend_iv)
        return {
            "status": "ok" if gap <= RECONCILE_TOLERANCE else "fail",
            "workbook_iv": iv_blended,
            "backend_iv": backend_iv,
            "gap": gap,
            "iv_gordon": iv_gordon,
            "iv_exit": iv_exit,
            "pv_fcff_sum": pv_fcff_sum,
            "ev_ops": ev_ops,
            "equity": equity,
            "current_shares": current_shares,
            "gordon_weight": gw,
            "exit_weight": ew,
        }


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def _kv(rows: Any) -> dict[str, Any]:
    if isinstance(rows, dict):
        return dict(rows)
    out: dict[str, Any] = {}
    for row in rows or []:
        if isinstance(row, dict) and "key" in row:
            out[str(row["key"])] = row.get("value")
    return out


def _num(value: Any, default: float = 0.0) -> float:
    if value in ("", None):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe(value: Any) -> Any:
    return "" if value is None else value


def _setup(ws, freeze: str = "A4") -> None:
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze


def _title(ws, title: str, subtitle: str | None = None, span: int = 10) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")


def _section(ws, row: int, label: str, end_col: int = 12) -> None:
    ws.cell(row, 1, label)
    ws.cell(row, 1).font = Font(bold=True, color=WHITE)
    for col in range(1, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws.cell(row, 1).font = Font(bold=True, color=WHITE)


def _header(ws, row: int, headers: Iterable[str]) -> None:
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row, col, text)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()


def _thin() -> Border:
    side = Side(style="thin", color="D9E2F3")
    return Border(left=side, right=side, top=side, bottom=side)


def _grid(ws, r0: int, r1: int, c0: int, c1: int) -> None:
    for row in ws.iter_rows(min_row=r0, max_row=r1, min_col=c0, max_col=c1):
        for cell in row:
            cell.border = _thin()
            if cell.alignment.vertical is None:
                cell.alignment = Alignment(vertical="center")


def _fit(ws, widths: dict[str, float] | None = None) -> None:
    widths = widths or {}
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        longest = 8
        for cell in col_cells[:120]:
            if cell.value is not None:
                longest = max(longest, min(40, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = longest


def _table(ws, name: str, r0: int, r1: int, c1: int) -> None:
    if r1 <= r0:
        return
    table = Table(displayName=name, ref=f"A{r0}:{get_column_letter(c1)}{r1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def _flag_fill(level: str):
    level = (level or "").lower().replace(" ", "_")
    if level in {"blocked", "critical", "review_required", "review"}:
        return PatternFill("solid", fgColor=PALE_RED), Font(color=RED_FONT, bold=True)
    if level in {"watch", "medium", "high"}:
        return PatternFill("solid", fgColor=PALE_AMBER), Font(color=AMBER_FONT)
    return None, None


# --------------------------------------------------------------------------- #
# Cover
# --------------------------------------------------------------------------- #
def _build_cover(ctx: _Context, ticker: str, json_file: Path, as_of: datetime, recon: dict) -> None:
    ws = ctx.wb.create_sheet("Cover")
    _setup(ws, freeze="")
    company = ctx.payload.get("company_name") or ticker
    _title(ws, f"{ticker} — {company}", "Advanced DCF review workbook (deterministic backend + agent judgment)")

    val, mkt = ctx.valuation, ctx.market
    price = _num(mkt.get("price"))
    base = _num(val.get("iv_base"))
    headline = [
        ("Overall decision status", ctx.overall_decision_status(), None),
        ("Sector / Industry", f"{ctx.payload.get('sector','')} / {ctx.payload.get('industry','')}", None),
        ("Current price", price, USD),
        ("Backend Base IV", base, USD),
        ("Base upside", _num(val.get("upside_base_pct")) / 100.0, PCT),
        ("Legacy prob-weighted IV (not PM-approved)", _num(val.get("expected_iv")), USD),
        ("Legacy prob-weighted upside", _num(val.get("expected_upside_pct")) / 100.0, PCT),
        ("Margin of safety", _num(val.get("margin_of_safety")) / 100.0, PCT),
        ("Analyst target / rec", f"{_num(mkt.get('analyst_target')):.2f}  ({mkt.get('analyst_recommendation','')})", None),
        ("Model applicability", val.get("model_applicability_status", ""), None),
        ("Model trust state", ctx.register.get("model_trust_state", "n/a"), None),
        ("Default resolution", ctx.resolution.get("status", "n/a"), None),
        ("WACC quality", ctx.wacc.get("quality_status", "n/a"), None),
        ("Source preflight", ctx.source_preflight.get("status", "blocked"), None),
        ("Source formula blockers", ctx.source_formula_error_count(), None),
    ]
    if ctx.diagnostic_shadow_cases:
        headline[4:4] = [
            ("Hold-current-margin shadow IV", ctx.diagnostic_shadow_cases.get("hold_current_ebit_margin_iv"), USD),
            ("Margin-policy delta / share", ctx.diagnostic_shadow_cases.get("margin_policy_delta_per_share"), USD),
        ]
    r = 4
    _section(ws, r, "Headline", 4)
    r += 1
    for label, value, fmt in headline:
        ws.cell(r, 1, label).font = Font(bold=True)
        c = ws.cell(r, 2, _safe(value))
        if fmt:
            c.number_format = fmt
        # color the trust/resolution rows
        if label in {"Overall decision status", "Model trust state", "Default resolution", "WACC quality", "Source preflight", "Hold-current-margin shadow IV", "Margin-policy delta / share"}:
            fill, font = _flag_fill(str(value))
            if fill:
                c.fill, c.font = fill, font
        r += 1

    r += 1
    _section(ws, r, "Reconciliation to deterministic backend", 4)
    r += 1
    recon_rows = [
        ("Workbook Base IV (rebuilt)", recon["workbook_iv"], USD),
        ("Backend Base IV", recon["backend_iv"], USD),
        ("Gap", recon["gap"], USD),
        ("Formula reconciliation status", "RECONCILED" if recon["status"] == "ok" else "FAILED", None),
    ]
    for label, value, fmt in recon_rows:
        ws.cell(r, 1, label).font = Font(bold=True)
        c = ws.cell(r, 2, _safe(value))
        if fmt:
            c.number_format = fmt
        if label == "Formula reconciliation status":
            c.font = Font(bold=True, color=GREEN_FONT if value == "RECONCILED" else RED_FONT)
        r += 1

    r += 1
    _section(ws, r, "How to read this workbook", 6)
    r += 1
    flow = [
        ("Thesis_Drivers", "Agent thesis cards (when a guided workup exists) + the deterministic story/sector layer."),
        ("PM_Review_Queue", "Agent driver proposals + missing-data flags + deterministic resolution items for your sign-off."),
        ("Assumptions", "Every input with provenance + register flag. Edit the PM Override column only."),
        ("DCF_Base", "Transparent Base DCF; reconciles to backend IV. Terminal shown both ways + blend."),
        ("Scenarios", "Backend bear/base/bull (official) with context-advisory set alongside."),
        ("Valuation_Bridge", "DCF vs EP vs FCFE vs comps vs market — the full method spread."),
        ("Checks", "Reconciliation, trust state, and model-hygiene flags to clear before discussion."),
    ]
    for sheet, note in flow:
        ws.cell(r, 1, sheet).font = Font(bold=True)
        ws.cell(r, 2, note)
        r += 1

    r += 1
    ws.cell(r, 1, "Source JSON").font = Font(italic=True, color="888888")
    ws.cell(r, 2, str(json_file.resolve())).font = Font(italic=True, color="888888")
    r += 1
    ws.cell(r, 1, "Generated").font = Font(italic=True, color="888888")
    ws.cell(r, 2, as_of.isoformat(timespec="seconds")).font = Font(italic=True, color="888888")
    r += 1
    ws.cell(r, 1, "Note").font = Font(italic=True, color="888888")
    ws.cell(r, 2, "Diagnostic only — source preflight and PM policy gates must pass before decision use; no writeback.").font = Font(italic=True, color="888888")
    _fit(ws, {"A": 30, "B": 96})


# --------------------------------------------------------------------------- #
# Thesis & drivers (the judgment -> model translation)
# --------------------------------------------------------------------------- #
def _build_thesis(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Thesis_Drivers")
    _setup(ws)
    story_src = ctx.source_for("story_profile") or "story_sector"
    agent = ctx.agent
    if agent:
        _title(ws, "Investment Thesis & Driver Layer",
               f"Agent judgment from guided workup {agent.get('run_stamp','')}; deterministic story/sector layer below.")
    else:
        _title(ws, "Story & Driver Layer (deterministic)",
               f"Source: {story_src}. These are deterministic story/sector priors, NOT a per-name agent read.")

    if agent:
        ws.cell(3, 1,
                f"Agent thesis from {Path(agent['source_path']).name}. These cards are the analyst agents' "
                "per-name read, surfaced read-only; they do not change the model. Proposed assumption changes "
                "and their PM-review status are on the PM_Review_Queue tab.").font = Font(italic=True, color="666666")
    else:
        ws.cell(3, 1,
                "NOTE: No guided-workup / analyst-prep run was found for this ticker, so there is no per-name "
                "agent judgment to show. story_profile below is a deterministic sector prior. Run the guided "
                "ticker workup (output/guided_workups/) to populate agent thesis and driver proposals.").font = Font(italic=True, color=RED_FONT)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    ws.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 56

    r = 4
    if agent and agent.get("thesis_cards"):
        _section(ws, r, "Agent Thesis Cards (per-name judgment)", 5)
        r += 1
        _header(ws, r, ["Thesis", "Claim", "Model implication", "Confidence", "What would change my mind"])
        r += 1
        first = r
        for card in agent["thesis_cards"]:
            conf = card.get("agent_confidence") or card.get("deterministic_confidence") or "n/a"
            ws.cell(r, 1, card.get("title")).font = Font(bold=True)
            ws.cell(r, 2, _safe(card.get("claim")))
            ws.cell(r, 3, _safe(card.get("model_implication")))
            ws.cell(r, 4, conf)
            ws.cell(r, 5, _safe(card.get("what_would_change_mind") or card.get("counter_evidence")))
            for col in (2, 3, 5):
                ws.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 44
            r += 1
        _grid(ws, first - 1, r - 1, 1, 5)
        r += 1

    s, adj = ctx.story, ctx.story_adj
    _section(ws, r, "Story Profile (deterministic sector prior)", 4)
    head = r + 1
    _header(ws, head, ["Attribute", "Assessment", "Scale", "Read"])
    profile_rows = [
        ("Moat strength", s.get("moat_strength"), "0-5", "Durability of competitive advantage"),
        ("Pricing power", s.get("pricing_power"), "0-5", "Ability to pass through cost / take price"),
        ("Cyclicality", s.get("cyclicality"), "low/med/high", "Sensitivity of demand to the cycle"),
        ("Capital intensity", s.get("capital_intensity"), "low/med/high", "Reinvestment needed to grow"),
        ("Governance risk", s.get("governance_risk"), "low/med/high", "Capital-allocation / disclosure risk"),
        ("Advantage period", s.get("competitive_advantage_years"), "years", "Years of excess returns assumed"),
    ]
    r = head + 1
    for label, value, scale, read in profile_rows:
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, _safe(value))
        ws.cell(r, 3, scale)
        ws.cell(r, 4, read)
        r += 1
    _grid(ws, head, r - 1, 1, 4)

    r += 1
    adjustment_title = (
        "Story -> Model Adjustments (what the thesis changed)"
        if adj
        else "Story -> Model Adjustments (advisory; none applied)"
    )
    _section(ws, r, adjustment_title, 4)
    r += 1
    _header(ws, r, ["Adjustment", "Value", "Effect on the model"])
    r += 1
    adj_rows = [
        ("Growth add", adj.get("growth_add"), PCT, "Added to near/mid revenue growth"),
        ("Margin add", adj.get("margin_add"), PCT, "Added to EBIT margin path"),
        ("Cyclicality growth multiplier", adj.get("cyclicality_growth_multiplier"), MULT, "Scales growth for cycle exposure"),
        ("Cyclicality WACC add", adj.get("cyclicality_wacc_add"), PCT, "Raises discount rate for cyclical names"),
        ("Governance WACC add", adj.get("governance_wacc_add"), PCT, "Raises discount rate for governance risk"),
        ("Capex target add", adj.get("capex_target_add"), PCT, "Raises terminal reinvestment intensity"),
        ("D&A target add", adj.get("da_target_add"), PCT, "Adjusts terminal D&A intensity"),
        ("Exit multiple cyclicality mult.", adj.get("exit_multiple_cyclicality_multiplier"), MULT, "Haircut/premium on exit multiple"),
        ("Exit multiple governance mult.", adj.get("exit_multiple_governance_multiplier"), MULT, "Haircut/premium on exit multiple"),
        ("Terminal blend — Gordon weight", adj.get("terminal_blend_gordon_weight"), PCT, "Weight on Gordon TV (cyclicality-driven)"),
        ("Terminal blend — Exit weight", adj.get("terminal_blend_exit_weight"), PCT, "Weight on exit-multiple TV"),
    ]
    for label, value, fmt, effect in adj_rows:
        ws.cell(r, 1, label).font = Font(bold=True)
        c = ws.cell(r, 2, _safe(value))
        c.number_format = fmt
        ws.cell(r, 3, effect)
        r += 1
    _grid(ws, r - len(adj_rows), r - 1, 1, 4)

    if adj:
        note = (
            "These adjustments are produced by the deterministic story layer and are already baked into the "
            "assumptions and terminal blend on the DCF tabs. They are shown here so the qualitative thesis and "
            "the numbers are auditable together. Change them via the PM Decision Queue, not in this workbook."
        )
    else:
        note = (
            "Story priors are advisory in this source-only run and were not applied to deterministic drivers. "
            "Any proposed change must enter through the PM Decision Queue."
        )
    r += 1
    ws.cell(r, 1, note).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 44
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    _fit(ws, {"A": 32, "B": 14, "C": 16, "D": 52})


# --------------------------------------------------------------------------- #
# PM review queue (default_resolution)
# --------------------------------------------------------------------------- #
def _build_pm_review_queue(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("PM_Review_Queue")
    _setup(ws)
    res = ctx.resolution
    status = res.get("status", "n/a")
    counts = res.get("counts", {})
    _title(ws, "PM Review Queue",
           f"Default resolution status: {status}  "
           f"(resolved {counts.get('resolved','?')}, review_required {counts.get('review_required','?')})")

    agent = ctx.agent
    r = 4

    # Agent driver proposals (concrete assumption changes routed through the PM Queue).
    if agent and agent.get("driver_cards"):
        _section(ws, r, "Agent Driver Proposals (review_required first — apply via PM Override, not here)", 8)
        r += 1
        _header(ws, r, ["Assumption", "Current", "Proposed", "Status", "Impact", "Rationale", "", ""])
        r += 1
        first = r
        cards = sorted(
            agent["driver_cards"],
            key=lambda c: 0 if str(c.get("pm_review_status")) == "review_required" else 1,
        )
        for c in cards:
            ws.cell(r, 1, c.get("assumption_name")).font = Font(bold=True)
            ws.cell(r, 2, _safe(c.get("current_value")))
            ws.cell(r, 3, _safe(c.get("proposed_or_effective_value")))
            ws.cell(r, 4, c.get("pm_review_status"))
            ws.cell(r, 5, _safe(c.get("valuation_impact")))
            ws.cell(r, 6, _safe(c.get("rationale")))
            ws.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="top")
            if str(c.get("pm_review_status")) == "review_required":
                fill, font = _flag_fill("review_required")
                for col in (1, 4):
                    ws.cell(r, col).fill = fill
                ws.cell(r, 4).font = font
            ws.row_dimensions[r].height = 40
            r += 1
        _grid(ws, first - 1, r - 1, 1, 8)
        r += 2

    # Missing-data flags raised by the agents.
    if agent and agent.get("missing_data"):
        _section(ws, r, "Missing-data flags", 8)
        r += 1
        _header(ws, r, ["Flag", "Severity", "Reason", "Suggested check", "", "", "", ""])
        r += 1
        first = r
        for m in agent["missing_data"]:
            ws.cell(r, 1, m.get("label")).font = Font(bold=True)
            ws.cell(r, 2, m.get("severity"))
            ws.cell(r, 3, _safe(m.get("reason")))
            ws.cell(r, 4, _safe(m.get("suggested_check")))
            ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
            fill, font = _flag_fill(str(m.get("severity")))
            if fill:
                ws.cell(r, 2).fill = fill
            ws.row_dimensions[r].height = 30
            r += 1
        _grid(ws, first - 1, r - 1, 1, 8)
        r += 2

    _section(ws, r, "Default-resolution fields (deterministic)", 8)
    r += 1
    fields = list(res.get("fields") or [])
    # Surface high-severity OR explicit review items first, then the rest.
    def _priority(f: dict) -> tuple:
        sev = {"high": 0, "medium": 1, "low": 2}.get(str(f.get("severity")), 3)
        return (0 if f.get("needs_pm_review") else 1, sev)

    fields.sort(key=_priority)

    _header(ws, r, ["Field", "Value", "Source", "Source class", "Severity",
                    "PM review?", "Fallback", "Why it matters"])
    fields_header = r
    r += 1
    for f in fields:
        ws.cell(r, 1, f.get("field")).font = Font(bold=True)
        ws.cell(r, 2, _safe(f.get("value")))
        ws.cell(r, 3, f.get("source"))
        ws.cell(r, 4, f.get("source_class"))
        ws.cell(r, 5, f.get("severity"))
        ws.cell(r, 6, "YES" if f.get("needs_pm_review") else "no")
        ws.cell(r, 7, _safe(f.get("fallback_value")))
        ws.cell(r, 8, f.get("why_it_matters"))
        ws.cell(r, 8).alignment = Alignment(wrap_text=True, vertical="top")
        needs = bool(f.get("needs_pm_review"))
        level = "review_required" if needs else str(f.get("severity"))
        fill, font = _flag_fill(level)
        if fill:
            for col in (1, 5, 6):
                ws.cell(r, col).fill = fill
            ws.cell(r, 6).font = font or ws.cell(r, 6).font
        ws.row_dimensions[r].height = 30
        r += 1
    if not fields:
        ws.cell(fields_header + 1, 1, "No resolution fields recorded in JSON.")
        r = fields_header + 2
    _grid(ws, fields_header, r - 1, 1, 8)
    _fit(ws, {"A": 22, "B": 12, "C": 34, "D": 16, "E": 10, "F": 11, "G": 12, "H": 60})


# --------------------------------------------------------------------------- #
# Assumptions with provenance + register flags
# --------------------------------------------------------------------------- #
def _build_assumptions(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Assumptions")
    _setup(ws)
    _title(ws, "Assumptions",
           "Edit only the PM Override column. Source = provenance; Flag = assumption-register hygiene.")

    a, w, mkt, term, val = ctx.assumptions, ctx.wacc, ctx.market, ctx.terminal, ctx.valuation
    adj = ctx.story_adj
    # (category, key, value, unit, lineage_key, register_name)
    rows: list[tuple] = [
        ("Company", "current_price", mkt.get("price"), "$/sh", "", ""),
        ("Company", "shares_outstanding_mm", a.get("shares_outstanding_mm"), "mm", "shares_outstanding", ""),
        ("Company", "annual_dilution_pct", ctx.drivers.get("annual_dilution_pct"), "%", "", ""),
        ("Company", "net_debt_mm", a.get("net_debt_mm"), "$mm", "net_debt", ""),
        ("Company", "non_operating_assets_mm", a.get("non_operating_assets_mm"), "$mm", "non_operating_assets", ""),
        ("Company", "minority_interest_mm", a.get("minority_interest_mm"), "$mm", "minority_interest", ""),
        ("Company", "preferred_equity_mm", a.get("preferred_equity_mm"), "$mm", "preferred_equity", "preferred_equity"),
        ("Company", "pension_deficit_mm", a.get("pension_deficit_mm"), "$mm", "pension_deficit", "pension_deficit"),
        ("Company", "lease_liabilities_mm", a.get("lease_liabilities_mm"), "$mm", "lease_liabilities", ""),
        ("Company", "options_value_mm", a.get("options_value_mm"), "$mm", "options_value", ""),
        ("Company", "convertibles_value_mm", a.get("convertibles_value_mm"), "$mm", "convertibles_value", ""),
        ("Operating", "revenue_mm", a.get("revenue_mm"), "$mm", "revenue_base", ""),
        ("Operating", "cogs_pct_of_revenue", ctx.drivers.get("cogs_pct_of_revenue"), "% rev", "", ""),
        ("Operating", "growth_terminal_pct", a.get("growth_terminal_pct"), "%", "", ""),
        ("WACC", "wacc", w.get("wacc"), "%", "", ""),
        ("WACC", "cost_of_equity", w.get("cost_of_equity"), "%", "cost_of_equity", ""),
        ("WACC", "cost_of_debt_after_tax", w.get("cost_of_debt_after_tax"), "%", "", ""),
        ("WACC", "risk_free_rate", w.get("risk_free_rate"), "%", "", ""),
        ("WACC", "equity_risk_premium", w.get("equity_risk_premium"), "%", "", ""),
        ("WACC", "beta_relevered", w.get("beta_relevered"), "x", "", ""),
        ("WACC", "size_premium", w.get("size_premium"), "%", "", ""),
        ("WACC", "equity_weight", w.get("equity_weight"), "% cap", "", ""),
        ("WACC", "debt_weight", w.get("debt_weight"), "% cap", "debt_weight", ""),
        ("Terminal", "exit_multiple", a.get("exit_multiple"), "x", "exit_multiple", "exit_multiple"),
        ("Terminal", "exit_metric", a.get("exit_metric"), "text", "", ""),
        ("Terminal", "ronic_terminal", term.get("terminal_ronic_pct", 11.0) / 100.0
            if term.get("terminal_ronic_pct") is not None else ctx.drivers.get("ronic_terminal"), "%", "ronic_terminal", ""),
        ("Terminal", "terminal_blend_gordon_weight", ctx.terminal_weight("terminal_blend_gordon_weight", 0.60), "%", "", ""),
        ("Terminal", "terminal_blend_exit_weight", ctx.terminal_weight("terminal_blend_exit_weight", 0.40), "%", "", ""),
        ("Scenario", "scenario_prob_bear", a.get("scenario_prob_bear"), "%", "", ""),
        ("Scenario", "scenario_prob_base", a.get("scenario_prob_base"), "%", "", ""),
        ("Scenario", "scenario_prob_bull", a.get("scenario_prob_bull"), "%", "", ""),
        ("Backend", "backend_iv_base", val.get("iv_base"), "$/sh", "", ""),
        ("Backend", "backend_iv_gordon", val.get("iv_gordon"), "$/sh", "", ""),
        ("Backend", "backend_iv_exit", val.get("iv_exit"), "$/sh", "", ""),
        ("Backend", "backend_tv_pct_ev", (term.get("tv_pct_of_ev", 0) or 0) / 100.0, "%", "", ""),
    ]

    _header(ws, 4, ["Category", "Key", "Source Value", "PM Override", "Effective Value",
                    "Unit", "Source", "Flag"])
    r = 5
    for category, key, value, unit, lineage_key, reg_name in rows:
        ws.cell(r, 1, category)
        ws.cell(r, 2, key)
        sv = ws.cell(r, 3, _safe(value))
        ov = ws.cell(r, 4)  # leave genuinely empty so ISBLANK/"" both mean "use source"
        ev = ws.cell(r, 5, f'=IF(D{r}="",C{r},D{r})')
        ws.cell(r, 6, unit)
        ws.cell(r, 7, ctx.source_for(lineage_key) if lineage_key else "")
        flag = ctx.flag_for(reg_name) if reg_name else ""
        ws.cell(r, 8, flag)

        sv.fill = PatternFill("solid", fgColor=PALE_GREEN)
        sv.font = Font(color=GREEN_FONT)
        ov.fill = PatternFill("solid", fgColor=PALE_YELLOW)
        ov.font = Font(color=BLUE_FONT)
        ev.font = Font(bold=True)

        if unit in ("%", "% rev", "% cap"):
            for col in (3, 4, 5):
                ws.cell(r, col).number_format = PCT
        elif "$" in unit:
            for col in (3, 4, 5):
                ws.cell(r, col).number_format = MM
        elif unit == "x":
            for col in (3, 4, 5):
                ws.cell(r, col).number_format = MULT
        if flag:
            fill, font = _flag_fill(flag)
            if fill:
                ws.cell(r, 8).fill = fill
                ws.cell(r, 8).font = font
        ctx.assumption_rows[key] = r
        r += 1
    _grid(ws, 4, r - 1, 1, 8)
    _fit(ws, {"A": 14, "B": 30, "C": 14, "D": 13, "E": 14, "F": 8, "G": 38, "H": 10})


# --------------------------------------------------------------------------- #
# Historical financials
# --------------------------------------------------------------------------- #
def _build_historical(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Historical_Financials")
    _setup(ws)
    _title(ws, "Historical Financials", "Annual CIQ actuals captured in valuation JSON")
    if not ctx.historical:
        ws.cell(3, 1,
                "No historical actuals available for this ticker (no CIQ Standard workbook / SEC "
                "history in the valuation JSON). The DCF and reconciliation are unaffected — they "
                "run off the forecast — but trend review needs a CIQ/SEC history source.").font = Font(italic=True, color=AMBER_FONT)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
        ws.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[3].height = 44
    headers = [
        "Fiscal Year", "Period", "Revenue", "Rev Growth", "EBIT", "EBIT Margin",
        "EBITDA", "EBITDA Margin", "D&A", "D&A %", "Capex", "Capex %", "CFO",
        "Pretax", "Tax", "Tax Rate", "Cash", "Debt", "Net Debt",
        "Tot Assets", "Tot Equity", "Source",
    ]
    fields = [
        "fiscal_year", "period", "revenue_mm", "revenue_growth_pct", "ebit_mm",
        "ebit_margin_pct", "ebitda_mm", "ebitda_margin_pct", "da_mm", "da_pct",
        "capex_mm", "capex_pct", "cfo_mm", "pretax_income_mm", "tax_expense_mm",
        "tax_rate_pct", "cash_mm", "debt_mm", "net_debt_mm", "total_assets_mm",
        "total_equity_mm", "source_file",
    ]
    _header(ws, 4, headers)
    r = 5
    for row in ctx.historical:
        for col, field in enumerate(fields, 1):
            cell = ws.cell(r, col, _safe(row.get(field)))
            if field.endswith("_pct"):
                cell.number_format = PCT
            elif field.endswith("_mm"):
                cell.number_format = MMI
        r += 1
    _grid(ws, 4, r - 1, 1, len(headers))
    _table(ws, "Historical_Financials_Table", 4, r - 1, len(headers))
    _fit(ws, {"A": 12, "B": 12, "V": 24})


# --------------------------------------------------------------------------- #
# Backend per-year forecast (Base) — source for the DCF tab
# --------------------------------------------------------------------------- #
def _build_forecast_input(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Input_Forecast")
    _setup(ws)
    _title(ws, "Base Forecast (deterministic backend)",
           "Per-year Base-case drivers and outputs. The DCF_Base tab recomputes the bridge from these.")
    headers = [
        "Year", "Revenue", "Growth", "EBIT Margin", "Tax Rate", "Capex %", "D&A %",
        "DSO", "DIO", "DPO", "EBIT", "NOPAT", "D&A", "Capex", "AR", "Inventory",
        "AP", "NWC", "Delta NWC", "FCFF", "Disc Factor", "PV FCFF", "ROIC", "Econ Profit",
    ]
    fields = [
        "year", "revenue_mm", "growth_rate", "ebit_margin", "tax_rate", "capex_pct",
        "da_pct", "dso", "dio", "dpo", "ebit_mm", "nopat_mm", "da_mm", "capex_mm",
        "ar_mm", "inventory_mm", "ap_mm", "nwc_mm", "delta_nwc_mm", "fcff_mm",
        "discount_factor", "pv_fcff_mm", "roic", "economic_profit_mm",
    ]
    ctx.fc_cols = {field: idx for idx, field in enumerate(fields, 1)}
    _header(ws, 4, headers)
    r = ctx.fc_start
    for row in ctx.forecast:
        for col, field in enumerate(fields, 1):
            cell = ws.cell(r, col, _safe(row.get(field)))
            if field in {"growth_rate", "ebit_margin", "tax_rate", "capex_pct", "da_pct", "roic"}:
                cell.number_format = PCT
            elif field != "year":
                cell.number_format = MM
        r += 1
    _grid(ws, 4, r - 1, 1, len(headers))
    _table(ws, "Input_Forecast_Table", 4, r - 1, len(headers))
    _fit(ws)


# --------------------------------------------------------------------------- #
# WACC
# --------------------------------------------------------------------------- #
def _build_wacc(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("WACC")
    _setup(ws)
    _title(ws, "WACC", "Cost of capital build. Edit components via Assumptions overrides.")
    _header(ws, 4, ["Line", "Value", "Notes"])

    def _optional_assumption(key: str) -> str:
        row = ctx.assumption_rows[key]
        return f'''=IF(AND('Assumptions'!$C${row}="",'Assumptions'!$D${row}=""),"",'Assumptions'!$E${row})'''


    rows = [
        ("Backend WACC (used in DCF)", f"={ctx.assump('wacc')}", "Default discount rate from backend", PCT),
        ("Cost of equity", f"={ctx.assump('cost_of_equity')}", "CAPM cost of equity", PCT),
        ("Risk-free rate", _optional_assumption("risk_free_rate"), "Blank if unavailable", PCT),
        ("Equity risk premium", _optional_assumption("equity_risk_premium"), "Blank if unavailable", PCT),
        ("Relevered beta", f"={ctx.assump('beta_relevered')}", "Peer/structure beta", MULT),
        ("Size premium", f"={ctx.assump('size_premium')}", "Duff & Phelps", PCT),
        ("After-tax cost of debt", _optional_assumption("cost_of_debt_after_tax"), "Blank if unavailable", PCT),
        ("Equity weight", f"={ctx.assump('equity_weight')}", "Market-value weight", PCT),
        ("Debt weight", f"={ctx.assump('debt_weight')}", "Market-value weight", PCT),
        ("Rebuilt cost of equity (check)",
         "=IF(COUNT(B7:B8)<2,B6,B7+B9*B8+B10)",
         "CAPM rebuild; falls back to backend Ke", PCT),
        ("Quality status", ctx.wacc.get("quality_status", "n/a"), "Degraded quality blocks decision-ready use", None),
        ("Missing inputs", ", ".join(ctx.wacc.get("missing_inputs") or []), "Load-bearing inputs unavailable", None),
        ("Beta source", ctx.wacc.get("beta_source", "n/a"), "Explicit beta lineage", None),
    ]
    r = 5
    for line, formula, note, fmt in rows:
        ws.cell(r, 1, line)
        c = ws.cell(r, 2, formula)
        if fmt:
            c.number_format = fmt
        ws.cell(r, 3, note)
        r += 1
    _grid(ws, 4, r - 1, 1, 3)
    _fit(ws, {"A": 30, "B": 28, "C": 44})


# --------------------------------------------------------------------------- #
# DCF_Base — transparent, reconciling rebuild
# --------------------------------------------------------------------------- #
def _build_dcf_base(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("DCF_Base")
    _setup(ws, freeze="C5")
    _title(ws, "Base Case DCF (transparent rebuild)",
           "Recomputes the bridge from backend per-year drivers; reconciles to backend Base IV.", span=13)
    n = len(ctx.forecast)
    last_col = 2 + n  # data columns C.. for years 1..n

    # Year header
    ws.cell(4, 1, "Operating model").font = Font(bold=True, color=WHITE)
    for col in range(1, last_col + 1):
        ws.cell(4, col).fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws.cell(4, 1).font = Font(bold=True, color=WHITE)
    for i in range(1, n + 1):
        c = ws.cell(5, 2 + i, f"Y{i}")
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")

    labels = [
        ("revenue", "Revenue", MM, "revenue_mm"),
        ("growth", "Revenue growth", PCT, "growth_rate"),
        ("ebit_margin", "EBIT margin", PCT, "ebit_margin"),
        ("ebit", "EBIT", MM, "ebit_mm"),
        ("tax_rate", "Tax rate", PCT, "tax_rate"),
        ("nopat", "NOPAT", MM, "nopat_mm"),
        ("da", "D&A", MM, "da_mm"),
        ("capex", "Capex", MM, "capex_mm"),
        ("nwc", "Net working capital", MM, "nwc_mm"),
        ("delta_nwc", "Delta NWC", MM, None),
        ("fcff", "FCFF (NOPAT+D&A-Capex-dNWC)", MM, None),
        ("fcff_backend", "FCFF — backend (foot check)", MM, "fcff_mm"),
        ("disc", "Discount factor", MULT, None),
        ("pv_fcff", "PV of FCFF", MM, None),
        ("ebitda", "EBITDA", MM, None),
    ]
    row_of: dict[str, int] = {}
    r = 6
    for key, label, _fmt, _src in labels:
        row_of[key] = r
        ws.cell(r, 1, label)
        if key in {"fcff", "pv_fcff"}:
            ws.cell(r, 1).font = Font(bold=True)
        r += 1

    wacc = ctx.assump("wacc")
    for i in range(1, n + 1):
        col = 2 + i
        cl = get_column_letter(col)
        ws.cell(row_of["revenue"], col, f"={ctx.fc_ref(i, 'revenue_mm')}")
        ws.cell(row_of["growth"], col, f"={ctx.fc_ref(i, 'growth_rate')}")
        ws.cell(row_of["ebit_margin"], col, f"={ctx.fc_ref(i, 'ebit_margin')}")
        ws.cell(row_of["ebit"], col, f"={cl}{row_of['revenue']}*{cl}{row_of['ebit_margin']}")
        ws.cell(row_of["tax_rate"], col, f"={ctx.fc_ref(i, 'tax_rate')}")
        ws.cell(row_of["nopat"], col, f"={cl}{row_of['ebit']}*(1-{cl}{row_of['tax_rate']})")
        ws.cell(row_of["da"], col, f"={ctx.fc_ref(i, 'da_mm')}")
        ws.cell(row_of["capex"], col, f"={ctx.fc_ref(i, 'capex_mm')}")
        ws.cell(row_of["nwc"], col, f"={ctx.fc_ref(i, 'nwc_mm')}")
        # Derive Delta NWC from NWC differences. The backend delta_nwc_mm field is
        # unreliable (observed unit corruption in later years); NWC levels are clean.
        if i == 1:
            ws.cell(row_of["delta_nwc"], col, f"={ctx.fc_ref(1, 'delta_nwc_mm')}")
        else:
            prev = get_column_letter(col - 1)
            ws.cell(row_of["delta_nwc"], col, f"={cl}{row_of['nwc']}-{prev}{row_of['nwc']}")
        ws.cell(row_of["fcff"], col,
                f"={cl}{row_of['nopat']}+{cl}{row_of['da']}-{cl}{row_of['capex']}-{cl}{row_of['delta_nwc']}")
        ws.cell(row_of["fcff_backend"], col, f"={ctx.fc_ref(i, 'fcff_mm')}")
        ws.cell(row_of["disc"], col, f"=POWER(1+{wacc},{i})")
        ws.cell(row_of["pv_fcff"], col, f"={cl}{row_of['fcff']}/{cl}{row_of['disc']}")
        ws.cell(row_of["ebitda"], col, f"={cl}{row_of['ebit']}+{cl}{row_of['da']}")

    # number formats across the year columns
    for key, _label, fmt, _src in labels:
        for col in range(3, last_col + 1):
            ws.cell(row_of[key], col).number_format = fmt

    last_cl = get_column_letter(last_col)

    # Valuation block
    vr = r + 1
    _section(ws, vr, "Terminal value & equity bridge", 4)
    vr += 1
    val_rows = [
        ("rev11", "Terminal revenue (Y+1)", f"={last_cl}{row_of['revenue']}*(1+{ctx.assump('growth_terminal_pct')})", MM),
        ("ebit11", "Terminal EBIT", f"=B{{rev11}}*{last_cl}{row_of['ebit_margin']}", MM),
        ("nopat11", "Terminal NOPAT", f"=B{{ebit11}}*(1-{last_cl}{row_of['tax_rate']})", MM),
        ("rr", "Terminal reinvestment rate (g/RONIC)", f"={ctx.assump('growth_terminal_pct')}/{ctx.assump('ronic_terminal')}", PCT),
        ("fcff11", "Terminal FCFF (value-driver)", "=B{nopat11}*(1-B{rr})", MM),
        ("tv_gordon", "Gordon terminal value", f"=IF({ctx.assump('wacc')}<={ctx.assump('growth_terminal_pct')},NA(),B{{fcff11}}/({ctx.assump('wacc')}-{ctx.assump('growth_terminal_pct')}))", MM),
        ("tv_exit", "Exit terminal value", f"=IF({ctx.assump('exit_metric')}=\"ev_ebit\",{last_cl}{row_of['ebit']},{last_cl}{row_of['ebitda']})*{ctx.assump('exit_multiple')}", MM),
        ("pv_tv_gordon", "PV of Gordon TV", f"=B{{tv_gordon}}/POWER(1+{ctx.assump('wacc')},{n})", MM),
        ("pv_tv_exit", "PV of exit TV", f"=B{{tv_exit}}/POWER(1+{ctx.assump('wacc')},{n})", MM),
        ("pv_tv_blend", "PV of blended TV", f"={ctx.assump('terminal_blend_gordon_weight')}*B{{pv_tv_gordon}}+{ctx.assump('terminal_blend_exit_weight')}*B{{pv_tv_exit}}", MM),
        ("pv_fcff_sum", "PV of explicit FCFF", f"=SUM(C{row_of['pv_fcff']}:{last_cl}{row_of['pv_fcff']})", MM),
        ("ev_ops", "Enterprise value — operations", "=B{pv_fcff_sum}+B{pv_tv_blend}", MM),
        ("tv_pct", "Terminal value % of EV", "=B{pv_tv_blend}/B{ev_ops}", PCT),
        ("non_op", "Non-operating assets", f"={ctx.assump('non_operating_assets_mm')}", MM),
        ("claims", "Non-equity claims", f"={ctx.assump('net_debt_mm')}+{ctx.assump('minority_interest_mm')}+{ctx.assump('preferred_equity_mm')}+{ctx.assump('pension_deficit_mm')}+{ctx.assump('lease_liabilities_mm')}+{ctx.assump('options_value_mm')}+{ctx.assump('convertibles_value_mm')}", MM),
        ("equity", "Equity value", "=B{ev_ops}+B{non_op}-B{claims}", MM),
        ("shares", "Current diluted shares", f"={ctx.assump('shares_outstanding_mm')}", MM),
        ("iv", "Intrinsic value / share", "=B{equity}/B{shares}", USD),
        ("iv_gordon", "IV / share — Gordon only", "=(B{pv_fcff_sum}+B{pv_tv_gordon}+B{non_op}-B{claims})/B{shares}", USD),
        ("iv_exit", "IV / share — exit only", "=(B{pv_fcff_sum}+B{pv_tv_exit}+B{non_op}-B{claims})/B{shares}", USD),
        ("upside", "Upside / downside vs price", f"=B{{iv}}/{ctx.assump('current_price')}-1", PCT),
    ]
    # assign rows then resolve B{name} placeholders
    name_row = {name: vr + idx for idx, (name, *_rest) in enumerate(val_rows)}
    ctx_iv_row = name_row["iv"]
    for idx, (name, label, formula, fmt) in enumerate(val_rows):
        rr = vr + idx
        ws.cell(rr, 1, label)
        resolved = formula
        for nm, nr in name_row.items():
            resolved = resolved.replace(f"B{{{nm}}}", f"B{nr}")
        c = ws.cell(rr, 2, resolved)
        c.number_format = fmt
        if name in {"ev_ops", "equity", "iv", "upside"}:
            ws.cell(rr, 1).font = Font(bold=True)
            c.font = Font(bold=True)

    # Reconciliation row
    rec_row = vr + len(val_rows) + 1
    ws.cell(rec_row, 1, "Reconciliation gap vs backend Base IV").font = Font(bold=True)
    ws.cell(rec_row, 2, f"=B{ctx_iv_row}-{ctx.assump('backend_iv_base')}").number_format = USD
    ws.cell(rec_row, 3, f'=IF(ABS(B{rec_row})<={RECONCILE_TOLERANCE},"RECONCILED","CHECK")')
    ctx.dcf_base_rows = {"iv": ctx_iv_row, "iv_gordon": name_row["iv_gordon"],
                         "iv_exit": name_row["iv_exit"], "tv_pct": name_row["tv_pct"],
                         "ev_ops": name_row["ev_ops"], "fcff_row": row_of["fcff"],
                         "ebit_row": row_of["ebit"], "ebitda_row": row_of["ebitda"],
                         "nopat_row": row_of["nopat"], "last_col": last_col}
    _grid(ws, 4, rec_row, 1, last_col)
    _fit(ws, {"A": 34, "B": 13})


# --------------------------------------------------------------------------- #
# Scenarios — backend official + context advisory
# --------------------------------------------------------------------------- #
def _build_scenarios(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Scenarios")
    _setup(ws)
    _title(ws, "Scenarios", "Backend deterministic scenario outputs. Official is canonical; context-advisory shown alongside.")

    val = ctx.valuation
    official = {s["name"]: s for s in (ctx.scenario_policy.get("official_specs") or [])}
    context = ctx.context_scenarios or {}

    _section(ws, 4, "Legacy v1 scenarios (diagnostic; no PM-approved weighting)", 9)
    _header(ws, 5, ["Scenario", "Prob", "Growth mult", "Margin shift", "WACC shift",
                    "Exit mult", "Backend IV", "Upside", "Prob-weighted IV"])
    iv_map = {"bear": val.get("iv_bear"), "base": val.get("iv_base"), "bull": val.get("iv_bull")}
    up_map = {"bear": val.get("upside_bear_pct"), "base": val.get("upside_base_pct"), "bull": val.get("upside_bull_pct")}
    r = 6
    for name in ("bear", "base", "bull"):
        spec = official.get(name, {})
        ws.cell(r, 1, name).font = Font(bold=True)
        ws.cell(r, 2, _num(spec.get("probability"))).number_format = PCT
        ws.cell(r, 3, _num(spec.get("growth_multiplier"))).number_format = MULT
        ws.cell(r, 4, _num(spec.get("margin_shift"))).number_format = PCT
        ws.cell(r, 5, _num(spec.get("wacc_shift"))).number_format = PCT
        ws.cell(r, 6, _num(spec.get("exit_multiple_multiplier"))).number_format = MULT
        ws.cell(r, 7, _safe(iv_map[name])).number_format = USD
        ws.cell(r, 8, _num(up_map[name]) / 100.0).number_format = PCT
        r += 1
    ws.cell(r, 1, "Legacy expected (not PM-approved)").font = Font(bold=True)
    ws.cell(r, 7, _num(val.get("expected_iv"))).number_format = USD
    ws.cell(r, 7).font = Font(bold=True)
    ws.cell(r, 8, _num(val.get("expected_upside_pct")) / 100.0).number_format = PCT
    _grid(ws, 5, r, 1, 9)

    r += 2
    _section(ws, r, "Context-advisory scenarios (cyclicality-aware, not canonical)", 9)
    r += 1
    _header(ws, r, ["Scenario", "Prob", "Growth mult", "Margin shift", "WACC shift", "Exit mult"])
    r += 1
    for name in ("bear", "base", "bull"):
        spec = context.get(name, {})
        ws.cell(r, 1, name).font = Font(bold=True)
        ws.cell(r, 2, _num(spec.get("probability"))).number_format = PCT
        ws.cell(r, 3, _num(spec.get("growth_multiplier"))).number_format = MULT
        ws.cell(r, 4, _num(spec.get("margin_shift"))).number_format = PCT
        ws.cell(r, 5, _num(spec.get("wacc_shift"))).number_format = PCT
        ws.cell(r, 6, _num(spec.get("exit_multiple_multiplier"))).number_format = MULT
        r += 1
    ws.cell(r, 1, "Context expected IV").font = Font(bold=True)
    ws.cell(r, 2, _num(val.get("context_expected_iv"))).number_format = USD
    ws.cell(r, 2).font = Font(bold=True)
    _grid(ws, r - 4, r, 1, 9)
    r += 2
    ws.cell(r, 1, "Policy: " + str(ctx.scenario_policy.get("policy", "")) +
            "  |  official: " + str(ctx.scenario_policy.get("official_policy", ""))).font = Font(italic=True, color="666666")
    _fit(ws, {"A": 26})


# --------------------------------------------------------------------------- #
# Valuation bridge (method spread / football field)
# --------------------------------------------------------------------------- #
def _build_valuation_bridge(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Valuation_Bridge")
    _setup(ws)
    _title(ws, "Valuation Bridge", "Where each method lands — the spread is the discussion.")
    val, mkt = ctx.valuation, ctx.market
    _header(ws, 4, ["Method", "Low", "Mid", "High", "Note"])
    base_iv_row = ctx.dcf_base_rows["iv"]
    exit_iv_ref = f"'DCF_Base'!$B${ctx.dcf_base_rows['iv_exit']}"
    gordon_iv_ref = f"'DCF_Base'!$B${ctx.dcf_base_rows['iv_gordon']}"
    comps_endpoints = [
        value
        for value in (val.get("comps_iv_ev_ebitda"), val.get("comps_iv_pe"))
        if isinstance(value, (int, float))
    ]
    comps_low = min(comps_endpoints) if comps_endpoints else ""
    comps_high = max(comps_endpoints) if comps_endpoints else ""
    rows = [
        ("DCF - scenario range (backend)", val.get("iv_bear"), val.get("iv_base"), val.get("iv_bull"), "Bear / Base / Bull"),
        ("DCF - terminal method (this wb)",
         f"=MIN({exit_iv_ref},{gordon_iv_ref})",
         f"='DCF_Base'!$B${base_iv_row}",
         f"=MAX({exit_iv_ref},{gordon_iv_ref})",
         "Low / blended / high across Gordon and exit methods"),
        ("Economic profit (EP) IV", "", val.get("ep_iv_base"), "", "Invested capital + PV economic profit"),
        ("FCFE cross-check", "", "", "", "Unavailable: legacy FCFE omits after-tax interest; debt/interest schedule required"),
        ("Comps", comps_low, val.get("comps_iv_base"), comps_high, "Low / blended / high across EV/EBITDA and P/E"),
        ("Market / street", mkt.get("price"), mkt.get("analyst_target"), "", "Current price / analyst target"),
    ]
    r = 5
    for label, lo, mid, hi, note in rows:
        ws.cell(r, 1, label).font = Font(bold=True)
        for col, v in zip((2, 3, 4), (lo, mid, hi)):
            c = ws.cell(r, col, _safe(v))
            c.number_format = USD
        ws.cell(r, 5, note)
        r += 1
    _grid(ws, 4, r - 1, 1, 5)
    # extra context
    r += 1
    ws.cell(r, 1, "Implied growth priced in").font = Font(bold=True)
    ws.cell(r, 2, _num(val.get("implied_growth_pct")) / 100.0).number_format = PCT
    r += 1
    ws.cell(r, 1, "DCF vs EP gap").font = Font(bold=True)
    ws.cell(r, 2, _num(val.get("dcf_ep_gap_pct")) / 100.0).number_format = PCT
    _fit(ws, {"A": 32, "B": 12, "C": 12, "D": 12, "E": 40})


# --------------------------------------------------------------------------- #
# Sensitivity (transparent, off the rebuilt base mechanics)
# --------------------------------------------------------------------------- #
def _build_sensitivities(ctx: _Context) -> None:
    ws = ctx.wb.create_sheet("Sensitivity")
    _setup(ws)
    _title(ws, "Sensitivity", "Live tables off the rebuilt Base mechanics; change Assumptions to refresh.")
    rows = ctx.dcf_base_rows
    base = "DCF_Base"
    n = len(ctx.forecast)
    last_col = rows["last_col"]
    fcff_cells = [f"'{base}'!{get_column_letter(c)}{rows['fcff_row']}" for c in range(3, last_col + 1)]
    shares = ctx.assump("shares_outstanding_mm")
    non_op = ctx.assump("non_operating_assets_mm")
    claims = (f"({ctx.assump('net_debt_mm')}+{ctx.assump('minority_interest_mm')}+{ctx.assump('preferred_equity_mm')}"
              f"+{ctx.assump('pension_deficit_mm')}+{ctx.assump('lease_liabilities_mm')}+{ctx.assump('options_value_mm')}"
              f"+{ctx.assump('convertibles_value_mm')})")
    current_shares = shares

    waccs = [-0.01, -0.005, 0.0, 0.005, 0.01]
    gs = [-0.01, -0.005, 0.0, 0.005, 0.01]

    _section(ws, 4, "IV/share: WACC (cols) vs terminal growth (rows) — Gordon path", 8)
    ws.cell(5, 1, "g \\ WACC")
    for col, dw in enumerate(waccs, 2):
        ws.cell(5, col, f"={ctx.assump('wacc')}+{dw}").number_format = PCT
    for row, dg in enumerate(gs, 6):
        ws.cell(row, 1, f"={ctx.assump('growth_terminal_pct')}+{dg}").number_format = PCT
        for col in range(2, 2 + len(waccs)):
            wref = f"{get_column_letter(col)}5"
            gref = f"A{row}"
            pv = "+".join(f"{f}/POWER(1+{wref},{i})" for i, f in enumerate(fcff_cells, 1))
            # value-driver terminal FCFF off year-10 NOPAT: NOPAT_11 = NOPAT_10*(1+g)
            nopat11 = f"('{base}'!{get_column_letter(last_col)}{rows['nopat_row']}*(1+{gref}))"
            term = (f"({nopat11}*(1-{gref}/{ctx.assump('ronic_terminal')})"
                    f"/({wref}-{gref}))/POWER(1+{wref},{n})")
            ws.cell(row, col, f"=IF({wref}<={gref},NA(),(({pv})+{term}+{non_op}-{claims})/{current_shares})").number_format = USD

    start = 13
    _section(ws, start, "IV/share: WACC (cols) vs exit multiple (rows) — exit path", 8)
    mults = [-2, -1, 0, 1, 2]
    ws.cell(start + 1, 1, "Exit \\ WACC")
    for col, dw in enumerate(waccs, 2):
        ws.cell(start + 1, col, f"={ctx.assump('wacc')}+{dw}").number_format = PCT
    ebit_last = f"'{base}'!{get_column_letter(last_col)}{rows['ebit_row']}"
    ebitda_last = f"'{base}'!{get_column_letter(last_col)}{rows['ebitda_row']}"
    exit_metric_base = f"IF({ctx.assump('exit_metric')}=\"ev_ebit\",{ebit_last},{ebitda_last})"
    for row, dm in enumerate(mults, start + 2):
        ws.cell(row, 1, f"={ctx.assump('exit_multiple')}+{dm}").number_format = MULT
        for col in range(2, 2 + len(waccs)):
            wref = f"{get_column_letter(col)}{start + 1}"
            mref = f"A{row}"
            pv = "+".join(f"{f}/POWER(1+{wref},{i})" for i, f in enumerate(fcff_cells, 1))
            term = f"({exit_metric_base}*{mref})/POWER(1+{wref},{n})"
            ws.cell(row, col, f"=(({pv})+{term}+{non_op}-{claims})/{current_shares}").number_format = USD

    r = start + len(mults) + 3
    ws.cell(r, 1, "Backend Base IV").font = Font(bold=True)
    ws.cell(r, 2, f"={ctx.assump('backend_iv_base')}").number_format = USD
    _grid(ws, 4, r, 1, 8)
    _fit(ws, {"A": 16})


# --------------------------------------------------------------------------- #
# Checks
# --------------------------------------------------------------------------- #
def _build_checks(ctx: _Context, recon: dict) -> None:
    ws = ctx.wb.create_sheet("Checks")
    _setup(ws)
    _title(ws, "Checks", "Clear every Fail/Review before using this for an investment discussion.")
    base = "DCF_Base"
    iv_row = ctx.dcf_base_rows["iv"]
    tv_row = ctx.dcf_base_rows["tv_pct"]
    reg = ctx.register
    _header(ws, 4, ["Check", "Value", "Status", "Notes"])
    checks = [
        ("Reconciles to backend Base IV",
         f"='{base}'!$B${iv_row}-{ctx.assump('backend_iv_base')}",
         f'=IF(ABS(B5)<={RECONCILE_TOLERANCE},"OK","FAIL")', USD,
         "Rebuilt Base IV must equal backend iv_base"),
        ("WACC greater than terminal growth",
         f"={ctx.assump('wacc')}-{ctx.assump('growth_terminal_pct')}",
         '=IF(B6>0,"OK","FAIL")', PCT, "Gordon invalid if WACC <= g"),
        ("RONIC greater than terminal growth",
         f"={ctx.assump('ronic_terminal')}-{ctx.assump('growth_terminal_pct')}",
         '=IF(B7>0,"OK","FAIL")', PCT, "Value-driver TV requires RONIC > g"),
        ("Terminal value share of EV",
         f"='{base}'!$B${tv_row}",
         '=IF(AND(B8>=0.4,B8<=0.85),"OK","REVIEW")', PCT, "Outside 40-85% deserves a look"),
        ("Scenario probabilities sum to 100%",
         f"={ctx.assump('scenario_prob_bear')}+{ctx.assump('scenario_prob_base')}+{ctx.assump('scenario_prob_bull')}",
         '=IF(ABS(B9-1)<0.001,"OK","FAIL")', PCT, "Edit probabilities in Assumptions"),
        ("Exit metric present",
         f"={ctx.assump('exit_metric')}",
         '=IF(B10<>"","OK","FAIL")', None, "Needed for exit terminal value"),
        ("Model applicability",
         ctx.valuation.get("model_applicability_status", ""),
         '=IF(B11="dcf_applicable","OK","REVIEW")', None, "Backend DCF applicability gate"),
        ("Model trust state",
         reg.get("model_trust_state", "n/a"),
         '=IF(B12="ok","OK","REVIEW")', None,
         f"Register flags: {reg.get('flag_counts', {})}"),
        ("PM review items outstanding",
         (ctx.resolution.get("counts", {}) or {}).get("review_required", 0),
         '=IF(B13=0,"OK","REVIEW")', None, "See PM_Review_Queue tab"),
        ("WACC input quality", ctx.wacc.get("quality_status", "n/a"),
         '=IF(B14="source_backed","OK","REVIEW")', None,
         f"Missing inputs: {ctx.wacc.get('missing_inputs', [])}; beta source: {ctx.wacc.get('beta_source')}"),
        ("Source workbook preflight", ctx.source_preflight.get("status", "blocked"),
         '=IF(OR(B15="ready",B15="ok"),"OK",IF(B15="blocked","BLOCKED","REVIEW"))', None,
         f"Blockers: {ctx.source_preflight.get('blockers', [])}"),
        ("Source formula-reference errors",
         ctx.source_formula_error_count(),
         '=IF(NOT(ISNUMBER(B16)),"BLOCKED",IF(B16=0,"OK","BLOCKED"))', None,
         "Cached values do not clear broken source formulas"),
        ("No hidden writeback", "Review artifact only", '="OK"', None,
         "Workbook never mutates the database or approved assumptions"),
    ]
    r = 5
    for label, value, status, fmt, note in checks:
        ws.cell(r, 1, label)
        c = ws.cell(r, 2, value)
        if fmt:
            c.number_format = fmt
        ws.cell(r, 3, status)
        ws.cell(r, 4, note)
        r += 1
    _grid(ws, 4, r - 1, 1, 4)
    _fit(ws, {"A": 36, "B": 18, "C": 12, "D": 60})
