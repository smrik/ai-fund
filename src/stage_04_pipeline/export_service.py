from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db.schema import create_tables, get_connection
from db.ticker_dossier import upsert_ticker_dossier_snapshot
from src.stage_04_pipeline.batch_funnel import load_saved_watchlist
from src.stage_04_pipeline.comps_dashboard import build_comps_dashboard_view
from src.stage_04_pipeline.dcf_audit import build_dcf_audit_view
from src.stage_04_pipeline.dossier_view import build_publishable_memo_context, build_research_board_view
from src.stage_04_pipeline.override_workbench import build_override_workbench
from src.stage_04_pipeline.report_archive import list_report_snapshots, load_report_snapshot
from src.stage_04_pipeline.ticker_dossier import build_ticker_dossier_from_export_payload, ticker_dossier_to_payload
from src.stage_04_pipeline.wacc_workbench import build_wacc_workbench
from src.utils import coerce_ticker, utc_now_iso

PROJECT_ROOT = Path(__file__).resolve().parents[2]
EXPORT_ROOT = PROJECT_ROOT / "data" / "exports" / "generated"
TICKER_EXPORT_TEMPLATE = PROJECT_ROOT / "templates" / "ticker_review.xlsx"
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F3864", end_color="1F3864")
SECTION_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
SUBTLE_FILL = PatternFill(fill_type="solid", start_color="F5F7FA", end_color="F5F7FA")


def _ensure_schema(conn) -> None:
    create_tables(conn)


def _coerce_ticker(ticker: str) -> str:
    return coerce_ticker(ticker)


def _now() -> str:
    return utc_now_iso()


def _json_dump(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def _artifact_row(
    *,
    artifact_key: str,
    artifact_role: str,
    title: str,
    path: Path,
    mime_type: str,
    is_primary: bool = False,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "artifact_key": artifact_key,
        "artifact_role": artifact_role,
        "title": title,
        "path": Path(path),
        "mime_type": mime_type,
        "is_primary": bool(is_primary),
        "metadata_json": json.dumps(metadata or {}, sort_keys=True),
    }


def _safe_sheet_title(value: str | None) -> str:
    return str(value or "").strip()


def _normalise_comps_analysis(comps: dict[str, Any] | None) -> dict[str, Any]:
    comps = dict(comps or {})
    target_vs_peers = comps.get("target_vs_peers") or {"target": {}, "peer_medians": {}, "deltas": {}}
    comparison_summary = comps.get("comparison_summary") or [
        {
            "metric": key,
            "label": key.replace("_", " ").title(),
            "target": value,
            "peer_median": (target_vs_peers.get("peer_medians") or {}).get(key),
            "delta": (target_vs_peers.get("deltas") or {}).get(key),
        }
        for key, value in (target_vs_peers.get("target") or {}).items()
    ]
    peer_table = comps.get("peer_table") or [
        {
            **row,
            "display_name": row.get("display_name") or row.get("name") or row.get("ticker"),
        }
        for row in comps.get("peers") or []
    ]
    return {
        "primary_metric": comps.get("primary_metric"),
        "peer_counts": comps.get("peer_counts") or {"raw": len(peer_table), "clean": len(peer_table)},
        "valuation_range": comps.get("valuation_range") or {},
        "valuation_by_metric_rows": comps.get("valuation_by_metric_rows") or [],
        "comparison_summary": comparison_summary,
        "peer_table": peer_table,
        "metric_status_rows": comps.get("metric_status_rows") or [],
        "football_field": comps.get("football_field") or {"ranges": [], "markers": [], "range_min": None, "range_max": None},
        "historical_multiples_summary": comps.get("historical_multiples_summary") or {"available": False, "metrics": {}},
        "operating_context": comps.get("operating_context") or {"target": {}, "peer_medians": {}, "peer_count": 0},
        "support_data_quality": comps.get("support_data_quality")
        or {
            "target_missing_fields": [],
            "peer_coverage": {},
            "valuation_metric_count": len(comps.get("valuation_by_metric_rows") or []),
            "common_patchups_needed": [],
        },
        "audit_flags": list(comps.get("audit_flags") or []),
        "notes": comps.get("notes") or "",
        "source_lineage": comps.get("source_lineage") or {},
        "similarity_method": comps.get("similarity_method"),
        "similarity_model": comps.get("similarity_model"),
        "weighting_formula": comps.get("weighting_formula"),
    }


def _empty_analyst_prep_payload(ticker: str, *, reason: str | None = None) -> dict[str, Any]:
    missing_data = []
    export_metadata: dict[str, Any] = {"status": "missing"}
    if reason:
        missing_data.append(
            {
                "flag_id": "analyst_prep_export_build_error",
                "label": "Analyst Prep unavailable",
                "severity": "high",
                "reason": reason,
                "suggested_check": "Run scripts/manual/run_analyst_prep_pack.py for diagnostics.",
                "source": "export_service",
            }
        )
        export_metadata["builder_error"] = reason
    return {
        "ticker": coerce_ticker(ticker),
        "generated_at": _now(),
        "source_quality": "missing",
        "sections": [],
        "thesis_cards": [],
        "driver_cards": [],
        "comps_card": None,
        "missing_data": missing_data,
        "segment_driver_rows": [],
        "evidence_packet_ids": [],
        "evidence_map": [],
        "conflict_groups": [],
        "export_metadata": export_metadata,
    }


def build_analyst_prep_export_payload(ticker: str) -> dict[str, Any]:
    from src.stage_04_pipeline.analyst_prep_pack import build_analyst_prep_payload

    return build_analyst_prep_payload(ticker)


def _safe_analyst_prep_export_payload(ticker: str) -> dict[str, Any]:
    try:
        return build_analyst_prep_export_payload(ticker)
    except Exception as exc:
        return _empty_analyst_prep_payload(ticker, reason=str(exc))


def _attach_ticker_dossier(
    payload: dict[str, Any],
    *,
    source_mode: str,
    snapshot_id: int | None = None,
) -> dict[str, Any]:
    dossier = build_ticker_dossier_from_export_payload(
        payload,
        source_mode=source_mode,
        snapshot_id=snapshot_id,
    )
    payload["ticker_dossier"] = ticker_dossier_to_payload(dossier)
    return payload


def _persist_attached_ticker_dossier(payload: dict[str, Any]) -> None:
    dossier_payload = payload.get("ticker_dossier")
    if not isinstance(dossier_payload, dict):
        return
    upsert_ticker_dossier_snapshot(dossier_payload, connection_factory=get_connection)


def _html_context_scalars_from_dossier(payload: dict[str, Any]) -> dict[str, Any]:
    dossier = payload.get("ticker_dossier")
    if not isinstance(dossier, dict):
        return {}
    latest = dossier.get("latest_snapshot") if isinstance(dossier.get("latest_snapshot"), dict) else {}
    identity = latest.get("company_identity") if isinstance(latest.get("company_identity"), dict) else {}
    market = latest.get("market_snapshot") if isinstance(latest.get("market_snapshot"), dict) else {}
    valuation = latest.get("valuation_snapshot") if isinstance(latest.get("valuation_snapshot"), dict) else {}
    metadata = dossier.get("export_metadata") if isinstance(dossier.get("export_metadata"), dict) else {}
    current_price = market.get("price")
    if current_price is None:
        current_price = valuation.get("current_price")
    return {
        "ticker": dossier.get("ticker"),
        "company_name": dossier.get("display_name") or identity.get("display_name"),
        "source_mode": metadata.get("source_mode"),
        "current_price": current_price,
        "base_iv": valuation.get("base_iv"),
        "expected_iv": valuation.get("expected_iv"),
        "as_of_date": dossier.get("as_of_date"),
        "snapshot_id": metadata.get("snapshot_id"),
        "ticker_dossier_contract_version": dossier.get("contract_version"),
        "ticker_dossier": dossier,
    }


def _apply_html_context_scalars(context: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    scalars = _html_context_scalars_from_dossier(payload)
    for key, value in scalars.items():
        if value is not None:
            context[key] = value

    valuation = dict(context.get("valuation") or {})
    for context_key, valuation_key in (
        ("current_price", "current_price"),
        ("base_iv", "iv_base"),
        ("expected_iv", "expected_iv"),
    ):
        value = scalars.get(context_key)
        if value is not None:
            valuation[valuation_key] = value
    context["valuation"] = valuation
    return context


def _clear_sheet(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.number_format = "General"
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal="general", vertical="bottom")


def _style_section_title(ws, cell_ref: str, title: str) -> None:
    ws[cell_ref] = title
    ws[cell_ref].font = SECTION_FONT
    ws[cell_ref].fill = SECTION_FILL


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]]) -> int:
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_row = start_row + 1
    for row_values in rows:
        for idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=data_row, column=idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        data_row += 1
    return data_row


def _autosize_columns(ws, widths: dict[int, float]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def _set_title(ws, title: str, merge_range: str = "A1:H1") -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if ":" in merge_range:
        ws.merge_cells(merge_range)


def _write_kv_rows(ws, start_row: int, rows: list[tuple[Any, Any]], *, label_col: int = 1, value_col: int = 2) -> int:
    row_idx = start_row
    for label, value in rows:
        ws.cell(row=row_idx, column=label_col, value=label).font = SECTION_FONT
        ws.cell(row=row_idx, column=label_col).fill = SUBTLE_FILL
        ws.cell(row=row_idx, column=value_col, value=value)
        row_idx += 1
    return row_idx


def _as_percent_change(numerator: Any, denominator: Any) -> float | None:
    try:
        numerator_float = float(numerator)
        denominator_float = float(denominator)
    except (TypeError, ValueError):
        return None
    if denominator_float == 0:
        return None
    return round((numerator_float / denominator_float - 1) * 100, 2)


def _review_identity(payload: dict[str, Any]) -> tuple[str, str]:
    ticker = coerce_ticker(str(payload.get("ticker") or ""))
    return ticker, str(payload.get("company_name") or ticker)


def _review_checks(payload: dict[str, Any], comps_analysis: dict[str, Any]) -> list[list[Any]]:
    market = payload.get("market") or {}
    valuation = payload.get("valuation") or {}
    terminal = payload.get("terminal") or {}
    ciq_lineage = payload.get("ciq_lineage") or {}
    source_lineage = payload.get("source_lineage") or {}
    assumptions = payload.get("assumptions") or {}
    default_resolution = payload.get("default_resolution") or {}

    rows: list[list[Any]] = []
    for flag, active in sorted((payload.get("health_flags") or {}).items()):
        if active:
            rows.append(["High", "DCF health", flag, "Inspect before relying on the valuation output."])

    tv_pct = terminal.get("tv_pct_of_ev")
    if tv_pct is not None:
        try:
            if float(tv_pct) >= 70:
                rows.append(["High", "Terminal value", f"Terminal value is {tv_pct}% of EV", "Pressure-test WACC, terminal growth, and exit multiple."])
        except (TypeError, ValueError):
            pass

    current_price = market.get("price") if market.get("price") is not None else valuation.get("current_price")
    base_iv = valuation.get("iv_base") if valuation.get("iv_base") is not None else valuation.get("base_iv")
    upside = _as_percent_change(base_iv, current_price)
    if upside is not None and abs(upside) >= 100:
        rows.append(["High", "Valuation range", f"Base IV differs from price by {upside}%", "Treat as a hypothesis until data quality and drivers are reviewed."])

    comps_base = (comps_analysis.get("valuation_range") or {}).get("base")
    comps_gap = _as_percent_change(base_iv, comps_base)
    if comps_gap is not None and abs(comps_gap) >= 50:
        rows.append(["High", "DCF vs comps", f"DCF base differs from comps base by {comps_gap}%", "Review whether the multiple set or terminal assumptions are driving the gap."])

    for flag in comps_analysis.get("audit_flags") or []:
        rows.append(["Medium", "Comps", flag, "Review peer set and source quality before using comps as a valuation anchor."])

    source_file = str(ciq_lineage.get("snapshot_source_file") or "")
    if "fallback" in source_file.lower():
        rows.append(["Medium", "Data source", f"Using {source_file}", "Refresh CIQ workbook if Cap IQ detail should be the source of record."])

    for field, source in sorted(source_lineage.items()):
        source_text = str(source or "")
        if "default" in source_text.lower():
            rows.append(["Medium", "Assumption source", f"{field}: {source_text}", "Replace default-backed fields with CIQ, filing, or PM-approved assumptions."])

    for item in default_resolution.get("fields") or []:
        if not item.get("needs_pm_review"):
            continue
        severity = str(item.get("severity") or "medium").title()
        rows.append(
            [
                severity,
                "Default resolution",
                f"{item.get('field')}: {item.get('source')} ({item.get('source_class')})",
                item.get("why_it_matters") or "Review before relying on the valuation output.",
            ]
        )

    missing_market = [field for field in ("analyst_price_target", "analyst_recommendation") if market.get(field) is None]
    if missing_market:
        rows.append(["Low", "Market data", f"Missing: {', '.join(missing_market)}", "Optional context gap; does not block deterministic valuation."])

    if not assumptions:
        rows.append(["High", "Assumptions", "No assumptions payload found", "Do not use workbook until exporter payload is fixed."])

    return rows or [["Info", "Review", "No active review checks generated", "Still inspect source lineage and assumptions manually."]]


def _populate_cover_sheet(workbook, payload: dict[str, Any], comps_analysis: dict[str, Any]) -> None:
    ws = workbook["Cover"] if "Cover" in workbook.sheetnames else workbook.create_sheet("Cover")
    _clear_sheet(ws)
    ticker, company_name = _review_identity(payload)
    valuation = payload.get("valuation") or {}
    market = payload.get("market") or {}
    ciq_lineage = payload.get("ciq_lineage") or {}

    _set_title(ws, f"{ticker} - PM Review Workbook")
    ws["A2"] = "Generated from the export JSON payload. Visible review tabs are rewritten during export to avoid stale template values."
    ws["A2"].alignment = Alignment(wrap_text=True)

    _style_section_title(ws, "A4", "Identity")
    _write_kv_rows(
        ws,
        5,
        [
            ("Company", company_name),
            ("Sector", payload.get("sector")),
            ("Generated At", payload.get("generated_at")),
            ("Current Price", market.get("price") if market.get("price") is not None else valuation.get("current_price")),
            ("Base IV", valuation.get("iv_base") if valuation.get("iv_base") is not None else valuation.get("base_iv")),
            ("Bear / Bull IV", f"{valuation.get('iv_bear')} / {valuation.get('iv_bull')}"),
            ("Snapshot Source", ciq_lineage.get("snapshot_source_file")),
            ("Peer Count", ciq_lineage.get("peer_count")),
        ],
    )

    _style_section_title(ws, "D4", "Top Review Checks")
    for row_idx, row in enumerate(_review_checks(payload, comps_analysis)[:8], start=5):
        ws.cell(row=row_idx, column=4, value=row[0]).font = SECTION_FONT
        ws.cell(row=row_idx, column=5, value=row[1])
        ws.cell(row=row_idx, column=6, value=row[2])
        ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True)

    _autosize_columns(ws, {1: 24, 2: 24, 4: 14, 5: 22, 6: 52})


def _populate_output_sheet(workbook, payload: dict[str, Any], comps_analysis: dict[str, Any]) -> None:
    ws = workbook["Output"] if "Output" in workbook.sheetnames else workbook.create_sheet("Output")
    _clear_sheet(ws)
    ticker, company_name = _review_identity(payload)
    market = payload.get("market") or {}
    valuation = payload.get("valuation") or {}
    terminal = payload.get("terminal") or {}
    assumptions = payload.get("assumptions") or {}
    scenarios = payload.get("scenarios") or {}

    _set_title(ws, f"{ticker} - Valuation Summary")
    ws["A2"] = company_name
    _style_section_title(ws, "A4", "Headline")
    current_price = market.get("price") if market.get("price") is not None else valuation.get("current_price")
    base_iv = valuation.get("iv_base") if valuation.get("iv_base") is not None else valuation.get("base_iv")
    _write_kv_rows(
        ws,
        5,
        [
            ("Current Price", current_price),
            ("Base IV", base_iv),
            ("Base Upside %", _as_percent_change(base_iv, current_price)),
            ("Expected IV", valuation.get("expected_iv")),
            ("Bear IV", valuation.get("iv_bear")),
            ("Bull IV", valuation.get("iv_bull")),
        ],
    )

    _style_section_title(ws, "D4", "Scenario Summary")
    _write_table(
        ws,
        5,
        ["Scenario", "IV", "Upside %", "Probability"],
        [
            [name.title(), scenario.get("iv"), scenario.get("upside_pct"), scenario.get("probability")]
            for name, scenario in scenarios.items()
            if isinstance(scenario, dict)
        ],
    )

    _style_section_title(ws, "A13", "Core Drivers")
    _write_kv_rows(
        ws,
        14,
        [
            ("Growth Near %", assumptions.get("growth_near_pct")),
            ("Growth Mid %", assumptions.get("growth_mid_pct")),
            ("EBIT Margin Start %", assumptions.get("ebit_margin_start_pct")),
            ("EBIT Margin Target %", assumptions.get("ebit_margin_target_pct")),
            ("WACC %", (payload.get("wacc") or {}).get("wacc_pct")),
            ("Terminal Growth %", terminal.get("terminal_growth_pct")),
            ("Terminal ROIC %", terminal.get("ronic_terminal_pct")),
            ("Exit Multiple", assumptions.get("exit_multiple")),
            ("TV % of EV", terminal.get("tv_pct_of_ev")),
        ],
    )

    valuation_range = comps_analysis.get("valuation_range") or {}
    _style_section_title(ws, "D13", "Comps Cross-Check")
    _write_kv_rows(
        ws,
        14,
        [
            ("Primary Metric", comps_analysis.get("primary_metric")),
            ("Comps Bear / Base / Bull", f"{valuation_range.get('bear')} / {valuation_range.get('base')} / {valuation_range.get('bull')}"),
            ("Comps Blended Base", valuation_range.get("blended_base")),
            ("DCF vs Comps Base %", _as_percent_change(base_iv, valuation_range.get("base"))),
            ("Peer Counts Raw / Clean", f"{(comps_analysis.get('peer_counts') or {}).get('raw')} / {(comps_analysis.get('peer_counts') or {}).get('clean')}"),
        ],
    )
    _autosize_columns(ws, {1: 24, 2: 18, 4: 24, 5: 24, 6: 16, 7: 16})


def _populate_assumptions_sheet(workbook, payload: dict[str, Any]) -> None:
    ws = workbook["Assumptions"] if "Assumptions" in workbook.sheetnames else workbook.create_sheet("Assumptions")
    _clear_sheet(ws)
    ticker, _ = _review_identity(payload)
    assumptions = payload.get("assumptions") or {}
    source_lineage = payload.get("source_lineage") or {}
    wacc = payload.get("wacc") or {}

    _set_title(ws, f"{ticker} - Assumptions Register")
    rows = []
    for field, value in sorted(assumptions.items()):
        lineage_key = field
        if field == "growth_near_pct":
            lineage_key = "revenue_growth_near"
        elif field == "growth_mid_pct":
            lineage_key = "revenue_growth_mid"
        elif field == "ebit_margin_start_pct":
            lineage_key = "ebit_margin_start"
        elif field == "ebit_margin_target_pct":
            lineage_key = "ebit_margin_target"
        rows.append([field, value, source_lineage.get(lineage_key), ""])
    if wacc:
        rows.append(["wacc_pct", wacc.get("wacc_pct"), source_lineage.get("wacc"), ""])
    _write_table(ws, 4, ["Field", "Value", "Source", "PM Notes"], rows)
    _autosize_columns(ws, {1: 28, 2: 18, 3: 30, 4: 42})


def _populate_dcf_sheet(workbook, payload: dict[str, Any], scenario_name: str) -> None:
    sheet_name = f"DCF_{scenario_name.title()}"
    ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    _clear_sheet(ws)
    ticker, _ = _review_identity(payload)
    scenarios = payload.get("scenarios") or {}
    scenario = scenarios.get(scenario_name.lower()) if isinstance(scenarios.get(scenario_name.lower()), dict) else {}
    forecast_rows = payload.get("forecast_bridge") if isinstance(payload.get("forecast_bridge"), list) else []

    _set_title(ws, f"{ticker} - DCF {scenario_name.title()}")
    _write_kv_rows(
        ws,
        3,
        [
            ("Scenario IV", scenario.get("iv") or (payload.get("valuation") or {}).get(f"iv_{scenario_name.lower()}")),
            ("Upside %", scenario.get("upside_pct")),
            ("Probability", scenario.get("probability")),
            ("Forecast Detail", "Base forecast bridge only" if scenario_name.lower() != "base" else "Base forecast bridge"),
        ],
    )
    _style_section_title(ws, "A9", "Forecast Bridge")
    _write_table(
        ws,
        10,
        ["Year", "Revenue", "Growth %", "EBIT Margin %", "EBIT", "NOPAT", "D&A", "Capex", "Delta NWC", "FCFF", "ROIC %"],
        [
            [
                row.get("year"),
                row.get("revenue_mm"),
                row.get("growth_pct"),
                row.get("ebit_margin_pct"),
                row.get("ebit_mm"),
                row.get("nopat_mm"),
                row.get("da_mm"),
                row.get("capex_mm"),
                row.get("delta_nwc_mm"),
                row.get("fcff_mm"),
                row.get("roic_pct"),
            ]
            for row in forecast_rows
        ],
    )
    _autosize_columns(ws, {1: 10, 2: 14, 3: 12, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 14, 11: 12})


def _populate_equity_bridge_sheet(workbook, payload: dict[str, Any]) -> None:
    ws = workbook["Equity_Bridge"] if "Equity_Bridge" in workbook.sheetnames else workbook.create_sheet("Equity_Bridge")
    _clear_sheet(ws)
    ticker, _ = _review_identity(payload)
    assumptions = payload.get("assumptions") or {}
    terminal = payload.get("terminal") or {}
    valuation = payload.get("valuation") or {}

    _set_title(ws, f"{ticker} - Equity Bridge")
    _style_section_title(ws, "A4", "Equity Adjustments")
    _write_kv_rows(
        ws,
        5,
        [
            ("Net Debt", assumptions.get("net_debt_mm")),
            ("Lease Liabilities", assumptions.get("lease_liabilities_mm")),
            ("Non-Operating Assets", assumptions.get("non_operating_assets_mm")),
            ("Preferred Equity", assumptions.get("preferred_equity_mm")),
            ("Minority Interest", assumptions.get("minority_interest_mm")),
            ("Pension Deficit", assumptions.get("pension_deficit_mm")),
        ],
    )

    _style_section_title(ws, "D4", "Terminal Bridge")
    _write_kv_rows(
        ws,
        5,
        [
            ("PV TV Blended", terminal.get("pv_tv_blended_mm")),
            ("PV TV Gordon", terminal.get("pv_tv_gordon_mm")),
            ("PV TV Exit", terminal.get("pv_tv_exit_mm")),
            ("TV % of EV", terminal.get("tv_pct_of_ev")),
            ("Terminal Method", terminal.get("method_used")),
            ("Gordon Formula Mode", terminal.get("gordon_formula_mode")),
        ],
        label_col=4,
        value_col=5,
    )

    _style_section_title(ws, "A14", "Per Share Output")
    _write_kv_rows(
        ws,
        15,
        [
            ("Current Price", valuation.get("current_price")),
            ("Bear IV", valuation.get("iv_bear")),
            ("Base IV", valuation.get("iv_base") if valuation.get("iv_base") is not None else valuation.get("base_iv")),
            ("Bull IV", valuation.get("iv_bull")),
            ("Expected IV", valuation.get("expected_iv")),
        ],
    )
    _autosize_columns(ws, {1: 24, 2: 18, 4: 24, 5: 18})


def _populate_sensitivity_sheet(workbook, payload: dict[str, Any]) -> None:
    ws = workbook["Sensitivity"] if "Sensitivity" in workbook.sheetnames else workbook.create_sheet("Sensitivity")
    _clear_sheet(ws)
    ticker, _ = _review_identity(payload)
    sensitivity = payload.get("sensitivity") or {}
    _set_title(ws, f"{ticker} - Sensitivity Analysis")

    next_row = 4
    _style_section_title(ws, f"A{next_row}", "Sensitivity Summary")
    next_row = _write_table(
        ws,
        next_row + 1,
        ["Grid", "Cell Count", "Min IV", "Max IV", "Spread"],
        [
            [row.get("grid"), row.get("cell_count"), row.get("min_iv"), row.get("max_iv"), row.get("spread")]
            for row in sensitivity.get("summary") or []
        ],
    )

    for grid_name in ("wacc_x_terminal_growth", "wacc_x_exit_multiple"):
        grid_rows = sensitivity.get(grid_name) if isinstance(sensitivity.get(grid_name), list) else []
        next_row += 2
        _style_section_title(ws, f"A{next_row}", grid_name.replace("_", " ").title())
        if not grid_rows:
            ws.cell(row=next_row + 1, column=1, value="No grid data available")
            next_row += 2
            continue
        headers = list(grid_rows[0].keys())
        next_row = _write_table(ws, next_row + 1, headers, [[row.get(header) for header in headers] for row in grid_rows])
    _autosize_columns(ws, {1: 26, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14})


def _populate_qoe_sheet(workbook, payload: dict[str, Any]) -> None:
    ws = workbook["QoE"] if "QoE" in workbook.sheetnames else workbook.create_sheet("QoE")
    _clear_sheet(ws)
    ticker, _ = _review_identity(payload)
    _set_title(ws, f"{ticker} - Quality Of Earnings / Data Quality")
    ciq_lineage = payload.get("ciq_lineage") or {}
    health_flags = payload.get("health_flags") or {}
    source_lineage = payload.get("source_lineage") or {}
    default_resolution = payload.get("default_resolution") or {}

    _style_section_title(ws, "A4", "Data Lineage")
    _write_kv_rows(
        ws,
        5,
        [
            ("Snapshot As Of", ciq_lineage.get("snapshot_as_of_date")),
            ("Snapshot Source", ciq_lineage.get("snapshot_source_file")),
            ("Peer Count", ciq_lineage.get("peer_count")),
            ("CIQ Source File", ciq_lineage.get("ciq_source_file")),
            ("CIQ Comps Source File", ciq_lineage.get("ciq_comps_source_file")),
        ],
    )

    _style_section_title(ws, "D4", "Health Flags")
    for row_idx, (flag, active) in enumerate(sorted(health_flags.items()), start=5):
        ws.cell(row=row_idx, column=4, value=flag).font = SECTION_FONT
        ws.cell(row=row_idx, column=5, value=active)

    _style_section_title(ws, "A13", "Default-Backed Sources")
    defaults = [[field, source] for field, source in sorted(source_lineage.items()) if "default" in str(source).lower()]
    next_row = _write_table(ws, 14, ["Field", "Source"], defaults)

    next_row += 2
    _style_section_title(ws, f"A{next_row}", "Default Resolution")
    _write_table(
        ws,
        next_row + 1,
        ["Field", "Value", "Source Class", "Severity", "Needs PM Review", "Why It Matters"],
        [
            [
                row.get("field"),
                row.get("value"),
                row.get("source_class"),
                row.get("severity"),
                row.get("needs_pm_review"),
                row.get("why_it_matters"),
            ]
            for row in default_resolution.get("fields") or []
        ],
    )
    _autosize_columns(ws, {1: 26, 2: 18, 3: 20, 4: 14, 5: 18, 6: 52})


def _populate_review_checks_sheet(workbook, payload: dict[str, Any], comps_analysis: dict[str, Any]) -> None:
    ws = workbook["Review Checks"] if "Review Checks" in workbook.sheetnames else workbook.create_sheet("Review Checks")
    _clear_sheet(ws)
    ticker, _ = _review_identity(payload)
    _set_title(ws, f"{ticker} - Review Checks")
    _write_table(ws, 4, ["Severity", "Area", "Finding", "Suggested PM Check"], _review_checks(payload, comps_analysis))
    _autosize_columns(ws, {1: 14, 2: 22, 3: 56, 4: 62})


def _populate_review_workbook_tabs(workbook, payload: dict[str, Any], comps_analysis: dict[str, Any]) -> None:
    _populate_cover_sheet(workbook, payload, comps_analysis)
    _populate_output_sheet(workbook, payload, comps_analysis)
    _populate_assumptions_sheet(workbook, payload)
    for scenario_name in ("base", "bear", "bull"):
        _populate_dcf_sheet(workbook, payload, scenario_name)
    _populate_equity_bridge_sheet(workbook, payload)
    _populate_sensitivity_sheet(workbook, payload)
    _populate_qoe_sheet(workbook, payload)
    _populate_review_checks_sheet(workbook, payload, comps_analysis)


def _populate_comps_sheet(workbook, ticker: str, company_name: str | None, market: dict[str, Any], comps_analysis: dict[str, Any]) -> None:
    ws = workbook["Comps"] if "Comps" in workbook.sheetnames else workbook.create_sheet("Comps")
    _clear_sheet(ws)

    source_lineage = comps_analysis.get("source_lineage") or {}
    valuation_range = comps_analysis.get("valuation_range") or {}
    peer_counts = comps_analysis.get("peer_counts") or {}

    ws["A1"] = f"{ticker} - Comparable Companies Appendix"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    metadata_pairs = [
        ("Company", company_name or ticker),
        ("As Of", source_lineage.get("as_of_date")),
        ("Source File", source_lineage.get("source_file")),
        ("Current Price", (market or {}).get("price")),
    ]
    for offset, (label, value) in enumerate(metadata_pairs):
        row = 2 + offset // 2
        col = 1 + (offset % 2) * 3
        ws.cell(row=row, column=col, value=label).font = SECTION_FONT
        ws.cell(row=row, column=col + 1, value=value)

    _style_section_title(ws, "A4", "Headline Valuation")
    headline_rows = [
        ["Primary Metric", comps_analysis.get("primary_metric")],
        ["Blended Base IV", valuation_range.get("blended_base")],
        ["Bear / Base / Bull", f"{valuation_range.get('bear')} / {valuation_range.get('base')} / {valuation_range.get('bull')}"],
        ["Peer Counts (raw / clean)", f"{peer_counts.get('raw')} / {peer_counts.get('clean')}"],
        ["Similarity Method", comps_analysis.get("similarity_method") or "market_cap_only"],
        ["Weighting Formula", comps_analysis.get("weighting_formula") or "market_cap_proximity_only"],
    ]
    for row_idx, (label, value) in enumerate(headline_rows, start=5):
        ws.cell(row=row_idx, column=1, value=label).fill = SUBTLE_FILL
        ws.cell(row=row_idx, column=1).font = SECTION_FONT
        ws.cell(row=row_idx, column=2, value=value)

    next_row = 12
    _style_section_title(ws, f"A{next_row}", "Valuation By Metric")
    next_row = _write_table(
        ws,
        next_row + 1,
        [
            "Metric",
            "Target Multiple",
            "Peer Median",
            "Bear Multiple",
            "Base Multiple",
            "Bull Multiple",
            "Bear IV",
            "Base IV",
            "Bull IV",
            "N Raw",
            "N Clean",
            "Primary",
        ],
        [
            [
                row.get("label"),
                row.get("target_multiple"),
                row.get("peer_median_multiple"),
                row.get("bear_multiple"),
                row.get("base_multiple"),
                row.get("bull_multiple"),
                row.get("bear_iv"),
                row.get("base_iv"),
                row.get("bull_iv"),
                row.get("n_raw"),
                row.get("n_clean"),
                "Yes" if row.get("is_primary") else "",
            ]
            for row in comps_analysis.get("valuation_by_metric_rows") or []
        ],
    )

    next_row += 1
    _style_section_title(ws, f"A{next_row}", "Target Vs Peer Benchmarks")
    next_row = _write_table(
        ws,
        next_row + 1,
        ["Metric", "Target", "Peer Median", "Delta"],
        [
            [row.get("label"), row.get("target"), row.get("peer_median"), row.get("delta")]
            for row in comps_analysis.get("comparison_summary") or []
        ],
    )

    next_row += 1
    _style_section_title(ws, f"A{next_row}", "Peer Table")
    _write_table(
        ws,
        next_row + 1,
        [
            "Ticker",
            "Company",
            "Similarity Score",
            "Model Weight",
            "Revenue LTM",
            "EBITDA LTM",
            "EBIT LTM",
            "Revenue Growth",
            "EBIT Margin",
            "Net Debt / EBITDA",
            "TEV / EBITDA LTM",
            "TEV / EBIT LTM",
            "P / E LTM",
        ],
        [
            [
                row.get("ticker"),
                row.get("display_name"),
                row.get("similarity_score"),
                row.get("model_weight"),
                row.get("revenue_ltm_mm"),
                row.get("ebitda_ltm_mm"),
                row.get("ebit_ltm_mm"),
                row.get("revenue_growth"),
                row.get("ebit_margin"),
                row.get("net_debt_to_ebitda"),
                row.get("tev_ebitda_ltm"),
                row.get("tev_ebit_ltm"),
                row.get("pe_ltm"),
            ]
            for row in comps_analysis.get("peer_table") or []
        ],
    )
    _autosize_columns(ws, {1: 24, 2: 28, 3: 14, 4: 14, 5: 14, 6: 14, 7: 18, 8: 16, 9: 14, 10: 12, 11: 12, 12: 12})


def _populate_comps_diagnostics_sheet(workbook, ticker: str, comps_analysis: dict[str, Any]) -> None:
    if "Comps Diagnostics" not in workbook.sheetnames:
        comps_index = workbook.sheetnames.index("Comps") + 1 if "Comps" in workbook.sheetnames else len(workbook.sheetnames)
        workbook.create_sheet("Comps Diagnostics", comps_index)
    ws = workbook["Comps Diagnostics"]
    _clear_sheet(ws)

    ws["A1"] = f"{ticker} - Comps Diagnostics"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    _style_section_title(ws, "A4", "Audit Flags")
    audit_flags = list(comps_analysis.get("audit_flags") or []) or ["None"]
    for idx, flag in enumerate(audit_flags, start=5):
        ws.cell(row=idx, column=1, value=flag)

    notes_row = 5 + len(audit_flags) + 1
    ws.cell(row=notes_row, column=1, value="Notes").font = SECTION_FONT
    ws.cell(row=notes_row, column=2, value=comps_analysis.get("notes"))

    next_row = notes_row + 2
    _style_section_title(ws, f"A{next_row}", "Metric Status")
    next_row = _write_table(
        ws,
        next_row + 1,
        ["Ticker", "Company", "Metric", "Label", "Raw Multiple", "Status"],
        [
            [
                row.get("ticker"),
                row.get("display_name"),
                row.get("metric"),
                row.get("label"),
                row.get("raw_multiple"),
                row.get("status"),
            ]
            for row in comps_analysis.get("metric_status_rows") or []
        ],
    )

    next_row += 1
    _style_section_title(ws, f"A{next_row}", "Football Field")
    next_row = _write_table(
        ws,
        next_row + 1,
        ["Label", "Bear", "Base", "Bull"],
        [
            [row.get("label"), row.get("bear"), row.get("base"), row.get("bull")]
            for row in (comps_analysis.get("football_field") or {}).get("ranges") or []
        ],
    )

    next_row += 1
    _style_section_title(ws, f"A{next_row}", "Historical Multiples Summary")
    metrics = (comps_analysis.get("historical_multiples_summary") or {}).get("metrics") or {}
    _write_table(
        ws,
        next_row + 1,
        ["Metric", "Current", "Median", "P25", "P75", "Current Percentile"],
        [
            [
                metric,
                payload.get("current"),
                (payload.get("summary") or {}).get("median"),
                (payload.get("summary") or {}).get("p25"),
                (payload.get("summary") or {}).get("p75"),
                (payload.get("summary") or {}).get("current_percentile"),
            ]
            for metric, payload in metrics.items()
        ],
    )
    _autosize_columns(ws, {1: 18, 2: 24, 3: 18, 4: 18, 5: 18, 6: 18})


def _sheet(workbook, title: str):
    ws = workbook[title] if title in workbook.sheetnames else workbook.create_sheet(title)
    _clear_sheet(ws)
    return ws


def _populate_analyst_prep_sheets(workbook, payload: dict[str, Any]) -> None:
    pack = payload.get("analyst_prep") or {}
    ticker = str(pack.get("ticker") or payload.get("ticker") or "")

    ws = _sheet(workbook, "Analyst_Prep")
    _set_title(ws, f"{ticker} - Analyst Prep Pack")
    _write_kv_rows(
        ws,
        4,
        [
            ("Generated At", pack.get("generated_at")),
            ("Source Quality", pack.get("source_quality")),
            ("Thesis Cards", len(pack.get("thesis_cards") or [])),
            ("Driver Cards", len(pack.get("driver_cards") or [])),
            ("Evidence Packets", len(pack.get("evidence_packet_ids") or [])),
            ("Missing Flags", len(pack.get("missing_data") or [])),
            ("Default Resolution Status", (pack.get("export_metadata") or {}).get("default_resolution_status")),
        ],
    )
    _style_section_title(ws, "A12", "Missing Data")
    _write_table(
        ws,
        13,
        ["Severity", "Label", "Reason", "Suggested Check"],
        [
            [
                row.get("severity"),
                row.get("label"),
                row.get("reason"),
                row.get("suggested_check"),
            ]
            for row in pack.get("missing_data") or []
        ]
        or [["Info", "No missing-data flags", "", ""]],
    )
    _autosize_columns(ws, {1: 24, 2: 26, 4: 14, 5: 26, 6: 56, 7: 42})

    ws = _sheet(workbook, "Thesis_Bridge")
    _set_title(ws, f"{ticker} - Thesis Bridge")
    _write_table(
        ws,
        4,
        [
            "Card ID",
            "Title",
            "Claim",
            "Business Evidence",
            "Model Implication",
            "Linked Fields",
            "Anchors",
            "What Would Change Mind",
        ],
        [
            [
                card.get("card_id"),
                card.get("title"),
                card.get("claim"),
                card.get("business_evidence_summary"),
                card.get("model_implication"),
                ", ".join(card.get("linked_assumption_fields") or []),
                ", ".join(card.get("evidence_anchor_ids") or []),
                card.get("what_would_change_mind"),
            ]
            for card in pack.get("thesis_cards") or []
        ],
    )
    _autosize_columns(ws, {1: 24, 2: 24, 3: 56, 4: 56, 5: 56, 6: 30, 7: 40, 8: 56})

    ws = _sheet(workbook, "Model_Driver_Map")
    _set_title(ws, f"{ticker} - Model Driver Map")
    _write_table(
        ws,
        4,
        ["Assumption", "Label", "Current", "Proposed / Effective", "Source", "Review Status", "Rationale", "Anchors"],
        [
            [
                card.get("assumption_name"),
                card.get("label"),
                card.get("current_value"),
                card.get("proposed_or_effective_value"),
                card.get("source"),
                card.get("pm_review_status"),
                card.get("rationale"),
                ", ".join(card.get("evidence_anchor_ids") or []),
            ]
            for card in pack.get("driver_cards") or []
        ],
    )
    _autosize_columns(ws, {1: 24, 2: 26, 3: 14, 4: 18, 5: 28, 6: 18, 7: 60, 8: 42})

    ws = _sheet(workbook, "Evidence_Map")
    _set_title(ws, f"{ticker} - Evidence Map")
    _write_table(
        ws,
        4,
        ["Anchor", "Packet ID", "Profile", "Kind", "Label", "Value", "Unit", "Source Quality", "Source Ref"],
        [
            [
                row.get("anchor_id"),
                row.get("packet_id"),
                row.get("profile_name"),
                row.get("kind"),
                row.get("label"),
                row.get("value"),
                row.get("unit"),
                row.get("source_quality"),
                row.get("source_ref"),
            ]
            for row in pack.get("evidence_map") or []
        ],
    )
    _autosize_columns(ws, {1: 34, 2: 12, 3: 22, 4: 22, 5: 40, 6: 18, 7: 12, 8: 18, 9: 42})

    ws = _sheet(workbook, "Comps_Judgment")
    _set_title(ws, f"{ticker} - Comps Judgment")
    comps_card = pack.get("comps_card") or {}
    _write_kv_rows(
        ws,
        4,
        [
            ("Peer Set Quality", comps_card.get("peer_set_quality")),
            ("Peer Count", comps_card.get("peer_count")),
            ("Primary Metric", comps_card.get("primary_metric")),
            ("Premium / Discount", comps_card.get("premium_discount_argument")),
            ("Exit Multiple Support", comps_card.get("exit_multiple_support")),
            ("Warnings", " | ".join(comps_card.get("warnings") or [])),
            ("Anchors", ", ".join(comps_card.get("evidence_anchor_ids") or [])),
        ],
    )
    _autosize_columns(ws, {1: 26, 2: 80})

    ws = _sheet(workbook, "Segment_Drivers")
    _set_title(ws, f"{ticker} - Segment Drivers")
    segment_rows = pack.get("segment_driver_rows") or []
    _write_table(
        ws,
        4,
        ["Segment", "Revenue Growth", "Margin", "Revenue Mix", "Source Ref", "Quality"],
        [
            [
                row.get("segment"),
                row.get("revenue_growth"),
                row.get("margin"),
                row.get("revenue_mix"),
                row.get("source_ref"),
                row.get("quality"),
            ]
            for row in segment_rows
        ]
        or [["Segment evidence missing", None, None, None, "segment_driver_rows", "missing"]],
    )
    _autosize_columns(ws, {1: 30, 2: 18, 3: 18, 4: 18, 5: 42, 6: 16})


def _copy_public_artifacts(artifacts: list[dict[str, Any]], bundle_dir: Path) -> list[dict[str, Any]]:
    copied: list[dict[str, Any]] = []
    assets_dir = bundle_dir / "assets"
    assets_dir.mkdir(parents=True, exist_ok=True)
    for artifact in artifacts:
        if artifact.get("path_mode") != "absolute":
            continue
        source_path = Path(str(artifact.get("path_value") or ""))
        if not source_path.exists() or not source_path.is_file():
            continue
        target_path = assets_dir / source_path.name
        shutil.copy2(source_path, target_path)
        copied.append(
            _artifact_row(
                artifact_key=str(artifact.get("artifact_key") or target_path.stem),
                artifact_role="sidecar_asset",
                title=str(artifact.get("title") or target_path.name),
                path=target_path,
                mime_type="application/octet-stream",
            )
        )
    return copied


def _render_html_report(context: dict[str, Any]) -> str:
    ticker = coerce_ticker(str(context.get("ticker") or ""))
    company_name = str(context.get("company_name") or ticker)
    source_mode = str(context.get("source_mode") or "loaded_backend_state")
    summary = str(context.get("summary") or "No publishable memo summary is available.")
    valuation = context.get("valuation") or {}
    current_price = valuation.get("current_price") or context.get("current_price") or "—"
    base_iv = valuation.get("iv_base") or valuation.get("base_iv") or context.get("base_iv") or "—"
    expected_iv = valuation.get("expected_iv") or context.get("expected_iv") or "—"
    artifacts = context.get("artifacts") or []
    artifact_links = "".join(
        f"<li>{item.get('title') or item.get('artifact_key')}</li>"
        for item in artifacts
    )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{ticker} Export</title>
  <style>
    body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px auto; max-width: 960px; line-height: 1.6; color: #111827; }}
    .hero {{ border-bottom: 1px solid #d1d5db; padding-bottom: 16px; margin-bottom: 24px; }}
    .metrics {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin: 16px 0 24px; }}
    .metric {{ border: 1px solid #d1d5db; border-radius: 8px; padding: 12px 14px; }}
    .label {{ font-size: 0.75rem; color: #6b7280; text-transform: uppercase; letter-spacing: 0.04em; }}
    .value {{ font-size: 1.2rem; font-weight: 700; margin-top: 4px; }}
    .summary {{ white-space: pre-wrap; }}
  </style>
</head>
<body>
  <section class="hero">
    <p>{source_mode}</p>
    <h1>{ticker} - {company_name}</h1>
  </section>
  <section class="metrics">
    <article class="metric"><div class="label">Current Price</div><div class="value">{current_price}</div></article>
    <article class="metric"><div class="label">Base IV</div><div class="value">{base_iv}</div></article>
    <article class="metric"><div class="label">Expected IV</div><div class="value">{expected_iv}</div></article>
  </section>
  <section>
    <h2>Summary</h2>
    <div class="summary">{summary}</div>
  </section>
  <section>
    <h2>Linked Public Artifacts</h2>
    <ul>{artifact_links or "<li>None</li>"}</ul>
  </section>
</body>
</html>
"""


def stage_power_query_workbook(ticker: str, payload: dict[str, Any], bundle_dir: Path) -> dict[str, Any]:
    bundle_dir = Path(bundle_dir)
    bundle_dir.mkdir(parents=True, exist_ok=True)
    template_path = Path(TICKER_EXPORT_TEMPLATE)
    if not template_path.exists():
        raise FileNotFoundError(f"Ticker export template not found: {template_path}")

    json_path = bundle_dir / f"{coerce_ticker(ticker)}_latest.json"
    _json_dump(json_path, payload)

    workbook_path = bundle_dir / f"{coerce_ticker(ticker)}_review.xlsx"
    shutil.copy2(template_path, workbook_path)

    workbook = load_workbook(workbook_path)
    if "Config" not in workbook.sheetnames:
        raise ValueError("Ticker export template must contain a Config sheet")
    workbook["Config"]["B2"] = str(json_path.resolve())
    comps_analysis = _normalise_comps_analysis(payload.get("comps_analysis"))
    _populate_review_workbook_tabs(workbook, payload, comps_analysis)
    _populate_analyst_prep_sheets(workbook, payload)
    _populate_comps_sheet(
        workbook,
        coerce_ticker(ticker),
        payload.get("company_name"),
        payload.get("market") or {},
        comps_analysis,
    )
    _populate_comps_diagnostics_sheet(
        workbook,
        coerce_ticker(ticker),
        comps_analysis,
    )
    workbook.save(workbook_path)

    manifest_path = bundle_dir / "manifest.json"
    _json_dump(
        manifest_path,
        {
            "ticker": coerce_ticker(ticker),
            "format": "xlsx",
            "template": str(template_path),
            "artifacts": ["excel_workbook", "power_query_json"],
        },
    )

    return {
        "primary_path": str(workbook_path),
        "bundle_dir": str(bundle_dir),
        "artifacts": [
            _artifact_row(
                artifact_key="excel_workbook",
                artifact_role="primary",
                title=f"{coerce_ticker(ticker)} review workbook",
                path=workbook_path,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                is_primary=True,
            ),
            _artifact_row(
                artifact_key="power_query_json",
                artifact_role="sidecar_data",
                title=f"{coerce_ticker(ticker)} export payload",
                path=json_path,
                mime_type="application/json",
            ),
        ],
    }


def build_html_export_bundle(ticker: str, context: dict[str, Any], bundle_dir: Path) -> dict[str, Any]:
    bundle_dir = Path(bundle_dir)
    bundle_dir.mkdir(parents=True, exist_ok=True)

    html_path = bundle_dir / f"{coerce_ticker(ticker).lower()}-memo.html"
    context_path = bundle_dir / "context.json"
    manifest_path = bundle_dir / "manifest.json"

    html_path.write_text(_render_html_report(context), encoding="utf-8")
    _json_dump(context_path, context)

    sidecars = _copy_public_artifacts(list(context.get("artifacts") or []), bundle_dir)
    artifacts = [
        _artifact_row(
            artifact_key="html_report",
            artifact_role="primary",
            title=f"{coerce_ticker(ticker)} memo export",
            path=html_path,
            mime_type="text/html",
            is_primary=True,
        ),
        _artifact_row(
            artifact_key="context_json",
            artifact_role="sidecar_context",
            title="Export context",
            path=context_path,
            mime_type="application/json",
        ),
        *sidecars,
    ]
    _json_dump(
        manifest_path,
        {
            "ticker": coerce_ticker(ticker),
            "format": "html",
            "artifacts": [
                {
                    "artifact_key": artifact["artifact_key"],
                    "title": artifact["title"],
                    "path": str(Path(artifact["path"]).resolve()),
                }
                for artifact in artifacts
            ],
        },
    )
    return {
        "primary_path": str(html_path),
        "bundle_dir": str(bundle_dir),
        "artifacts": artifacts,
    }


def _export_row_from_db(row: Any) -> dict[str, Any]:
    return {
        "export_id": row["export_id"],
        "scope": row["scope"],
        "ticker": row["ticker"],
        "status": row["status"],
        "export_format": row["export_format"],
        "source_mode": row["source_mode"],
        "template_strategy": row["template_strategy"],
        "title": row["title"],
        "bundle_dir": row["bundle_dir"],
        "primary_artifact_key": row["primary_artifact_key"],
        "created_by": row["created_by"],
        "snapshot_id": row["snapshot_id"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "metadata": json.loads(row["metadata_json"] or "{}"),
    }


def register_export_bundle(
    *,
    scope: str,
    export_format: str,
    source_mode: str,
    template_strategy: str,
    bundle_dir: Path,
    primary_artifact_key: str,
    artifacts: list[dict[str, Any]],
    ticker: str | None = None,
    created_by: str = "api",
    title: str | None = None,
    snapshot_id: int | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    export_id = uuid4().hex
    bundle_dir = Path(bundle_dir)
    bundle_dir.mkdir(parents=True, exist_ok=True)
    ticker_value = coerce_ticker(ticker) if ticker else None
    now = _now()

    with get_connection() as conn:
        _ensure_schema(conn)
        conn.execute(
            """
            INSERT INTO generated_exports (
                export_id, scope, ticker, status, export_format, source_mode,
                template_strategy, title, bundle_dir, primary_artifact_key,
                created_by, snapshot_id, created_at, updated_at, metadata_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                export_id,
                scope,
                ticker_value,
                "completed",
                export_format,
                source_mode,
                template_strategy,
                title or f"{scope} {export_format} export",
                str(bundle_dir.resolve()),
                primary_artifact_key,
                created_by,
                snapshot_id,
                now,
                now,
                json.dumps(metadata or {}, sort_keys=True),
            ],
        )
        for artifact in artifacts:
            path = Path(artifact["path"])
            size_bytes = path.stat().st_size if path.exists() and path.is_file() else None
            conn.execute(
                """
                INSERT INTO generated_export_artifacts (
                    export_id, artifact_key, artifact_role, title, path,
                    mime_type, size_bytes, is_primary, created_at, metadata_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    export_id,
                    artifact["artifact_key"],
                    artifact["artifact_role"],
                    artifact["title"],
                    str(path.resolve()),
                    artifact["mime_type"],
                    size_bytes,
                    1 if artifact.get("is_primary") else 0,
                    now,
                    artifact.get("metadata_json") or "{}",
                ],
            )
        conn.commit()
    return load_export(export_id) or {"export_id": export_id}


def list_exports(*, ticker: str | None = None, scope: str | None = None, limit: int = 25) -> list[dict[str, Any]]:
    query = [
        "SELECT * FROM generated_exports",
        "WHERE 1=1",
    ]
    params: list[Any] = []
    if ticker:
        query.append("AND ticker = ?")
        params.append(coerce_ticker(ticker))
    if scope:
        query.append("AND scope = ?")
        params.append(scope)
    query.append("ORDER BY created_at DESC LIMIT ?")
    params.append(max(int(limit), 1))

    with get_connection() as conn:
        _ensure_schema(conn)
        rows = conn.execute(" ".join(query), params).fetchall()
    return [load_export(row["export_id"]) or _export_row_from_db(row) for row in rows]


def load_export(export_id: str) -> dict[str, Any] | None:
    with get_connection() as conn:
        _ensure_schema(conn)
        row = conn.execute(
            "SELECT * FROM generated_exports WHERE export_id = ? LIMIT 1",
            [export_id],
        ).fetchone()
        if row is None:
            return None
        artifacts = conn.execute(
            """
            SELECT * FROM generated_export_artifacts
            WHERE export_id = ?
            ORDER BY is_primary DESC, artifact_key ASC
            """,
            [export_id],
        ).fetchall()
    payload = _export_row_from_db(row)
    payload["artifacts"] = [
        {
            "artifact_key": artifact["artifact_key"],
            "artifact_role": artifact["artifact_role"],
            "title": artifact["title"],
            "path": artifact["path"],
            "mime_type": artifact["mime_type"],
            "size_bytes": artifact["size_bytes"],
            "is_primary": bool(artifact["is_primary"]),
            "metadata": json.loads(artifact["metadata_json"] or "{}"),
        }
        for artifact in artifacts
    ]
    return payload


def resolve_export_artifact_path(export_id: str, artifact_key: str | None = None) -> Path:
    payload = load_export(export_id)
    if payload is None:
        raise FileNotFoundError(f"Unknown export id: {export_id}")
    if artifact_key is None:
        artifact_key = payload.get("primary_artifact_key")
    for artifact in payload.get("artifacts") or []:
        if artifact["artifact_key"] == artifact_key:
            return Path(str(artifact["path"]))
    raise FileNotFoundError(f"Artifact {artifact_key!r} not found for export {export_id}")


def _build_current_ticker_payload(ticker: str) -> dict[str, Any]:
    ticker = coerce_ticker(ticker)
    workbench = build_override_workbench(ticker)
    dcf = build_dcf_audit_view(ticker)
    comps = build_comps_dashboard_view(ticker)
    wacc = build_wacc_workbench(ticker, apply_overrides=True)
    research = build_research_board_view(ticker)

    assumption_map = {
        row["field"]: row.get("effective_value")
        for row in workbench.get("fields") or []
    }
    source_lineage = {
        row["field"]: row.get("effective_source")
        for row in workbench.get("fields") or []
        if row.get("effective_source")
    }
    scenario_map = {
        str(row.get("scenario") or "").lower(): {
            "probability": row.get("probability"),
            "iv": row.get("intrinsic_value"),
            "upside_pct": row.get("upside_pct"),
        }
        for row in dcf.get("scenario_summary") or []
        if row.get("scenario")
    }
    comps_target = (comps.get("target_vs_peers") or {}).get("target") or {}
    peer_medians = (comps.get("target_vs_peers") or {}).get("peer_medians") or {}
    valuation_ciq_lineage = dict(workbench.get("ciq_lineage") or {})
    comps_lineage = comps.get("source_lineage") or {}
    ciq_lineage = {
        **valuation_ciq_lineage,
        "snapshot_source_file": comps_lineage.get("source_file") or valuation_ciq_lineage.get("snapshot_source_file"),
        "snapshot_as_of_date": comps_lineage.get("as_of_date") or valuation_ciq_lineage.get("snapshot_as_of_date"),
        "peer_count": (comps.get("peer_counts") or {}).get("clean") or valuation_ciq_lineage.get("peer_count"),
        "comps_source_file": comps_lineage.get("source_file") or valuation_ciq_lineage.get("comps_source_file"),
        "comps_as_of_date": comps_lineage.get("as_of_date") or valuation_ciq_lineage.get("comps_as_of_date"),
    }
    payload = {
        "$schema_version": "1.0",
        "generated_at": _now(),
        "ticker": ticker,
        "company_name": workbench.get("company_name") or research.get("company_name") or ticker,
        "sector": workbench.get("sector"),
        "market": {
            "price": workbench.get("current_price"),
            "analyst_target": research.get("analyst_target"),
            "analyst_recommendation": research.get("tracker", {}).get("pm_action"),
            "num_analysts": None,
        },
        "assumptions": {
            "growth_near_pct": assumption_map.get("revenue_growth_near"),
            "growth_mid_pct": assumption_map.get("revenue_growth_mid"),
            "ebit_margin_start_pct": assumption_map.get("ebit_margin_start"),
            "ebit_margin_target_pct": assumption_map.get("ebit_margin_target"),
            "tax_rate_start_pct": assumption_map.get("tax_rate_start"),
            "tax_rate_target_pct": assumption_map.get("tax_rate_target"),
            "capex_pct": assumption_map.get("capex_pct_start"),
            "da_pct": assumption_map.get("da_pct_start"),
            "dso_start": assumption_map.get("dso_start"),
            "dio_start": assumption_map.get("dio_start"),
            "dpo_start": assumption_map.get("dpo_start"),
            "exit_multiple": assumption_map.get("exit_multiple"),
            "net_debt_mm": assumption_map.get("net_debt"),
            "non_operating_assets_mm": assumption_map.get("non_operating_assets"),
            "lease_liabilities_mm": assumption_map.get("lease_liabilities"),
            "minority_interest_mm": assumption_map.get("minority_interest"),
            "preferred_equity_mm": assumption_map.get("preferred_equity"),
            "pension_deficit_mm": assumption_map.get("pension_deficit"),
        },
        "wacc": {
            "wacc": wacc.get("effective_preview", {}).get("wacc"),
            "cost_of_equity": wacc.get("effective_preview", {}).get("cost_of_equity"),
            "equity_weight": wacc.get("effective_preview", {}).get("equity_weight"),
            "debt_weight": wacc.get("effective_preview", {}).get("debt_weight"),
            "method": (wacc.get("current_selection") or {}).get("selected_method"),
        },
        "valuation": {
            "iv_bear": scenario_map.get("bear", {}).get("iv"),
            "iv_base": scenario_map.get("base", {}).get("iv"),
            "iv_bull": scenario_map.get("bull", {}).get("iv"),
            "expected_iv": dcf.get("ev_bridge", {}).get("intrinsic_value_per_share"),
            "base_iv": dcf.get("ev_bridge", {}).get("intrinsic_value_per_share"),
            "current_price": workbench.get("current_price"),
        },
        "scenarios": scenario_map,
        "sensitivity": dcf.get("sensitivity") or {},
        "terminal": dcf.get("terminal_bridge") or {},
        "health_flags": dcf.get("health_flags") or {},
        "forecast_bridge": dcf.get("forecast_bridge") or [],
        "source_lineage": source_lineage,
        "default_resolution": workbench.get("default_resolution") or {},
        "ciq_lineage": ciq_lineage,
        "comps_detail": {
            "target": comps_target,
            "peers": comps.get("peers") or [],
            "medians": peer_medians,
        },
        "comps_analysis": _normalise_comps_analysis(comps),
        "research": research,
        "analyst_prep": _safe_analyst_prep_export_payload(ticker),
    }
    return _attach_ticker_dossier(payload, source_mode="loaded_backend_state")


def _build_snapshot_ticker_payload(ticker: str) -> tuple[dict[str, Any], int]:
    ticker = coerce_ticker(ticker)
    snapshots = list_report_snapshots(ticker, limit=1)
    if not snapshots:
        raise FileNotFoundError(f"No archived snapshot found for {ticker}")
    snapshot_id = int(snapshots[0]["id"])
    snapshot = load_report_snapshot(snapshot_id) or {}
    memo = snapshot.get("memo") or {}
    dashboard_snapshot = snapshot.get("dashboard_snapshot") or {}
    dcf = dashboard_snapshot.get("dcf_audit") or {}
    comps = dashboard_snapshot.get("comps_view") or {}
    payload = {
            "$schema_version": "1.0",
            "generated_at": _now(),
            "ticker": ticker,
            "company_name": snapshot.get("company_name") or memo.get("company_name") or ticker,
            "sector": snapshot.get("sector") or memo.get("sector"),
            "market": {
                "price": snapshot.get("current_price"),
                "analyst_target": None,
                "analyst_recommendation": snapshot.get("action"),
                "num_analysts": None,
            },
            "assumptions": {},
            "wacc": {},
            "valuation": {
                "iv_bear": memo.get("valuation", {}).get("bear"),
                "iv_base": memo.get("valuation", {}).get("base"),
                "iv_bull": memo.get("valuation", {}).get("bull"),
                "expected_iv": memo.get("valuation", {}).get("base"),
                "base_iv": snapshot.get("base_iv"),
                "current_price": snapshot.get("current_price"),
            },
            "scenarios": {
                str(row.get("scenario") or "").lower(): {
                    "probability": row.get("probability"),
                    "iv": row.get("intrinsic_value"),
                    "upside_pct": row.get("upside_pct"),
                }
                for row in dcf.get("scenario_summary") or []
                if row.get("scenario")
            },
            "sensitivity": dcf.get("sensitivity") or {},
            "terminal": dcf.get("terminal_bridge") or {},
            "health_flags": dcf.get("health_flags") or {},
            "forecast_bridge": dcf.get("forecast_bridge") or [],
            "source_lineage": {},
            "ciq_lineage": {
                "snapshot_source_file": (comps.get("source_lineage") or {}).get("source_file"),
                "snapshot_as_of_date": (comps.get("source_lineage") or {}).get("as_of_date"),
                "peer_count": (comps.get("peer_counts") or {}).get("clean"),
            },
            "comps_detail": {
                "target": (comps.get("target_vs_peers") or {}).get("target") or {},
                "peers": comps.get("peers") or [],
                "medians": (comps.get("target_vs_peers") or {}).get("peer_medians") or {},
            },
            "comps_analysis": _normalise_comps_analysis(comps),
            "snapshot": snapshot,
        }
    return (
        _attach_ticker_dossier(payload, source_mode="latest_snapshot", snapshot_id=snapshot_id),
        snapshot_id,
    )


def _build_html_context(ticker: str, source_mode: str) -> tuple[dict[str, Any], int | None]:
    ticker = coerce_ticker(ticker)
    if source_mode == "latest_snapshot":
        payload, snapshot_id = _build_snapshot_ticker_payload(ticker)
        memo = payload.get("snapshot", {}).get("memo") or {}
        publishable = build_publishable_memo_context(ticker)
        context = _apply_html_context_scalars(
            {
                "ticker": ticker,
                "company_name": payload.get("company_name"),
                "source_mode": source_mode,
                "current_price": payload.get("valuation", {}).get("current_price"),
                "base_iv": payload.get("valuation", {}).get("iv_base"),
                "expected_iv": payload.get("valuation", {}).get("expected_iv"),
                "summary": publishable.get("memo_content") or memo.get("one_liner") or memo.get("variant_thesis_prompt") or "",
                "valuation": payload.get("valuation") or {},
                "artifacts": publishable.get("artifacts") or [],
                "ticker_dossier": payload.get("ticker_dossier"),
            },
            payload,
        )
        return (context, snapshot_id)
    payload = _build_current_ticker_payload(ticker)
    publishable = build_publishable_memo_context(ticker)
    research = payload.get("research") or {}
    context = _apply_html_context_scalars(
        {
            "ticker": ticker,
            "company_name": payload.get("company_name"),
            "source_mode": source_mode,
            "current_price": payload.get("valuation", {}).get("current_price"),
            "base_iv": payload.get("valuation", {}).get("iv_base"),
            "expected_iv": payload.get("valuation", {}).get("expected_iv"),
            "summary": publishable.get("memo_content") or research.get("publishable_memo_preview") or "",
            "valuation": payload.get("valuation") or {},
            "artifacts": publishable.get("artifacts") or [],
            "ticker_dossier": payload.get("ticker_dossier"),
        },
        payload,
    )
    return (context, None)


def _ticker_bundle_dir(ticker: str, export_format: str) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return EXPORT_ROOT / "ticker" / coerce_ticker(ticker) / f"{stamp}-{export_format}-{uuid4().hex[:8]}"


def _watchlist_bundle_dir(export_format: str) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return EXPORT_ROOT / "watchlist" / f"{stamp}-{export_format}-{uuid4().hex[:8]}"


def run_ticker_export(
    *,
    ticker: str,
    export_format: str,
    source_mode: str,
    template_strategy: str | None = None,
    created_by: str = "api",
) -> dict[str, Any]:
    ticker = coerce_ticker(ticker)
    if export_format == "xlsx":
        snapshot_id = None
        if source_mode == "latest_snapshot":
            payload, snapshot_id = _build_snapshot_ticker_payload(ticker)
        else:
            payload = _build_current_ticker_payload(ticker)
        _persist_attached_ticker_dossier(payload)
        bundle_dir = _ticker_bundle_dir(ticker, export_format)
        staged = stage_power_query_workbook(ticker, payload, bundle_dir)
        return register_export_bundle(
            scope="ticker",
            export_format="xlsx",
            source_mode=source_mode,
            template_strategy=template_strategy or "power_query",
            ticker=ticker,
            created_by=created_by,
            title=f"{ticker} Excel export",
            bundle_dir=bundle_dir,
            primary_artifact_key="excel_workbook",
            artifacts=staged["artifacts"],
            snapshot_id=snapshot_id,
            metadata={"source_label": source_mode},
        )

    if export_format == "html":
        context, snapshot_id = _build_html_context(ticker, source_mode)
        _persist_attached_ticker_dossier(context)
        bundle_dir = _ticker_bundle_dir(ticker, export_format)
        staged = build_html_export_bundle(ticker, context, bundle_dir)
        return register_export_bundle(
            scope="ticker",
            export_format="html",
            source_mode=source_mode,
            template_strategy=template_strategy or "html_bundle",
            ticker=ticker,
            created_by=created_by,
            title=f"{ticker} HTML export",
            bundle_dir=bundle_dir,
            primary_artifact_key="html_report",
            artifacts=staged["artifacts"],
            snapshot_id=snapshot_id,
            metadata={"source_label": source_mode},
        )

    raise ValueError(f"Unsupported export format: {export_format}")


def _build_watchlist_html_bundle(rows: list[dict[str, Any]], bundle_dir: Path) -> dict[str, Any]:
    bundle_dir.mkdir(parents=True, exist_ok=True)
    html_path = bundle_dir / "watchlist-summary.html"
    json_path = bundle_dir / "watchlist.json"
    html_rows = "".join(
        f"<tr><td>{row.get('ticker') or ''}</td><td>{row.get('company_name') or ''}</td><td>{row.get('expected_upside_pct') or row.get('upside_base_pct') or '—'}</td></tr>"
        for row in rows[:25]
    )
    html_path.write_text(
        "<html><body><h1>Watchlist Export</h1><table><tr><th>Ticker</th><th>Company</th><th>Upside</th></tr>"
        + html_rows
        + "</table></body></html>",
        encoding="utf-8",
    )
    _json_dump(json_path, rows)
    return {
        "artifacts": [
            _artifact_row(
                artifact_key="html_report",
                artifact_role="primary",
                title="Watchlist HTML summary",
                path=html_path,
                mime_type="text/html",
                is_primary=True,
            ),
            _artifact_row(
                artifact_key="watchlist_json",
                artifact_role="sidecar_data",
                title="Watchlist rows",
                path=json_path,
                mime_type="application/json",
            ),
        ]
    }


def run_watchlist_export(
    *,
    export_format: str,
    source_mode: str,
    shortlist_size: int = 10,
    created_by: str = "api",
) -> dict[str, Any]:
    if source_mode != "saved_watchlist":
        raise ValueError("Watchlist exports currently support only saved_watchlist source mode")
    view = load_saved_watchlist(shortlist_size=shortlist_size)
    rows = list(view.get("rows") or [])
    bundle_dir = _watchlist_bundle_dir(export_format)
    bundle_dir.mkdir(parents=True, exist_ok=True)
    if export_format == "html":
        staged = _build_watchlist_html_bundle(rows, bundle_dir)
        return register_export_bundle(
            scope="batch",
            export_format="html",
            source_mode=source_mode,
            template_strategy="python_generated",
            created_by=created_by,
            title="Watchlist HTML export",
            bundle_dir=bundle_dir,
            primary_artifact_key="html_report",
            artifacts=staged["artifacts"],
            metadata={"shortlist_size": shortlist_size},
        )

    if export_format == "xlsx":
        from src.stage_02_valuation.batch_runner import export_to_excel

        workbook_path = bundle_dir / "watchlist-export.xlsx"
        export_to_excel(rows, workbook_path)
        staged = {
            "artifacts": [
                _artifact_row(
                    artifact_key="excel_workbook",
                    artifact_role="primary",
                    title="Watchlist workbook",
                    path=workbook_path,
                    mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    is_primary=True,
                )
            ]
        }
        return register_export_bundle(
            scope="batch",
            export_format="xlsx",
            source_mode=source_mode,
            template_strategy="python_generated",
            created_by=created_by,
            title="Watchlist Excel export",
            bundle_dir=bundle_dir,
            primary_artifact_key="excel_workbook",
            artifacts=staged["artifacts"],
            metadata={"shortlist_size": shortlist_size},
        )

    raise ValueError(f"Unsupported export format: {export_format}")


def list_saved_exports(*, ticker: str | None = None, scope: str | None = None, limit: int = 25) -> list[dict[str, Any]]:
    return list_exports(ticker=ticker, scope=scope, limit=limit)


def load_saved_export(export_id: str) -> dict[str, Any] | None:
    return load_export(export_id)


def resolve_export_download_path(export_id: str, artifact_key: str | None = None) -> Path:
    return resolve_export_artifact_path(export_id, artifact_key=artifact_key)
