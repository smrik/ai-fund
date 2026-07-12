"""Deterministic professional-model artifact review and PM controls.

This module is the business/state layer behind the FastAPI transport.  It
discovers only run-bound artifacts under the configured output root, verifies
their identity, exposes bounded workbook slices, and persists immutable PM
review events.  It intentionally contains no valuation formulas.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from hashlib import sha256
from io import BytesIO
import json
import math
import os
from pathlib import Path, PureWindowsPath
import re
from typing import Any, Iterable, Mapping, Sequence
from uuid import uuid4

from config import DB_PATH, OUTPUT_DIR, ROOT_DIR
from db.schema import create_tables, get_connection
from db.loader import (
    PROFESSIONAL_MODEL_REVIEW_FINGERPRINT_VERSION as REVIEW_FINGERPRINT_VERSION,
)

from src.stage_04_pipeline.professional_model_review_contract import (
    SEMANTIC_QA_CHECK_IDS,
    ProfessionalModelReviewContractError,
    build_decision_semantic_qa_verification,
    enrich_pm_driver_requirement,
    normalize_preview_review_context,
)

PROFESSIONAL_MODEL_ROOT = OUTPUT_DIR / "professional_models"
PROFESSIONAL_MODEL_REBUILD_ROOT = OUTPUT_DIR / "professional_model_rebuilds"
WORKBOOK_NAME_TEMPLATE = "{ticker}_professional_model_v2.xlsx"
REVIEW_EVIDENCE_NAME = "review_evidence.json"
MANIFEST_NAME = "manifest.json"
QA_REPORT_NAME = "qa_report.json"
FINAL_SIGNOFF_KEY = "final_signoff"
FINAL_SIGNOFF_SCOPE = "workbook"

MIN_FULL_LIFECYCLE_MODEL_RUN_ID = 4
QA_SCHEMA_VERSION = "professional_model_qa_v2"
_READABLE_QA_SCHEMA_VERSIONS = {"professional_model_qa_v1", QA_SCHEMA_VERSION}
MAX_SHEET_ROWS = 200
MAX_SHEET_COLUMNS = 50
MAX_SHEET_CELLS = 5_000
MAX_CELL_TEXT = 8_000
MAX_COMMENT_TEXT = 4_000
MAX_SHEET_TEXT_BYTES = 4_000_000
MAX_MANIFEST_BYTES = 16 * 1024 * 1024
MAX_QA_REPORT_BYTES = 4 * 1024 * 1024
MAX_REVIEW_EVIDENCE_BYTES = 4 * 1024 * 1024

_TICKER_RE = re.compile(r"^[A-Z][A-Z0-9.-]{0,14}$")
_SHA256_RE = re.compile(r"^[0-9a-f]{64}$")
_WINDOWS_ABSOLUTE_RE = re.compile(r"^[A-Za-z]:[\\/]")
_TIMESTAMPED_VALUATION_RE = re.compile(
    r"(?:^|[-_])\d{8}T\d{6}Z-valuation\.json$",
    re.IGNORECASE,
)
_ALLOWED_REPORTED_STATUSES = {
    "BLOCKED",
    "AVAILABLE",
    "READY",
    "DECISION_READY",
    "PARTIAL",
    "FULL",
}
_READY_REPORTED_STATUSES = {"AVAILABLE", "READY", "DECISION_READY", "FULL"}
_ALLOWED_CHECK_STATUSES = {"PASS", "BLOCKED", "FAIL", "REVIEW", "NEEDS_PM_REVIEW", "UNVERIFIED", "PARTIAL"}

_REQUIRED_RENDERER_QA_CHECK_IDS = {
    "model_readiness",
    "source_preflight",
    "scenario_completeness",
    "forecast_completeness",
    "balance_sheet",
    "cash_flow_tie",
    "valuation_bridge",
    "dcf_decision_gate",
    "valuation_input_gate",
    "scenario_selector",
    "calculation_verification",
    "pm_driver_approvals",
    "optional_modules",
}
_REQUIRED_QA_CHECK_IDS = _REQUIRED_RENDERER_QA_CHECK_IDS | set(SEMANTIC_QA_CHECK_IDS)
_REQUIRED_PARITY_CHECK_IDS = _REQUIRED_RENDERER_QA_CHECK_IDS - {"model_readiness"}
_REQUIRED_QA_CHECK_CELLS = {
    "model_readiness": ("Checks", "C5"),
    "source_preflight": ("Checks", "C6"),
    "scenario_completeness": ("Checks", "C7"),
    "forecast_completeness": ("Checks", "C8"),
    "balance_sheet": ("Checks", "C9"),
    "cash_flow_tie": ("Checks", "C10"),
    "valuation_bridge": ("Checks", "C11"),
    "dcf_decision_gate": ("Checks", "C12"),
    "valuation_input_gate": ("Checks", "C13"),
    "scenario_selector": ("Checks", "C14"),
    "calculation_verification": ("Checks", "C15"),
    "pm_driver_approvals": ("Checks", "C16"),
    "optional_modules": ("Checks", "C17"),
}
_VERIFIED_RECALCULATION_REASON_CODES = {
    "recalculated_model_available",
    "recalculated_model_blocked",
    "recalculated_model_partial",
    "recalculated_model_ready",
}

_PACKAGE_WARNING_REQUIREMENTS: dict[str, tuple[str, str, str]] = {
    "segments_unavailable": (
        "segments",
        "finance_data",
        "Provide source-backed segment history and an approved segment driver set.",
    ),
    "consensus_unavailable": (
        "consensus",
        "data",
        "Provide a frozen, as-of-matched consensus snapshot.",
    ),
    "fcfe_unavailable": (
        "fcfe",
        "finance",
        "Provide the integrated debt and after-tax interest schedule required for FCFE.",
    ),
    "sotp_unavailable": (
        "sotp",
        "finance_data",
        "Provide normalized segment evidence before enabling SOTP.",
    ),
}


class ProfessionalModelError(RuntimeError):
    """Base error for professional-model review operations."""


class ProfessionalModelNotFoundError(FileNotFoundError, ProfessionalModelError):
    """Raised when no complete root artifact tuple exists."""


class ProfessionalModelConflictError(ProfessionalModelError):
    """Raised when current state contradicts or invalidates an action."""


class ProfessionalModelValidationError(ValueError, ProfessionalModelError):
    """Raised for invalid tickers, sheet bounds, or review requests."""


@dataclass(frozen=True)
class ProfessionalModelArtifacts:
    ticker: str
    model_run_id: int
    artifact_dir: Path
    workbook_path: Path
    manifest_path: Path
    qa_report_path: Path
    manifest: dict[str, Any]
    review_evidence: dict[str, Any] | None
    qa_report: dict[str, Any]
    workbook_hash: str
    qa_report_hash: str
    review_evidence_file_hash: str | None
    workbook_bytes: int
    issues: tuple[dict[str, str], ...]
    forecast_periods: tuple[str, ...] = ()


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _validate_ticker(value: str) -> str:
    ticker = str(value or "").strip().upper()
    if not _TICKER_RE.fullmatch(ticker) or ".." in ticker:
        raise ProfessionalModelValidationError("invalid ticker")
    return ticker


def _clean_text(value: Any, field_name: str, *, max_length: int = 4_000) -> str:
    text = str(value or "").strip()
    if not text:
        raise ProfessionalModelValidationError(f"{field_name} is required")
    if len(text) > max_length:
        raise ProfessionalModelValidationError(
            f"{field_name} must be at most {max_length} characters"
        )
    return text


def _canonical_json(value: Any) -> str:
    def _default(item: Any) -> str:
        if isinstance(item, (date, datetime)):
            return item.isoformat()
        raise TypeError(f"cannot serialize {type(item)!r}")

    return json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        allow_nan=False,
        default=_default,
    )


def _canonical_hash(value: Any) -> str:
    return sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _load_json_object(path: Path, label: str) -> dict[str, Any]:
    payload, _ = _load_json_snapshot(path, label)
    return payload


def _load_json_snapshot(path: Path, label: str) -> tuple[dict[str, Any], str]:
    """Parse and hash one immutable byte snapshot of a JSON artifact."""
    try:
        def _reject_constant(value: str) -> None:
            raise ValueError(f"non-finite JSON constant: {value}")

        raw = path.read_bytes()
        payload = json.loads(
            raw.decode("utf-8"),
            parse_constant=_reject_constant,
        )
    except (OSError, UnicodeDecodeError, ValueError) as exc:
        raise ProfessionalModelConflictError(f"invalid {label}") from exc
    if not isinstance(payload, dict):
        raise ProfessionalModelConflictError(f"{label} must be a JSON object")
    return payload, sha256(raw).hexdigest()


def _is_under(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
    except ValueError:
        return False
    return True


def _trusted_child(root: Path, *parts: str) -> Path:
    resolved_root = root.resolve()
    candidate = resolved_root.joinpath(*parts).resolve()
    if not _is_under(candidate, resolved_root):
        raise ProfessionalModelValidationError("resolved path escapes trusted root")
    return candidate


def _artifact_issue(code: str, detail: str) -> dict[str, str]:
    return {
        "code": code,
        "detail": _redact_workbook_text(str(detail), limit=500),
    }


def _safe_hash(value: Any) -> str | None:
    candidate = str(value or "").strip().lower()
    return candidate if _SHA256_RE.fullmatch(candidate) else None


def _read_workbook_identity(snapshot: bytes) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(snapshot), read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        cover = workbook["Cover"] if "Cover" in workbook.sheetnames else None
        scenarios = (
            workbook["Scenarios"]
            if "Scenarios" in workbook.sheetnames
            else None
        )
        forecast_periods = tuple(
            str(scenarios.cell(4, column).value or "").strip()
            for column in range(6, 11)
        ) if scenarios is not None else ()
        return {
            "sheet_names": sheet_names,
            "cover_ticker": cover["B3"].value if cover is not None else None,
            "cover_source_hash": cover["B9"].value if cover is not None else None,
            "cover_run_id": cover["B10"].value if cover is not None else None,
            "calculation_marker": cover["B12"].value if cover is not None else None,
            "reported_status": cover["B11"].value if cover is not None else None,
            "forecast_periods": forecast_periods,
        }
    finally:
        workbook.close()




def _validate_review_evidence(
    payload: Mapping[str, Any],
    *,
    ticker: str,
    model_run_id: int,
    manifest: Mapping[str, Any],
    workbook_hash: str,
) -> list[str]:
    issues: list[str] = []
    if payload.get("schema_version") != "1.0.0":
        issues.append("review_evidence_schema_unknown")
    if str(payload.get("ticker") or "").upper() != ticker:
        issues.append("review_evidence_ticker_mismatch")
    try:
        evidence_run_id = int(payload.get("model_run_id"))
    except (TypeError, ValueError):
        evidence_run_id = None
    if evidence_run_id != model_run_id:
        issues.append("review_evidence_run_mismatch")

    identity = payload.get("artifact_identity")
    identity = identity if isinstance(identity, Mapping) else {}
    expected_identity = {
        "source_sha256": _safe_hash(manifest.get("source_hash")),
        "model_input_sha256": _safe_hash(manifest.get("model_input_hash")),
        "workbook_sha256": workbook_hash,
        "result_sha256": _safe_hash(manifest.get("result_hash")),
    }
    for field_name, expected in expected_identity.items():
        if _safe_hash(identity.get(field_name)) != expected:
            issues.append(f"review_evidence_{field_name}_mismatch")

    events = payload.get("consumed_approval_events")
    if not isinstance(events, list) or not events:
        issues.append("review_evidence_approval_events_missing")
        events = []
    identities: set[tuple[str, str]] = set()
    event_ids: set[int] = set()
    event_requirement_hashes: dict[tuple[str, str], str] = {}
    for item in events:
        if not isinstance(item, Mapping):
            issues.append("review_evidence_event_not_object")
            continue
        scope = str(item.get("scope") or "")
        approval_key = str(item.get("approval_key") or "")
        try:
            event_id = int(item.get("event_id"))
        except (TypeError, ValueError):
            event_id = 0
        values = item.get("reviewed_values")
        scope_parts = scope.split(":")
        valid_identity = bool(
            len(scope_parts) == 3
            and scope_parts[0] == "scenario_driver"
            and scope_parts[1] in {"Base", "Upside", "Downside"}
            and approval_key
            == f"pmq:{scope_parts[1]}:{scope_parts[2]}"
        )
        if (
            not valid_identity
            or event_id <= 0
            or _safe_hash(item.get("reviewed_value_fingerprint")) is None
            or not isinstance(values, list)
            or len(values) != 5
            or any(
                isinstance(value, bool)
                or not isinstance(value, (int, float))
                or not math.isfinite(float(value))
                for value in (values if isinstance(values, list) else [])
            )
        ):
            issues.append("review_evidence_event_invalid")
            continue
        approval_identity = item.get("approval_artifact_identity")
        approval_identity = (
            approval_identity if isinstance(approval_identity, Mapping) else {}
        )
        try:
            approval_run_id = int(approval_identity.get("model_run_id"))
        except (TypeError, ValueError):
            approval_run_id = 0
        approval_hash_fields = (
            "source_sha256",
            "model_input_sha256",
            "result_sha256",
            "workbook_sha256",
        )
        if approval_run_id != model_run_id or any(
            _safe_hash(approval_identity.get(field)) is None
            for field in approval_hash_fields
        ):
            issues.append("review_evidence_approval_artifact_identity_invalid")
            continue
        if _safe_hash(approval_identity.get("source_sha256")) != expected_identity["source_sha256"]:
            issues.append("review_evidence_approval_source_identity_mismatch")
        requirement_hash = _safe_hash(item.get("requirement_hash"))
        requirement_contract = item.get("requirement_contract")
        if not isinstance(requirement_contract, Mapping):
            issues.append("review_evidence_requirement_contract_missing")
            continue
        requirement_content = {
            key: value
            for key, value in requirement_contract.items()
            if key != "requirement_hash"
        }
        contract_artifact_identity = requirement_contract.get("artifact_identity")
        if (
            requirement_hash is None
            or requirement_contract.get("requirement_hash") != requirement_hash
            or _canonical_hash(requirement_content) != requirement_hash
            or requirement_contract.get("scope") != scope
            or requirement_contract.get("approval_key") != approval_key
            or requirement_contract.get("approvable") is not True
            or not isinstance(contract_artifact_identity, Mapping)
            or any(
                contract_artifact_identity.get(field) != approval_identity.get(field)
                for field in approval_hash_fields
            )
        ):
            issues.append("review_evidence_requirement_contract_invalid")
            continue
        try:
            review_context = normalize_preview_review_context(item.get("review_context"))
        except ProfessionalModelReviewContractError:
            issues.append("review_evidence_review_context_invalid")
            continue
        if (
            not review_context.get("source_ref")
            or not review_context.get("method")
            or not review_context.get("as_of")
            or not review_context.get("evidence_locator")
        ):
            issues.append("review_evidence_review_context_incomplete")
            continue
        identity_key = (scope, approval_key)
        expected_fingerprint = _canonical_hash(
            {
                "approval_key": approval_key,
                "fingerprint_version": REVIEW_FINGERPRINT_VERSION,
                "reviewed_values": values,
                "scope": scope,
                "requirement_hash": requirement_hash,
                "review_context": review_context,
            }
        )
        if item.get("reviewed_value_fingerprint") != expected_fingerprint:
            issues.append("review_evidence_event_fingerprint_mismatch")
            continue
        if identity_key in identities or event_id in event_ids:
            issues.append("review_evidence_event_duplicate")
        event_requirement_hashes[identity_key] = requirement_hash
        identities.add(identity_key)
        event_ids.add(event_id)
    raw_inventory = payload.get("required_approval_identities")
    normalized_inventory: list[dict[str, str]] = []
    inventory_identities: set[tuple[str, str]] = set()
    if not isinstance(raw_inventory, list) or not raw_inventory:
        issues.append("review_evidence_required_inventory_missing")
    else:
        for item in raw_inventory:
            if not isinstance(item, Mapping):
                issues.append("review_evidence_required_inventory_invalid")
                continue
            requirement_hash = _safe_hash(item.get("requirement_hash"))
            scope = str(item.get("scope") or "")
            approval_key = str(item.get("approval_key") or "")
            identity = (scope, approval_key)
            if (
                not scope
                or not approval_key
                or requirement_hash is None
                or identity in inventory_identities
            ):
                issues.append("review_evidence_required_inventory_invalid")
                continue
            inventory_identities.add(identity)
            normalized_inventory.append(
                {
                    "scope": scope,
                    "approval_key": approval_key,
                    "requirement_hash": requirement_hash,
                }
            )
    normalized_inventory.sort(
        key=lambda item: (item["scope"], item["approval_key"])
    )
    if any(
        event_requirement_hashes.get(identity) != item["requirement_hash"]
        for identity, item in zip(sorted(inventory_identities), normalized_inventory)
    ):
        issues.append("review_evidence_requirement_inventory_mismatch")
    if identities != inventory_identities:
        issues.append("review_evidence_approval_inventory_incomplete")
    if _safe_hash(payload.get("required_approval_inventory_hash")) != _canonical_hash(
        normalized_inventory
    ):
        issues.append("review_evidence_required_inventory_hash_mismatch")
    try:
        event_count = int(payload.get("approval_event_count"))
    except (TypeError, ValueError):
        event_count = -1
    if event_count != len(events):
        issues.append("review_evidence_event_count_mismatch")
    if _safe_hash(payload.get("approval_set_hash")) != _canonical_hash(events):
        issues.append("review_evidence_approval_set_hash_mismatch")
    exact_content = {
        key: value for key, value in payload.items() if key != "review_evidence_hash"
    }
    if _safe_hash(payload.get("review_evidence_hash")) != _canonical_hash(exact_content):
        issues.append("review_evidence_hash_mismatch")
    return sorted(set(issues))
def discover_professional_model_artifacts(
    ticker: str,
    *,
    artifact_root: str | Path | None = None,
) -> ProfessionalModelArtifacts:
    """Discover and cross-check the newest numeric model run.

    Discovery never falls back to an older run when the newest directory is
    incomplete or contradictory.
    """
    ticker = _validate_ticker(ticker)
    root = Path(artifact_root or PROFESSIONAL_MODEL_ROOT).resolve()
    ticker_dir = _trusted_child(root, ticker)
    if not ticker_dir.is_dir():
        raise ProfessionalModelNotFoundError(
            f"professional model not found for ticker={ticker}"
        )
    run_dirs = sorted(
        (
            item
            for item in ticker_dir.iterdir()
            if item.is_dir() and item.name.isdigit() and int(item.name) > 0
        ),
        key=lambda item: int(item.name),
        reverse=True,
    )
    if not run_dirs:
        raise ProfessionalModelNotFoundError(
            f"professional model run not found for ticker={ticker}"
        )

    artifact_dir = run_dirs[0].resolve()
    if not _is_under(artifact_dir, root):
        raise ProfessionalModelConflictError("artifact directory escapes trusted root")
    model_run_id = int(artifact_dir.name)
    workbook_path = _trusted_child(
        artifact_dir,
        WORKBOOK_NAME_TEMPLATE.format(ticker=ticker),
    )
    manifest_path = _trusted_child(artifact_dir, MANIFEST_NAME)
    qa_report_path = _trusted_child(artifact_dir, QA_REPORT_NAME)
    missing = [
        name
        for name, path in (
            (workbook_path.name, workbook_path),
            (MANIFEST_NAME, manifest_path),
            (QA_REPORT_NAME, qa_report_path),
        )
        if not path.is_file()
    ]
    if missing:
        raise ProfessionalModelNotFoundError(
            "newest professional model run is incomplete: " + ", ".join(missing)
        )

    if manifest_path.stat().st_size > MAX_MANIFEST_BYTES:
        raise ProfessionalModelConflictError(
            "professional model manifest exceeds the supported size"
        )
    if qa_report_path.stat().st_size > MAX_QA_REPORT_BYTES:
        raise ProfessionalModelConflictError(
            "professional model QA report exceeds the supported size"
        )

    manifest, _ = _load_json_snapshot(manifest_path, "professional model manifest")
    qa_report, qa_report_hash = _load_json_snapshot(
        qa_report_path, "professional model QA report"
    )
    try:
        workbook_snapshot = workbook_path.read_bytes()
    except OSError as exc:
        raise ProfessionalModelConflictError("professional model workbook is unreadable") from exc
    workbook_hash = sha256(workbook_snapshot).hexdigest()
    workbook_bytes = len(workbook_snapshot)
    issues: list[dict[str, str]] = []

    if str(manifest.get("ticker") or "").upper() != ticker:
        issues.append(_artifact_issue("manifest_ticker_mismatch", "Manifest ticker does not match the requested ticker."))
    if str(qa_report.get("ticker") or "").upper() != ticker:
        issues.append(_artifact_issue("qa_ticker_mismatch", "QA ticker does not match the requested ticker."))
    if qa_report.get("schema_version") not in _READABLE_QA_SCHEMA_VERSIONS:
        issues.append(
            _artifact_issue("qa_schema_unknown", "QA schema version is not supported.")
        )
    try:
        qa_run_id = int(qa_report.get("source_run_id"))
    except (TypeError, ValueError):
        qa_run_id = None
    if qa_run_id != model_run_id:
        issues.append(_artifact_issue("run_id_mismatch", "QA source run does not match the artifact directory."))

    manifest_source_hash = _safe_hash(manifest.get("source_hash"))
    qa_source_hash = _safe_hash(qa_report.get("source_sha256"))
    if manifest_source_hash is None or qa_source_hash is None:
        issues.append(_artifact_issue("source_hash_unverified", "A required source SHA-256 is blank or invalid."))
    elif manifest_source_hash != qa_source_hash:
        issues.append(_artifact_issue("source_hash_mismatch", "Manifest and QA source hashes differ."))

    for field in ("model_input_hash", "result_hash", "manifest_hash"):
        if _safe_hash(manifest.get(field)) is None:
            issues.append(_artifact_issue(f"{field}_unverified", f"Manifest {field} is blank or invalid."))

    qa_workbook_hash = _safe_hash(qa_report.get("workbook_sha256"))
    if qa_workbook_hash is None:
        issues.append(_artifact_issue("workbook_hash_unverified", "QA workbook SHA-256 is blank or invalid."))
    elif qa_workbook_hash != workbook_hash:
        issues.append(_artifact_issue("workbook_hash_mismatch", "QA workbook SHA-256 does not match the exact file."))
    try:
        qa_workbook_bytes = int(qa_report.get("workbook_bytes"))
    except (TypeError, ValueError):
        qa_workbook_bytes = None
    if qa_workbook_bytes != workbook_bytes:
        issues.append(_artifact_issue("workbook_size_mismatch", "QA workbook size does not match the exact file."))

    manifest_blockers = sorted(str(item) for item in (manifest.get("blockers") or []))
    qa_blockers = sorted(str(item) for item in (qa_report.get("remaining_blockers") or []))
    if manifest_blockers != qa_blockers:
        issues.append(_artifact_issue("blocker_set_mismatch", "Manifest and QA blocker sets differ."))
    manifest_warnings = sorted(str(item) for item in (manifest.get("warnings") or []))
    qa_warnings = sorted(str(item) for item in (qa_report.get("warnings") or []))
    if manifest_warnings != qa_warnings:
        issues.append(_artifact_issue("warning_set_mismatch", "Manifest and QA warning sets differ."))

    raw_manifest_hash = _safe_hash(manifest.get("manifest_hash"))
    raw_manifest_content = {
        key: value for key, value in manifest.items() if key != "manifest_hash"
    }
    if (
        raw_manifest_hash is not None
        and _canonical_hash(raw_manifest_content) != raw_manifest_hash
    ):
        issues.append(
            _artifact_issue(
                "manifest_hash_mismatch",
                "Manifest SHA-256 does not match its exact persisted canonical content.",
            )
        )
    try:
        from src.contracts.professional_financial_model import WorkbookManifest

        # Validate structure independently from the persisted self-hash. New
        # optional contract fields may default to null after an older manifest
        # was signed; the raw canonical hash above remains the authority for the
        # exact persisted document.
        structural_manifest = dict(manifest)
        structural_manifest["manifest_hash"] = None
        WorkbookManifest.model_validate(structural_manifest)
    except Exception as exc:
        issues.append(_artifact_issue("manifest_contract_invalid", str(exc)))

    try:
        workbook_identity = _read_workbook_identity(workbook_snapshot)
    except Exception as exc:
        workbook_identity = {}
        issues.append(_artifact_issue("workbook_unreadable", str(exc)))
    expected_sheet_order = list(manifest.get("sheet_order") or [])
    qa_sheet_order = list((qa_report.get("integrity") or {}).get("sheet_order") or [])
    workbook_sheet_order = list(workbook_identity.get("sheet_names") or [])
    if not expected_sheet_order or expected_sheet_order != qa_sheet_order or expected_sheet_order != workbook_sheet_order:
        issues.append(_artifact_issue("sheet_order_mismatch", "Workbook, manifest, and QA sheet orders must match exactly."))
    forecast_periods = tuple(workbook_identity.get("forecast_periods") or ())
    if (
        len(forecast_periods) != 5
        or any(not item for item in forecast_periods)
        or len(set(forecast_periods)) != 5
    ):
        issues.append(_artifact_issue("forecast_periods_unverified", "The exact five-period review axis is unavailable."))
    try:
        cover_run_id = int(workbook_identity.get("cover_run_id"))
    except (TypeError, ValueError):
        cover_run_id = None
    if cover_run_id != model_run_id:
        issues.append(_artifact_issue("workbook_run_id_mismatch", "Cover run ID does not match the artifact directory."))
    if str(workbook_identity.get("cover_ticker") or "").upper() != ticker:
        issues.append(
            _artifact_issue("workbook_ticker_mismatch", "Cover ticker does not match.")
        )
    if _safe_hash(workbook_identity.get("cover_source_hash")) != manifest_source_hash:
        issues.append(
            _artifact_issue("workbook_source_hash_mismatch", "Cover source SHA-256 does not match.")
        )
    qa_calculation = qa_report.get("calculation")
    qa_calculation = qa_calculation if isinstance(qa_calculation, Mapping) else {}
    if str(workbook_identity.get("calculation_marker") or "").upper() != str(
        qa_calculation.get("calculation_marker") or ""
    ).upper():
        issues.append(_artifact_issue("workbook_calculation_marker_mismatch", "Cover and QA calculation markers differ."))
    if str(workbook_identity.get("reported_status") or "").upper() != str(
        qa_report.get("model_status") or ""
    ).upper():
        issues.append(_artifact_issue("workbook_status_mismatch", "Cover and QA model statuses differ."))

    qa_workbook_field = str(qa_report.get("workbook") or "").strip()
    if qa_workbook_field:
        qa_workbook_path = Path(qa_workbook_field)
        if not qa_workbook_path.is_absolute():
            qa_workbook_path = ROOT_DIR / qa_workbook_path
        try:
            qa_workbook_resolved = qa_workbook_path.resolve()
        except OSError:
            qa_workbook_resolved = qa_workbook_path
        if qa_workbook_resolved != workbook_path.resolve():
            issues.append(_artifact_issue("qa_workbook_identity_mismatch", "QA workbook identity does not name the exact root artifact."))
    review_evidence_path = _trusted_child(artifact_dir, REVIEW_EVIDENCE_NAME)
    review_evidence: dict[str, Any] | None = None
    review_evidence_file_hash: str | None = None
    if review_evidence_path.is_file():
        if review_evidence_path.stat().st_size > MAX_REVIEW_EVIDENCE_BYTES:
            raise ProfessionalModelConflictError(
                "professional model review evidence exceeds the supported size"
            )
        try:
            review_evidence, review_evidence_file_hash = _load_json_snapshot(
                review_evidence_path,
                "professional model review evidence",
            )
        except ProfessionalModelError as exc:
            issues.append(
                _artifact_issue("review_evidence_invalid", str(exc))
            )
        if review_evidence is not None:
            evidence_issue_codes = _validate_review_evidence(
                review_evidence,
                ticker=ticker,
                model_run_id=model_run_id,
                manifest=manifest,
                workbook_hash=workbook_hash,
            )
            for code in evidence_issue_codes:
                issues.append(
                    _artifact_issue(code, "Review approval evidence is invalid.")
                )
            if evidence_issue_codes:
                review_evidence = None
    has_pm_review_blockers = any(
        blocker.startswith("pm_approval_required:")
        for blocker in manifest_blockers
    )
    if not has_pm_review_blockers and review_evidence is None:
        issues.append(
            _artifact_issue(
                "review_evidence_missing",
                "A model without PM driver blockers requires a signed approval-evidence sidecar.",
            )
        )

    return ProfessionalModelArtifacts(
        ticker=ticker,
        model_run_id=model_run_id,
        artifact_dir=artifact_dir,
        workbook_path=workbook_path,
        manifest_path=manifest_path,
        qa_report_path=qa_report_path,
        manifest=manifest,
        review_evidence=review_evidence,
        qa_report=qa_report,
        workbook_hash=workbook_hash,
        qa_report_hash=qa_report_hash,
        review_evidence_file_hash=review_evidence_file_hash,
        workbook_bytes=workbook_bytes,
        issues=tuple(issues),
        forecast_periods=forecast_periods,
    )
def _qa_contract_payload() -> dict[str, Any]:
    return {
        "schema_version": QA_SCHEMA_VERSION,
        "required_check_ids": sorted(_REQUIRED_QA_CHECK_IDS),
        "renderer_check_ids": sorted(_REQUIRED_RENDERER_QA_CHECK_IDS),
        "semantic_check_ids": sorted(SEMANTIC_QA_CHECK_IDS),
        "parity_check_ids": sorted(_REQUIRED_PARITY_CHECK_IDS),
        "required_check_cells": {
            check_id: {"sheet": sheet, "coordinate": coordinate}
            for check_id, (sheet, coordinate) in sorted(_REQUIRED_QA_CHECK_CELLS.items())
        },
        "decision_positive_statuses": ["PASS"],
        "unknown_missing_extra_duplicate_policy": "fail_closed",
    }




def _cached_check_cell_reasons(
    artifacts: ProfessionalModelArtifacts,
    qa_check_by_id: Mapping[str, Mapping[str, Any]],
) -> list[str]:
    reasons: list[str] = []
    raw_check_cells = artifacts.manifest.get("check_cells")
    check_cells: dict[str, tuple[str, str]] = {}
    if not isinstance(raw_check_cells, list):
        return ["manifest_check_cells_missing"]
    for item in raw_check_cells:
        if not isinstance(item, Mapping):
            reasons.append("manifest_check_cell_not_object")
            continue
        check_id = str(item.get("check_id") or "").strip()
        sheet = str(item.get("sheet") or "").strip()
        cell = str(item.get("cell") or "").strip().upper()
        if not check_id or check_id in check_cells:
            reasons.append("manifest_check_cell_identity_invalid")
            continue
        check_cells[check_id] = (sheet, cell)
        if item.get("contract_version") != "1.0.0":
            reasons.append("manifest_check_cell_contract_unknown")
    if set(check_cells) != set(_REQUIRED_QA_CHECK_CELLS):
        reasons.append("manifest_check_cell_set_invalid")
    for check_id, expected_location in _REQUIRED_QA_CHECK_CELLS.items():
        if check_cells.get(check_id) != expected_location:
            reasons.append(f"manifest_check_cell_location_invalid:{check_id}")

    try:
        workbook_snapshot = artifacts.workbook_path.read_bytes()
    except OSError:
        return sorted(set([*reasons, "workbook_check_snapshot_unreadable"]))
    if sha256(workbook_snapshot).hexdigest() != artifacts.workbook_hash:
        return sorted(set([*reasons, "workbook_changed_during_check_verification"]))

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(workbook_snapshot), read_only=True, data_only=True)
    try:
        for check_id, (sheet_name, coordinate) in _REQUIRED_QA_CHECK_CELLS.items():
            qa_check = qa_check_by_id.get(check_id)
            if qa_check is None or check_cells.get(check_id) != (sheet_name, coordinate):
                continue
            if sheet_name not in workbook.sheetnames:
                reasons.append(f"workbook_check_sheet_missing:{check_id}")
                continue
            sheet = workbook[sheet_name]
            status_cell = sheet[coordinate]
            cached_status = str(status_cell.value or "").strip().upper()
            cached_difference = sheet.cell(status_cell.row, 4).value
            cached_tolerance = sheet.cell(status_cell.row, 5).value
            if cached_status != str(qa_check.get("status") or "").strip().upper():
                reasons.append(f"workbook_check_status_mismatch:{check_id}")
            if cached_difference != qa_check.get("difference_or_count"):
                reasons.append(f"workbook_check_value_mismatch:{check_id}")
            if cached_tolerance != qa_check.get("tolerance_or_expected"):
                reasons.append(f"workbook_check_tolerance_mismatch:{check_id}")
    except Exception:
        reasons.append("workbook_check_cells_unreadable")
    finally:
        workbook.close()
    return sorted(set(reasons))


def _calculation_verification(
    artifacts: ProfessionalModelArtifacts,
) -> dict[str, Any]:
    qa = artifacts.qa_report
    calculation = qa.get("calculation") if isinstance(qa.get("calculation"), dict) else {}
    integrity = qa.get("integrity") if isinstance(qa.get("integrity"), dict) else {}
    recalculation = artifacts.manifest.get("recalculation_state")
    recalculation = recalculation if isinstance(recalculation, dict) else {}
    reasons: list[str] = []

    if qa.get("schema_version") != QA_SCHEMA_VERSION:
        reasons.append("qa_schema_not_full_lifecycle")
    marker = str(calculation.get("calculation_marker") or "").upper()
    if marker != "CALCULATED":
        reasons.append("calculation_marker_not_calculated")
    if not str(calculation.get("cache_engine") or "").strip():
        reasons.append("cache_engine_missing")
    if not str(calculation.get("cache_merge") or "").strip():
        reasons.append("cache_merge_evidence_missing")
    formula_errors = integrity.get("formula_errors")
    if formula_errors != []:
        reasons.append("formula_errors_unknown_or_present")

    def _nonnegative_int(value: Any) -> int | None:
        if isinstance(value, bool):
            return None
        try:
            number = int(value)
        except (TypeError, ValueError):
            return None
        return number if number >= 0 else None

    formula_count = _nonnegative_int(integrity.get("formula_count"))
    nonblank = _nonnegative_int(integrity.get("formula_results_nonblank"))
    intentional = _nonnegative_int(
        integrity.get("formula_results_blank_intentional_gates")
    )
    if formula_count is None or formula_count <= 0:
        reasons.append("formula_count_unknown")
    if nonblank is None or intentional is None or formula_count is None:
        reasons.append("formula_cache_coverage_unknown")
    elif nonblank + intentional != formula_count:
        reasons.append("formula_cache_coverage_mismatch")

    recalculation_status = str(recalculation.get("status") or "").lower()
    if recalculation_status not in {"available", "blocking"}:
        reasons.append("recalculation_status_unknown")
    recalculation_reason = str(recalculation.get("reason_code") or "").strip()
    if recalculation_reason not in _VERIFIED_RECALCULATION_REASON_CODES:
        reasons.append("recalculation_reason_not_verified")
    blockers = {
        str(item).strip().lower()
        for item in artifacts.manifest.get("blockers") or []
    }
    if {"recalculation_not_run", "recalculation_not_verified"} & blockers:
        reasons.append("recalculation_blocker_present")

    raw_checks = qa.get("checks")
    qa_status_by_id: dict[str, str] = {}
    qa_check_by_id: dict[str, Mapping[str, Any]] = {}
    if not isinstance(raw_checks, list):
        reasons.append("qa_checks_missing")
    else:
        for item in raw_checks:
            if not isinstance(item, Mapping):
                reasons.append("qa_check_not_object")
                continue
            check_id = str(item.get("check_id") or "").strip()
            status = str(item.get("status") or "").strip().upper()
            if not check_id or check_id in qa_status_by_id:
                reasons.append("qa_check_identity_invalid")
                continue
            qa_status_by_id[check_id] = status
            qa_check_by_id[check_id] = item
            if status not in _ALLOWED_CHECK_STATUSES:
                reasons.append("qa_check_status_unknown")
        if set(qa_status_by_id) != _REQUIRED_QA_CHECK_IDS:
            reasons.append("qa_check_set_incomplete")

    reasons.extend(_cached_check_cell_reasons(artifacts, qa_check_by_id))
    parity_status_by_id: dict[str, str] = {}
    parity_by_id: dict[str, Mapping[str, Any]] = {}
    for item in artifacts.manifest.get("parity_results") or []:
        if not isinstance(item, Mapping):
            reasons.append("manifest_parity_not_object")
            continue
        check_id = str(item.get("check_id") or "").strip()
        status = str(item.get("status") or "").strip().upper()
        if not check_id or check_id in parity_status_by_id:
            reasons.append("manifest_parity_identity_invalid")
            continue
        parity_status_by_id[check_id] = status
        parity_by_id[check_id] = item
    def _parity_values_match(check_id: str) -> bool:
        qa_check = qa_check_by_id[check_id]
        manifest_check = parity_by_id[check_id]
        if check_id in {"scenario_completeness", "forecast_completeness"}:
            return (
                qa_check.get("difference_or_count")
                == qa_check.get("tolerance_or_expected")
                and manifest_check.get("difference") in (0, 0.0)
                and manifest_check.get("tolerance") in (0, 0.0)
            )
        return (
            qa_check.get("difference_or_count") == manifest_check.get("difference")
            and qa_check.get("tolerance_or_expected") == manifest_check.get("tolerance")
        )

    expected_parity_ids = _REQUIRED_PARITY_CHECK_IDS
    if set(parity_status_by_id) != expected_parity_ids:
        reasons.append("manifest_parity_set_incomplete")
    elif any(
        qa_status_by_id.get(check_id) != status
        for check_id, status in parity_status_by_id.items()
    ):
        reasons.append("qa_manifest_parity_mismatch")
    elif any(
        not _parity_values_match(check_id)
        for check_id in expected_parity_ids
    ):
        reasons.append("qa_manifest_parity_values_mismatch")
    if _safe_hash(qa.get("workbook_sha256")) != artifacts.workbook_hash:
        reasons.append("workbook_hash_not_verified")
    if artifacts.issues:
        reasons.append("artifact_identity_contradictory")

    return {
        "verified": not reasons,
        "status": "VERIFIED" if not reasons else "UNVERIFIED",
        "reasons": sorted(set(reasons)),
        "qa_contract": _qa_contract_payload(),
        "calculation_marker": marker or None,
        "cache_engine": calculation.get("cache_engine"),
        "cache_merge": calculation.get("cache_merge"),
        "native_excel": calculation.get("native_excel"),
        "recalculation_state": recalculation,
        "formula_count": formula_count,
        "formula_results_nonblank": nonblank,
        "formula_results_blank_intentional_gates": intentional,
        "formula_errors": formula_errors if isinstance(formula_errors, list) else None,
    }


def _blocker_group(blocker: str) -> str:
    code = str(blocker or "").strip().lower()
    if code.startswith("pm_approval_required:"):
        return "pm_review"
    if (
        code.startswith("source_or_pm_required:segment")
        or code.startswith("sotp_")
        or code.startswith("historical:source_dependent_modules")
        or code.startswith("segment")
        or code.startswith("consensus")
        or code.startswith("fcfe")
    ):
        return "package"
    if code.startswith("source") or code.startswith("historical:"):
        return "source"
    if "formula" in code or "recalculation" in code or "forecast_completeness" in code or "scenario_completeness" in code:
        return "mechanical"
    if code.startswith("balance_sheet") or code.startswith("cash_flow") or code.startswith("accounting"):
        return "accounting"
    if code.startswith("wacc") or code.startswith("market_data"):
        return "market_data"
    if code.startswith("valuation") or code.startswith("dcf") or code.startswith("missing_valuation"):
        return "valuation"
    return "unknown"


def _blocker_root_cause_id(blocker: str) -> str:
    code = str(blocker or "").strip().lower()
    if re.fullmatch(r"(?:source_)?formula_reference_errors:\d+", code) or re.fullmatch(
        r"source_formula_errors:\d+", code
    ):
        return "source_formula_reference_errors"
    if code in {
        "source_preflight_blocked",
        "sotp_source_preflight_blocked",
        "historical:source_dependent_modules_unavailable",
    }:
        return "source_preflight"
    if code.startswith("wacc_degraded:") or code.startswith("market_data"):
        return "wacc_and_market_data_evidence"
    if code.startswith("source_or_pm_required:segment"):
        return "segment_source_evidence"
    if code.startswith("pm_approval_required:"):
        return "pm_driver:" + code.removeprefix("pm_approval_required:")
    return code or "blank_blocker"


def group_professional_model_blockers(blockers: Iterable[Any]) -> dict[str, Any]:
    grouped: dict[str, list[str]] = defaultdict(list)
    root_causes: dict[str, dict[str, Any]] = {}
    raw_total = 0
    for raw in blockers:
        raw_total += 1
        blocker = str(raw or "").strip() or "blank_blocker"
        group = _blocker_group(blocker) if blocker != "blank_blocker" else "unknown"
        grouped[group].append(blocker)
        root_cause_id = _blocker_root_cause_id(blocker)
        root = root_causes.setdefault(
            root_cause_id,
            {
                "root_cause_id": root_cause_id,
                "symptoms": set(),
                "groups": set(),
                "occurrence_count": 0,
                "approvable": group == "pm_review",
            },
        )
        root["symptoms"].add(blocker)
        root["groups"].add(group)
        root["occurrence_count"] += 1
        root["approvable"] = bool(root["approvable"] and group == "pm_review")

    order = (
        "mechanical",
        "source",
        "accounting",
        "market_data",
        "valuation",
        "pm_review",
        "package",
        "unknown",
    )
    normalized = {
        name: sorted(set(grouped.get(name, [])))
        for name in order
        if grouped.get(name)
    }
    normalized_root_causes = [
        {
            **root,
            "symptoms": sorted(root["symptoms"]),
            "groups": sorted(root["groups"]),
        }
        for _, root in sorted(root_causes.items())
    ]
    return {
        "groups": normalized,
        "counts": {name: len(values) for name, values in normalized.items()},
        "total": sum(len(values) for values in normalized.values()),
        "raw_total": raw_total,
        "normalized_root_cause_count": len(normalized_root_causes),
        "root_causes": normalized_root_causes,
    }


def _normalize_checks(qa_report: Mapping[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    raw_checks = qa_report.get("checks")
    if not isinstance(raw_checks, list):
        return [], ["checks_missing"]
    checks: list[dict[str, Any]] = []
    issues: list[str] = []
    seen: set[str] = set()
    for raw in raw_checks:
        if not isinstance(raw, dict):
            issues.append("check_not_object")
            continue
        check_id = str(raw.get("check_id") or "").strip()
        status = str(raw.get("status") or "").strip().upper()
        if not check_id:
            issues.append("check_id_blank")
            continue
        if check_id in seen:
            issues.append(f"duplicate_check:{check_id}")
        seen.add(check_id)
        if status not in _ALLOWED_CHECK_STATUSES:
            issues.append(f"unknown_check_status:{check_id}:{status or 'blank'}")
        checks.append(
            {
                "check_id": check_id,
                "status": status or None,
                "difference_or_count": raw.get("difference_or_count"),
                "tolerance_or_expected": raw.get("tolerance_or_expected"),
            }
        )
    return checks, issues


def _review_requirement_from_blocker(blocker: str) -> dict[str, Any] | None:
    parts = blocker.split(":", 3)
    if len(parts) != 4 or parts[0] != "pm_approval_required":
        return None
    scenario = parts[1].strip()
    driver = parts[2].strip()
    approval_key = parts[3].strip()
    if not scenario or not driver or not approval_key:
        return None
    return {
        "approval_key": approval_key,
        "scope": f"scenario_driver:{scenario}:{driver}",
        "scenario": scenario,
        "driver": driver,
        "blocker": blocker,
        "required_value_shape": {"type": "number_array", "length": 5},
        "value_source": "pm_submitted_exact_path",
    }




def _expected_driver_approval_identities() -> set[tuple[str, str]]:
    from src.stage_02_valuation.integrated_financial_forecast import DRIVER_SPECS

    return {
        (
            f"scenario_driver:{scenario}:{driver}",
            f"pmq:{scenario}:{driver}",
        )
        for scenario in ("Base", "Upside", "Downside")
        for driver in DRIVER_SPECS
    }
def _load_review_events(ticker: str) -> list[dict[str, Any]]:
    from db.loader import list_professional_model_review_events

    with get_connection() as conn:
        return list_professional_model_review_events(conn, ticker=ticker)


def _event_stale_reasons(
    event: Mapping[str, Any], artifacts: ProfessionalModelArtifacts
) -> list[str]:
    reasons: list[str] = []
    scope = str(event.get("approval_scope") or "")
    expected = {
        "model_run_id": artifacts.model_run_id,
        "source_hash": _safe_hash(artifacts.manifest.get("source_hash")),
        "input_hash": _safe_hash(artifacts.manifest.get("model_input_hash")),
        "result_hash": _safe_hash(artifacts.manifest.get("result_hash")),
        "workbook_hash": artifacts.workbook_hash,
    }
    if scope == FINAL_SIGNOFF_SCOPE or not scope.startswith("scenario_driver:"):
        expected.update(
            {
                "manifest_hash": _safe_hash(artifacts.manifest.get("manifest_hash")),
                "qa_hash": artifacts.qa_report_hash,
                "review_evidence_hash": artifacts.review_evidence_file_hash,
            }
        )
    for field, value in expected.items():
        observed = event.get(field)
        if field == "model_run_id":
            try:
                observed = int(observed)
            except (TypeError, ValueError):
                observed = None
        else:
            observed = str(observed or "").lower()
        if observed != value:
            reasons.append(f"{field}_changed")

    metadata = event.get("metadata")
    metadata = metadata if isinstance(metadata, Mapping) else {}
    if metadata.get("fingerprint_version") != REVIEW_FINGERPRINT_VERSION:
        reasons.append("fingerprint_version_changed")
    fingerprint_payload = {
        "approval_key": event.get("approval_key"),
        "fingerprint_version": REVIEW_FINGERPRINT_VERSION,
        "reviewed_values": event.get("reviewed_values"),
        "scope": scope,
    }
    if scope.startswith("scenario_driver:"):
        fingerprint_payload["requirement_hash"] = metadata.get("requirement_hash")
        fingerprint_payload["review_context"] = metadata.get("review_context") or {}
    expected_fingerprint = _canonical_hash(fingerprint_payload)
    if event.get("reviewed_value_fingerprint") != expected_fingerprint:
        reasons.append("reviewed_value_fingerprint_changed")
    return sorted(set(reasons))


def _decorate_review_events(
    events: Sequence[dict[str, Any]], artifacts: ProfessionalModelArtifacts
) -> list[dict[str, Any]]:
    latest_by_key: dict[tuple[int, str, str], int] = {}
    for event in events:
        key = (
            int(event.get("model_run_id") or 0),
            str(event.get("approval_scope") or ""),
            str(event.get("approval_key") or ""),
        )
        latest_by_key[key] = max(latest_by_key.get(key, 0), int(event["event_id"]))
    decorated: list[dict[str, Any]] = []
    for event in events:
        item = dict(event)
        key = (
            int(item.get("model_run_id") or 0),
            str(item.get("approval_scope") or ""),
            str(item.get("approval_key") or ""),
        )
        stale_reasons = _event_stale_reasons(item, artifacts)
        item["superseded"] = int(item["event_id"]) != latest_by_key.get(key)
        item["stale"] = bool(stale_reasons)
        item["stale_reasons"] = stale_reasons
        decorated.append(item)
    return decorated


def _approval_artifact_identity(event: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "model_run_id": int(event.get("model_run_id") or 0),
        "source_sha256": _safe_hash(event.get("source_hash")),
        "model_input_sha256": _safe_hash(event.get("input_hash")),
        "result_sha256": _safe_hash(event.get("result_hash")),
        "workbook_sha256": _safe_hash(event.get("workbook_hash")),
    }



def _requirements_for_artifact(
    artifacts: ProfessionalModelArtifacts,
) -> tuple[list[dict[str, Any]], list[str]]:
    blockers = [str(item) for item in (artifacts.manifest.get("blockers") or [])]
    pm_blockers = [
        blocker for blocker in blockers
        if blocker.startswith("pm_approval_required:")
    ]
    requirements: list[dict[str, Any]] = []
    issues: list[str] = []
    if pm_blockers:
        for blocker in pm_blockers:
            requirement = _review_requirement_from_blocker(blocker)
            if requirement is None:
                issues.append("malformed_pm_approval_blocker")
                continue
            if requirement["scenario"] not in {"Base", "Upside", "Downside"}:
                issues.append("unknown_pm_approval_scenario")
            if requirement["approval_key"] != (
                f"pmq:{requirement['scenario']}:{requirement['driver']}"
            ):
                issues.append("pm_approval_key_mismatch")
            requirement_contract = enrich_pm_driver_requirement(
                ticker=artifacts.ticker,
                model_run_id=artifacts.model_run_id,
                artifact_hashes=_hash_payload(artifacts),
                forecast_periods=artifacts.forecast_periods,
                blocker_row={
                    "row": blockers.index(blocker) + 5,
                    "sheet": "PM_Review_Queue",
                    "code": blocker,
                },
                scenario=requirement["scenario"],
                driver=requirement["driver"],
                approval_key=requirement["approval_key"],
                scope=requirement["scope"],
                canonical_hash_fn=_canonical_hash,
            )
            requirement = {
                **requirement,
                **requirement_contract,
                "requirement_contract": requirement_contract,
                "value_source": "pm_submitted_exact_path",
            }
            requirements.append(requirement)
    else:
        evidence = artifacts.review_evidence or {}
        for item in evidence.get("consumed_approval_events") or []:
            if not isinstance(item, Mapping):
                issues.append("review_evidence_event_not_object")
                continue
            scope = str(item.get("scope") or "")
            parts = scope.split(":", 2)
            if len(parts) != 3 or parts[0] != "scenario_driver":
                issues.append("review_evidence_scope_invalid")
                continue
            scenario, driver = parts[1], parts[2]
            approval_key = str(item.get("approval_key") or "")
            if (
                scenario not in {"Base", "Upside", "Downside"}
                or approval_key != f"pmq:{scenario}:{driver}"
            ):
                issues.append("review_evidence_requirement_invalid")
                continue
            requirement_contract = item.get("requirement_contract")
            if not isinstance(requirement_contract, Mapping):
                issues.append("review_evidence_requirement_contract_missing")
                continue
            requirement_hash = _safe_hash(requirement_contract.get("requirement_hash"))
            requirement_content = {
                key: value
                for key, value in requirement_contract.items()
                if key != "requirement_hash"
            }
            if (
                requirement_hash != _canonical_hash(requirement_content)
                or requirement_contract.get("approval_key") != approval_key
                or requirement_contract.get("scope") != scope
                or requirement_contract.get("scenario") != scenario
                or requirement_contract.get("driver") != driver
            ):
                issues.append("review_evidence_requirement_contract_invalid")
                continue
            requirements.append(
                {
                    "approval_key": approval_key,
                    "requirement_contract": dict(requirement_contract),
                    **dict(requirement_contract),
                    "scope": scope,
                    "scenario": scenario,
                    "driver": driver,
                    "blocker": None,
                    "required_value_shape": {"type": "number_array", "length": 5},
                    "value_source": "rebuild_review_evidence",
                    "consumed_event_id": int(item.get("event_id") or 0),
                    "consumed_fingerprint": item.get(
                        "reviewed_value_fingerprint"
                    ),
                    "consumed_reviewed_values": item.get("reviewed_values"),
                    "consumed_artifact_identity": item.get("approval_artifact_identity"),
                }
            )
        if not requirements:
            issues.append("pm_review_positive_list_missing")

    by_identity: dict[tuple[str, str], dict[str, Any]] = {}
    for requirement in requirements:
        identity = (requirement["scope"], requirement["approval_key"])
        if identity in by_identity:
            issues.append("duplicate_pm_approval_requirement")
        by_identity[identity] = requirement
    ordered = sorted(
        by_identity.values(),
        key=lambda item: (
            item["scenario"],
            item["driver"],
            item["scope"],
            item["approval_key"],
        ),
    )
    if len(ordered) != len(pm_blockers) and pm_blockers:
        issues.append("pm_approval_requirement_count_mismatch")
    return ordered, sorted(set(issues))


def _review_state_for_artifact(
    artifacts: ProfessionalModelArtifacts,
    *,
    events: Sequence[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    requirements, requirement_issues = _requirements_for_artifact(artifacts)
    artifact_requirement_identities = {
        (str(item.get("scope") or ""), str(item.get("approval_key") or ""))
        for item in requirements
    }
    expected_requirement_identities = _expected_driver_approval_identities()
    contract_issues = {
        str(issue)
        for item in requirements
        for issue in (item.get("contract_issues") or [])
    }
    if artifact_requirement_identities != expected_requirement_identities:
        contract_issues.add("runtime_driver_inventory_mismatch")
    contract_compatible = bool(requirements) and not contract_issues and all(
        item.get("approvable") is True for item in requirements
    )
    review_contract = {
        "version": "professional-model-review-requirement-v1",
        "compatible": contract_compatible,
        "regeneration_required": not contract_compatible,
        "issues": sorted(contract_issues),
        "artifact_requirement_count": len(artifact_requirement_identities),
        "runtime_requirement_count": len(expected_requirement_identities),
        "missing_identities": [
            {"scope": scope, "approval_key": key}
            for scope, key in sorted(expected_requirement_identities - artifact_requirement_identities)
        ],
        "unexpected_identities": [
            {"scope": scope, "approval_key": key}
            for scope, key in sorted(artifact_requirement_identities - expected_requirement_identities)
        ],
    }
    raw_events = (
        list(events)
        if events is not None
        else _load_review_events(artifacts.ticker)
    )
    events = _decorate_review_events(raw_events, artifacts)
    latest_current: dict[tuple[str, str], dict[str, Any]] = {}
    latest_approved: dict[tuple[str, str], dict[str, Any]] = {}
    for event in events:
        if int(event.get("model_run_id") or 0) != artifacts.model_run_id:
            continue
        if event.get("approval_scope") == FINAL_SIGNOFF_SCOPE:
            continue
        event_key = (
            str(event.get("approval_scope") or ""),
            str(event.get("approval_key") or ""),
        )
        current = latest_current.get(event_key)
        if current is None or int(event["event_id"]) > int(current["event_id"]):
            latest_current[event_key] = event
        if event.get("event_type") == "approve" and event.get("state") == "approved":
            approved = latest_approved.get(event_key)
            if approved is None or int(event["event_id"]) > int(approved["event_id"]):
                latest_approved[event_key] = event

    requirement_rows: list[dict[str, Any]] = []
    for requirement in requirements:
        current = latest_current.get(
            (requirement["scope"], requirement["approval_key"])
        )
        requirement_stale_reasons = (
            list(current.get("stale_reasons") or []) if current else []
        )
        if current is not None:
            event_metadata = current.get("metadata")
            event_metadata = event_metadata if isinstance(event_metadata, Mapping) else {}
            if event_metadata.get("requirement_hash") != requirement.get("requirement_hash"):
                requirement_stale_reasons.append("requirement_contract_changed")
            if (
                current.get("event_type") == "approve"
                and event_metadata.get("review_context_complete") is not True
            ):
                requirement_stale_reasons.append("approval_evidence_context_incomplete")
        approved_event = latest_approved.get(
            (requirement["scope"], requirement["approval_key"])
        )
        consumed_event_id = requirement.get("consumed_event_id")
        if consumed_event_id is not None:
            requirement_stale_reasons = [
                reason
                for reason in requirement_stale_reasons
                if reason
                not in {"input_hash_changed", "result_hash_changed", "workbook_hash_changed"}
            ]
        if current is not None and consumed_event_id is not None:
            if int(current.get("event_id") or 0) != int(consumed_event_id):
                requirement_stale_reasons.append("consumed_event_superseded")
            if current.get("reviewed_value_fingerprint") != requirement.get(
                "consumed_fingerprint"
            ):
                requirement_stale_reasons.append(
                    "consumed_fingerprint_changed"
                )
            if _canonical_json(current.get("reviewed_values")) != _canonical_json(
                requirement.get("consumed_reviewed_values")
            ):
                requirement_stale_reasons.append(
                    "consumed_reviewed_values_changed"
                )
            if _canonical_json(
                _approval_artifact_identity(current)
            ) != _canonical_json(requirement.get("consumed_artifact_identity")):
                requirement_stale_reasons.append(
                    "consumed_artifact_identity_changed"
                )
        requirement_stale_reasons = sorted(
            set(requirement_stale_reasons)
        )
        if current is None:
            status = "pending"
        elif requirement_stale_reasons:
            status = "stale"
        elif current.get("event_type") == "approve" and current.get("state") == "approved":
            status = "approved"
        elif current.get("event_type") == "preview" and current.get("state") == "previewed":
            status = "previewed"
        elif current.get("event_type") == "reject" and current.get("state") == "rejected":
            status = "rejected"
        else:
            status = "unknown"
        current_artifact_identity = (
            _approval_artifact_identity(current) if current else None
        )
        approval_identity_fingerprint = (
            _canonical_hash(
                {
                    "approval_key": requirement["approval_key"],
                    "scope": requirement["scope"],
                    "reviewed_values": current.get("reviewed_values"),
                    "artifact_identity": current_artifact_identity,
                }
            ) if current else None
        )
        current_metadata = current.get("metadata") if current else {}
        current_metadata = current_metadata if isinstance(current_metadata, Mapping) else {}
        review_context = current_metadata.get("review_context")
        review_context = review_context if isinstance(review_context, Mapping) else {}
        current_values = current.get("reviewed_values") if current else None
        proposed_path = (
            list(current_values)
            if isinstance(current_values, list) and len(current_values) == 5
            else None
        )
        approved_values = approved_event.get("reviewed_values") if approved_event else None
        approved_path = (
            list(approved_values)
            if isinstance(approved_values, list) and len(approved_values) == 5
            else None
        )
        applied_values = requirement.get("consumed_reviewed_values")
        applied_path = (
            list(applied_values)
            if isinstance(applied_values, list) and len(applied_values) == 5
            else None
        )
        latest_event = (
            {
                "event_id": current.get("event_id"),
                "event_type": current.get("event_type"),
                "state": current.get("state"),
                "reviewer": current.get("actor"),
                "rationale": current.get("rationale"),
                "timestamp": current.get("created_at"),
                "superseded": bool(current.get("superseded")),
                "stale_reasons": requirement_stale_reasons,
            }
            if current else None
        )
        requirement_rows.append(
            {
                **requirement,
                "status": status,
                "artifact_current_path": requirement.get("artifact_current_path"),
                "current_path": requirement.get("artifact_current_path"),
                "proposed_path": proposed_path,
                "proposed_path_status": status if proposed_path is not None else "not_provided",
                "approved_path": approved_path,
                "approved_path_status": (
                    "approved" if approved_path is not None and not approved_event.get("stale") else
                    "stale" if approved_path is not None else "not_approved"
                ),
                "applied_path": applied_path,
                "applied_path_status": "applied" if applied_path is not None else "not_applied",
                "source_ref": review_context.get("source_ref") or requirement.get("source_ref"),
                "method": review_context.get("method") or requirement.get("method"),
                "as_of": review_context.get("as_of") or requirement.get("as_of"),
                "materiality": review_context.get("materiality", requirement.get("materiality")),
                "impact": review_context.get("impact", requirement.get("impact")),
                "evidence_locator": review_context.get("evidence_locator") or requirement.get("evidence_locator"),
                "downstream_dependencies": review_context.get("downstream_dependencies") or requirement.get("downstream_dependencies", []),
                "review_context": dict(review_context),
                "latest_event": latest_event,
                "latest_event_type": current.get("event_type") if current else None,
                "reviewer": current.get("actor") if current else None,
                "rationale": current.get("rationale") if current else None,
                "timestamp": current.get("created_at") if current else None,
                "stale_reason": requirement_stale_reasons[0] if requirement_stale_reasons else None,
                "reviewed_values": current.get("reviewed_values") if current else None,
                "reviewed_value_fingerprint": current.get("reviewed_value_fingerprint") if current else None,
                "approval_artifact_identity": current_artifact_identity,
                "approval_identity_fingerprint": approval_identity_fingerprint,
                "current_event_id": current.get("event_id") if current else None,
                "actor": current.get("actor") if current else None,
                "reviewed_at": current.get("created_at") if current else None,
                "stale_reasons": requirement_stale_reasons,
            }
        )

    signoff_events = [
        event
        for event in events
        if event.get("approval_scope") == FINAL_SIGNOFF_SCOPE
        and event.get("approval_key") == FINAL_SIGNOFF_KEY
        and int(event.get("model_run_id") or 0) == artifacts.model_run_id
    ]
    signoff = max(signoff_events, key=lambda item: int(item["event_id"]), default=None)
    latest_review_event_id = max(
        (row.get("current_event_id") or 0 for row in requirement_rows),
        default=0,
    )
    signoff_current = bool(
        signoff
        and not signoff.get("stale")
        and not signoff.get("superseded")
        and signoff.get("event_type") == "signoff"
        and signoff.get("state") == "signed_off"
        and int(signoff.get("event_id") or 0) > latest_review_event_id
    )
    signoff_stale_reasons = list(signoff.get("stale_reasons") or []) if signoff else []
    if signoff and signoff.get("superseded"):
        signoff_stale_reasons.append("superseded_by_later_signoff")
    if (
        signoff
        and int(signoff.get("event_id") or 0) <= latest_review_event_id
    ):
        signoff_stale_reasons.append("review_event_after_signoff")
    signoff_stale_reasons = sorted(set(signoff_stale_reasons))

    return {
        "requirements": requirement_rows,
        "review_contract": review_contract,
        "events": events,
        "counts": dict(Counter(row["status"] for row in requirement_rows)),
        "required_count": len(requirement_rows),
        "approved_count": sum(row["status"] == "approved" for row in requirement_rows),
        "all_approved": bool(requirement_rows)
        and not requirement_issues
        and contract_compatible
        and all(
            row["status"] == "approved" for row in requirement_rows
        ),
        "evidence_issues": requirement_issues,
        "signoff": {
            "status": "signed_off" if signoff_current else ("stale" if signoff_stale_reasons else "pending"),
            "current": signoff_current,
            "event_id": signoff.get("event_id") if signoff else None,
            "actor": signoff.get("actor") if signoff_current else None,
            "signed_at": signoff.get("created_at") if signoff_current else None,
            "workbook_hash": signoff.get("workbook_hash") if signoff_current else None,
            "stale_reasons": signoff_stale_reasons,
        },
    }


def _package_requirements(
    blocker_groups: Mapping[str, Any],
    warnings: Sequence[str],
) -> list[dict[str, Any]]:
    requirements: dict[str, dict[str, Any]] = {}
    for warning in warnings:
        details = _PACKAGE_WARNING_REQUIREMENTS.get(str(warning))
        if details is None:
            continue
        key, owner, remediation = details
        requirements[key] = {
            "requirement": key,
            "status": "INCOMPLETE",
            "owner": owner,
            "remediation": remediation,
            "evidence": [str(warning)],
        }
    for blocker in (blocker_groups.get("groups") or {}).get("package", []):
        lower = blocker.lower()
        if "segment" in lower:
            key, owner = "segments", "finance_data"
            remediation = "Provide source-backed segment history and approved segment drivers."
        elif "sotp" in lower:
            key, owner = "sotp", "finance_data"
            remediation = "Provide normalized segment evidence before enabling SOTP."
        elif "consensus" in lower:
            key, owner = "consensus", "data"
            remediation = "Provide a frozen, as-of-matched consensus snapshot."
        elif "fcfe" in lower:
            key, owner = "fcfe", "finance"
            remediation = "Provide an integrated debt and after-tax interest schedule."
        else:
            key, owner = "source_dependent_modules", "finance_data"
            remediation = "Complete every source-dependent professional-model module."
        entry = requirements.setdefault(
            key,
            {
                "requirement": key,
                "status": "INCOMPLETE",
                "owner": owner,
                "remediation": remediation,
                "evidence": [],
            },
        )
        entry["evidence"] = sorted(set([*entry.get("evidence", []), blocker]))
    return sorted(requirements.values(), key=lambda item: item["requirement"])


def _evaluate_state(
    artifacts: ProfessionalModelArtifacts,
    calculation: Mapping[str, Any],
    review_state: Mapping[str, Any],
) -> dict[str, Any]:
    qa = artifacts.qa_report
    blockers = [str(item) for item in (artifacts.manifest.get("blockers") or [])]
    blocker_groups = group_professional_model_blockers(blockers)
    warnings = sorted(str(item) for item in (artifacts.manifest.get("warnings") or []))
    checks, check_issues = _normalize_checks(qa)
    state_issues = [item["code"] for item in artifacts.issues]
    state_issues.extend(check_issues)
    reported_status = str(qa.get("model_status") or "").strip().upper()
    state_issues.extend(review_state.get("evidence_issues") or [])
    if reported_status not in _ALLOWED_REPORTED_STATUSES:
        state_issues.append(f"unknown_reported_status:{reported_status or 'blank'}")
    decision_ready = qa.get("decision_ready")
    if not isinstance(decision_ready, bool):
        state_issues.append("decision_readiness_unknown")
    semantic_qa = build_decision_semantic_qa_verification(
        qa.get("checks"),
        required_core_ids=_REQUIRED_RENDERER_QA_CHECK_IDS,
        canonical_hash_fn=_canonical_hash,
    )
    review_contract = review_state.get("review_contract") or {}

    core_failure_groups = {
        group
        for group in ("mechanical", "source", "accounting", "market_data", "valuation", "unknown")
        if (blocker_groups.get("counts") or {}).get(group, 0) > 0
    }
    if not semantic_qa.get("verified"):
        core_failure_groups.add("semantic_qa")
    if review_contract.get("compatible") is not True:
        core_failure_groups.add("review_contract")
    pm_open = not bool(review_state.get("all_approved"))
    check_group_map = {
        "source_preflight": "source",
        "scenario_completeness": "mechanical",
        "forecast_completeness": "mechanical",
        "balance_sheet": "accounting",
        "cash_flow_tie": "accounting",
        "valuation_bridge": "valuation",
        "dcf_decision_gate": "valuation",
        "valuation_input_gate": "valuation",
        "scenario_selector": "mechanical",
        "calculation_verification": "mechanical",
        **{check_id: "semantic_qa" for check_id in SEMANTIC_QA_CHECK_IDS},
    }
    check_failures: list[dict[str, Any]] = []
    dependency_check_failures: list[dict[str, Any]] = []
    for check in checks:
        status = check.get("status")
        check_id = str(check.get("check_id"))
        if (
            status == "PASS"
            or check_id in {"model_readiness", "pm_driver_approvals", "optional_modules"}
        ):
            continue
        if check_id in {"dcf_decision_gate", "valuation_bridge"} and pm_open:
            if not (blocker_groups.get("groups") or {}).get("valuation", []):
                dependency_check_failures.append(
                    {
                        "check_id": check_id,
                        "status": status,
                        "group": "valuation",
                        "dependency": "pm_driver_approvals",
                    }
                )
                continue
        group = check_group_map.get(check_id, "unknown")
        core_failure_groups.add(group)
        check_failures.append({"check_id": check_id, "status": status, "group": group})

    package_requirements = _package_requirements(blocker_groups, warnings)
    missing_approvals = [
        row for row in review_state.get("requirements", [])
        if row.get("status") != "approved"
    ]
    all_approvals = bool(review_state.get("all_approved"))
    package_complete = not package_requirements
    identity_verified = not artifacts.issues
    calculation_verified = bool(calculation.get("verified"))
    reported_ready = bool(decision_ready is True and reported_status in _READY_REPORTED_STATUSES)
    if state_issues or not calculation_verified or not identity_verified:
        normalized_state = "UNVERIFIED"
    elif core_failure_groups:
        normalized_state = "BLOCKED"
    elif not all_approvals:
        normalized_state = "NEEDS_PM_REVIEW"
    elif not package_complete:
        normalized_state = "PARTIAL"
    elif not reported_ready:
        normalized_state = "BLOCKED"
        core_failure_groups.add("reported_decision_readiness")
    elif not bool((review_state.get("signoff") or {}).get("current")):
        normalized_state = "NEEDS_PM_REVIEW"
    else:
        normalized_state = "FULL"

    group_requirements = (
        (
            "artifact_identity",
            "engineering",
            "Regenerate a complete root artifact tuple whose ticker, run, sheets, and hashes agree.",
            identity_verified and not state_issues,
            [item["code"] for item in artifacts.issues],
        ),
        (
            "calculation_verification",
            "engineering",
            "Run deterministic recalculation/cache-parity QA and publish its evidence.",
            calculation_verified,
            list(calculation.get("reasons") or []),
        ),
        (
            "mechanical_integrity",
            "engineering",
            "Clear formula, recalculation, scenario, and forecast-completeness failures.",
            "mechanical" not in core_failure_groups,
            (blocker_groups.get("groups") or {}).get("mechanical", []),
        ),
        (
            "source_evidence",
            "data",
            "Repair source formula references and bind complete run-exact evidence.",
            "source" not in core_failure_groups,
            (blocker_groups.get("groups") or {}).get("source", []),
        ),
        (
            "accounting_integrity",
            "finance",
            "Clear balance-sheet, cash-flow, and accounting checks with source-backed mechanics.",
            "accounting" not in core_failure_groups,
            (blocker_groups.get("groups") or {}).get("accounting", []),
        ),
        (
            "market_data_and_wacc",
            "finance_data",
            "Replace degraded WACC or market-data fallbacks with reviewed source-backed inputs.",
            "market_data" not in core_failure_groups,
            (blocker_groups.get("groups") or {}).get("market_data", []),
        ),
        (
            "valuation_gates",
            "finance",
            "Clear DCF input and EV-to-equity bridge gates in deterministic model code.",
            "valuation" not in core_failure_groups and not dependency_check_failures,
            [
                *(blocker_groups.get("groups") or {}).get("valuation", []),
                *[row["check_id"] for row in check_failures if row["group"] == "valuation"],
                *[row["check_id"] for row in dependency_check_failures],
            ],
        ),
        (
            "review_contract_integrity",
            "finance_engineering",
            "Regenerate a versioned driver-review contract whose inventory, periods, units, and evidence fields are complete.",
            review_contract.get("compatible") is True,
            review_contract,
        ),
        (
            "semantic_qa_positive_list",
            "finance_engineering",
            "Publish exact PASS evidence for WACC methodology/parity, fully diluted share basis, and as-of alignment.",
            semantic_qa.get("verified") is True,
            semantic_qa,
        ),
    )
    full_requirements = [
        {
            "requirement": name,
            "status": (
                "PENDING"
                if name == "valuation_gates" and dependency_check_failures
                else "PASS"
                if passed
                else ("UNKNOWN" if name == "artifact_identity" and state_issues else "FAIL")
            ),
            "owner": owner,
            "remediation": remediation,
            "evidence": evidence,
        }
        for name, owner, remediation, passed, evidence in group_requirements
    ]
    full_requirements.extend(
        [
            {
                "requirement": "pm_driver_approvals",
                "status": "PASS" if all_approvals else "PENDING",
                "owner": "pm",
                "remediation": "Preview and approve every exact five-year scenario-driver path; stale fingerprints never pass.",
                "evidence": {
                    "required": review_state.get("required_count", 0),
                    "approved": review_state.get("approved_count", 0),
                    "open_keys": [row["approval_key"] for row in missing_approvals],
                },
            },
            {
                "requirement": "required_package_modules",
                "status": "PASS" if package_complete else "INCOMPLETE",
                "owner": "finance_data",
                "remediation": "Complete every positive-listed package module required for FULL.",
                "evidence": package_requirements,
            },
            {
                "requirement": "reported_decision_readiness",
                "status": (
                    "PASS" if reported_ready else (
                        "PENDING"
                        if missing_approvals or package_requirements or core_failure_groups
                        else "FAIL"
                    )
                ),
                "owner": "finance_engineering",
                "remediation": "Publish QA that positively states decision readiness after preceding gates pass.",
                "evidence": {
                    "reported_workbook_status": reported_status or None,
                    "decision_ready": decision_ready if isinstance(decision_ready, bool) else None,
                },
            },
            {
                "requirement": "exact_hash_signoff",
                "status": "PASS" if (review_state.get("signoff") or {}).get("current") else "PENDING",
                "owner": "pm",
                "remediation": "Sign off only after every other gate passes, using the exact workbook SHA-256.",
                "evidence": review_state.get("signoff"),
            },
        ]
    )
    pre_signoff_ready = bool(
        identity_verified
        and calculation_verified
        and not state_issues
        and not core_failure_groups
        and all_approvals
        and package_complete
        and reported_ready
    )
    return {
        "normalized_state": normalized_state,
        "reported_workbook_status": reported_status or None,
        "decision_readiness": normalized_state == "FULL",
        "reported_decision_ready": decision_ready if isinstance(decision_ready, bool) else None,
        "state_issues": sorted(set(state_issues)),
        "core_failure_groups": sorted(core_failure_groups),
        "check_failures": check_failures,
        "checks": checks,
        "decision_semantic_qa_verification": semantic_qa,
        "review_contract": review_contract,
        "blockers": blocker_groups,
        "warnings": warnings,
        "package_requirements": package_requirements,
        "full_state_requirements": full_requirements,
        "pre_signoff_ready": pre_signoff_ready,
        "permitted_actions": {
            "view_summary": True,
            "view_sheets": identity_verified,
            "download": identity_verified,
            "review_preview": identity_verified
            and calculation_verified
            and any(
                row.get("approvable") is True for row in review_state.get("requirements", [])
            ),
            "review_approve": identity_verified
            and calculation_verified
            and any(
                row.get("status") == "previewed"
                for row in review_state.get("requirements", [])
            ),
            "review_reject": identity_verified
            and calculation_verified
            and any(
                row.get("approvable") is True
                and row.get("current_event_id") is not None
                for row in review_state.get("requirements", [])
            ),
            "signoff": pre_signoff_ready,
            "rebuild": identity_verified,
        },
    }


def _sheet_audit_findings(artifacts: ProfessionalModelArtifacts) -> dict[str, Any]:
    raw = artifacts.qa_report.get("sheet_audit_findings")
    if isinstance(raw, list):
        return {
            "available": True,
            "findings": [item for item in raw if isinstance(item, dict)],
        }
    return {"available": False, "findings": []}


def _sheet_summaries(artifacts: ProfessionalModelArtifacts) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    classification_counts = Counter(
        str(item.get("sheet"))
        for item in artifacts.manifest.get("cell_classifications") or []
        if isinstance(item, dict) and item.get("sheet")
    )
    mapping_counts = Counter(
        str(item.get("sheet"))
        for item in artifacts.manifest.get("line_cell_mappings") or []
        if isinstance(item, dict) and item.get("sheet")
    )
    blank_results = (artifacts.qa_report.get("integrity") or {}).get(
        "blank_results_by_sheet"
    ) or {}
    workbook = load_workbook(artifacts.workbook_path, data_only=False, read_only=False)
    try:
        summaries: list[dict[str, Any]] = []
        for index, worksheet in enumerate(workbook.worksheets):
            nonempty = formulas = comments = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    nonempty += cell.value is not None
                    formulas += cell.data_type == "f"
                    comments += cell.comment is not None
            hidden_rows = sum(
                bool(worksheet.row_dimensions[row].hidden)
                for row in range(1, worksheet.max_row + 1)
            )
            hidden_columns = sum(
                bool(worksheet.column_dimensions[column].hidden)
                for column in worksheet.column_dimensions
            )
            summaries.append(
                {
                    "name": worksheet.title,
                    "index": index,
                    "dimensions": {
                        "range": worksheet.calculate_dimension(),
                        "max_row": worksheet.max_row,
                        "max_column": worksheet.max_column,
                    },
                    "visibility": worksheet.sheet_state,
                    "hidden_row_count": hidden_rows,
                    "hidden_column_count": hidden_columns,
                    "merged_range_count": len(worksheet.merged_cells.ranges),
                    "formula_count": formulas,
                    "nonempty_cell_count": nonempty,
                    "comment_count": comments,
                    "manifest_classification_count": classification_counts[worksheet.title],
                    "manifest_line_mapping_count": mapping_counts[worksheet.title],
                    "cached_blank_result_count": blank_results.get(worksheet.title, 0),
                }
            )
        return summaries
    finally:
        workbook.close()


def _hash_payload(artifacts: ProfessionalModelArtifacts) -> dict[str, Any]:
    return {
        "source_sha256": _safe_hash(artifacts.manifest.get("source_hash")),
        "model_input_sha256": _safe_hash(artifacts.manifest.get("model_input_hash")),
        "result_sha256": _safe_hash(artifacts.manifest.get("result_hash")),
        "manifest_sha256": _safe_hash(artifacts.manifest.get("manifest_hash")),
        "workbook_sha256": artifacts.workbook_hash,
        "qa_report_sha256": artifacts.qa_report_hash,
        "review_evidence_sha256": artifacts.review_evidence_file_hash,
    }


def build_professional_model_summary(
    ticker: str,
    *,
    artifact_root: str | Path | None = None,
) -> dict[str, Any]:
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    calculation = _calculation_verification(artifacts)
    try:
        review_state = _review_state_for_artifact(artifacts)
        persistence_issue = None
    except Exception as exc:
        review_state = {
            "requirements": [],
            "events": [],
            "counts": {},
            "required_count": 0,
            "approved_count": 0,
            "all_approved": False,
            "signoff": {"status": "unknown", "current": False},
        }
        persistence_issue = str(exc)
    evaluation = _evaluate_state(artifacts, calculation, review_state)
    if persistence_issue:
        evaluation["normalized_state"] = "UNVERIFIED"
        evaluation["decision_readiness"] = False
        evaluation["state_issues"] = sorted(
            set([*evaluation.get("state_issues", []), "review_persistence_unavailable"])
        )
    qa = artifacts.qa_report
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "normalized_state": evaluation["normalized_state"],
        "reported_workbook_status": evaluation["reported_workbook_status"],
        "decision_readiness": evaluation["decision_readiness"],
        "reported_decision_ready": evaluation["reported_decision_ready"],
        "artifact_identity": {
            "verified": not artifacts.issues,
            "issues": list(artifacts.issues),
            "workbook_filename": artifacts.workbook_path.name,
            "workbook_bytes": artifacts.workbook_bytes,
            "renderer_version": artifacts.manifest.get("renderer_version"),
            "workbook_schema_version": artifacts.manifest.get("workbook_schema_version"),
            "qa_schema_version": qa.get("schema_version"),
        },
        "hashes": _hash_payload(artifacts),
        "calculation_verification": calculation,
        "decision_semantic_qa_verification": evaluation["decision_semantic_qa_verification"],
        "blockers": evaluation["blockers"],
        "warnings": evaluation["warnings"],
        "full_state_requirements": evaluation["full_state_requirements"],
        "checks": evaluation["checks"],
        "integrity": qa.get("integrity") or {},
        "valuation_diagnostics": qa.get("diagnostic_valuation") or {},
        "ev_to_equity_bridge": qa.get("bridge") or {},
        "sheets": _sheet_summaries(artifacts),
        "sheet_audit": _sheet_audit_findings(artifacts),
        "review": {
            "required_count": review_state.get("required_count", 0),
            "approved_count": review_state.get("approved_count", 0),
            "counts": review_state.get("counts", {}),
            "signoff": review_state.get("signoff"),
            "contract": review_state.get("review_contract"),
        },
        "download_url": f"/api/tickers/{artifacts.ticker}/professional-model/download",
        "permitted_actions": evaluation["permitted_actions"],
        "state_issues": evaluation["state_issues"],
    }


def build_professional_model_review_payload(
    ticker: str,
    *,
    artifact_root: str | Path | None = None,
) -> dict[str, Any]:
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    calculation = _calculation_verification(artifacts)
    review_state = _review_state_for_artifact(artifacts)
    evaluation = _evaluate_state(artifacts, calculation, review_state)
    audit_events = list(review_state["events"])
    recent_audit_events = audit_events[-500:]
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "normalized_state": evaluation["normalized_state"],
        "hashes": _hash_payload(artifacts),
        "requirements": review_state["requirements"],
        "review_contract": review_state["review_contract"],
        "decision_semantic_qa_verification": evaluation["decision_semantic_qa_verification"],
        "counts": review_state["counts"],
        "required_count": review_state["required_count"],
        "approved_count": review_state["approved_count"],
        "signoff": review_state["signoff"],
        "audit_events": recent_audit_events,
        "audit_event_page": {
            "total": len(audit_events),
            "returned": len(recent_audit_events),
            "truncated": len(recent_audit_events) < len(audit_events),
        },
        "permitted_actions": evaluation["permitted_actions"],
    }


def _require_verified_identity(artifacts: ProfessionalModelArtifacts) -> None:
    if artifacts.issues:
        raise ProfessionalModelConflictError(
            "professional model artifact identity is unverified"
        )

def _require_full_lifecycle_run(
    artifacts: ProfessionalModelArtifacts, *, action: str
) -> None:
    if artifacts.model_run_id < MIN_FULL_LIFECYCLE_MODEL_RUN_ID:
        raise ProfessionalModelConflictError(
            f"professional-model {action} requires regenerated run 4 or later"
        )



def _require_actionable_artifact(artifacts: ProfessionalModelArtifacts) -> None:
    _require_verified_identity(artifacts)
    _require_full_lifecycle_run(artifacts, action="review action")
    if not _calculation_verification(artifacts).get("verified"):
        raise ProfessionalModelConflictError(
            "professional model calculation evidence is unverified"
        )


def _refresh_actionable_artifact(
    expected: ProfessionalModelArtifacts,
    *,
    artifact_root: str | Path | None,
) -> ProfessionalModelArtifacts:
    """Rehash the immutable run tuple immediately before an event append."""
    current = discover_professional_model_artifacts(
        expected.ticker,
        artifact_root=artifact_root,
    )
    _require_actionable_artifact(current)
    if (
        current.model_run_id != expected.model_run_id
        or _hash_payload(current) != _hash_payload(expected)
    ):
        raise ProfessionalModelConflictError(
            "professional model artifact changed during review"
        )
    return current


def _require_review_requirement(
    artifacts: ProfessionalModelArtifacts, approval_key: str
) -> dict[str, Any]:
    key = _clean_text(approval_key, "approval_key", max_length=300)
    requirements, issues = _requirements_for_artifact(artifacts)
    if issues:
        raise ProfessionalModelConflictError(
            "professional-model approval inventory is invalid"
        )
    for requirement in requirements:
        if requirement["approval_key"] == key:
            if requirement.get("approvable") is not True:
                raise ProfessionalModelConflictError(
                    "professional-model review contract requires artifact regeneration"
                )
            return requirement
    raise ProfessionalModelValidationError(
        "approval key is not required by the current artifact"
    )


def _validate_reviewed_values(
    requirement: Mapping[str, Any], reviewed_values: Any
) -> list[float]:
    if not isinstance(reviewed_values, list) or len(reviewed_values) != 5:
        raise ProfessionalModelValidationError(
            "reviewed_values must contain exactly five annual numbers"
        )
    normalized: list[float] = []
    for value in reviewed_values:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ProfessionalModelValidationError("reviewed_values must be numeric")
        number = float(value)
        if not math.isfinite(number):
            raise ProfessionalModelValidationError("reviewed_values must be finite")
        normalized.append(number)
    try:
        from src.stage_02_valuation.integrated_financial_forecast import (
            DRIVER_SPECS,
            DriverPath,
        )

        driver = str(requirement["driver"])
        spec = DRIVER_SPECS[driver]
        DriverPath(
            key=driver,
            values=tuple(normalized),
            unit=spec.unit,
            source_ref="professional_model_review_preview",
            method="pm_submitted_exact_path",
            queue_item_id="professional_model_review_preview",
        )
    except KeyError as exc:
        raise ProfessionalModelValidationError("unknown professional-model driver") from exc
    except ValueError as exc:
        raise ProfessionalModelValidationError(str(exc)) from exc
    return normalized


def _reviewed_value_fingerprint(
    requirement: Mapping[str, Any],
    reviewed_values: Sequence[float],
    *,
    review_context: Mapping[str, Any] | None = None,
) -> str:
    return _canonical_hash(
        {
            "fingerprint_version": REVIEW_FINGERPRINT_VERSION,
            "approval_key": requirement["approval_key"],
            "scope": requirement["scope"],
            "reviewed_values": list(reviewed_values),
            "requirement_hash": requirement.get("requirement_hash"),
            "review_context": dict(review_context or {}),
        }
    )


def _event_base(
    artifacts: ProfessionalModelArtifacts,
    *,
    approval_key: str,
    approval_scope: str,
    reviewed_values: Any,
    reviewed_value_fingerprint: str,
    actor: str,
    rationale: str | None,
    event_type: str,
    state: str,
    parent_event_id: int | None,
    supersedes_event_id: int | None,
    metadata: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    hashes = _hash_payload(artifacts)
    return {
        "created_at": _now(),
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "approval_key": approval_key,
        "approval_scope": approval_scope,
        "event_type": event_type,
        "state": state,
        "reviewed_values": reviewed_values,
        "reviewed_value_fingerprint": reviewed_value_fingerprint,
        "input_hash": hashes["model_input_sha256"],
        "result_hash": hashes["result_sha256"],
        "source_hash": hashes["source_sha256"],
        "manifest_hash": hashes["manifest_sha256"],
        "workbook_hash": hashes["workbook_sha256"],
        "qa_hash": hashes["qa_report_sha256"],
        "review_evidence_hash": hashes["review_evidence_sha256"],
        "actor": actor,
        "rationale": rationale,
        "parent_event_id": parent_event_id,
        "supersedes_event_id": supersedes_event_id,
        "metadata": {
            "fingerprint_version": REVIEW_FINGERPRINT_VERSION,
            **dict(metadata or {}),
        },
    }


def preview_professional_model_review(
    ticker: str,
    *,
    approval_key: str,
    reviewed_values: Any,
    actor: str,
    rationale: str | None = None,
    artifact_root: str | Path | None = None,
    review_context: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    _require_actionable_artifact(artifacts)
    requirement = _require_review_requirement(artifacts, approval_key)
    values = _validate_reviewed_values(requirement, reviewed_values)
    actor = _clean_text(actor, "actor", max_length=200)
    rationale = str(rationale).strip() if rationale is not None else None
    try:
        normalized_review_context = normalize_preview_review_context(review_context)
    except ProfessionalModelReviewContractError as exc:
        raise ProfessionalModelValidationError(str(exc)) from exc
    review_context_complete = bool(
        normalized_review_context.get("source_ref")
        and normalized_review_context.get("method")
        and normalized_review_context.get("as_of")
        and normalized_review_context.get("evidence_locator")
    )
    fingerprint = _reviewed_value_fingerprint(
        requirement,
        values,
        review_context=normalized_review_context,
    )
    from db.loader import (
        insert_professional_model_review_event,
        list_professional_model_review_events,
        load_professional_model_review_event,
    )

    with get_connection() as conn:
        create_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        artifacts = _refresh_actionable_artifact(
            artifacts,
            artifact_root=artifact_root,
        )
        existing = list_professional_model_review_events(
            conn,
            ticker=artifacts.ticker,
            model_run_id=artifacts.model_run_id,
            approval_key=requirement["approval_key"],
            approval_scope=requirement["scope"],
        )
        supersedes = existing[-1]["event_id"] if existing else None
        event_id = insert_professional_model_review_event(
            conn,
            _event_base(
                artifacts,
                approval_key=requirement["approval_key"],
                approval_scope=requirement["scope"],
                reviewed_values=values,
                reviewed_value_fingerprint=fingerprint,
                actor=actor,
                rationale=rationale,
                event_type="preview",
                state="previewed",
                parent_event_id=None,
                supersedes_event_id=supersedes,
                metadata={
                    "value_origin": "pm_submitted_exact_path",
                    "requirement_hash": requirement.get("requirement_hash"),
                    "review_context": normalized_review_context,
                    "review_context_complete": review_context_complete,
                },
            ),
        )
        conn.commit()
        event = load_professional_model_review_event(conn, event_id)
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "preview_id": event_id,
        "approval_key": requirement["approval_key"],
        "scope": requirement["scope"],
        "scenario": requirement["scenario"],
        "driver": requirement["driver"],
        "reviewed_values": values,
        "reviewed_value_fingerprint": fingerprint,
        "hashes": _hash_payload(artifacts),
        "state": "previewed",
        "requirement_hash": requirement.get("requirement_hash"),
        "unit": requirement.get("unit"),
        "forecast_periods": requirement.get("forecast_periods"),
        "review_context": normalized_review_context,
        "review_context_complete": review_context_complete,
        "previewed_at": event.get("created_at") if event else None,
        "approval_allowed": review_context_complete,
    }


def approve_professional_model_review(
    ticker: str,
    *,
    preview_id: int,
    reviewed_value_fingerprint: str,
    actor: str,
    rationale: str | None = None,
    artifact_root: str | Path | None = None,
) -> dict[str, Any]:
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    _require_actionable_artifact(artifacts)
    actor = _clean_text(actor, "actor", max_length=200)
    supplied_fingerprint = str(reviewed_value_fingerprint or "").strip().lower()
    if not _SHA256_RE.fullmatch(supplied_fingerprint):
        raise ProfessionalModelValidationError(
            "reviewed_value_fingerprint must be SHA-256"
        )
    rationale = str(rationale).strip() if rationale is not None else None
    from db.loader import (
        insert_professional_model_review_event,
        list_professional_model_review_events,
        load_professional_model_review_event,
    )

    with get_connection() as conn:
        create_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        artifacts = _refresh_actionable_artifact(
            artifacts,
            artifact_root=artifact_root,
        )
        preview = load_professional_model_review_event(conn, int(preview_id))
        if preview is None or preview.get("event_type") != "preview":
            conn.rollback()
            raise ProfessionalModelConflictError(
                "approval requires a persisted preview"
            )
        if (
            preview.get("ticker") != artifacts.ticker
            or int(preview.get("model_run_id") or 0) != artifacts.model_run_id
        ):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "preview belongs to a different model artifact"
            )
        requirement = _require_review_requirement(
            artifacts, str(preview.get("approval_key"))
        )
        preview_metadata = preview.get("metadata")
        preview_metadata = preview_metadata if isinstance(preview_metadata, Mapping) else {}
        try:
            preview_review_context = normalize_preview_review_context(
                preview_metadata.get("review_context")
            )
        except ProfessionalModelReviewContractError as exc:
            conn.rollback()
            raise ProfessionalModelConflictError(
                "preview review context is invalid"
            ) from exc
        if (
            preview_metadata.get("requirement_hash") != requirement.get("requirement_hash")
            or preview_metadata.get("review_context_complete") is not True
        ):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "approval requires complete evidence for the current review requirement"
            )
        expected_fingerprint = _reviewed_value_fingerprint(
            requirement,
            preview.get("reviewed_values") or [],
            review_context=preview_review_context,
        )
        if (
            supplied_fingerprint != expected_fingerprint
            or preview.get("reviewed_value_fingerprint") != expected_fingerprint
        ):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "reviewed-value fingerprint changed after preview"
            )
        stale_reasons = _event_stale_reasons(preview, artifacts)
        if stale_reasons:
            conn.rollback()
            raise ProfessionalModelConflictError(
                "preview is stale: " + ", ".join(stale_reasons)
            )
        existing = list_professional_model_review_events(
            conn,
            ticker=artifacts.ticker,
            model_run_id=artifacts.model_run_id,
            approval_key=requirement["approval_key"],
            approval_scope=requirement["scope"],
        )
        if not existing or int(existing[-1]["event_id"]) != int(preview_id):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "preview was superseded by a later review event"
            )
        event_id = insert_professional_model_review_event(
            conn,
            _event_base(
                artifacts,
                approval_key=requirement["approval_key"],
                approval_scope=requirement["scope"],
                reviewed_values=preview["reviewed_values"],
                reviewed_value_fingerprint=expected_fingerprint,
                actor=actor,
                rationale=rationale,
                event_type="approve",
                state="approved",
                parent_event_id=int(preview_id),
                supersedes_event_id=int(preview_id),
                metadata={
                    **dict(preview_metadata),
                    "preview_event_id": int(preview_id),
                    "requirement_hash": requirement.get("requirement_hash"),
                    "review_context": preview_review_context,
                },
            ),
        )
        conn.commit()
        event = load_professional_model_review_event(conn, event_id)
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "event_id": event_id,
        "preview_id": int(preview_id),
        "approval_key": requirement["approval_key"],
        "scope": requirement["scope"],
        "requirement_hash": requirement.get("requirement_hash"),
        "review_context": preview_review_context,
        "reviewed_values": preview["reviewed_values"],
        "reviewed_value_fingerprint": expected_fingerprint,
        "hashes": _hash_payload(artifacts),
        "state": "approved",
        "approved_at": event.get("created_at") if event else None,
        "actor": actor,
    }


def reject_professional_model_review(
    ticker: str,
    *,
    approval_key: str,
    actor: str,
    rationale: str,
    artifact_root: str | Path | None = None,
) -> dict[str, Any]:
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    _require_actionable_artifact(artifacts)
    requirement = _require_review_requirement(artifacts, approval_key)
    actor = _clean_text(actor, "actor", max_length=200)
    rationale = _clean_text(rationale, "rationale", max_length=4_000)
    from db.loader import (
        insert_professional_model_review_event,
        list_professional_model_review_events,
        load_professional_model_review_event,
    )

    with get_connection() as conn:
        create_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        artifacts = _refresh_actionable_artifact(
            artifacts,
            artifact_root=artifact_root,
        )
        existing = list_professional_model_review_events(
            conn,
            ticker=artifacts.ticker,
            model_run_id=artifacts.model_run_id,
            approval_key=requirement["approval_key"],
            approval_scope=requirement["scope"],
        )
        prior = existing[-1] if existing else None
        values = list(prior.get("reviewed_values") or []) if prior else []
        if prior is None or len(values) != 5:
            conn.rollback()
            raise ProfessionalModelConflictError(
                "rejection requires a persisted five-value preview or approval"
            )
        prior_metadata = prior.get("metadata")
        prior_metadata = prior_metadata if isinstance(prior_metadata, Mapping) else {}
        if prior_metadata.get("requirement_hash") != requirement.get("requirement_hash"):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "rejection target is stale for the current review requirement"
            )
        try:
            prior_review_context = normalize_preview_review_context(
                prior_metadata.get("review_context")
            )
        except ProfessionalModelReviewContractError as exc:
            conn.rollback()
            raise ProfessionalModelConflictError("rejection target context is invalid") from exc
        fingerprint = _reviewed_value_fingerprint(
            requirement,
            values,
            review_context=prior_review_context,
        )
        event_id = insert_professional_model_review_event(
            conn,
            _event_base(
                artifacts,
                approval_key=requirement["approval_key"],
                approval_scope=requirement["scope"],
                reviewed_values=values,
                reviewed_value_fingerprint=fingerprint,
                actor=actor,
                rationale=rationale,
                event_type="reject",
                state="rejected",
                parent_event_id=prior.get("event_id") if prior else None,
                supersedes_event_id=prior.get("event_id") if prior else None,
                metadata={
                    **dict(prior_metadata),
                    "requirement_hash": requirement.get("requirement_hash"),
                    "review_context": prior_review_context,
                    "rejected_event_id": prior.get("event_id"),
                },
            ),
        )
        conn.commit()
        event = load_professional_model_review_event(conn, event_id)
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "event_id": event_id,
        "approval_key": requirement["approval_key"],
        "scope": requirement["scope"],
        "state": "rejected",
        "rationale": rationale,
        "rejected_at": event.get("created_at") if event else None,
        "actor": actor,
    }


def signoff_professional_model(
    ticker: str,
    *,
    workbook_sha256: str,
    actor: str,
    rationale: str,
    artifact_root: str | Path | None = None,
) -> dict[str, Any]:
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    supplied_hash = str(workbook_sha256 or "").strip().lower()
    if supplied_hash != artifacts.workbook_hash:
        raise ProfessionalModelConflictError(
            "sign-off workbook hash is stale or incorrect"
        )
    actor = _clean_text(actor, "actor", max_length=200)
    rationale = _clean_text(rationale, "rationale", max_length=4_000)
    from db.loader import (
        insert_professional_model_review_event,
        list_professional_model_review_events,
        load_professional_model_review_event,
    )

    with get_connection() as conn:
        create_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        artifacts = _refresh_actionable_artifact(
            artifacts,
            artifact_root=artifact_root,
        )
        current_events = list_professional_model_review_events(
            conn,
            ticker=artifacts.ticker,
            model_run_id=artifacts.model_run_id,
        )
        calculation = _calculation_verification(artifacts)
        review_state = _review_state_for_artifact(
            artifacts,
            events=current_events,
        )
        evaluation = _evaluate_state(artifacts, calculation, review_state)
        if not evaluation.get("pre_signoff_ready"):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "final sign-off is blocked until every other FULL-state requirement passes"
            )
        if artifacts.review_evidence_file_hash is None:
            conn.rollback()
            raise ProfessionalModelConflictError(
                "final sign-off requires exact review-evidence bytes"
            )
        reviewed_values = {
            "workbook_sha256": artifacts.workbook_hash,
            "qa_report_sha256": artifacts.qa_report_hash,
            "review_evidence_sha256": artifacts.review_evidence_file_hash,
        }
        fingerprint = _canonical_hash(
            {
                "fingerprint_version": REVIEW_FINGERPRINT_VERSION,
                "approval_key": FINAL_SIGNOFF_KEY,
                "scope": FINAL_SIGNOFF_SCOPE,
                "reviewed_values": reviewed_values,
            }
        )
        parent = max(
            (
                int(event["event_id"])
                for event in current_events
                if event.get("state") == "approved"
            ),
            default=None,
        )
        prior_signoff = max(
            (
                event
                for event in current_events
                if event.get("approval_key") == FINAL_SIGNOFF_KEY
                and event.get("approval_scope") == FINAL_SIGNOFF_SCOPE
            ),
            key=lambda item: int(item["event_id"]),
            default=None,
        )
        event_id = insert_professional_model_review_event(
            conn,
            _event_base(
                artifacts,
                approval_key=FINAL_SIGNOFF_KEY,
                approval_scope=FINAL_SIGNOFF_SCOPE,
                reviewed_values=reviewed_values,
                reviewed_value_fingerprint=fingerprint,
                actor=actor,
                rationale=rationale,
                event_type="signoff",
                state="signed_off",
                parent_event_id=parent,
                supersedes_event_id=(
                    prior_signoff.get("event_id") if prior_signoff else None
                ),
            ),
        )
        try:
            confirmed = _refresh_actionable_artifact(
                artifacts,
                artifact_root=artifact_root,
            )
        except ProfessionalModelError:
            conn.rollback()
            raise
        artifacts = confirmed
        conn.commit()
        event = load_professional_model_review_event(conn, event_id)
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "event_id": event_id,
        "state": "signed_off",
        "normalized_state": "FULL",
        "workbook_sha256": artifacts.workbook_hash,
        "qa_report_sha256": artifacts.qa_report_hash,
        "review_evidence_sha256": artifacts.review_evidence_file_hash,
        "signoff_fingerprint": fingerprint,
        "signed_at": event.get("created_at") if event else None,
        "actor": actor,
        "rationale": rationale,
    }


def resolve_professional_model_download(
    ticker: str,
    *,
    expected_workbook_sha256: str,
    expected_model_run_id: int,
    artifact_root: str | Path | None = None,
) -> tuple[Path, dict[str, Any]]:
    """Return the exact verified workbook and public identity; never writes."""
    expected_hash = str(expected_workbook_sha256 or "").strip().lower()
    if _safe_hash(expected_hash) is None:
        raise ProfessionalModelValidationError(
            "expected_workbook_sha256 must be SHA-256"
        )
    try:
        expected_run_id = int(expected_model_run_id)
    except (TypeError, ValueError) as exc:
        raise ProfessionalModelValidationError(
            "expected_model_run_id must be a positive integer"
        ) from exc
    if expected_run_id <= 0:
        raise ProfessionalModelValidationError(
            "expected_model_run_id must be a positive integer"
        )
    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    if artifacts.issues:
        raise ProfessionalModelConflictError(
            "professional model download identity is unverified"
        )
    _require_full_lifecycle_run(artifacts, action="download")
    if expected_run_id != artifacts.model_run_id or expected_hash != artifacts.workbook_hash:
        raise ProfessionalModelConflictError(
            "professional model download expectation is stale or mismatched"
        )
    return artifacts.workbook_path, {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "filename": artifacts.workbook_path.name,
        "workbook_sha256": artifacts.workbook_hash,
        "workbook_bytes": artifacts.workbook_bytes,
    }


def _redact_workbook_text(value: str, *, limit: int) -> str:
    """Return workbook text without leaking embedded local filesystem paths."""
    text = value[:limit]
    stripped = text.strip()
    is_web_uri = bool(re.match(r"^https?://", stripped, re.I))
    path_probe = re.sub(r"https?://\S+", "", stripped, flags=re.I)
    path_like = bool(
        not is_web_uri
        and (
            _WINDOWS_ABSOLUTE_RE.search(path_probe)
            or re.search(r"(?<![A-Za-z0-9_])\\\\[^\\\s]+\\[^\\\s]+", path_probe)
            or re.search(r"(?<![:/A-Za-z0-9_])/(?:[^/\s]+/)+[^/\s]+", path_probe)
            or re.search(r"\[[^\]]+\.(?:xlsx?|xlsm|csv|json)\]", path_probe, re.I)
            or re.search(r"(?<![A-Za-z0-9_])[A-Za-z]:[\\/]", path_probe)
        )
    )
    return "[redacted filesystem path]" if path_like else text


def _public_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bytes):
        return "[binary value]"
    return _redact_workbook_text(str(value), limit=MAX_CELL_TEXT)


def _window_intersects_range(
    bounds: tuple[int, int, int, int],
    *,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> bool:
    min_column, min_row, max_column, max_row = bounds
    return not (
        max_row < start_row
        or min_row > end_row
        or max_column < start_column
        or min_column > end_column
    )


def build_professional_model_sheet_payload(
    ticker: str,
    sheet_name: str,
    *,
    start_row: int = 1,
    start_column: int = 1,
    row_limit: int = 100,
    column_limit: int = 20,
    artifact_root: str | Path | None = None,
) -> dict[str, Any]:
    """Return a bounded formula-and-cache view of one manifest-listed sheet."""
    try:
        start_row = int(start_row)
        start_column = int(start_column)
        row_limit = int(row_limit)
        column_limit = int(column_limit)
    except (TypeError, ValueError) as exc:
        raise ProfessionalModelValidationError(
            "sheet pagination values must be integers"
        ) from exc
    if start_row < 1 or start_column < 1:
        raise ProfessionalModelValidationError(
            "start_row and start_column must be positive"
        )
    if not 1 <= row_limit <= MAX_SHEET_ROWS:
        raise ProfessionalModelValidationError(
            f"row_limit must be between 1 and {MAX_SHEET_ROWS}"
        )
    if not 1 <= column_limit <= MAX_SHEET_COLUMNS:
        raise ProfessionalModelValidationError(
            f"column_limit must be between 1 and {MAX_SHEET_COLUMNS}"
        )
    if row_limit * column_limit > MAX_SHEET_CELLS:
        raise ProfessionalModelValidationError(
            f"requested window exceeds {MAX_SHEET_CELLS} cells"
        )
    requested_sheet = str(sheet_name or "").strip()
    if not requested_sheet or len(requested_sheet) > 100:
        raise ProfessionalModelValidationError("invalid sheet_name")

    artifacts = discover_professional_model_artifacts(ticker, artifact_root=artifact_root)
    manifest_sheets = [str(item) for item in artifacts.manifest.get("sheet_order") or []]
    _require_verified_identity(artifacts)
    if requested_sheet not in manifest_sheets:
        raise ProfessionalModelNotFoundError(
            f"professional model sheet not found: {requested_sheet}"
        )

    from openpyxl import load_workbook
    from openpyxl.utils.cell import get_column_letter, range_boundaries

    try:
        workbook_snapshot = artifacts.workbook_path.read_bytes()
    except OSError as exc:
        raise ProfessionalModelConflictError(
            "professional model workbook is unreadable"
        ) from exc
    if sha256(workbook_snapshot).hexdigest() != artifacts.workbook_hash:
        raise ProfessionalModelConflictError(
            "professional model changed during sheet retrieval"
        )
    formula_book = load_workbook(
        BytesIO(workbook_snapshot),
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    cached_book = load_workbook(
        BytesIO(workbook_snapshot),
        read_only=False,
        data_only=True,
        keep_links=False,
    )
    try:
        if (
            requested_sheet not in formula_book.sheetnames
            or requested_sheet not in cached_book.sheetnames
        ):
            raise ProfessionalModelConflictError(
                "workbook sheet order contradicts the manifest"
            )
        formula_sheet = formula_book[requested_sheet]
        cached_sheet = cached_book[requested_sheet]
        max_row = max(int(formula_sheet.max_row or 0), 1)
        max_column = max(int(formula_sheet.max_column or 0), 1)
        end_row = min(start_row + row_limit - 1, max_row)
        end_column = min(start_column + column_limit - 1, max_column)
        in_bounds = start_row <= max_row and start_column <= max_column

        classification_index: dict[str, dict[str, Any]] = {}
        for item in artifacts.manifest.get("cell_classifications") or []:
            if isinstance(item, Mapping) and item.get("sheet") == requested_sheet:
                coordinate = str(item.get("cell") or "")
                if coordinate:
                    classification_index[coordinate] = {
                        "kind": item.get("kind"),
                        "contract_version": item.get("contract_version"),
                    }
        lineage_index: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for item in artifacts.manifest.get("line_cell_mappings") or []:
            if isinstance(item, Mapping) and item.get("sheet") == requested_sheet:
                coordinate = str(item.get("cell") or "")
                if coordinate:
                    lineage_index[coordinate].append(
                        {
                            "canonical_key": item.get("canonical_key"),
                            "period_key": item.get("period_key"),
                            "scenario_key": item.get("scenario_key"),
                            "contract_version": item.get("contract_version"),
                        }
                    )

        merged_ranges: list[str] = []
        merged_by_coordinate: dict[str, str] = {}
        if in_bounds:
            for merged in formula_sheet.merged_cells.ranges:
                merged_text = str(merged)
                bounds = range_boundaries(merged_text)
                if not _window_intersects_range(
                    bounds,
                    start_row=start_row,
                    end_row=end_row,
                    start_column=start_column,
                    end_column=end_column,
                ):
                    continue
                merged_ranges.append(merged_text)
                min_column, min_row, max_column_range, max_row_range = bounds
                for row_index in range(
                    max(start_row, min_row), min(end_row, max_row_range) + 1
                ):
                    for column_index in range(
                        max(start_column, min_column),
                        min(end_column, max_column_range) + 1,
                    ):
                        merged_by_coordinate[
                            f"{get_column_letter(column_index)}{row_index}"
                        ] = merged_text


        cells: list[dict[str, Any]] = []
        if in_bounds:
            for row_index in range(start_row, end_row + 1):
                row_hidden = bool(formula_sheet.row_dimensions[row_index].hidden)
                for column_index in range(start_column, end_column + 1):
                    coordinate = f"{get_column_letter(column_index)}{row_index}"
                    formula_cell = formula_sheet[coordinate]
                    cached_cell = cached_sheet[coordinate]
                    raw_value = formula_cell.value
                    formula = (
                        _redact_workbook_text(raw_value, limit=MAX_CELL_TEXT)
                        if isinstance(raw_value, str) and raw_value.startswith("=")
                        else None
                    )
                    comment = formula_cell.comment
                    cells.append(
                        {
                            "coordinate": coordinate,
                            "cached_value": _public_cell_value(cached_cell.value),
                            "formula": formula,
                            "number_format": _redact_workbook_text(
                                str(formula_cell.number_format or "General"),
                                limit=500,
                            ),
                            "comment": (
                                {
                                    "author": _redact_workbook_text(
                                        str(comment.author or ""), limit=200
                                    ),
                                    "text": _redact_workbook_text(
                                        str(comment.text or ""), limit=MAX_COMMENT_TEXT
                                    ),
                                }
                                if comment is not None
                                else None
                            ),
                            "lineage": lineage_index.get(coordinate, []),
                            "classification": classification_index.get(coordinate),
                            "row_hidden": row_hidden,
                            "column_hidden": bool(
                                formula_sheet.column_dimensions[
                                    get_column_letter(column_index)
                                ].hidden
                            ),
                            "merged_range": merged_by_coordinate.get(coordinate),
                        }
                    )

        cells_payload_bytes = len(_canonical_json(cells).encode("utf-8"))
        if cells_payload_bytes > MAX_SHEET_TEXT_BYTES:
            raise ProfessionalModelValidationError(
                "requested sheet window exceeds the response-size cap; request a smaller window"
            )

        returned_rows = end_row - start_row + 1 if in_bounds else 0
        returned_columns = end_column - start_column + 1 if in_bounds else 0
        next_row = end_row + 1 if in_bounds and end_row < max_row else None
        next_column = (
            end_column + 1 if in_bounds and end_column < max_column else None
        )
        return {
            "ticker": artifacts.ticker,
            "model_run_id": artifacts.model_run_id,
            "workbook_sha256": artifacts.workbook_hash,
            "sheet_name": requested_sheet,
            "dimensions": {
                "range": formula_sheet.calculate_dimension(),
                "max_row": max_row,
                "max_column": max_column,
            },
            "pagination": {
                "payload_bytes": cells_payload_bytes,
                "payload_bytes_cap": MAX_SHEET_TEXT_BYTES,
                "start_row": start_row,
                "start_column": start_column,
                "row_limit": row_limit,
                "column_limit": column_limit,
                "returned_rows": returned_rows,
                "returned_columns": returned_columns,
                "returned_cells": len(cells),
                "next_row": next_row,
                "next_column": next_column,
                "has_more_rows": next_row is not None,
                "has_more_columns": next_column is not None,
            },
            "row_visibility": [
                {
                    "row": row_index,
                    "hidden": bool(formula_sheet.row_dimensions[row_index].hidden),
                }
                for row_index in (
                    range(start_row, end_row + 1) if in_bounds else []
                )
            ],
            "column_visibility": [
                {
                    "column": column_index,
                    "letter": get_column_letter(column_index),
                    "hidden": bool(
                        formula_sheet.column_dimensions[
                            get_column_letter(column_index)
                        ].hidden
                    ),
                }
                for column_index in (
                    range(start_column, end_column + 1) if in_bounds else []
                )
            ],
            "merged_ranges": sorted(merged_ranges),
            "cells": cells,
            "sheet_audit": _sheet_audit_findings(artifacts),
        }
    finally:
        formula_book.close()
        cached_book.close()


# Compatibility alias for callers that use the transport-oriented verb.
get_professional_model_sheet = build_professional_model_sheet_payload


def _current_db_path() -> Path:
    override = os.getenv("ALPHA_POD_DB_PATH")
    path = Path(override).resolve() if override else Path(DB_PATH).resolve()
    if not path.is_file():
        raise ProfessionalModelConflictError(
            "professional-model rebuild database is unavailable"
        )
    return path


def _resolve_rebuild_source_workbook(
    artifacts: ProfessionalModelArtifacts,
) -> Path:
    raw = str(artifacts.qa_report.get("source_workbook") or "").strip()
    if not raw:
        raise ProfessionalModelConflictError(
            "QA report does not identify a source workbook"
        )
    source = Path(raw)
    source = source.resolve() if source.is_absolute() else (ROOT_DIR / source).resolve()
    if not _is_under(source, ROOT_DIR) or not source.is_file():
        raise ProfessionalModelConflictError(
            "source workbook is outside the trusted repository or unavailable"
        )
    expected_hash = _safe_hash(artifacts.manifest.get("source_hash"))
    if expected_hash is None or _sha256_file(source) != expected_hash:
        raise ProfessionalModelConflictError(
            "source workbook no longer matches the run-bound SHA-256"
        )
    return source


def _resolve_frozen_valuation_json(
    artifacts: ProfessionalModelArtifacts,
) -> tuple[Path, dict[str, str]]:
    from src.stage_04_pipeline.professional_model_adapter import (
        load_frozen_valuation_document,
    )

    output_root = Path(OUTPUT_DIR).resolve()
    source_hash = str(artifacts.manifest.get("source_hash") or "")
    candidates: list[tuple[str, str, Path, str]] = []
    for directory, child_dirs, filenames in os.walk(output_root):
        child_dirs[:] = sorted(
            name
            for name in child_dirs
            if name not in {"professional_models", "professional_model_rebuilds"}
        )
        directory_path = Path(directory)
        if not _is_under(directory_path, output_root):
            continue
        for filename in sorted(filenames):
            match = re.search(
                r"(\d{8}T\d{6}Z)-valuation\.json$",
                filename,
                re.IGNORECASE,
            )
            if match is None or "latest" in filename.lower():
                continue
            candidate = (directory_path / filename).resolve()
            if not _is_under(candidate, output_root) or not candidate.is_file():
                continue
            try:
                load_frozen_valuation_document(
                    candidate,
                    ticker=artifacts.ticker,
                    requested_run_id=artifacts.model_run_id,
                    expected_source_hash=source_hash,
                )
            except Exception:
                continue
            relative_name = candidate.relative_to(output_root).as_posix()
            candidates.append(
                (match.group(1).upper(), relative_name, candidate, _sha256_file(candidate))
            )
    if not candidates:
        raise ProfessionalModelConflictError(
            "no explicit frozen valuation JSON matches the active source run"
        )
    candidates.sort(key=lambda item: (item[0], item[1]))
    newest_timestamp = candidates[-1][0]
    newest = [item for item in candidates if item[0] == newest_timestamp]
    if len({item[3] for item in newest}) > 1:
        raise ProfessionalModelConflictError(
            "multiple distinct frozen valuation inputs share the newest timestamp"
        )
    selected = newest[-1]
    return selected[2], {
        "sha256": selected[3],
        "timestamp": selected[0],
    }


def _build_approved_forecast_bundle(
    artifacts: ProfessionalModelArtifacts,
    *,
    db_path: Path,
) -> tuple[Any | None, dict[str, Any]]:
    review_state = _review_state_for_artifact(artifacts)
    required_count = int(review_state.get("required_count") or 0)
    approved_count = int(review_state.get("approved_count") or 0)
    counts = {
        "required": required_count,
        "approved": approved_count,
        "consumed": 0,
        "consumed_approval_events": [],
    }
    if (
        required_count == 0
        or approved_count != required_count
        or not review_state.get("all_approved")
    ):
        return None, counts

    from src.contracts.professional_financial_model import PeriodType
    from src.stage_02_valuation.integrated_financial_forecast import (
        DRIVER_SPECS,
        DriverPath,
        ScenarioDriverSet,
        build_complete_scenario_forecasts,
    )
    from src.stage_02_valuation.integrated_financial_model import (
        build_historical_financial_model_from_sqlite,
    )

    historical = build_historical_financial_model_from_sqlite(
        db_path,
        ticker=artifacts.ticker,
        run_id=artifacts.model_run_id,
    )
    fiscal_periods = [
        period
        for period in historical.result.period_axis.periods
        if period.period_type is PeriodType.FISCAL_YEAR
    ]
    if not fiscal_periods:
        raise ProfessionalModelConflictError(
            "approved rebuild requires a fiscal-year historical anchor"
        )

    by_scenario: dict[str, list[Any]] = defaultdict(list)
    for row in review_state.get("requirements") or []:
        if row.get("status") != "approved":
            raise ProfessionalModelConflictError(
                "approved rebuild bundle changed during assembly"
            )
        scenario = str(row.get("scenario") or "")
        driver = str(row.get("driver") or "")
        values = row.get("reviewed_values")
        event_id = int(row.get("current_event_id") or 0)
        reviewed_value_fingerprint = str(
            row.get("reviewed_value_fingerprint") or ""
        )
        approval_artifact_identity = row.get("approval_artifact_identity")
        requirement_hash = _safe_hash(row.get("requirement_hash"))
        requirement_contract = row.get("requirement_contract")
        review_context = row.get("review_context")
        expected_approval_artifact_identity = {
            "model_run_id": artifacts.model_run_id,
            "source_sha256": _safe_hash(artifacts.manifest.get("source_hash")),
            "model_input_sha256": _safe_hash(artifacts.manifest.get("model_input_hash")),
            "result_sha256": _safe_hash(artifacts.manifest.get("result_hash")),
            "workbook_sha256": artifacts.workbook_hash,
        }
        if (
            scenario not in {"Base", "Upside", "Downside"}
            or driver not in DRIVER_SPECS
            or not isinstance(values, list)
            or len(values) != 5
            or event_id <= 0
            or not _SHA256_RE.fullmatch(reviewed_value_fingerprint)
            or approval_artifact_identity != expected_approval_artifact_identity
            or requirement_hash is None
            or not isinstance(requirement_contract, Mapping)
            or requirement_hash != requirement_contract.get("requirement_hash")
            or requirement_hash != _canonical_hash({
                key: value for key, value in requirement_contract.items()
                if key != "requirement_hash"
            })
            or not isinstance(review_context, Mapping)
        ):
            raise ProfessionalModelConflictError(
                "approved review event cannot be converted into a complete driver path"
            )
        spec = DRIVER_SPECS[driver]
        source_ref = f"professional_model_review_event:{event_id}"
        method = "pm_approved_exact_path_v1"
        driver_fingerprint = DriverPath.fingerprint_for(
            key=driver,
            values=values,
            unit=spec.unit,
            source_ref=source_ref,
            method=method,
        )
        by_scenario[scenario].append(
            DriverPath(
                key=driver,
                values=tuple(float(item) for item in values),
                unit=spec.unit,
                source_ref=source_ref,
                method=method,
                approval_ref=f"professional_model_review_event:{event_id}",
                current_driver_fingerprint=driver_fingerprint,
                approved_driver_fingerprint=driver_fingerprint,
            )
        )
        counts["consumed_approval_events"].append(
            {
                "scope": row["scope"],
                "approval_key": row["approval_key"],
                "event_id": event_id,
                "reviewed_value_fingerprint": reviewed_value_fingerprint,
                "reviewed_values": [float(item) for item in values],
                "approval_artifact_identity": approval_artifact_identity,
                "requirement_hash": requirement_hash,
                "requirement_contract": dict(requirement_contract),
                "review_context": dict(review_context),
            }
        )

    scenarios = tuple(
        ScenarioDriverSet(
            scenario_key=scenario,
            paths=tuple(by_scenario.get(scenario, [])),
            parent_scenario_key=None if scenario == "Base" else "Base",
        )
        for scenario in ("Base", "Upside", "Downside")
    )
    bundle = build_complete_scenario_forecasts(
        historical,
        scenarios,
        last_historical_period_end=fiscal_periods[-1].end_date,
    )
    counts["consumed"] = approved_count
    counts["consumed_approval_events"] = sorted(
        counts["consumed_approval_events"],
        key=lambda item: (item["scope"], item["approval_key"]),
    )
    return bundle, counts


def _persist_rebuild_review_evidence(
    artifacts: ProfessionalModelArtifacts,
    *,
    consumed_events: Sequence[Mapping[str, Any]],
    built_manifest: Mapping[str, Any],
    built_workbook_hash: str,
    built_output_dir: Path,
) -> dict[str, Any] | None:
    if not consumed_events:
        return None
    expected_identities = _expected_driver_approval_identities()
    consumed_by_identity = {
        (str(item.get("scope") or ""), str(item.get("approval_key") or "")): item
        for item in consumed_events
    }
    if set(consumed_by_identity) != expected_identities:
        raise ProfessionalModelConflictError(
            "rebuild approval evidence does not cover the required driver inventory"
        )

    from db.loader import list_professional_model_review_events

    with get_connection() as conn:
        conn.execute("BEGIN IMMEDIATE")
        current_events = list_professional_model_review_events(
            conn,
            ticker=artifacts.ticker,
            model_run_id=artifacts.model_run_id,
        )
        current_state = _review_state_for_artifact(
            artifacts,
            events=current_events,
        )
        if not current_state.get("all_approved"):
            conn.rollback()
            raise ProfessionalModelConflictError(
                "professional-model approval inventory changed during rebuild"
            )
        rows_by_identity = {
            (row["scope"], row["approval_key"]): row
            for row in current_state.get("requirements") or []
        }
        for identity, consumed in consumed_by_identity.items():
            current = rows_by_identity.get(identity)
            if (
                current is None
                or current.get("status") != "approved"
                or int(current.get("current_event_id") or 0)
                != int(consumed.get("event_id") or 0)
                or current.get("reviewed_value_fingerprint")
                != consumed.get("reviewed_value_fingerprint")
                or _canonical_json(current.get("reviewed_values"))
                != _canonical_json(consumed.get("reviewed_values"))
                or _canonical_json(current.get("approval_artifact_identity"))
                != _canonical_json(consumed.get("approval_artifact_identity"))
                or current.get("requirement_hash") != consumed.get("requirement_hash")
                or _canonical_json(current.get("requirement_contract"))
                != _canonical_json(consumed.get("requirement_contract"))
                or _canonical_json(current.get("review_context"))
                != _canonical_json(consumed.get("review_context"))
            ):
                conn.rollback()
                raise ProfessionalModelConflictError(
                    "a consumed professional-model approval changed during rebuild"
                )

        normalized_events = sorted(
            [dict(item) for item in consumed_events],
            key=lambda item: (item["scope"], item["approval_key"]),
        )
        required_inventory = [
            {
                "scope": scope,
                "approval_key": approval_key,
                "requirement_hash": consumed_by_identity[(scope, approval_key)]["requirement_hash"],
            }
            for scope, approval_key in sorted(expected_identities)
        ]
        payload: dict[str, Any] = {
            "schema_version": "1.0.0",
            "fingerprint_version": REVIEW_FINGERPRINT_VERSION,
            "ticker": artifacts.ticker,
            "model_run_id": artifacts.model_run_id,
            "artifact_identity": {
                "source_sha256": built_manifest.get("source_hash"),
                "model_input_sha256": built_manifest.get("model_input_hash"),
                "result_sha256": built_manifest.get("result_hash"),
                "workbook_sha256": built_workbook_hash,
            },
            "approval_event_count": len(normalized_events),
            "required_approval_identities": required_inventory,
            "required_approval_inventory_hash": _canonical_hash(required_inventory),
            "approval_set_hash": _canonical_hash(normalized_events),
            "consumed_approval_events": normalized_events,
        }
        payload["review_evidence_hash"] = _canonical_hash(payload)
        evidence_path = _trusted_child(
            built_output_dir,
            REVIEW_EVIDENCE_NAME,
        )
        temporary_path = _trusted_child(
            built_output_dir,
            f".{REVIEW_EVIDENCE_NAME}.tmp",
        )
        temporary_path.write_text(
            json.dumps(payload, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        temporary_path.replace(evidence_path)
        evidence_file_hash = _sha256_file(evidence_path)
        conn.commit()
    return {
        "review_evidence_sha256": evidence_file_hash,
        "review_evidence_hash": payload["review_evidence_hash"],
        "approval_set_hash": payload["approval_set_hash"],
        "required_approval_inventory_hash": payload["required_approval_inventory_hash"],
        "approval_event_count": len(normalized_events),
    }


def rebuild_professional_model(
    ticker: str,
    *,
    model_run_id: int | None = None,
    actor: str = "api",
    rationale: str | None = None,
    tracker_run_id: str | None = None,
    artifact_root: str | Path | None = None,
    rebuild_root: str | Path | None = None,
) -> dict[str, Any]:
    """Build a run-bound workbook in staging; recalculation/QA stays separate."""
    artifacts = discover_professional_model_artifacts(
        ticker,
        artifact_root=artifact_root,
    )
    _require_verified_identity(artifacts)
    _require_full_lifecycle_run(artifacts, action="rebuild")
    if model_run_id is not None:
        try:
            requested_model_run_id = int(model_run_id)
        except (TypeError, ValueError) as exc:
            raise ProfessionalModelValidationError(
                "model_run_id must be a positive integer"
            ) from exc
        if requested_model_run_id <= 0:
            raise ProfessionalModelValidationError(
                "model_run_id must be a positive integer"
            )
        if requested_model_run_id != artifacts.model_run_id:
            raise ProfessionalModelConflictError(
                "requested model run is not the active professional-model artifact"
            )

    actor = _clean_text(actor, "actor", max_length=200)
    rationale = str(rationale).strip() if rationale is not None else None
    if rationale is not None and len(rationale) > 4_000:
        raise ProfessionalModelValidationError(
            "rationale must be at most 4000 characters"
        )
    job_id = str(tracker_run_id or uuid4().hex).strip()
    if (
        not re.fullmatch(r"[A-Za-z0-9_-]{1,100}", job_id)
        or ".." in job_id
    ):
        raise ProfessionalModelValidationError("invalid tracker_run_id")

    db_path = _current_db_path()
    source_workbook = _resolve_rebuild_source_workbook(artifacts)
    valuation_json, valuation_identity = _resolve_frozen_valuation_json(artifacts)
    forecast_bundle, approval_counts = _build_approved_forecast_bundle(
        artifacts,
        db_path=db_path,
    )

    staging_root = Path(
        rebuild_root or PROFESSIONAL_MODEL_REBUILD_ROOT
    ).resolve()
    if rebuild_root is None and not _is_under(staging_root, OUTPUT_DIR):
        raise ProfessionalModelConflictError(
            "configured rebuild staging root escapes the output directory"
        )
    job_root = _trusted_child(staging_root, job_id)
    staging_root.mkdir(parents=True, exist_ok=True)
    try:
        job_root.mkdir(exist_ok=False)
    except FileExistsError as exc:
        raise ProfessionalModelConflictError(
            "rebuild staging directory is already reserved"
        ) from exc

    from src.stage_04_pipeline.professional_model_adapter import (
        build_professional_model_v2,
    )

    built = build_professional_model_v2(
        ticker=artifacts.ticker,
        db_path=db_path,
        run_id=artifacts.model_run_id,
        workbook_path=source_workbook,
        valuation_json=valuation_json,
        output_dir=job_root,
        forecast_bundle=forecast_bundle,
    )
    if (
        not _is_under(built.output_dir, job_root)
        or not _is_under(built.workbook_path, job_root)
        or not _is_under(built.manifest_path, job_root)
    ):
        raise ProfessionalModelConflictError(
            "stage-04 builder returned an artifact outside the staging root"
        )
    manifest = built.manifest.model_dump(mode="json")
    workbook_hash = _sha256_file(built.workbook_path)
    workbook_bytes = built.workbook_path.stat().st_size
    review_evidence = _persist_rebuild_review_evidence(
        artifacts,
        consumed_events=approval_counts["consumed_approval_events"],
        built_manifest=manifest,
        built_workbook_hash=workbook_hash,
        built_output_dir=built.output_dir,
    )
    return {
        "ticker": artifacts.ticker,
        "model_run_id": artifacts.model_run_id,
        "tracker_run_id": job_id,
        "status": "built_calculation_pending",
        "normalized_state": "UNVERIFIED",
        "decision_readiness": False,
        "artifact_identity": {
            "source_sha256": manifest.get("source_hash"),
            "model_input_sha256": manifest.get("model_input_hash"),
            "result_sha256": manifest.get("result_hash"),
            "manifest_sha256": manifest.get("manifest_hash"),
            "workbook_sha256": workbook_hash,
            "workbook_bytes": workbook_bytes,
        },
        "calculation_verification": {
            "status": "PENDING",
            "verified": False,
            "reason": "stage04_builder_requires_downstream_recalculation_and_qa",
        },
        "review_inputs": {
            "required": approval_counts["required"],
            "approved": approval_counts["approved"],
            "consumed": approval_counts["consumed"],
            "evidence": review_evidence,
            "mode": (
                "pm_approved_exact_paths"
                if approval_counts["consumed"]
                else "diagnostic_pm_gated_paths"
            ),
        },
        "frozen_inputs": {
            "source_sha256": artifacts.manifest.get("source_hash"),
            "valuation_json_sha256": valuation_identity["sha256"],
            "valuation_timestamp": valuation_identity["timestamp"],
        },
        "blocker_count": len(manifest.get("blockers") or []),
        "warning_count": len(manifest.get("warnings") or []),
        "blockers": sorted(str(item) for item in manifest.get("blockers") or []),
        "warnings": sorted(str(item) for item in manifest.get("warnings") or []),
        "promotion_ready": False,
        "required_next_action": "run_isolated_recalculation_and_publish_qa",
        "requested_by": actor,
        "rationale": rationale,
    }
