from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml

from config import CONFIG_PATH, ROOT_DIR
from db.schema import create_tables, get_connection
from src.contracts.assumption_policy import (
    DamodaranPolicyDraft,
    ValuationPolicy,
    ValuationPolicyGlobalDefaults,
    ValuationPolicyPreview,
)


DAMODARAN_DROP_DIR = ROOT_DIR / "data" / "damodaran_drop"


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _config_wacc_defaults() -> dict[str, float]:
    if not CONFIG_PATH.exists():
        return {"risk_free_rate": 0.045, "equity_risk_premium": 0.05}
    data = yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}
    wacc = data.get("wacc_params") or {}
    return {
        "risk_free_rate": float(wacc.get("risk_free_rate", 0.045)),
        "equity_risk_premium": float(wacc.get("equity_risk_premium", 0.05)),
    }


def _sector_defaults() -> dict[str, dict[str, float]]:
    from src.stage_02_valuation.input_assembler import SECTOR_DEFAULTS

    return {sector: dict(defaults) for sector, defaults in SECTOR_DEFAULTS.items()}


def bootstrap_valuation_policy() -> ValuationPolicy:
    return ValuationPolicy(
        actor="config_bootstrap",
        global_defaults=ValuationPolicyGlobalDefaults(**_config_wacc_defaults()),
        sector_defaults=_sector_defaults(),
        source_ref="config/config.yaml",
    )


def load_current_valuation_policy() -> ValuationPolicy:
    try:
        with get_connection() as conn:
            create_tables(conn)
            from db.loader import load_latest_valuation_policy_version

            row = load_latest_valuation_policy_version(conn)
    except Exception:
        row = None
    if not row:
        return bootstrap_valuation_policy()
    return ValuationPolicy(
        policy_id=row["id"],
        created_at=row["created_at"],
        actor=row["actor"],
        global_defaults=row["global_defaults"],
        sector_defaults=row["sector_defaults"],
        source_ref=row.get("source_ref"),
        notes=row.get("notes"),
    )


def save_valuation_policy(policy: ValuationPolicy, *, actor: str = "api") -> ValuationPolicy:
    payload = policy.model_copy(update={"actor": actor, "created_at": _now()})
    with get_connection() as conn:
        create_tables(conn)
        from db.loader import insert_valuation_policy_version

        policy_id = insert_valuation_policy_version(
            conn,
            {
                "created_at": payload.created_at,
                "actor": actor,
                "global_defaults_json": payload.global_defaults.model_dump_json(),
                "sector_defaults_json": json.dumps(payload.sector_defaults, separators=(",", ":")),
                "source_ref": payload.source_ref,
                "notes": payload.notes,
            },
        )
    return payload.model_copy(update={"policy_id": policy_id})


def preview_valuation_policy_edits(
    *,
    global_defaults: dict[str, Any] | None = None,
    sector_defaults: dict[str, dict[str, float]] | None = None,
) -> ValuationPolicyPreview:
    current = load_current_valuation_policy()
    proposed = current.model_copy(
        update={
            "global_defaults": ValuationPolicyGlobalDefaults(
                **{**current.global_defaults.model_dump(), **(global_defaults or {})}
            ),
            "sector_defaults": {
                **current.sector_defaults,
                **(sector_defaults or {}),
            },
            "created_at": _now(),
        }
    )
    changed: dict[str, dict[str, Any]] = {}
    if current.global_defaults != proposed.global_defaults:
        for key, value in proposed.global_defaults.model_dump().items():
            prior = current.global_defaults.model_dump().get(key)
            if prior != value:
                changed[f"global_defaults.{key}"] = {"prior": prior, "new": value}
    for sector, defaults in proposed.sector_defaults.items():
        prior_defaults = current.sector_defaults.get(sector, {})
        if prior_defaults != defaults:
            changed[f"sector_defaults.{sector}"] = {"prior": prior_defaults, "new": defaults}
    return ValuationPolicyPreview(current_policy=current, proposed_policy=proposed, changed_fields=changed)


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        text = str(value).strip().replace("%", "")
        number = float(text)
        if "%" in str(value) or number > 1.0:
            return number / 100.0
        return number
    except (TypeError, ValueError):
        return None


def _normalise_field(value: str) -> str:
    cleaned = str(value).strip().lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "rf": "risk_free_rate",
        "riskfree": "risk_free_rate",
        "risk_free": "risk_free_rate",
        "erp": "equity_risk_premium",
        "equity_risk_premium": "equity_risk_premium",
        "implied_premium": "equity_risk_premium",
        "terminal_growth_rate": "terminal_growth",
    }
    return aliases.get(cleaned, cleaned)


def _source_kind(path: Path) -> str:
    name = path.name.lower()
    if "erp" in name or "impl" in name:
        return "equity_risk_premium"
    if "risk" in name or "ctry" in name:
        return "country_risk"
    if "rating" in name or "spread" in name:
        return "synthetic_rating"
    return "damodaran"


def _drafts_from_rows(path: Path, rows: list[dict[str, Any]]) -> list[DamodaranPolicyDraft]:
    drafts: list[DamodaranPolicyDraft] = []
    for idx, row in enumerate(rows, start=1):
        lowered = {str(k).strip().lower(): v for k, v in row.items()}
        field_raw = lowered.get("field") or lowered.get("assumption") or lowered.get("metric")
        value_raw = lowered.get("value") or lowered.get("premium") or lowered.get("rate")
        if not field_raw:
            for key in row:
                normalised = _normalise_field(key)
                if normalised in {"risk_free_rate", "equity_risk_premium", "terminal_growth"}:
                    field_raw = normalised
                    value_raw = row[key]
                    break
        field = _normalise_field(field_raw or "")
        value = _as_float(value_raw)
        if field not in {"risk_free_rate", "equity_risk_premium", "terminal_growth"} or value is None:
            continue
        row_key = str(lowered.get("date") or lowered.get("year") or idx)
        drafts.append(
            DamodaranPolicyDraft(
                source_file=path.name,
                source_kind=_source_kind(path),
                row_key=row_key,
                field=field,
                value=value,
                unit=lowered.get("unit") or "rate",
                source_date=lowered.get("date") or lowered.get("as_of_date"),
                raw={str(k): v for k, v in row.items()},
            )
        )
    return drafts


def _drafts_from_csv(path: Path) -> list[DamodaranPolicyDraft]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return _drafts_from_rows(path, list(csv.DictReader(handle)))


def _drafts_from_xlsx(path: Path) -> list[DamodaranPolicyDraft]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    mapped = [
        {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers) and headers[idx]}
        for row in rows[1:]
    ]
    return _drafts_from_rows(path, mapped)


def parse_damodaran_drop_folder(drop_dir: Path = DAMODARAN_DROP_DIR) -> dict[str, Any]:
    drop_dir.mkdir(parents=True, exist_ok=True)
    parsed: list[DamodaranPolicyDraft] = []
    rejected: list[dict[str, str]] = []
    for path in sorted(drop_dir.iterdir()):
        if path.is_dir():
            continue
        if path.suffix.lower() not in {".csv", ".xlsx"}:
            rejected.append({"source_file": path.name, "reason": "unsupported file type; use csv or xlsx"})
            continue
        try:
            if path.suffix.lower() == ".xlsx":
                parsed.extend(_drafts_from_xlsx(path))
            else:
                parsed.extend(_drafts_from_csv(path))
        except Exception as exc:
            rejected.append({"source_file": path.name, "reason": str(exc)})
    with get_connection() as conn:
        create_tables(conn)
        from db.loader import load_damodaran_policy_drafts, upsert_damodaran_policy_drafts

        count = upsert_damodaran_policy_drafts(conn, [draft.model_dump() for draft in parsed])
        drafts = [
            DamodaranPolicyDraft(
                draft_id=row["id"],
                created_at=row["created_at"],
                source_file=row["source_file"],
                source_kind=row["source_kind"],
                row_key=row["row_key"],
                field=row["field"],
                value=row["value"],
                unit=row.get("unit"),
                source_date=row.get("source_date"),
                status=row.get("status", "draft"),
                raw=row.get("raw") or {},
            ).model_dump()
            for row in load_damodaran_policy_drafts(conn, status="draft")
        ]
    return {
        "drop_dir": str(drop_dir),
        "parsed_count": count,
        "drafts": drafts,
        "rejected_files": rejected,
    }
