"""Contract tests for the additive professional-model workbook renderer."""

from __future__ import annotations

from dataclasses import replace
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
import pytest

from scripts.manual.recalculate_excel_isolated import (
    FormulaCellEvidence,
    formula_text_hash,
)
from src.contracts.professional_financial_model import CellKind, PROFESSIONAL_WORKBOOK_SHEETS
from src.stage_04_pipeline.professional_model_workbook import (
    ComparableCompany,
    HistoricalSourceCell,
    ModelLine,
    NormalizedProfessionalWorkbookPayload,
    ScenarioForecast,
    SourceWorkbookRun,
    TypedAvailability,
    render_professional_model_workbook,
)


SCENARIOS = ("base", "upside", "downside")


def _payload(*, source_status: str = "ready") -> NormalizedProfessionalWorkbookPayload:
    historical_periods = ("FY22", "FY23", "FY24", "FY25", "LTM")
    forecast_periods = ("FY26E", "FY27E", "FY28E", "FY29E", "FY30E")
    sheets = (
        "Income_Statement",
        "Balance_Sheet",
        "Cash_Flow",
        "Working_Capital",
        "PP&E_Intangibles",
        "Debt_Cash_Interest",
        "Capital_Allocation",
        "Taxes",
        "Shares_EPS",
    )
    core_keys = {
        0: ("revenue", "Revenue"),
        1: ("ebit", "EBIT"),
        2: ("depreciation_amortization", "D&A"),
        3: ("capex", "Capital Expenditures"),
        4: ("change_in_net_working_capital", "Change in Net Working Capital"),
        5: ("cash", "Cash and Equivalents"),
        6: ("diluted_shares", "Diluted Shares"),
        7: ("total_assets", "Total Assets"),
        8: ("total_liabilities_and_equity", "Total Liabilities and Equity"),
        9: ("ending_cash", "Ending Cash per Cash Flow Statement"),
        10: ("cf.unlevered_fcf", "Integrated Unlevered Free Cash Flow"),
        11: ("cf.net_change_cash", "Net Change in Cash"),
    }
    lines: list[ModelLine] = []
    for index in range(72):
        sheet = sheets[index % len(sheets)]
        key, label = core_keys.get(index, (f"{sheet.lower()}_line_{index:02d}", f"Model line {index:02d}"))
        if key in {"cf.unlevered_fcf", "cf.net_change_cash"}:
            sheet = "Cash_Flow"
        value_index = 5 if key == "ending_cash" else 7 if key == "total_liabilities_and_equity" else index
        history = tuple(
            HistoricalSourceCell(
                period_key=period,
                value=float(1000 + value_index * 17 + period_index * 23),
                source_sheet="Financial Statements",
                source_cell=f"{chr(66 + period_index)}{index + 3}",
                source_row_id=f"fs:{index + 3}",
                source_formula=f"=SUM({chr(66 + period_index)}{index + 4}:{chr(66 + period_index)}{index + 5})"
                if index % 4 == 0
                else None,
            )
            for period_index, period in enumerate(historical_periods)
        )
        forecasts = tuple(
            ScenarioForecast(
                scenario_key=scenario,
                values=tuple(
                    (
                        period,
                        float(1200 + value_index * 19 + scenario_index * 71 + period_index * 31),
                    )
                    for period_index, period in enumerate(forecast_periods)
                ),
            )
            for scenario_index, scenario in enumerate(SCENARIOS)
        )
        lines.append(
            ModelLine(
                canonical_key=key,
                label=label,
                sheet=sheet,
                unit="USD mm" if key != "diluted_shares" else "mm shares",
                historical=history,
                scenario_forecasts=forecasts,
            )
        )

    return NormalizedProfessionalWorkbookPayload(
        ticker="MSFT",
        company_name="Microsoft Corporation",
        as_of_date=date(2026, 3, 31),
        currency="USD",
        unit_convention="USD mm except per-share data",
        source=SourceWorkbookRun(
            source_file="MSFT_Standard.xlsx",
            source_path=r"data\exports\MSFT_Standard.xlsx",
            source_hash="a" * 64,
            run_id=4,
            parser_version="ibm_standard_v4",
            status=source_status,
            fact_count=8601,
            formula_error_count=0 if source_status == "ready" else 24,
        ),
        historical_periods=historical_periods,
        forecast_periods=forecast_periods,
        lines=tuple(lines),
        valuation_inputs={
            "risk_free_rate": 0.0425,
            "beta": 0.94,
            "equity_risk_premium": 0.0500,
            "pre_tax_cost_of_debt": 0.043,
            "tax_rate": 0.18,
            "debt_value": 125_432.0,
            "equity_value": 2_860_000.0,
            "terminal_growth": 0.0275,
            "net_debt": 47_204.0,
            "current_basic_shares": 7428.4347,
            "current_fully_diluted_shares": 7450.0,
            "dcf_discount_exponent_1": 0.1260273973,
            "dcf_discount_exponent_2": 0.7506849315,
            "dcf_discount_exponent_3": 1.7520547945,
            "dcf_discount_exponent_4": 2.7534246575,
            "dcf_discount_exponent_5": 3.7534246575,
            "dcf_terminal_discount_exponent": 4.2520547945,
            "dcf_annual_fy26_fcff_base": 100.0,
            "dcf_ytd_fcff_base": 75.0,
            "dcf_stub_fcff_base": 25.0,
            "dcf_annual_fy26_fcff_upside": 110.0,
            "dcf_ytd_fcff_upside": 82.0,
            "dcf_stub_fcff_upside": 28.0,
            "dcf_annual_fy26_fcff_downside": 90.0,
            "dcf_ytd_fcff_downside": 70.0,
            "dcf_stub_fcff_downside": 20.0,
            "dcf_wacc_base": 0.08,
            "dcf_terminal_growth_base": 0.02,
            "dcf_wacc_upside": 0.075,
            "dcf_terminal_growth_upside": 0.025,
            "dcf_wacc_downside": 0.09,
            "dcf_terminal_growth_downside": 0.01,
            "current_price": 385.10,
            "diagnostic_comps_ev_ebitda_per_share": 518.384,
            "diagnostic_comps_ev_ebit_per_share": None,
            "diagnostic_comps_pe_per_share": 475.0856,
            "diagnostic_economic_profit_per_share": 217.69,
            "diagnostic_v1_gordon_per_share": 194.41,
            "diagnostic_v1_exit_per_share": 375.12,
            "diagnostic_reverse_dcf_implied_growth_pct_points": 16.9,
            "bridge_cash": 32_105.0,
            "bridge_short_term_investments": 40_000.0,
            "bridge_long_term_investments": 6_123.0,
            "bridge_gross_debt": 125_432.0,
            "bridge_total_borrowings": 100_000.0,
            "bridge_short_term_borrowings": 5_000.0,
            "bridge_current_long_term_debt": 5_000.0,
            "bridge_long_term_debt": 90_000.0,
            "bridge_current_lease_liabilities": 5_000.0,
            "bridge_long_term_lease_liabilities": 20_432.0,
            "bridge_lease_liabilities": 25_432.0,
            "bridge_minority_interest": 0.0,
            "bridge_pension_liability": 12_345.0,
        },
        availability={
            "segments": TypedAvailability("unavailable", "segment_data_absent", "No source-backed segment history."),
            "consensus": TypedAvailability("unavailable", "consensus_absent", "No frozen consensus snapshot."),
            "sotp": TypedAvailability("unavailable", "segment_data_absent", "SOTP requires segment evidence."),
        },
        backend_checks={
            "scenario.formula_first_gate": "PASS",
            "scenario.policy_gate": "PASS",
            "fcfe.state": "UNAVAILABLE",
            "fcfe.reason_code": "integrated_debt_interest_schedule_required",
            "fcfe.detail": "FCFE remains unavailable until the integrated debt, interest, tax, and net-borrowing schedules tie.",
            **(
                {
                    f"source.formula_error.{index:03d}.cell": f"Detailed Comps!{column}{row}"
                    for index, (row, column) in enumerate(
                        ((row, column) for row in range(3, 11) for column in ("AC", "AG", "AK")),
                        start=1,
                    )
                }
                if source_status != "ready"
                else {}
            ),
        },
        blockers=() if source_status == "ready" else ("source_formula_errors",),
    )


def _formula_count(workbook) -> int:
    return sum(
        1
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )


def test_renderer_builds_full_formula_driven_workbook_and_manifest(tmp_path: Path) -> None:
    path = tmp_path / "MSFT_professional_model.xlsx"
    manifest = render_professional_model_workbook(_payload(), path)
    workbook = load_workbook(path, data_only=False)

    assert tuple(workbook.sheetnames) == PROFESSIONAL_WORKBOOK_SHEETS
    assert manifest.sheet_order == PROFESSIONAL_WORKBOOK_SHEETS
    assert manifest.ticker == "MSFT"
    assert manifest.source_hash == "a" * 64
    assert _formula_count(workbook) >= 1100
    assert len(manifest.line_cell_mappings) == 72 * (5 + 5) * 3
    assert {mapping.scenario_key for mapping in manifest.line_cell_mappings} == set(SCENARIOS)
    assert len({(item.canonical_key, item.scenario_key, item.period_key) for item in manifest.line_cell_mappings}) == len(
        manifest.line_cell_mappings
    )

    historical = workbook["Historical_Data"]
    assert historical["C5"].comment is not None
    assert "Financial Statements!" in historical["C5"].comment.text
    assert "Run 4" in historical["C5"].comment.text
    assert workbook["Cover"]["B9"].value == "a" * 64
    assert workbook["Cover"]["B10"].value == 4
    assert workbook["Cover"]["B11"].value == "=Checks!$C$5"

    formulas = [
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert not any("[" in formula or "]" in formula for formula in formulas)
    assert not any(any(error in formula for error in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!")) for formula in formulas)
    assert not workbook._external_links
    expected_formula_hash = formula_text_hash(
        FormulaCellEvidence(
            sheet=worksheet.title,
            cell=cell.coordinate,
            formula_text=str(cell.value),
            cache_populated=False,
        )
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    assert manifest.expected_formula_text_hash == expected_formula_hash

    defined_names = set(workbook.defined_names)
    assert {"Model_Status", "Selected_Scenario", "Source_Hash", "Source_Run_ID", "WACC_Base", "DCF_Base_Per_Share"} <= defined_names
    assert {check.check_id for check in manifest.check_cells} >= {
        "source_preflight",
        "scenario_completeness",
        "forecast_completeness",
        "balance_sheet",
        "cash_flow_tie",
        "valuation_bridge",
    }

    rendered_history = next(
        item
        for item in manifest.line_cell_mappings
        if item.canonical_key == "revenue" and item.scenario_key == "base" and item.period_key == "FY22"
    )
    assert workbook[rendered_history.sheet][rendered_history.cell].comment is None

    integrated_fcf = next(
        item
        for item in manifest.line_cell_mappings
        if item.canonical_key == "cf.unlevered_fcf"
        and item.scenario_key == "base"
        and item.period_key == "FY26E"
    )
    dcf_formula = workbook["DCF"]["C14"].value
    assert integrated_fcf.cell in dcf_formula.replace("$", "")
    assert integrated_fcf.sheet in dcf_formula

    fy25 = next(
        item
        for item in manifest.line_cell_mappings
        if item.canonical_key == "revenue" and item.scenario_key == "base" and item.period_key == "FY25"
    )
    ltm = next(
        item
        for item in manifest.line_cell_mappings
        if item.canonical_key == "revenue" and item.scenario_key == "base" and item.period_key == "LTM"
    )
    fy26 = next(
        item
        for item in manifest.line_cell_mappings
        if item.canonical_key == "revenue" and item.scenario_key == "base" and item.period_key == "FY26E"
    )
    forecast_column = "".join(filter(str.isalpha, fy26.cell))
    forecast_row = int("".join(filter(str.isdigit, fy26.cell)))
    change_formula = workbook[fy26.sheet][f"{forecast_column}{forecast_row + 1}"].value
    assert f"/{fy25.cell}" in change_formula
    assert f"/{ltm.cell}" not in change_formula

    scenario_classification = next(
        item for item in manifest.cell_classifications if item.sheet == "Scenarios" and item.cell == "F5"
    )
    assert scenario_classification.kind is CellKind.CHECK
    assert workbook["Scenarios"]["F5"].comment is not None
    assert "calculated backend output" in workbook["Scenarios"]["F5"].comment.text.lower()
    assert workbook["Scenarios"]["A1"].value == "Scenario forecast outputs"

    assert workbook["Checks"]["D9"].value.startswith("=IF(")
    assert workbook["Checks"]["D9"].value.count("ABS(") == 15
    assert "COUNT(" in workbook["Checks"]["D9"].value
    assert workbook["Checks"]["D10"].value.startswith("=IF(")
    assert workbook["Checks"]["D10"].value.count("ABS(") == 15
    assert "COUNT(" in workbook["Checks"]["D10"].value
    assert "COUNTIF(" in workbook["Checks"]["C7"].value
    assert workbook["Checks"]["D8"].value.startswith("=COUNT(")

    assumption_rows = {
        workbook["Assumptions"][f"A{row}"].value: row
        for row in range(8, workbook["Assumptions"].max_row + 1)
    }
    frozen_keys = (
        "bridge_cash",
        "bridge_short_term_borrowings",
        "diagnostic_comps_ev_ebitda_per_share",
        "diagnostic_reverse_dcf_implied_growth_pct_points",
    )
    for key in frozen_keys:
        row = assumption_rows[key]
        classification = next(
            item
            for item in manifest.cell_classifications
            if item.sheet == "Assumptions" and item.cell == f"B{row}"
        )
        assert classification.kind is CellKind.SOURCE

    net_claims_formula = workbook["DCF"]["C22"].value.replace("$", "")
    for key in (
        "bridge_cash",
        "bridge_short_term_investments",
        "bridge_long_term_investments",
        "bridge_short_term_borrowings",
        "bridge_current_long_term_debt",
        "bridge_long_term_debt",
        "bridge_current_lease_liabilities",
        "bridge_long_term_lease_liabilities",
        "bridge_minority_interest",
        "bridge_pension_liability",
    ):
        assert f"B{assumption_rows[key]}" in net_claims_formula
    for reference_only in (
        "bridge_gross_debt",
        "bridge_total_borrowings",
        "bridge_lease_liabilities",
        "net_debt",
    ):
        assert f"B{assumption_rows[reference_only]}" not in net_claims_formula

    assert workbook["DCF"]["C30"].value == '=IF($B$3="AVAILABLE","END_YEAR","")'
    assert workbook["DCF"]["C31"].value.startswith('=IF(AND($B$3="AVAILABLE"')
    assert workbook["Valuation"]["A6"].value == "Diagnostic comps: EV / EBITDA"
    assert workbook["Valuation"]["E6"].value == "DIAGNOSTIC"
    assert workbook["Valuation"]["E7"].value == "UNAVAILABLE"
    assert workbook["Valuation"]["A12"].value == "Diagnostic reverse-DCF implied growth"
    assert workbook["Valuation"]["E13"].value == "UNAVAILABLE"
    assert "integrated_debt_interest_schedule_required" in workbook["Valuation"]["F13"].value
    assert workbook["Valuation"]["E14"].value == "UNAVAILABLE"
    assert workbook["Valuation"]["E15"].value == "UNAVAILABLE"
    assert workbook["Valuation"]["E33"].value == "REFERENCE ONLY; not counted"
    assert workbook["Valuation"]["B34"].value.startswith("=")
    assert 'DCF!$B$3="AVAILABLE"' not in workbook["Valuation"]["B34"].value
    assert workbook["Summary"]["A17"].value == "Diagnostic comps: EV / EBITDA"
    assert workbook["Summary"]["A23"].value == "Diagnostic reverse-DCF implied growth"
    assert workbook["Summary"]["A25"].value == "FCFE cross-check"
    assert workbook["Summary"]["D27"].value == "UNAVAILABLE"
    assert workbook["Cover"]["B12"].value == "UNVERIFIED"
    assert "SHA-bound verification sidecar" in workbook["Cover"]["B13"].value
    assert "explicit positive evidence" in workbook["Cover"]["B22"].value
    assert "Calculation_Status" in workbook.defined_names
    assert "recalculation_not_run" in manifest.blockers

    workbook_text = "|".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    banned_key = "comps_" + "implied_price"
    banned_name = "Comps_Implied_" + "Per_Share"
    assert banned_key not in workbook_text
    assert banned_name not in set(workbook.defined_names)


def test_renderer_is_deterministic_and_fails_closed(tmp_path: Path) -> None:
    first_path = tmp_path / "first.xlsx"
    second_path = tmp_path / "second.xlsx"
    payload = _payload(source_status="blocked")
    first_manifest = render_professional_model_workbook(payload, first_path)
    second_manifest = render_professional_model_workbook(payload, second_path)

    assert first_manifest.line_cell_mappings == second_manifest.line_cell_mappings
    assert first_manifest.defined_names == second_manifest.defined_names
    assert first_manifest.check_cells == second_manifest.check_cells
    assert first_manifest.manifest_hash == second_manifest.manifest_hash
    assert "source_preflight_blocked" in first_manifest.blockers

    workbook = load_workbook(first_path, data_only=False)
    assert workbook["Cover"]["B11"].value == "=Checks!$C$5"
    assert workbook["Summary"]["B4"].value == "=Checks!$C$5"
    assert workbook["Checks"]["C5"].value.startswith("=IF(")
    assert workbook["Checks"]["C6"].value == "BLOCKED"
    assert workbook["SOTP"]["B4"].value == "BLOCKING"
    assert workbook["Segment_Build"]["B4"].value == "UNAVAILABLE"
    assert workbook["Consensus_Bridge"]["B4"].value == "UNAVAILABLE"
    assert workbook["DCF"]["B3"].value == "AVAILABLE"
    assert workbook["DCF"]["E3"].value == "INELIGIBLE"
    assert workbook["Sensitivities"]["B3"].value == "AVAILABLE"
    assert workbook["Sensitivities"]["C7"].value.startswith('=IF($B$3<>"AVAILABLE","",')
    assert ',0)' not in workbook["Sensitivities"]["C7"].value
    assert 'DCF!$B$3="AVAILABLE"' in workbook["Valuation"]["B5"].value
    assert 'DCF!$B$3="AVAILABLE"' in workbook["Summary"]["B10"].value
    assert 'DCF!$B$3="AVAILABLE"' in workbook["Summary"]["D10"].value
    assert workbook["Valuation"]["B6"].value.startswith("='Assumptions'!")
    assert 'Cover!$B$8' not in workbook["Valuation"]["B6"].value
    assert workbook["Summary"]["B17"].value.startswith("='Assumptions'!")
    assert 'Cover!$B$8' not in workbook["Summary"]["B17"].value
    assert workbook["Summary"]["C10"].value.startswith("=IF(ISNUMBER('Assumptions'!")
    assert "ISNUMBER(" in workbook["Summary"]["B32"].value
    assert 'DCF!$B$3' not in workbook["Summary"]["C10"].value
    assert workbook["Summary"]["B32"].value.startswith("=IFERROR(")
    assert "MATCH(" in workbook["Summary"]["B32"].value
    assert 'DCF!$B$3' not in workbook["Summary"]["B32"].value
    for coordinate in ("B21", "B24", "B27", "B29", "B30", "B31", "B33", "B34", "B35", "B36"):
        assert 'DCF!$B$3="AVAILABLE"' not in workbook["Valuation"][coordinate].value
    for coordinate in ("B20", "B32", "B37"):
        assert 'DCF!$B$3="AVAILABLE"' in workbook["Valuation"][coordinate].value
    assert ',"")' in workbook["DCF"]["C25"].value
    assert ',0)' not in workbook["DCF"]["C25"].value
    assert workbook["Checks"]["C11"].value.startswith('=IF(DCF!$B$3<>"AVAILABLE","BLOCKED"')
    assert workbook["Checks"]["C12"].value == "BLOCKED"
    queue_codes = {
        workbook["PM_Review_Queue"][f"C{row}"].value: workbook["PM_Review_Queue"][f"D{row}"].value
        for row in range(5, workbook["PM_Review_Queue"].max_row + 1)
    }
    assert "Detailed Comps!AC3" in queue_codes["source.formula_error.001.cell"]
    assert "Detailed Comps!AK10" in queue_codes["source.formula_error.024.cell"]
    assert "calculation_cache_verification" in queue_codes
    assert "recalculation_not_run" not in queue_codes


def test_comps_blank_denominators_are_count_gated_and_evidence_is_partial(tmp_path: Path) -> None:
    payload = replace(
        _payload(),
        comparables=(
            ComparableCompany(
                ticker="BLNK",
                company_name="Blank Denominator Corp.",
                enterprise_value=100.0,
                equity_value=90.0,
                revenue=None,
                ebitda=None,
                net_income=None,
                share_price=10.0,
            ),
        ),
    )
    path = tmp_path / "blank-comps.xlsx"
    render_professional_model_workbook(payload, path)
    workbook = load_workbook(path, data_only=False)

    assert workbook["Comps"]["B3"].value == "PARTIAL"
    assert "Frozen exact-run/cached peer evidence" in workbook["Comps"]["C3"].value
    assert "LTM/NTM basis and as-of comparability are not independently verified" in workbook["Comps"]["C3"].value
    for coordinate in ("I6", "J6", "K6"):
        formula = workbook["Comps"][coordinate].value
        assert "IFERROR" in formula and "ISNUMBER" in formula
        assert '"NM"' in formula
    for coordinate in ("I7", "J7", "K7"):
        formula = workbook["Comps"][coordinate].value
        assert formula.startswith("=IF(COUNT(")
        assert "MEDIAN(" in formula
        assert ',"")' in formula
    assert "#NUM!" not in "|".join(
        str(cell.value)
        for row in workbook["Comps"].iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_degraded_wacc_and_pm_drivers_block_dcf_outputs(tmp_path: Path) -> None:
    payload = replace(
        _payload(),
        blockers=(
            "wacc_degraded:degraded_fallback",
            "pm_approval_required:Base:revenue_growth:queue:Base",
        ),
    )
    path = tmp_path / "degraded.xlsx"
    render_professional_model_workbook(payload, path)
    workbook = load_workbook(path, data_only=False)

    assert workbook["WACC"]["B3"].value == "DEGRADED"
    assert "wacc_degraded:degraded_fallback" in workbook["WACC"]["B4"].value
    assert workbook["DCF"]["B3"].value == "AVAILABLE"
    assert workbook["DCF"]["E3"].value == "NEEDS_PM_REVIEW"
    assert "pm_approval_required" in workbook["DCF"]["E4"].value
    assert ',"")' in workbook["DCF"]["C25"].value
    assert ',0)' not in workbook["DCF"]["C25"].value
    assert workbook["Valuation"]["E5"].value == "AVAILABLE / NEEDS_PM_REVIEW"


def test_missing_bridge_component_blocks_dcf_and_names_exact_gap(tmp_path: Path) -> None:
    payload = _payload()
    valuation_inputs = dict(payload.valuation_inputs)
    del valuation_inputs["bridge_pension_liability"]
    broken = replace(payload, valuation_inputs=valuation_inputs)
    path = tmp_path / "missing-bridge.xlsx"
    manifest = render_professional_model_workbook(broken, path)
    workbook = load_workbook(path, data_only=False)

    assert workbook["DCF"]["B3"].value == "BLOCKED"
    assert "bridge_pension_liability" in workbook["DCF"]["B4"].value
    assert workbook["DCF"]["C25"].value.endswith(',"")')
    assert any(
        blocker.startswith("missing_dcf_valuation_inputs:")
        and "bridge_pension_liability" in blocker
        for blocker in manifest.blockers
    )


def test_renderer_rejects_incomplete_scenario_forecast(tmp_path: Path) -> None:
    payload = _payload()
    line = payload.lines[0]
    broken_line = ModelLine(
        canonical_key=line.canonical_key,
        label=line.label,
        sheet=line.sheet,
        unit=line.unit,
        historical=line.historical,
        scenario_forecasts=line.scenario_forecasts[:-1],
    )
    broken = NormalizedProfessionalWorkbookPayload(
        **{**payload.__dict__, "lines": (broken_line, *payload.lines[1:])}
    )
    with pytest.raises(ValueError, match="base, upside, and downside"):
        render_professional_model_workbook(broken, tmp_path / "broken.xlsx")


def test_workflow_truth_table_and_selector_fail_closed_controls(tmp_path: Path) -> None:
    path = tmp_path / "truth-table.xlsx"
    render_professional_model_workbook(_payload(), path)
    workbook = load_workbook(path, data_only=False)

    aggregate = workbook["Checks"]["C5"].value
    assert '"FULL"' in aggregate
    assert '"NEEDS_PM_REVIEW"' in aggregate
    assert '"UNVERIFIED"' in aggregate
    assert '"PARTIAL"' in aggregate
    assert 'COUNTIF(C6:C17,"PASS")=ROWS(C6:C17)' in aggregate
    assert workbook["Checks"]["C14"].value == (
        '=IF(OR(Assumptions!$B$4="base",Assumptions!$B$4="upside",'
        'Assumptions!$B$4="downside"),"PASS","FAIL")'
    )
    assert workbook["Checks"]["C15"].value == "UNVERIFIED"
    assert workbook["Cover"]["B12"].value == "UNVERIFIED"
    assert workbook["Summary"]["C5"].value.endswith('),"PASS","FAIL")')
    for coordinate in ("B31", "B32", "B33", "B34", "B35"):
        formula = workbook["Summary"][coordinate].value
        assert formula.startswith("=IFERROR(")
        assert '$C$5="PASS"' in formula
    assert workbook["Summary"]["B64"].value == "PARTIAL"
    assert "no narrative has been invented" in workbook["Summary"]["C64"].value


def test_source_presentation_price_evidence_and_summary_paths_are_visible(tmp_path: Path) -> None:
    payload = _payload()
    line = payload.lines[0]
    fact = line.historical[0]
    enriched_fact = replace(
        fact,
        raw_value="1,234.0",
        normalized_value=1234.0,
        derived_value=None,
        transformation_rule="parse_numeric_and_scale",
        unit="USD mm",
        unit_kind="currency",
        scale=1_000_000.0,
        currency="USD",
        period_type="fiscal_year",
        period_end=date(2022, 6, 30),
        upstream_dependencies=("source:revenue",),
        downstream_dependencies=("is.gross_profit", "wc.accounts_receivable"),
    )
    enriched_line = replace(
        line,
        historical=(enriched_fact, *line.historical[1:]),
    )
    enriched_payload = replace(
        payload,
        lines=(enriched_line, *payload.lines[1:]),
        current_price_source="CIQ frozen snapshot | MSFT_Standard.xlsx | run 4 | stock_price",
        current_price_as_of=date(2026, 3, 31),
    )
    path = tmp_path / "source-presentation.xlsx"
    render_professional_model_workbook(enriched_payload, path)
    workbook = load_workbook(path, data_only=False)

    sources = workbook["Sources"]
    headers = {
        sources.cell(13, column).value: column
        for column in range(1, sources.max_column + 1)
    }
    assert {
        "Raw source value",
        "Normalized value",
        "Derived value",
        "Transformation / normalization",
        "Unit kind",
        "Period type",
        "Formula error",
        "Downstream dependencies",
    } <= set(headers)
    source_row = next(
        row
        for row in range(14, sources.max_row + 1)
        if sources.cell(row, 1).value == line.canonical_key
        and sources.cell(row, 2).value == fact.period_key
    )
    assert sources.cell(source_row, headers["Raw source value"]).value == "1,234.0"
    assert sources.cell(source_row, headers["Normalized value"]).value == 1234.0
    assert sources.cell(source_row, headers["Transformation / normalization"]).value == "parse_numeric_and_scale"
    assert "wc.accounts_receivable" in sources.cell(source_row, headers["Downstream dependencies"]).value
    assumption_rows = {
        workbook["Assumptions"][f"A{row}"].value: row
        for row in range(8, workbook["Assumptions"].max_row + 1)
    }
    price_cell = workbook["Assumptions"][f"B{assumption_rows['current_price']}"]
    assert "stock_price" in price_cell.comment.text
    assert "2026-03-31" in price_cell.comment.text
    reverse_row = assumption_rows["diagnostic_reverse_dcf_implied_growth_pct_points"]
    assert workbook["Assumptions"][f"B{reverse_row}"].number_format == '0.0"%"'
    assert workbook["Summary"]["B43"].value == "Revenue"
    assert workbook["Summary"]["B44"].value == "EBIT margin"
    assert workbook["Summary"]["H43"].value in {"NEEDS_PM_REVIEW", "UNVERIFIED"}
    assert len(workbook["Summary"]._charts) == 1


def test_valuation_input_scopes_preserve_intrinsic_dcf_and_blank_price(tmp_path: Path) -> None:
    payload = _payload()

    no_price_inputs = dict(payload.valuation_inputs)
    del no_price_inputs["current_price"]
    no_price_path = tmp_path / "no-price.xlsx"
    render_professional_model_workbook(
        replace(payload, valuation_inputs=no_price_inputs),
        no_price_path,
    )
    no_price = load_workbook(no_price_path, data_only=False)
    assert no_price["WACC"]["B3"].value == "AVAILABLE"
    assert no_price["DCF"]["B3"].value == "AVAILABLE"
    assert no_price["Checks"]["C13"].value == "PARTIAL"
    assert no_price["DCF"]["C26"].value == '=""'
    assert no_price["Summary"]["C10"].value == '=""'
    assert "ISNUMBER(" in no_price["Summary"]["B32"].value

    no_terminal_inputs = dict(payload.valuation_inputs)
    del no_terminal_inputs["terminal_growth"]
    no_terminal_path = tmp_path / "no-terminal.xlsx"
    render_professional_model_workbook(
        replace(payload, valuation_inputs=no_terminal_inputs),
        no_terminal_path,
    )
    no_terminal = load_workbook(no_terminal_path, data_only=False)
    assert no_terminal["WACC"]["B3"].value == "AVAILABLE"
    assert no_terminal["DCF"]["B3"].value == "BLOCKED"
    assert no_terminal["Checks"]["C13"].value == "BLOCKED"

    no_reference_inputs = dict(payload.valuation_inputs)
    del no_reference_inputs["bridge_gross_debt"]
    no_reference_path = tmp_path / "no-reference.xlsx"
    render_professional_model_workbook(
        replace(payload, valuation_inputs=no_reference_inputs),
        no_reference_path,
    )
    no_reference = load_workbook(no_reference_path, data_only=False)
    assert no_reference["WACC"]["B3"].value == "AVAILABLE"
    assert no_reference["DCF"]["B3"].value == "AVAILABLE"
    assert no_reference["Checks"]["C13"].value == "PARTIAL"
    assert '"PARTIAL"' in no_reference["Checks"]["C11"].value
    assert no_reference["Valuation"]["B33"].value == '=""'
    assert all(
        "=None" not in str(cell.value) and cell.value != "=0"
        for worksheet in no_reference.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_missing_core_bridge_evidence_blanks_component_and_aggregate(tmp_path: Path) -> None:
    payload = _payload()
    valuation_inputs = dict(payload.valuation_inputs)
    del valuation_inputs["bridge_pension_liability"]
    path = tmp_path / "missing-core-bridge.xlsx"
    render_professional_model_workbook(
        replace(payload, valuation_inputs=valuation_inputs),
        path,
    )
    workbook = load_workbook(path, data_only=False)

    assert workbook["DCF"]["B3"].value == "BLOCKED"
    assert workbook["Valuation"]["B30"].value == '=""'
    assert workbook["Valuation"]["B31"].value == '=""'
    assert workbook["Valuation"]["B29"].value.startswith("='Assumptions'!")


def test_model_output_links_preserve_missing_values_as_blanks(tmp_path: Path) -> None:
    payload = _payload()
    line = payload.lines[0]
    base = line.scenario_forecasts[0]
    missing_base = replace(
        base,
        values=((base.values[0][0], None), *base.values[1:]),
    )
    missing_line = replace(
        line,
        scenario_forecasts=(missing_base, *line.scenario_forecasts[1:]),
    )
    path = tmp_path / "missing-output.xlsx"
    manifest = render_professional_model_workbook(
        replace(payload, lines=(missing_line, *payload.lines[1:])),
        path,
    )
    workbook = load_workbook(path, data_only=False)
    mapping = next(
        item
        for item in manifest.line_cell_mappings
        if item.canonical_key == line.canonical_key
        and item.scenario_key == "base"
        and item.period_key == payload.forecast_periods[0]
    )
    assert workbook[mapping.sheet][mapping.cell].value.startswith("=IF(ISNUMBER(")
    assert '"BLOCKED"' in workbook["Checks"]["C9"].value
    assert '"BLOCKED"' in workbook["Checks"]["C10"].value
