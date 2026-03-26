from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook


def create_ibm_style_workbook(path: str | Path, include_unknown: bool = False, break_anchor: bool = False) -> Path:
    """Create a minimal CIQ workbook fixture following IBM_Standard anchor contract."""
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_hidden = wb.active
    ws_hidden.title = "_CIQHiddenCacheSheet"

    fs = wb.create_sheet("Financial Statements")
    fs["A1"] = "Select Language" if not break_anchor else "Wrong Anchor"
    fs["A2"] = "Ticker"
    fs["C2"] = "NYSE:TEST"
    fs["A9"] = "INCOME STATEMENT - USD IN MILLIONS"
    fs["A10"] = "Period Date"
    fs["D10"] = date(2023, 12, 31)
    fs["E10"] = date(2024, 12, 31)
    fs["F10"] = date(2025, 12, 31)
    fs["A11"] = "Calculation Type"
    fs["D11"] = "REP"
    fs["E11"] = "REP"
    fs["F11"] = "LTM"

    fs["A12"] = "Total Revenues"
    fs["D12"] = 1000
    fs["E12"] = 1100
    fs["F12"] = 1200

    fs["A39"] = "Operating Income"
    fs["D39"] = 100
    fs["E39"] = 121
    fs["F39"] = 150

    fs["A48"] = "EBT Excl Unusual Items"
    fs["D48"] = 90
    fs["E48"] = 110
    fs["F48"] = 140

    fs["A62"] = "Income Tax Expense"
    fs["D62"] = 18
    fs["E62"] = 22
    fs["F62"] = 28

    fs["A84"] = "Weighted Avg. Diluted Shares Out."
    fs["D84"] = 10
    fs["E84"] = 10.5
    fs["F84"] = 11

    fs["A120"] = "Capital Expenditure"
    fs["D120"] = -50
    fs["E120"] = -55
    fs["F120"] = -60

    fs["A121"] = "Depreciation & Amort."
    fs["D121"] = 20
    fs["E121"] = 21
    fs["F121"] = 22

    fs["A200"] = "Total Debt"
    fs["D200"] = 300
    fs["E200"] = 280
    fs["F200"] = 260

    fs["A201"] = "Cash and Equivalents"
    fs["D201"] = 100
    fs["E201"] = 120
    fs["F201"] = 140

    fs["A202"] = "ROIC"
    fs["D202"] = 12
    fs["E202"] = 13
    fs["F202"] = 14

    fs["A203"] = "FCF Yield"
    fs["D203"] = 4.0
    fs["E203"] = 4.2
    fs["F203"] = 4.5

    fs["A204"] = "Total Debt/"
    fs["D204"] = 2.5
    fs["E204"] = 2.3
    fs["F204"] = 2.1

    if include_unknown:
        fs["A205"] = "My Custom KPI"
        fs["D205"] = 1
        fs["E205"] = 2
        fs["F205"] = 3

    cs = wb.create_sheet("Common Size")
    cs["A1"] = "COMMON SIZE PERCENTAGE"
    cs["A10"] = "Period Date"
    cs["D10"] = date(2023, 12, 31)
    cs["E10"] = date(2024, 12, 31)
    cs["F10"] = date(2025, 12, 31)
    cs["A11"] = "Calculation Type"
    cs["D11"] = "REP"
    cs["E11"] = "REP"
    cs["F11"] = "LTM"
    cs["A12"] = "Operating Margin"
    cs["D12"] = 10
    cs["E12"] = 11
    cs["F12"] = 12

    dc = wb.create_sheet("Detailed Comps")
    dc["A1"] = "COMPARABLE COMPANY ANALYSIS"
    dc["A5"] = "Ticker"
    dc["C5"] = "Name"
    dc["D5"] = "Stock Price"
    dc["E5"] = "TEV/Revenue"
    dc["A7"] = "NYSE:PEER"
    dc["C7"] = "Peer Co"
    dc["D7"] = 100
    dc["E7"] = 2.0
    dc["A8"] = "NYSE:TEST"
    dc["C8"] = "Test Co"
    dc["D8"] = 120
    dc["E8"] = 2.2

    sc = wb.create_sheet("Summary Comps")
    sc["A1"] = "COMPARABLE COMPANY ANALYSIS"
    sc["A4"] = "Ticker"
    sc["C4"] = "Trading Data and Size"
    sc["D4"] = "Stock Price"
    sc["E4"] = "Market Cap"
    sc["A5"] = "NYSE:PEER"
    sc["C5"] = "Peer Co"
    sc["D5"] = 99
    sc["E5"] = 10000
    sc["A6"] = "NYSE:TEST"
    sc["C6"] = "Test Co"
    sc["D6"] = 120
    sc["E6"] = 12000

    wb.save(out)
    return out
