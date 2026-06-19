from __future__ import annotations

import json
import tempfile
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from src.stage_04_pipeline.advanced_dcf_model import (
    RECONCILE_TOLERANCE,
    build_advanced_dcf_model,
    refresh_model_data,
)

EXPECTED_SHEETS = [
    "Cover",
    "Thesis_Drivers",
    "PM_Review_Queue",
    "Assumptions",
    "Historical_Financials",
    "Input_Forecast",
    "WACC",
    "DCF_Base",
    "Scenarios",
    "Valuation_Bridge",
    "Sensitivity",
    "Checks",
]


def _workspace_tempdir(name: str) -> Path:
    root = Path(tempfile.gettempdir()) / "advanced-dcf-model"
    root.mkdir(parents=True, exist_ok=True)
    path = root / f"{name}-{uuid4().hex}"
    path.mkdir(parents=True)
    return path


def _sample_payload() -> dict:
    """A small but self-consistent payload whose Base IV reconciles by construction."""
    wacc = 0.08
    historical = []
    for year in range(2017, 2027):
        revenue = 8_000 + (year - 2017) * 250
        ebit = revenue * 0.10
        historical.append(
            {
                "period": f"{year}-03-31",
                "fiscal_year": f"FY{str(year)[-2:]}",
                "source": "ciq_standard_workbook",
                "source_file": "TEST_Standard.xlsx",
                "revenue_mm": revenue,
                "revenue_growth_pct": 0.03,
                "ebit_mm": ebit,
                "ebit_margin_pct": 0.10,
                "ebitda_mm": ebit + 100,
                "ebitda_margin_pct": 0.112,
                "da_mm": 100,
                "da_pct": 100 / revenue,
                "capex_mm": 90,
                "capex_pct": 90 / revenue,
                "cfo_mm": 700,
                "pretax_income_mm": 750,
                "tax_expense_mm": 150,
                "tax_rate_pct": 0.20,
                "cash_mm": 500,
                "debt_mm": 1_000,
                "net_debt_mm": 500,
                "total_assets_mm": 4_000,
                "total_equity_mm": 1_500,
            }
        )

    forecast = []
    pv_fcff_sum = 0.0
    prev_nwc = 380.0
    for year in range(1, 11):
        revenue = 10_000 * (1.03**year)
        ebit = revenue * 0.12
        da = revenue * 0.015
        capex = revenue * 0.012
        nopat = ebit * 0.78
        nwc = revenue * 0.04
        delta_nwc = nwc - prev_nwc
        prev_nwc = nwc
        fcff = nopat + da - capex - delta_nwc
        pv = fcff / (1 + wacc) ** year
        pv_fcff_sum += pv
        forecast.append(
            {
                "year": year,
                "revenue_mm": revenue,
                "growth_rate": 0.03,
                "ebit_margin": 0.12,
                "tax_rate": 0.22,
                "capex_pct": 0.012,
                "da_pct": 0.015,
                "dso": 60,
                "dio": 10,
                "dpo": 30,
                "ebit_mm": ebit,
                "nopat_mm": nopat,
                "da_mm": da,
                "capex_mm": capex,
                "ar_mm": revenue * 60 / 365,
                "inventory_mm": revenue * 10 / 365,
                "ap_mm": revenue * 30 / 365,
                "nwc_mm": nwc,
                "delta_nwc_mm": delta_nwc,
                "fcff_mm": fcff,
                "discount_factor": (1 + wacc) ** year,
                "pv_fcff_mm": pv,
                "roic": 0.15,
                "economic_profit_mm": 300,
            }
        )

    # Terminal block computed exactly as the DCF_Base formulas do (value-driver
    # Gordon + exit multiple), so iv_base reconciles to the rebuild.
    gordon_w, exit_w = 0.6, 0.4
    g_term, ronic, exit_multiple = 0.025, 0.11, 10.0
    y10 = forecast[-1]
    nopat11 = y10["nopat_mm"] * (1 + g_term)
    fcff11 = nopat11 * (1 - g_term / ronic)
    tv_gordon = fcff11 / (wacc - g_term)
    ebitda10 = y10["ebit_mm"] + y10["da_mm"]  # exit_metric = ev_ebitda
    tv_exit = ebitda10 * exit_multiple
    pv_tv_gordon = tv_gordon / (1 + wacc) ** 10
    pv_tv_exit = tv_exit / (1 + wacc) ** 10
    pv_tv_blended = pv_tv_gordon * gordon_w + pv_tv_exit * exit_w
    non_op, claims, shares = 50.0, 500.0, 100.0
    ev_ops = pv_fcff_sum + pv_tv_blended
    iv_base = (ev_ops + non_op - claims) / shares
    iv_gordon = (pv_fcff_sum + pv_tv_gordon + non_op - claims) / shares
    iv_exit = (pv_fcff_sum + pv_tv_exit + non_op - claims) / shares

    return {
        "ticker": "TEST",
        "company_name": "Test Company",
        "sector": "Industrials",
        "industry": "Consulting Services",
        "historical_financials": historical,
        "story_profile": {
            "moat_strength": 3,
            "pricing_power": 3,
            "cyclicality": "high",
            "capital_intensity": "high",
            "governance_risk": "medium",
            "competitive_advantage_years": 6,
        },
        "story_adjustments": {
            "growth_add": 0.0,
            "margin_add": 0.0,
            "cyclicality_growth_multiplier": 0.9,
            "cyclicality_wacc_add": 0.01,
            "governance_wacc_add": 0.0,
            "capex_target_add": 0.01,
            "da_target_add": 0.005,
            "terminal_blend_gordon_weight": gordon_w,
            "terminal_blend_exit_weight": exit_w,
            "exit_multiple_cyclicality_multiplier": 0.9,
            "exit_multiple_governance_multiplier": 1.0,
        },
        "default_resolution": {
            "status": "review_required",
            "counts": {"resolved": 5, "review_required": 1},
            "fields": [
                {
                    "field": "exit_multiple",
                    "value": 10.0,
                    "source": "ciq_comps_fallback",
                    "source_class": "ciq",
                    "fallback_value": 11.0,
                    "severity": "high",
                    "needs_pm_review": False,
                    "why_it_matters": "Drives the exit terminal value.",
                },
                {
                    "field": "pension_deficit",
                    "value": 0.0,
                    "source": "default",
                    "source_class": "unproven_zero",
                    "fallback_value": 0.0,
                    "severity": "medium",
                    "needs_pm_review": True,
                    "why_it_matters": "Could be a hidden claim on equity.",
                },
            ],
        },
        "scenario_policy": {
            "policy": "context_advisory_v1",
            "official_policy": "fixed_default",
            "official_specs": [
                {"name": "bear", "probability": 0.2, "growth_multiplier": 0.8, "margin_shift": -0.02, "wacc_shift": 0.01, "exit_multiple_multiplier": 0.9},
                {"name": "base", "probability": 0.6, "growth_multiplier": 1.0, "margin_shift": 0.0, "wacc_shift": 0.0, "exit_multiple_multiplier": 1.0},
                {"name": "bull", "probability": 0.2, "growth_multiplier": 1.2, "margin_shift": 0.02, "wacc_shift": -0.01, "exit_multiple_multiplier": 1.1},
            ],
        },
        "context_scenarios": {
            "bear": {"probability": 0.2, "growth_multiplier": 0.73, "margin_shift": -0.027, "wacc_shift": 0.0165, "exit_multiple_multiplier": 0.85},
            "base": {"probability": 0.6, "growth_multiplier": 1.0, "margin_shift": 0.0, "wacc_shift": 0.0, "exit_multiple_multiplier": 1.0},
            "bull": {"probability": 0.2, "growth_multiplier": 1.27, "margin_shift": 0.027, "wacc_shift": -0.0135, "exit_multiple_multiplier": 1.15},
        },
        "assumption_register_summary": {
            "model_trust_state": "watch",
            "flag_counts": {"none": 24, "watch": 9, "review_required": 0, "critical": 0},
            "max_flag_level": "watch",
            "has_critical": False,
            "flagged_entries": [
                {"assumption_name": "exit_multiple", "flag_level": "watch"},
            ],
        },
        "drivers_raw": {
            "shares_outstanding": shares * 1e6,
            "annual_dilution_pct": 0.0,
            "ronic_terminal": 0.11,
            "cogs_pct_of_revenue": 0.46,
            "non_operating_assets": non_op * 1e6,
            "net_debt": claims * 1e6,
        },
        "excel_flat": {
            "metadata": [{"key": "ticker", "value": "TEST"}],
            "assumptions": [
                {"key": "revenue_mm", "value": 10_000},
                {"key": "growth_terminal_pct", "value": 0.025},
                {"key": "exit_multiple", "value": 10.0},
                {"key": "exit_metric", "value": "ev_ebitda"},
                {"key": "net_debt_mm", "value": claims},
                {"key": "shares_outstanding_mm", "value": shares},
                {"key": "non_operating_assets_mm", "value": non_op},
                {"key": "minority_interest_mm", "value": 0},
                {"key": "preferred_equity_mm", "value": 0},
                {"key": "pension_deficit_mm", "value": 0},
                {"key": "lease_liabilities_mm", "value": 0},
                {"key": "options_value_mm", "value": 0},
                {"key": "convertibles_value_mm", "value": 0},
                {"key": "scenario_prob_bear", "value": 0.2},
                {"key": "scenario_prob_base", "value": 0.6},
                {"key": "scenario_prob_bull", "value": 0.2},
            ],
            "wacc": [
                {"key": "wacc", "value": wacc},
                {"key": "cost_of_equity", "value": 0.09},
                {"key": "cost_of_debt", "value": 0.05},
                {"key": "risk_free_rate", "value": 0.04},
                {"key": "equity_risk_premium", "value": 0.05},
                {"key": "beta_relevered", "value": 1.0},
                {"key": "size_premium", "value": 0.0},
                {"key": "equity_weight", "value": 0.8},
                {"key": "debt_weight", "value": 0.2},
            ],
            "valuation": [
                {"key": "iv_base", "value": round(iv_base, 2)},
                {"key": "iv_gordon", "value": round(iv_gordon, 2)},
                {"key": "iv_exit", "value": round(iv_exit, 2)},
                {"key": "expected_iv", "value": round(iv_base * 1.05, 2)},
                {"key": "context_expected_iv", "value": round(iv_base * 1.08, 2)},
                {"key": "iv_bear", "value": round(iv_base * 0.7, 2)},
                {"key": "iv_bull", "value": round(iv_base * 1.5, 2)},
                {"key": "upside_base_pct", "value": 26.3},
                {"key": "upside_bear_pct", "value": -23.5},
                {"key": "upside_bull_pct", "value": 108.5},
                {"key": "expected_upside_pct", "value": 32.8},
                {"key": "margin_of_safety", "value": 20.8},
                {"key": "ep_iv_base", "value": round(iv_base * 1.2, 2)},
                {"key": "fcfe_iv_base", "value": round(iv_base * 1.25, 2)},
                {"key": "dcf_ep_gap_pct", "value": 21.8},
                {"key": "implied_growth_pct", "value": -2.9},
                {"key": "comps_iv_ev_ebitda", "value": 80},
                {"key": "comps_iv_base", "value": 95},
                {"key": "comps_iv_pe", "value": 115},
                {"key": "model_applicability_status", "value": "dcf_applicable"},
            ],
            "terminal": [
                {"key": "pv_tv_gordon_mm", "value": pv_tv_gordon},
                {"key": "pv_tv_exit_mm", "value": pv_tv_exit},
                {"key": "pv_tv_blended_mm", "value": pv_tv_blended},
                {"key": "tv_pct_of_ev", "value": round(100 * pv_tv_blended / ev_ops, 1)},
                {"key": "terminal_growth_pct", "value": 2.5},
                {"key": "terminal_ronic_pct", "value": 11.0},
                {"key": "non_operating_assets_mm", "value": non_op},
                {"key": "non_equity_claims_mm", "value": claims},
                {"key": "ev_operations_mm", "value": round(ev_ops, 1)},
            ],
            "market": [
                {"key": "price", "value": 75},
                {"key": "analyst_target", "value": 90},
                {"key": "analyst_recommendation", "value": "hold"},
            ],
            "scenarios": [{"scenario": "base", "probability": 0.6, "iv": round(iv_base, 2), "upside_pct": 0.33}],
            "source_lineage": [
                {"key": "exit_multiple", "value": "ciq_comps_fallback|story_sector"},
                {"key": "net_debt", "value": "ciq"},
            ],
            "health_flags": [{"key": "model_applicability_status", "value": "dcf_applicable"}],
            "historical_financials": historical,
            "forecast": forecast,
            "comps_peers": [{"ticker": "PEER", "display_name": "Peer Co", "ev_ebitda": 9.5}],
            "comps_valuation": [{"key": "comps_iv_base", "value": 95}],
        },
    }


def _build(tmp_name: str = "build"):
    tmp = _workspace_tempdir(tmp_name)
    json_path = tmp / "TEST_latest.json"
    out_path = tmp / "TEST_advanced_dcf_model.xlsx"
    json_path.write_text(json.dumps(_sample_payload()), encoding="utf-8")
    result = build_advanced_dcf_model("TEST", json_path=json_path, output_path=out_path)
    return result, out_path


def test_build_creates_judgment_aware_workbook():
    result, out_path = _build()
    assert result == out_path
    wb = load_workbook(result, data_only=False)
    assert wb.sheetnames == EXPECTED_SHEETS
    # judgment layer present
    assert wb["Thesis_Drivers"]["A1"].value.startswith("Story & Driver Layer")
    assert wb["PM_Review_Queue"].max_row >= 6
    # provenance + override mechanics
    assert wb["Assumptions"]["E5"].value == '=IF(D5="",C5,D5)'
    assert wb["Assumptions"]["D5"].value is None  # override genuinely empty, not ""
    # DCF references the backend per-year forecast
    assert wb["DCF_Base"]["C6"].value == "='Input_Forecast'!$B$5"


def test_no_broken_references_and_formula_density():
    result, _ = _build("refs")
    wb = load_workbook(result, data_only=False)
    formulas = [
        cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert len(formulas) > 200
    assert not any("#REF!" in f or "#NAME?" in f for f in formulas)


def test_refresh_preserves_model_edits_and_overrides():
    """A refresh swaps data but keeps the PM's formula sheets, overrides, added tabs."""
    tmp = _workspace_tempdir("refresh")
    json_path = tmp / "TEST_latest.json"
    model_path = tmp / "TEST_model.xlsx"
    json_path.write_text(json.dumps(_sample_payload()), encoding="utf-8")
    build_advanced_dcf_model("TEST", json_path=json_path, output_path=model_path)

    # Simulate PM edits: a custom sheet, a custom cell on a model sheet, an override.
    wb = load_workbook(model_path)
    wb.create_sheet("My_Notes")["A1"] = "custom analysis"
    wb["Sensitivity"]["A30"] = "PM custom row"
    aws = wb["Assumptions"]
    ovr_row = next(r for r in range(5, aws.max_row + 1) if aws.cell(r, 2).value == "exit_multiple")
    aws.cell(ovr_row, 4).value = 9.0
    wb.save(model_path)

    out = tmp / "TEST_model_refreshed.xlsx"
    refresh_model_data(model_path, ticker="TEST", json_path=json_path, output_path=out)

    wb2 = load_workbook(out)
    assert "My_Notes" in wb2.sheetnames
    assert wb2["My_Notes"]["A1"].value == "custom analysis"
    assert wb2["Sensitivity"]["A30"].value == "PM custom row"
    aws2 = wb2["Assumptions"]
    ov = next(aws2.cell(r, 4).value for r in range(5, aws2.max_row + 1) if aws2.cell(r, 2).value == "exit_multiple")
    assert ov == 9.0
    # data sheets still present and model sheets intact
    assert set(EXPECTED_SHEETS).issubset(set(wb2.sheetnames))


def test_agent_judgment_surfaced_when_guided_workup_present():
    """When a guided-workup JSON exists, its thesis + driver cards appear, read-only."""
    tmp = _workspace_tempdir("agent")
    json_path = tmp / "TEST_latest.json"
    json_path.write_text(json.dumps(_sample_payload()), encoding="utf-8")
    workup = tmp / "TEST-20260615T000000Z.json"
    workup.write_text(json.dumps({
        "ticker": "TEST",
        "run_stamp": "20260615T000000Z",
        "queue_decisions": [{"item_id": 1, "action": "deferred", "reason": "non_interactive"}],
        "analyst_prep": {
            "thesis_cards": [
                {"title": "Valuation Setup", "claim": "Base IV above price.",
                 "model_implication": "Review growth and exit multiple.",
                 "deterministic_confidence": "medium",
                 "what_would_change_mind": "A stale refresh."},
            ],
            "driver_cards": [
                {"assumption_name": "exit_multiple", "label": "Exit Multiple",
                 "current_value": 10.0, "proposed_or_effective_value": 11.0,
                 "source": "ciq", "rationale": "PM Queue item 88 proposes a higher exit.",
                 "valuation_impact": None, "pm_review_status": "review_required"},
            ],
            "missing_data": [
                {"label": "Segment evidence missing", "severity": "medium",
                 "reason": "No segment rows found.", "suggested_check": "Pull 10-K segments."},
            ],
        },
    }), encoding="utf-8")

    out = tmp / "TEST_model.xlsx"
    build_advanced_dcf_model("TEST", json_path=json_path, output_path=out, guided_workup_path=workup)
    wb = load_workbook(out)

    thesis = wb["Thesis_Drivers"]
    assert thesis["A1"].value.startswith("Investment Thesis")
    thesis_text = " ".join(
        str(c.value) for row in thesis.iter_rows() for c in row if c.value
    )
    assert "Valuation Setup" in thesis_text and "Review growth and exit multiple." in thesis_text

    queue = wb["PM_Review_Queue"]
    queue_text = " ".join(
        str(c.value) for row in queue.iter_rows() for c in row if c.value
    )
    assert "Agent Driver Proposals" in queue_text
    assert "exit_multiple" in queue_text and "PM Queue item 88" in queue_text
    assert "Segment evidence missing" in queue_text


def test_builds_without_historicals():
    """Names lacking CIQ/SEC history still build (historicals only enrich a display tab)."""
    tmp = _workspace_tempdir("no-hist")
    payload = _sample_payload()
    payload["historical_financials"] = []
    payload["excel_flat"]["historical_financials"] = []
    json_path = tmp / "TEST_latest.json"
    json_path.write_text(json.dumps(payload), encoding="utf-8")
    out = build_advanced_dcf_model("TEST", json_path=json_path, output_path=tmp / "TEST.xlsx")
    wb = load_workbook(out)
    assert "Historical_Financials" in wb.sheetnames
    hist_text = " ".join(
        str(c.value) for row in wb["Historical_Financials"].iter_rows() for c in row if c.value
    )
    assert "No historical actuals available" in hist_text


def test_build_fails_when_not_reconciled():
    """If the backend iv_base disagrees with the rebuild, the export must refuse."""
    tmp = _workspace_tempdir("recon-fail")
    payload = _sample_payload()
    for row in payload["excel_flat"]["valuation"]:
        if row["key"] == "iv_base":
            row["value"] = row["value"] + 25.0  # force a divergence
    json_path = tmp / "TEST_latest.json"
    json_path.write_text(json.dumps(payload), encoding="utf-8")
    with pytest.raises(ValueError, match="reconcil"):
        build_advanced_dcf_model("TEST", json_path=json_path, output_path=tmp / "x.xlsx")


def test_formula_chain_reconciles_to_backend_iv():
    """End-to-end: recalculate the workbook and assert Base IV ties to backend.

    Skips cleanly when the optional ``formulas`` engine is not installed so the
    offline CI suite stays green.
    """
    formulas = pytest.importorskip("formulas")
    import warnings

    result, out_path = _build("recalc")
    payload = _sample_payload()
    backend_iv = next(
        r["value"] for r in payload["excel_flat"]["valuation"] if r["key"] == "iv_base"
    )

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        model = formulas.ExcelModel().loads(str(out_path)).finish()
        sol = model.calculate()

    wb = load_workbook(out_path)
    ws = wb["DCF_Base"]
    iv_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value == "Intrinsic value / share"
    )
    key = f"'[{out_path.name}]DCF_BASE'!B{iv_row}"
    workbook_iv = float(sol[key].value[0, 0])
    assert abs(workbook_iv - backend_iv) <= RECONCILE_TOLERANCE
