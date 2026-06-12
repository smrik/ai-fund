from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from db.schema import create_tables


def _workspace_tempdir(name: str) -> Path:
    root = Path.cwd() / ".tmp-tests" / "export-service"
    root.mkdir(parents=True, exist_ok=True)
    path = root / f"{name}-{uuid4().hex}"
    path.mkdir(parents=True)
    return path


def _temp_conn_factory(db_path: Path):
    def _factory():
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        create_tables(conn)
        return conn

    return _factory


def _payload_with_attached_dossier(source_mode: str, snapshot_id: int | None = None) -> dict:
    return {
        "ticker": "IBM",
        "company_name": "Legacy Export Name",
        "source_mode": source_mode,
        "generated_at": "2026-04-01T00:00:00+00:00",
        "valuation": {"current_price": 10.0, "iv_base": 20.0, "expected_iv": 30.0},
        "research": {"publishable_memo_preview": "Legacy research memo."},
        "snapshot": {"memo": {"one_liner": "Legacy snapshot memo."}},
        "ticker_dossier": {
            "contract_name": "TickerDossier",
            "contract_version": "1.0.0",
            "ticker": "IBM",
            "as_of_date": "2026-04-30",
            "display_name": "Canonical Machines",
            "currency": "USD",
            "latest_snapshot": {
                "company_identity": {"ticker": "IBM", "display_name": "Canonical Machines", "sector": "Canonical Sector"},
                "market_snapshot": {"as_of_date": "2026-04-30", "price": 111.0},
                "valuation_snapshot": {"base_iv": 155.0, "expected_iv": 166.0, "current_price": 112.0},
                "historical_series": {},
                "qoe_snapshot": {"present": False, "score": None, "flags": []},
                "comps_snapshot": {},
                "source_lineage": {},
            },
            "loaded_backend_state": {"backend_name": "test", "source_mode": source_mode},
            "source_lineage": {},
            "export_metadata": {"source_mode": source_mode, "snapshot_id": snapshot_id},
            "optional_overlays": {},
        },
    }


def test_stage_power_query_workbook_copies_template_and_points_json_path(monkeypatch):
    from src.stage_04_pipeline import export_service

    tmp_path = _workspace_tempdir("export-workbook")
    template_path = tmp_path / "ticker_review.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"
    ws["A1"] = "Setting"
    ws["B1"] = "Value"
    ws["A2"] = "json_path"
    ws["B2"] = "C:\\placeholder.json"
    wb.create_named_range("json_path", ws, "$B$2")
    wb.save(template_path)

    bundle_dir = tmp_path / "bundle"
    bundle_dir.mkdir()

    monkeypatch.setattr(export_service, "TICKER_EXPORT_TEMPLATE", template_path)

    payload = {"ticker": "IBM", "valuation": {"iv_base": 123.0}}
    result = export_service.stage_power_query_workbook("IBM", payload, bundle_dir)

    workbook_path = Path(result["primary_path"])
    json_path = bundle_dir / "IBM_latest.json"

    assert workbook_path.exists()
    assert json_path.exists()
    assert json.loads(json_path.read_text(encoding="utf-8"))["ticker"] == "IBM"

    staged = load_workbook(workbook_path)
    assert staged["Config"]["B2"].value == str(json_path.resolve())
    assert list(staged.defined_names["json_path"].destinations) == [("Config", "$B$2")]
    assert result["artifacts"][0]["artifact_key"] == "excel_workbook"


def test_stage_power_query_workbook_populates_comps_tabs(monkeypatch):
    from src.stage_04_pipeline import export_service

    tmp_path = _workspace_tempdir("export-workbook-comps")
    template_path = tmp_path / "ticker_review.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"
    ws["A2"] = "json_path"
    ws["B2"] = "C:\\placeholder.json"
    wb.create_named_range("json_path", ws, "$B$2")
    wb.create_sheet("Comps")
    wb.save(template_path)

    bundle_dir = tmp_path / "bundle"
    bundle_dir.mkdir()
    monkeypatch.setattr(export_service, "TICKER_EXPORT_TEMPLATE", template_path)

    payload = {
        "ticker": "IBM",
        "company_name": "International Business Machines",
        "market": {"price": 260.0},
        "comps_analysis": {
            "primary_metric": "tev_ebitda_ltm",
            "peer_counts": {"raw": 5, "clean": 4},
            "valuation_range": {"bear": 190.0, "base": 202.0, "bull": 216.0, "blended_base": 205.0},
            "valuation_by_metric_rows": [
                {
                    "metric": "tev_ebitda_ltm",
                    "label": "TEV / EBITDA LTM",
                    "target_multiple": 3.0,
                    "peer_median_multiple": 17.0,
                    "bear_multiple": 15.0,
                    "base_multiple": 17.0,
                    "bull_multiple": 19.0,
                    "bear_iv": 190.0,
                    "base_iv": 202.0,
                    "bull_iv": 216.0,
                    "n_raw": 5,
                    "n_clean": 4,
                    "is_primary": True,
                }
            ],
            "comparison_summary": [
                {"metric": "revenue_growth", "label": "Revenue Growth", "target": 0.03, "peer_median": 0.05, "delta": -0.02}
            ],
            "peer_table": [
                {
                    "ticker": "ACN",
                    "display_name": "Accenture",
                    "similarity_score": 0.8,
                    "model_weight": 0.4,
                    "revenue_ltm_mm": 70000.0,
                    "ebitda_ltm_mm": 13000.0,
                    "ebit_ltm_mm": 10500.0,
                    "tev_ebitda_ltm": 17.0,
                }
            ],
            "metric_status_rows": [
                {"ticker": "ACN", "metric": "tev_ebitda_ltm", "label": "TEV / EBITDA LTM", "raw_multiple": 17.0, "status": "included"}
            ],
            "football_field": {
                "ranges": [{"label": "TEV / EBITDA LTM", "bear": 190.0, "base": 202.0, "bull": 216.0}],
                "markers": [{"label": "Current Price", "value": 260.0, "type": "spot"}],
            },
            "historical_multiples_summary": {
                "metrics": {"pe_trailing": {"current": 20.0, "summary": {"median": 18.0, "current_percentile": 0.65}}}
            },
            "operating_context": {
                "target": {"revenue_ltm_mm": 60000.0, "ebitda_ltm_mm": 11428.6, "ebit_ltm_mm": 8450.0},
                "peer_medians": {"ebit_margin": 0.18, "net_debt_to_ebitda": 0.6},
                "peer_count": 1,
            },
            "support_data_quality": {
                "target_missing_fields": [],
                "peer_coverage": {"ebitda_ltm_mm": {"present": 1, "total": 1, "ratio": 1.0}},
                "valuation_metric_count": 1,
                "common_patchups_needed": [],
            },
            "audit_flags": ["Outliers removed from tev_ebitda_ltm: MSFT"],
            "notes": "primary=tev_ebitda_ltm",
            "source_lineage": {"as_of_date": "2026-03-01", "source_file": "IBM_comps.xlsx"},
        },
    }

    result = export_service.stage_power_query_workbook("IBM", payload, bundle_dir)
    staged = load_workbook(Path(result["primary_path"]))

    assert "Comps Diagnostics" in staged.sheetnames
    assert "IBM" in str(staged["Comps"]["A1"].value)
    assert staged["Comps"]["B5"].value == "tev_ebitda_ltm"
    assert staged["Comps"]["A20"].value == "Peer Table"
    assert staged["Comps"]["E21"].value == "Revenue LTM"
    assert staged["Comps Diagnostics"]["A4"].value == "Audit Flags"
    assert staged["Comps Diagnostics"]["A10"].value == "Ticker"


def test_stage_power_query_workbook_populates_analyst_prep_sheets(monkeypatch):
    from src.stage_04_pipeline import export_service

    tmp_path = _workspace_tempdir("export-workbook-analyst-prep")
    template_path = tmp_path / "ticker_review.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"
    ws["A2"] = "json_path"
    ws["B2"] = "C:\\placeholder.json"
    wb.save(template_path)
    bundle_dir = tmp_path / "bundle"
    bundle_dir.mkdir()
    monkeypatch.setattr(export_service, "TICKER_EXPORT_TEMPLATE", template_path)

    payload = {
        "ticker": "IBM",
        "company_name": "International Business Machines",
        "analyst_prep": {
            "ticker": "IBM",
            "generated_at": "2026-06-07T00:00:00Z",
            "source_quality": "real",
            "thesis_cards": [
                {
                    "card_id": "IBM:valuation_setup",
                    "title": "Valuation Setup",
                    "claim": "Base IV is 150.",
                    "business_evidence_summary": "DCF support.",
                    "model_implication": "Review WACC.",
                    "linked_assumption_fields": ["wacc"],
                    "evidence_anchor_ids": ["deterministic:dcf:base_iv"],
                    "what_would_change_mind": "Fresh CIQ.",
                }
            ],
            "driver_cards": [
                {
                    "assumption_name": "wacc",
                    "label": "WACC",
                    "current_value": 0.09,
                    "proposed_or_effective_value": 0.095,
                    "source": "wacc_peer_beta",
                    "pm_review_status": "review_required",
                    "rationale": "Review method.",
                    "evidence_anchor_ids": ["deterministic:assumption:wacc"],
                }
            ],
            "comps_card": {
                "peer_set_quality": "partial",
                "peer_count": 3,
                "primary_metric": "tev_ebitda_ltm",
                "premium_discount_argument": "Target trades at a discount.",
                "exit_multiple_support": "Review exit multiple.",
                "warnings": ["public fallback"],
                "evidence_anchor_ids": ["deterministic:comps:peer_set"],
            },
            "missing_data": [
                {
                    "flag_id": "segment_data_missing",
                    "label": "Segment evidence missing",
                    "severity": "medium",
                    "reason": "No segment rows.",
                    "suggested_check": "Refresh CIQ.",
                }
            ],
            "segment_driver_rows": [],
            "evidence_packet_ids": [7],
            "evidence_map": [
                {
                    "anchor_id": "packet:7:fact:growth",
                    "packet_id": 7,
                    "profile_name": "company_analysis",
                    "kind": "packet_fact",
                    "label": "Growth",
                    "value": 8.0,
                    "unit": "%",
                    "source_quality": "real",
                    "source_ref": "ciq",
                }
            ],
            "export_metadata": {"default_resolution_status": "review_required"},
        },
    }

    result = export_service.stage_power_query_workbook("IBM", payload, bundle_dir)
    staged = load_workbook(Path(result["primary_path"]))

    for sheet_name in [
        "Analyst_Prep",
        "Thesis_Bridge",
        "Model_Driver_Map",
        "Evidence_Map",
        "Comps_Judgment",
        "Segment_Drivers",
    ]:
        assert sheet_name in staged.sheetnames
    assert staged["Analyst_Prep"]["B5"].value == "real"
    assert staged["Thesis_Bridge"]["B5"].value == "Valuation Setup"
    assert staged["Model_Driver_Map"]["A5"].value == "wacc"
    assert staged["Evidence_Map"]["A5"].value == "packet:7:fact:growth"
    assert staged["Segment_Drivers"]["A5"].value == "Segment evidence missing"


def test_stage_power_query_workbook_rewrites_visible_review_tabs(monkeypatch):
    from src.stage_04_pipeline import export_service

    tmp_path = _workspace_tempdir("export-workbook-visible-tabs")
    template_path = tmp_path / "ticker_review.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"
    ws["A2"] = "json_path"
    ws["B2"] = "C:\\placeholder.json"
    wb.create_named_range("json_path", ws, "$B$2")
    for sheet_name in [
        "Cover",
        "Output",
        "Assumptions",
        "DCF_Base",
        "DCF_Bear",
        "DCF_Bull",
        "Equity_Bridge",
        "Sensitivity",
        "QoE",
        "Comps",
    ]:
        sheet = wb.create_sheet(sheet_name)
        sheet["A1"] = "IBM stale template value"
    wb.save(template_path)

    bundle_dir = tmp_path / "bundle"
    bundle_dir.mkdir()
    monkeypatch.setattr(export_service, "TICKER_EXPORT_TEMPLATE", template_path)

    payload = {
        "ticker": "VRRM",
        "company_name": "Verra Mobility Corporation",
        "sector": "Industrials",
        "generated_at": "2026-06-07T08:00:00+00:00",
        "market": {"price": 4.31, "analyst_price_target": None},
        "valuation": {"current_price": 4.31, "iv_bear": 30.43, "iv_base": 59.49, "iv_bull": 244.94, "expected_iv": 59.49},
        "scenarios": {"base": {"iv": 59.49, "upside_pct": 1280.3}},
        "assumptions": {
            "growth_near_pct": 10.76,
            "growth_mid_pct": 7.53,
            "ebit_margin_start_pct": 24.66,
            "exit_multiple": 17.14,
            "net_debt_mm": 123.0,
        },
        "wacc": {"wacc_pct": 5.68},
        "terminal": {
            "terminal_growth_pct": 3.5,
            "ronic_terminal_pct": 15.0,
            "tv_pct_of_ev": 77.9,
            "pv_tv_blended_mm": 8367.76,
            "method_used": "blend",
        },
        "forecast_bridge": [{"year": 1, "revenue_mm": 1084.8, "growth_pct": 10.76, "fcff_mm": 191.94}],
        "sensitivity": {"summary": [{"grid": "wacc_x_terminal_growth", "cell_count": 9, "min_iv": 42.1, "max_iv": 143.38, "spread": 101.28}]},
        "health_flags": {"tv_high_flag": True},
        "source_lineage": {"revenue_growth_near": "yfinance|story_sector", "exit_multiple": "default|story_sector"},
        "default_resolution": {
            "status": "review_required_high",
            "fields": [
                {
                    "field": "exit_multiple",
                    "value": 17.14,
                    "source": "default|story_sector",
                    "source_class": "missing_default",
                    "severity": "high",
                    "needs_pm_review": True,
                    "why_it_matters": "Directly affects terminal value and equity value.",
                }
            ],
        },
        "ciq_lineage": {"snapshot_source_file": "public_market_yfinance_fallback", "peer_count": 9},
        "comps_analysis": {
            "primary_metric": "tev_ebitda_ltm",
            "peer_counts": {"raw": 9, "clean": 9},
            "valuation_range": {"bear": 20.0, "base": 27.44, "bull": 38.0, "blended_base": 27.44},
            "audit_flags": ["No CIQ comps detail available; using public market yfinance fallback comps"],
        },
    }

    result = export_service.stage_power_query_workbook("VRRM", payload, bundle_dir)
    staged = load_workbook(Path(result["primary_path"]))

    visible_review_sheets = [
        "Cover",
        "Output",
        "Assumptions",
        "DCF_Base",
        "DCF_Bear",
        "DCF_Bull",
        "Equity_Bridge",
        "Sensitivity",
        "QoE",
        "Review Checks",
    ]
    for sheet_name in visible_review_sheets:
        values = [str(cell.value) for row in staged[sheet_name].iter_rows() for cell in row if cell.value is not None]
        assert any("VRRM" in value for value in values), sheet_name
        assert all("IBM stale template value" not in value for value in values), sheet_name

    assert staged["Cover"]["A1"].value == "VRRM - PM Review Workbook"
    assert staged["Output"]["B7"].value == 1280.28
    assert staged["Assumptions"]["A5"].value == "ebit_margin_start_pct"
    assert staged["Review Checks"]["A5"].value == "High"
    assert "Terminal value is 77.9% of EV" in staged["Review Checks"]["C6"].value
    review_values = [cell.value for row in staged["Review Checks"].iter_rows() for cell in row if cell.value is not None]
    assert "Default resolution" in review_values
    qoe_values = [cell.value for row in staged["QoE"].iter_rows() for cell in row if cell.value is not None]
    assert "Default Resolution" in qoe_values
    assert "missing_default" in qoe_values


def test_build_html_export_bundle_writes_primary_and_sidecar_assets():
    from src.stage_04_pipeline import export_service

    tmp_path = _workspace_tempdir("export-html")
    bundle_dir = tmp_path / "bundle"
    bundle_dir.mkdir()
    public_asset = tmp_path / "chart.png"
    public_asset.write_bytes(b"png")

    context = {
        "ticker": "IBM",
        "company_name": "International Business Machines",
        "source_mode": "latest_snapshot",
        "summary": "Public thesis summary.",
        "artifacts": [
            {
                "artifact_key": "public_chart",
                "title": "DCF Chart",
                "path_mode": "absolute",
                "path_value": str(public_asset),
                "artifact_type": "export_png",
            }
        ],
    }

    result = export_service.build_html_export_bundle("IBM", context, bundle_dir)

    primary_path = Path(result["primary_path"])
    manifest_path = bundle_dir / "manifest.json"
    context_path = bundle_dir / "context.json"
    copied_asset = bundle_dir / "assets" / public_asset.name

    assert primary_path.exists()
    assert "Public thesis summary." in primary_path.read_text(encoding="utf-8")
    assert manifest_path.exists()
    assert context_path.exists()
    assert copied_asset.exists()
    assert {artifact["artifact_key"] for artifact in result["artifacts"]} == {"html_report", "context_json", "public_chart"}


def test_build_html_context_prefers_attached_canonical_dossier_for_current_and_snapshot(monkeypatch):
    from src.stage_04_pipeline import export_service

    monkeypatch.setattr(export_service, "_build_snapshot_ticker_payload", lambda ticker: (_payload_with_attached_dossier("latest_snapshot", 7), 7))
    monkeypatch.setattr(export_service, "_build_current_ticker_payload", lambda ticker: _payload_with_attached_dossier("loaded_backend_state"))
    monkeypatch.setattr(
        export_service,
        "build_publishable_memo_context",
        lambda ticker: {"memo_content": "Canonical memo content.", "artifacts": [{"artifact_key": "chart"}]},
    )

    snapshot_context, snapshot_id = export_service._build_html_context("IBM", "latest_snapshot")
    current_context, current_snapshot_id = export_service._build_html_context("IBM", "loaded_backend_state")

    for context in (snapshot_context, current_context):
        assert context["company_name"] == "Canonical Machines"
        assert context["current_price"] == 111.0
        assert context["base_iv"] == 155.0
        assert context["expected_iv"] == 166.0
        assert context["as_of_date"] == "2026-04-30"
        assert context["ticker_dossier_contract_version"] == "1.0.0"
        assert context["valuation"]["current_price"] == 111.0
        assert context["valuation"]["iv_base"] == 155.0
        assert context["valuation"]["expected_iv"] == 166.0
        assert context["summary"] == "Canonical memo content."
        assert context["ticker_dossier"]["display_name"] == "Canonical Machines"

    assert snapshot_context["source_mode"] == "latest_snapshot"
    assert snapshot_context["snapshot_id"] == 7
    assert snapshot_id == 7
    assert current_context["source_mode"] == "loaded_backend_state"
    assert current_snapshot_id is None


def test_build_current_ticker_payload_preserves_valuation_input_lineage(monkeypatch):
    from src.stage_04_pipeline import export_service

    monkeypatch.setattr(export_service, "_attach_ticker_dossier", lambda payload, source_mode, snapshot_id=None: payload)
    monkeypatch.setattr(
        export_service,
        "build_override_workbench",
        lambda ticker: {
            "available": True,
            "ticker": "IBM",
            "company_name": "International Business Machines",
            "sector": "Technology",
            "current_price": 100.0,
            "fields": [
                {"field": "exit_multiple", "effective_value": 14.0, "effective_source": "public_market_yfinance_fallback_tev_ebitda_ltm"},
                {"field": "revenue_growth_near", "effective_value": 0.08, "effective_source": "ciq_consensus"},
            ],
            "ciq_lineage": {
                "public_comps_fallback_used": True,
                "public_comps_fallback_source_file": "public_market_yfinance_fallback",
                "public_comps_fallback_peer_count": 3,
                "snapshot_source_file": None,
            },
            "default_resolution": {"status": "ok", "fields": []},
        },
    )
    monkeypatch.setattr(
        export_service,
        "build_dcf_audit_view",
        lambda ticker: {
            "scenario_summary": [{"scenario": "Base", "intrinsic_value": 150.0, "upside_pct": 50.0, "probability": 0.6}],
            "ev_bridge": {"intrinsic_value_per_share": 150.0},
            "sensitivity": {},
            "terminal_bridge": {},
            "health_flags": {},
            "forecast_bridge": [],
        },
    )
    monkeypatch.setattr(
        export_service,
        "build_comps_dashboard_view",
        lambda ticker: {
            "source_lineage": {"source_file": "public_market_yfinance_fallback", "as_of_date": "2026-06-07"},
            "peer_counts": {"clean": 3},
            "target_vs_peers": {"target": {}, "peer_medians": {"tev_ebitda_ltm": 14.0}},
            "peers": [],
        },
    )
    monkeypatch.setattr(export_service, "build_wacc_workbench", lambda ticker, apply_overrides=True: {"effective_preview": {}})
    monkeypatch.setattr(export_service, "build_research_board_view", lambda ticker: {"company_name": "IBM", "tracker": {}})
    monkeypatch.setattr(
        export_service,
        "build_analyst_prep_export_payload",
        lambda ticker: {
            "ticker": ticker,
            "generated_at": "2026-06-07T00:00:00+00:00",
            "source_quality": "missing",
            "sections": [],
            "thesis_cards": [],
            "driver_cards": [],
            "comps_card": None,
            "missing_data": [],
            "segment_driver_rows": [],
            "evidence_packet_ids": [],
            "evidence_map": [],
            "conflict_groups": [],
            "export_metadata": {"status": "test"},
        },
    )

    payload = export_service._build_current_ticker_payload("IBM")

    assert payload["source_lineage"]["exit_multiple"] == "public_market_yfinance_fallback_tev_ebitda_ltm"
    assert payload["ciq_lineage"]["public_comps_fallback_used"] is True
    assert payload["ciq_lineage"]["public_comps_fallback_peer_count"] == 3
    assert payload["ciq_lineage"]["comps_source_file"] == "public_market_yfinance_fallback"
    assert payload["default_resolution"]["status"] == "ok"


def test_register_export_bundle_persists_export_and_artifacts(monkeypatch):
    from src.stage_04_pipeline import export_service

    tmp_path = _workspace_tempdir("export-registry")
    db_path = tmp_path / "exports.db"
    monkeypatch.setattr(export_service, "get_connection", _temp_conn_factory(db_path))

    bundle_dir = tmp_path / "bundle"
    bundle_dir.mkdir()
    primary = bundle_dir / "report.html"
    primary.write_text("<html>ok</html>", encoding="utf-8")
    sidecar = bundle_dir / "context.json"
    sidecar.write_text("{}", encoding="utf-8")

    export_row = export_service.register_export_bundle(
        scope="ticker",
        export_format="html",
        source_mode="latest_snapshot",
        template_strategy="html_bundle",
        ticker="IBM",
        created_by="api",
        title="IBM Memo Export",
        bundle_dir=bundle_dir,
        primary_artifact_key="html_report",
        artifacts=[
            {
                "artifact_key": "html_report",
                "artifact_role": "primary",
                "title": "IBM Memo",
                "path": primary,
                "mime_type": "text/html",
                "is_primary": True,
            },
            {
                "artifact_key": "context_json",
                "artifact_role": "sidecar",
                "title": "Context",
                "path": sidecar,
                "mime_type": "application/json",
                "is_primary": False,
            },
        ],
        metadata={"source_label": "Latest snapshot"},
    )

    assert export_row["ticker"] == "IBM"
    assert export_row["primary_artifact_key"] == "html_report"

    listed = export_service.list_exports(ticker="IBM")
    loaded = export_service.load_export(export_row["export_id"])

    assert len(listed) == 1
    assert listed[0]["export_id"] == export_row["export_id"]
    assert loaded is not None
    assert loaded["artifacts"][0]["artifact_key"] == "html_report"
    assert Path(export_service.resolve_export_artifact_path(export_row["export_id"])).name == "report.html"
    assert Path(export_service.resolve_export_artifact_path(export_row["export_id"], "context_json")).name == "context.json"


def test_ticker_exports_persist_current_and_snapshot_dossiers(monkeypatch):
    from src.stage_04_pipeline import export_service
    from src.stage_04_pipeline.ticker_dossier import build_ticker_dossier_from_export_payload, ticker_dossier_to_payload

    tmp_path = _workspace_tempdir("export-dossier")
    db_path = tmp_path / "exports.db"
    monkeypatch.setattr(export_service, "get_connection", _temp_conn_factory(db_path))
    monkeypatch.setattr(export_service, "_ticker_bundle_dir", lambda ticker, export_format: tmp_path / f"{ticker}-{export_format}")

    def _payload(source_mode: str, snapshot_id: int | None = None) -> dict:
        payload = {
            "ticker": "IBM",
            "company_name": "International Business Machines",
            "industry": "IT Services",
            "exchange": "NYSE",
            "generated_at": "2026-04-30T12:00:00+00:00",
            "market": {"price": 260.0},
            "valuation": {"iv_base": 202.0, "expected_iv": 205.0},
            "ciq_lineage": {"snapshot_as_of_date": "2026-04-30"},
            "forecast_bridge": [{"year": 2027, "fcff": 100.0}],
            "historical_series": {"revenue": [{"period": "2025", "value": 1000.0}]},
            "qoe": {
                "qoe_score": 2.0,
                "qoe_flag": "red",
                "deterministic": {"signal_scores": {"dso": "amber"}},
                "llm": {
                    "dcf_ebit_override_pending": True,
                    "revenue_recognition_flags": ["Channel stuffing risk"],
                    "auditor_flags": [],
                },
            },
        }
        dossier = build_ticker_dossier_from_export_payload(payload, source_mode=source_mode, snapshot_id=snapshot_id)
        payload["ticker_dossier"] = ticker_dossier_to_payload(dossier)
        return payload

    def _stage_workbook(ticker, payload, bundle_dir):
        bundle_dir.mkdir(parents=True, exist_ok=True)
        primary = bundle_dir / "workbook.xlsx"
        primary.write_bytes(b"xlsx")
        return {
            "artifacts": [
                {
                    "artifact_key": "excel_workbook",
                    "artifact_role": "primary",
                    "title": "Workbook",
                    "path": primary,
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "is_primary": True,
                }
            ]
        }

    def _stage_html(ticker, context, bundle_dir):
        bundle_dir.mkdir(parents=True, exist_ok=True)
        primary = bundle_dir / "report.html"
        primary.write_text("<html>ok</html>", encoding="utf-8")
        return {
            "artifacts": [
                {
                    "artifact_key": "html_report",
                    "artifact_role": "primary",
                    "title": "Report",
                    "path": primary,
                    "mime_type": "text/html",
                    "is_primary": True,
                }
            ]
        }

    monkeypatch.setattr(export_service, "_build_snapshot_ticker_payload", lambda ticker: (_payload("latest_snapshot", 7), 7))
    monkeypatch.setattr(export_service, "_build_html_context", lambda ticker, source_mode: (_payload("loaded_backend_state"), None))
    monkeypatch.setattr(export_service, "stage_power_query_workbook", _stage_workbook)
    monkeypatch.setattr(export_service, "build_html_export_bundle", _stage_html)

    export_service.run_ticker_export(ticker="IBM", export_format="xlsx", source_mode="latest_snapshot")
    export_service.run_ticker_export(ticker="IBM", export_format="html", source_mode="loaded_backend_state")

    conn = _temp_conn_factory(db_path)()
    rows = conn.execute(
        """
        SELECT source_mode, source_key
        FROM ticker_dossier_snapshots
        ORDER BY source_mode
        """
    ).fetchall()

    assert [(row["source_mode"], row["source_key"]) for row in rows] == [
        ("latest_snapshot", "snapshot:7"),
        ("loaded_backend_state", "asof:2026-04-30"),
    ]

    payloads = [
        json.loads(row["payload_json"])
        for row in conn.execute(
            """
            SELECT payload_json
            FROM ticker_dossier_snapshots
            ORDER BY source_mode
            """
        ).fetchall()
    ]
    for dossier_payload in payloads:
        latest = dossier_payload["latest_snapshot"]
        assert latest["company_identity"]["industry"] == "IT Services"
        assert latest["company_identity"]["exchange"] == "NYSE"
        assert latest["qoe_snapshot"]["present"] is True
        assert latest["qoe_snapshot"]["flags"] == ["red", "dso:amber", "Channel stuffing risk", "dcf_ebit_override_pending"]
        assert latest["historical_series"]["revenue"] == [{"period": "2025", "value": 1000.0}]
