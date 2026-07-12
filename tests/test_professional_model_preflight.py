from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from scripts.manual import professional_model_preflight as preflight
from tests.ciq_test_utils import create_ibm_style_workbook


def test_resolve_workbook_rejects_missing_explicit_path(tmp_path: Path) -> None:
    with pytest.raises(preflight.PreflightError, match="workbook does not exist"):
        preflight.resolve_workbook_path(
            ticker="MSFT",
            workbook_path=tmp_path / "missing.xlsx",
            search_roots=(tmp_path,),
        )


def test_resolve_workbook_rejects_ambiguous_discovery(tmp_path: Path) -> None:
    (tmp_path / "MSFT_Standard.xlsx").write_bytes(b"one")
    other = tmp_path / "other"
    other.mkdir()
    (other / "MSFT_Standard.xlsx").write_bytes(b"two")

    with pytest.raises(preflight.PreflightError, match="multiple workbook candidates"):
        preflight.resolve_workbook_path(
            ticker="MSFT",
            workbook_path=None,
            search_roots=(tmp_path, other),
        )


def test_inspect_workbook_cells_reports_cached_excel_errors(tmp_path: Path) -> None:
    path = tmp_path / "errors.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "=1+1"
    ws["A2"] = "#REF!"
    wb.save(path)

    result = preflight.inspect_workbook_cells(path)

    assert result["formula_count"] == 1
    assert result["error_count"] == 1
    assert result["errors"][0]["cell"] == "A2"


def test_inspect_workbook_cells_reports_formula_reference_errors_even_with_no_cached_error(
    tmp_path: Path,
) -> None:
    path = tmp_path / "formula-errors.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Comps"
    ws["A1"] = "=SUM(1,#REF!)"
    wb.save(path)

    result = preflight.inspect_workbook_cells(path)

    assert result["formula_count"] == 1
    assert result["formula_error_count"] == 1
    assert result["cached_error_count"] == 0
    assert result["error_count"] == 1
    assert result["errors"] == [
        {
            "sheet": "Detailed Comps",
            "cell": "A1",
            "value": "=SUM(1,#REF!)",
            "kind": "formula_reference_error",
        }
    ]


def test_build_manifest_rejects_ticker_mismatch(tmp_path: Path) -> None:
    workbook = create_ibm_style_workbook(tmp_path / "TEST_Standard.xlsx")

    with pytest.raises(preflight.PreflightError, match="ticker mismatch"):
        preflight.build_preflight_manifest(
            ticker="MSFT",
            workbook_path=workbook,
            db_path=tmp_path / "missing.db",
        )


def test_build_manifest_pins_matching_ingest_run(tmp_path: Path) -> None:
    workbook = create_ibm_style_workbook(tmp_path / "TEST_Standard.xlsx")
    payload = preflight.parse_ciq_workbook(workbook)
    db_path = tmp_path / "alpha_pod.db"
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE ciq_ingest_runs (
                id INTEGER PRIMARY KEY,
                source_file TEXT,
                file_hash TEXT,
                ticker TEXT,
                parser_version TEXT,
                ingest_ts TEXT,
                status TEXT,
                as_of_date TEXT
            )
            """
        )
        conn.execute(
            """
            INSERT INTO ciq_ingest_runs
                (id, source_file, file_hash, ticker, parser_version, ingest_ts, status, as_of_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                7,
                workbook.name,
                payload.file_hash,
                "TEST",
                payload.parser_version,
                "2026-07-11T20:00:00Z",
                "completed",
                "2026-03-31",
            ),
        )

    manifest = preflight.build_preflight_manifest(
        ticker="TEST",
        workbook_path=workbook,
        db_path=db_path,
    )

    assert manifest["source"]["sha256"] == payload.file_hash
    assert manifest["source"]["ingest_status"] == "matched"
    assert manifest["source"]["run_id"] == 7
    assert manifest["parser"]["rows_parsed"] == payload.rows_parsed


def test_write_manifest_uses_fingerprint_directory(tmp_path: Path) -> None:
    manifest = {"ticker": "MSFT", "source": {"sha256": "a" * 64}}

    path = preflight.write_manifest(manifest, output_root=tmp_path)

    assert path == tmp_path / "MSFT" / ("a" * 12) / "preflight_fingerprint.json"
    assert json.loads(path.read_text(encoding="utf-8")) == manifest
