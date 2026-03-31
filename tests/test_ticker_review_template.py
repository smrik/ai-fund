from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _sheet_contains_value(ws, expected: str) -> bool:
    needle = expected.upper()
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if needle in str(value or "").upper():
                return True
    return False


def test_canonical_ticker_review_template_has_comps_diagnostics_tab():
    workbook_path = Path("templates/ticker_review.xlsx")
    wb = load_workbook(workbook_path)

    assert "Comps" in wb.sheetnames
    assert "Comps Diagnostics" in wb.sheetnames

    comps = wb["Comps"]
    diagnostics = wb["Comps Diagnostics"]

    assert _sheet_contains_value(comps, "Comparable Companies")
    assert _sheet_contains_value(comps, "Headline Valuation")
    assert _sheet_contains_value(comps, "Peer Table")
    assert _sheet_contains_value(diagnostics, "Comps Diagnostics")
    assert _sheet_contains_value(diagnostics, "Audit Flags")
    assert _sheet_contains_value(diagnostics, "Metric Status")
