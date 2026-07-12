from __future__ import annotations

from pathlib import Path

import pytest

from ciq.workbook_parser import parse_ciq_workbook
from tests.ciq_test_utils import create_ibm_style_workbook


MSFT_STANDARD_WORKBOOK = Path("data/exports/MSFT_Standard.xlsx")


def test_valuation_snapshot_ignores_common_size_duplicate_metric_rows(tmp_path: Path) -> None:
    workbook_path = create_ibm_style_workbook(tmp_path / "TEST_Standard.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    cs = wb["Common Size"]
    cs["A13"] = "Total Revenues"
    cs["D13"] = 1.0
    cs["E13"] = 1.0
    cs["F13"] = 1.0
    wb.save(workbook_path)

    payload = parse_ciq_workbook(workbook_path)
    snapshot = payload.valuation_snapshot

    assert snapshot["revenue_mm"] == 1200
    assert snapshot["ebit_margin"] == pytest.approx(150 / 1200)
    assert snapshot["revenue_cagr_3yr"] == pytest.approx(0.10)


def test_committed_cleandata_workbook_has_plausible_ciq_growth_snapshot() -> None:
    payload = parse_ciq_workbook("ciq/templates/ciq_cleandata.xlsx")
    snapshot = payload.valuation_snapshot

    assert snapshot["ticker"] == payload.ticker
    assert snapshot["revenue_mm"] is not None
    assert snapshot["revenue_cagr_3yr"] is not None
    assert -1.0 < snapshot["revenue_cagr_3yr"] < 1.0


def test_da_prefers_total_cash_flow_da_over_ambiguous_zero_row(tmp_path: Path) -> None:
    workbook_path = create_ibm_style_workbook(tmp_path / "TEST_Standard.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    fs = wb["Financial Statements"]
    fs["D121"], fs["E121"], fs["F121"] = 0, 0, 0
    fs["A122"] = "D&A For EBITDA"
    fs["D122"], fs["E122"], fs["F122"] = 31, 33, 35
    fs["A123"] = "Depreciation & Amort., Total"
    fs["D123"], fs["E123"], fs["F123"] = 31, 33, 35
    wb.save(workbook_path)

    payload = parse_ciq_workbook(workbook_path)

    assert payload.valuation_snapshot["da_mm"] == 35
    assert payload.valuation_snapshot["da_pct_avg_3yr"] == pytest.approx((33 / 1100 + 31 / 1000) / 2)


def test_long_form_facts_retain_exact_source_cell_and_formula_state(tmp_path: Path) -> None:
    workbook_path = create_ibm_style_workbook(tmp_path / "TEST_Standard.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    wb["Financial Statements"]["D12"] = "=500+500"
    wb.save(workbook_path)

    payload = parse_ciq_workbook(workbook_path)
    facts = {
        row["a1_locator"]: row
        for row in payload.long_form_records
        if row["sheet_name"] == "Financial Statements" and row["row_label"] == "Total Revenues"
    }

    formula_fact = facts["D12"]
    assert formula_fact["row_index"] == 12
    assert formula_fact["column_index"] == 4
    assert formula_fact["source_row_id"] == "Financial Statements!12"
    assert formula_fact["cell_locator"] == "Financial Statements!D12"
    assert formula_fact["formula_text"] == "=500+500"
    assert formula_fact["cached_value"] is None
    assert formula_fact["has_formula"] is True
    assert formula_fact["has_cached_value"] is False
    assert formula_fact["formula_status"] == "formula_cache_missing"
    assert formula_fact["formula_error"] is None
    assert formula_fact["cached_error"] is None

    literal_fact = facts["F12"]
    assert literal_fact["row_index"] == 12
    assert literal_fact["cell_locator"] == "Financial Statements!F12"
    assert literal_fact["formula_text"] is None
    assert literal_fact["cached_value"] == 1200
    assert literal_fact["formula_status"] == "literal"


def test_duplicate_logical_rows_remain_distinct_by_source_row_and_cell(tmp_path: Path) -> None:
    workbook_path = create_ibm_style_workbook(tmp_path / "DUPLICATE_ROWS.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    fs = wb["Financial Statements"]
    fs["A206"] = "Total Revenues"
    fs["D206"], fs["E206"], fs["F206"] = 900, 1000, 1100
    wb.save(workbook_path)

    payload = parse_ciq_workbook(workbook_path)
    latest_revenue_facts = [
        row
        for row in payload.long_form_records
        if row["sheet_name"] == "Financial Statements"
        and row["row_label"] == "Total Revenues"
        and row["period_date"] == "2025-12-31"
    ]

    assert len(latest_revenue_facts) == 2
    assert {row["row_index"] for row in latest_revenue_facts} == {12, 206}
    assert {row["a1_locator"] for row in latest_revenue_facts} == {"F12", "F206"}
    assert len({row["cell_locator"] for row in latest_revenue_facts}) == 2
    # The additive provenance fields must not change the existing v1 snapshot choice.
    assert payload.valuation_snapshot["revenue_mm"] == 1200


def test_msft_source_formula_reference_failures_are_cell_identified() -> None:
    assert MSFT_STANDARD_WORKBOOK.exists(), "The tracked MSFT Standard fixture is required for this regression"

    payload = parse_ciq_workbook(MSFT_STANDARD_WORKBOOK)
    expected_cells = {
        f"{column}{row}"
        for row in range(3, 11)
        for column in ("AC", "AG", "AK")
    }
    detailed_facts = {
        row["a1_locator"]: row
        for row in payload.long_form_records
        if row["sheet_name"] == "Detailed Comps"
    }
    formula_error_cells = {
        locator
        for locator, row in detailed_facts.items()
        if row["formula_error"] == "#REF!"
    }

    assert formula_error_cells == expected_cells
    for locator in expected_cells:
        fact = detailed_facts[locator]
        assert fact["row_index"] == int("".join(filter(str.isdigit, locator)))
        assert fact["cell_locator"] == f"Detailed Comps!{locator}"
        assert fact["formula_text"].startswith("=_xll.ciqfunctions.udf.CIQ(")
        assert "#REF!" in fact["formula_text"]
        assert fact["cached_value"] is not None
        assert fact["has_formula"] is True
        assert fact["has_cached_value"] is True
        assert fact["formula_status"] == "formula_error"
        assert fact["cached_error"] is None

    representative = detailed_facts["AG3"]
    assert representative["value_num"] == pytest.approx(float(representative["cached_value"]))
