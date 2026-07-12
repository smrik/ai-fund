"""Unit tests for src/stage_02_valuation/json_exporter.py — no network, no DB."""
from __future__ import annotations

import json
import tempfile
from pathlib import Path

import pytest

from src.stage_02_valuation.json_exporter import (
    build_excel_flat_tables,
    build_historical_financials_from_ciq_workbook,
    build_nested_structure,
    export_ticker_json,
    _json_default,
)


# ── Fixtures ─────────────────────────────────────────────────────────────────

MINIMAL_RESULT: dict = {
    "ticker": "IBM",
    "company_name": "International Business Machines",
    "sector": "Technology",
    "industry": "Information Technology Services",
    "price": 260.0,
    "market_cap_mm": 24_000.0,
    "ev_mm": 30_000.0,
    "pe_trailing": 20.0,
    "pe_forward": 18.0,
    "ev_ebitda": 12.0,
    "revenue_mm": 62_000.0,
    "growth_near": 6.5,
    "growth_mid": 4.0,
    "ebit_margin_used": 14.2,
    "capex_pct_used": 2.1,
    "da_pct_used": 3.5,
    "tax_rate_used": 18.0,
    "dso_used": 47.0,
    "dio_used": 6.5,
    "dpo_used": 26.0,
    "exit_multiple_used": 12.5,
    "exit_metric_used": "ebitda",
    "net_debt_mm": 22_000.0,
    "ronic_terminal_used": 10.0,
    "non_operating_assets_used_mm": 500.0,
    "minority_interest_used_mm": 0.0,
    "preferred_equity_used_mm": 0.0,
    "pension_deficit_used_mm": 3_000.0,
    "lease_liabilities_used_mm": 800.0,
    "options_value_used_mm": 0.0,
    "convertibles_value_used_mm": 0.0,
    "wacc": 8.5,
    "cost_of_equity": 10.2,
    "beta_raw": 0.72,
    "beta_unlevered": 0.60,
    "beta_relevered": 0.70,
    "size_premium": 0.5,
    "equity_weight": 70.0,
    "peers_used": "ACN, ORCL, MSFT",
    "iv_bear": 140.0,
    "iv_base": 240.0,
    "iv_bull": 380.0,
    "expected_iv": 250.0,
    "upside_base_pct": -7.7,
    "upside_bear_pct": -46.2,
    "upside_bull_pct": 46.2,
    "expected_upside_pct": -3.8,
    "margin_of_safety": -8.3,
    "iv_gordon": 220.0,
    "iv_exit": 240.0,
    "iv_blended": 230.0,
    "ep_iv_base": 235.0,
    "fcfe_iv_base": 228.0,
    "dcf_ep_gap_pct": 2.1,
    "comps_iv_ev_ebitda": 200.0,
    "comps_iv_ev_ebit": 195.0,
    "comps_iv_pe": 210.0,
    "comps_iv_base": 202.0,
    "implied_growth_pct": 7.2,
    "model_applicability_status": "dcf_applicable",
    "tv_pct_of_ev": 68.5,
    "tv_gordon_mm": 180_000.0,
    "tv_exit_mm": 210_000.0,
    "tv_blended_mm": 195_000.0,
    "pv_tv_gordon_mm": 90_000.0,
    "pv_tv_exit_mm": 105_000.0,
    "pv_tv_blended_mm": 97_000.0,
    "terminal_growth_pct": 2.5,
    "terminal_ronic_pct": 10.0,
    "gordon_formula_mode": "nopat_reinvestment",
    "ev_operations_mm": 28_000.0,
    "ev_total_mm": 28_500.0,
    "non_operating_assets_mm": 500.0,
    "non_equity_claims_mm": 25_800.0,
    "tv_high_flag": False,
    "health_tv_extreme_flag": False,
    "tv_method_fallback_flag": False,
    "roic_consistency_flag": False,
    "nwc_driver_quality_flag": False,
    "health_terminal_growth_guardrail_flag": False,
    "health_terminal_ronic_guardrail_flag": False,
    "health_terminal_denominator_guardrail_flag": False,
    "health_fcff_interest_contamination_flag": False,
    "ep_reconcile_flag": False,
    "scenario_prob_bear": 0.20,
    "scenario_prob_base": 0.60,
    "scenario_prob_bull": 0.20,
    "context_expected_iv": 248.0,
    "context_expected_upside_pct": -4.6,
    "context_scenario_policy_json": json.dumps({
        "policy": "context_advisory_v1",
        "official_policy": "fixed_default",
        "ticker": "IBM",
        "probability_source": "fixed_default",
        "context_inputs": {
            "cyclicality": "medium",
            "capital_intensity": "medium",
            "governance_risk": "low",
            "moat_strength": 4,
            "pricing_power": 4,
            "maturity": "middle",
            "driver_disagreement_width": 1.0,
            "shock_width": 1.0,
        },
        "official_specs": [
            {"name": "bear", "probability": 0.2, "growth_multiplier": 0.8},
            {"name": "base", "probability": 0.6, "growth_multiplier": 1.0},
            {"name": "bull", "probability": 0.2, "growth_multiplier": 1.2},
        ],
        "context_specs": [
            {"name": "bear", "probability": 0.2, "growth_multiplier": 0.82},
            {"name": "base", "probability": 0.6, "growth_multiplier": 1.0},
            {"name": "bull", "probability": 0.2, "growth_multiplier": 1.24},
        ],
    }),
    "driver_consensus_json": json.dumps([
        {
            "field": "revenue_growth_near",
            "current_value": 0.065,
            "suggested_value": 0.07,
            "suggested_range_low": 0.06,
            "suggested_range_high": 0.08,
            "source_count": 2,
            "agreement_level": "medium",
            "disagreement_flag": False,
            "official_action": "none",
            "sources": ["company", "industry"],
        }
    ]),
    "revenue_source": "ciq_snapshot",
    "growth_source": "ciq_consensus",
    "growth_source_detail": "CIQ FY1 consensus",
    "revenue_period_type": "ltm",
    "growth_period_type": "fy1",
    "revenue_alignment_flag": "aligned",
    "revenue_data_quality_flag": "ok",
    "ebit_margin_source": "ciq_snapshot",
    "capex_source": "ciq_snapshot",
    "da_source": "ciq_snapshot",
    "tax_source": "company_etr",
    "dso_source": "ciq_long_form",
    "dio_source": "ciq_long_form",
    "dpo_source": "ciq_long_form",
    "exit_multiple_source": "ciq_comps_fwd",
    "net_debt_source": "ciq_snapshot",
    "shares_source": "ciq_snapshot",
    "cost_of_equity_source": "yfinance_capm",
    "debt_weight_source": "yfinance_capm",
    "ronic_terminal_source": "sector_default",
    "invested_capital_source": "ciq_snapshot",
    "non_operating_assets_source": "ciq_snapshot",
    "minority_interest_source": "ciq_snapshot",
    "preferred_equity_source": "ciq_snapshot",
    "pension_deficit_source": "ciq_snapshot",
    "lease_liabilities_source": "ciq_snapshot",
    "options_value_source": "ciq_snapshot",
    "convertibles_value_source": "ciq_snapshot",
    "story_profile_source": "default",
    "default_resolution_json": json.dumps({
        "status": "review_required_high",
        "fields": [
            {
                "field": "exit_multiple",
                "value": 12.5,
                "source": "default",
                "source_class": "missing_default",
                "needs_pm_review": True,
            }
        ],
    }),
    "ciq_snapshot_used": True,
    "ciq_run_id": 42,
    "ciq_source_file": "IBM_ciq.xlsx",
    "ciq_as_of_date": "2026-03-01",
    "ciq_comps_used": True,
    "ciq_comps_run_id": 43,
    "ciq_comps_source_file": "IBM_comps.xlsx",
    "ciq_comps_as_of_date": "2026-03-01",
    "ciq_peer_count": 4,
    "public_comps_fallback_used": False,
    "public_comps_fallback_source_file": None,
    "public_comps_fallback_peer_count": None,
    "analyst_target": 310.0,
    "analyst_recommendation": "buy",
    "num_analysts": 20,
    "drivers_json": json.dumps({
        "revenue_base": 62_000_000_000,
        "revenue_growth_near": 0.065,
        "revenue_growth_mid": 0.04,
        "revenue_growth_terminal": 0.025,
        "ebit_margin_start": 0.142,
        "ebit_margin_target": 0.16,
        "capex_pct_start": 0.021,
        "da_pct_start": 0.035,
        "tax_rate_start": 0.18,
        "tax_rate_target": 0.20,
        "dso_start": 47.0,
        "dso_target": 45.0,
        "dio_start": 6.5,
        "dio_target": 6.0,
        "dpo_start": 26.0,
        "dpo_target": 28.0,
        "exit_multiple": 12.5,
        "exit_metric": "ebitda",
        "net_debt": 22_000_000_000,
        "shares_outstanding": 910_000_000,
        "ronic_terminal": 0.10,
        "invested_capital_start": 40_000_000_000,
        "wacc": 0.085,
        "risk_free_rate": 0.045,
        "equity_risk_premium": 0.05,
        "cost_of_debt": 0.04,
        "non_operating_assets": 500_000_000,
        "minority_interest": 0,
        "preferred_equity": 0,
        "pension_deficit": 3_000_000_000,
        "lease_liabilities": 800_000_000,
        "options_value": 0,
        "convertibles_value": 0,
    }),
    "forecast_bridge_json": json.dumps([
        {
            "year": i,
            "revenue": 62_000_000_000 * (1.065 ** i),
            "growth_rate": 0.065,
            "ebit_margin": 0.142,
            "fcff": 5_000_000_000,
            "pv_fcff": 4_500_000_000,
        }
        for i in range(1, 11)
    ]),
    "story_profile_json": json.dumps({"narrative": "AI transition", "moat": "hybrid_cloud"}),
    "story_adjustments_json": json.dumps({"revenue_adj": 0.0}),
}

MINIMAL_QOE: dict = {
    "ticker": "IBM",
    "sector": "Technology",
    "qoe_score": 3,
    "qoe_flag": "amber",
    "sloan_accruals_ratio": 0.05,
    "cash_conversion": 0.80,
    "dso_current": 47.0,
    "dso_baseline": 44.0,
    "dso_baseline_source": "ciq_history",
    "dso_drift": 3.0,
    "dio_current": 6.5,
    "dio_baseline": 6.0,
    "dio_baseline_source": "ciq_history",
    "dio_drift": 0.5,
    "dpo_current": 26.0,
    "dpo_baseline": 25.0,
    "dpo_baseline_source": "ciq_history",
    "dpo_drift": 1.0,
    "capex_da_ratio": 0.65,
    "signal_scores": {
        "accruals": "green",
        "cash_conversion": "green",
        "dso": "green",
        "dio": "green",
        "dpo": "green",
        "capex_da": "amber",
    },
    "accruals_thresholds": {"amber": 0.08, "red": 0.15},
}

MINIMAL_COMPS: dict = {
    "target": {
        "ticker": "IBM",
        "market_cap_mm": 24_000.0,
        "tev_mm": 30_000.0,
        "revenue_ltm_mm": 62_000.0,
        "ebitda_ltm_mm": 10_000.0,
        "ebit_ltm_mm": 8_800.0,
        "eps_ltm": 9.5,
        "tev_ebitda_ltm": 3.0,
        "tev_ebitda_fwd": 11.0,
        "tev_ebit_ltm": 3.4,
        "tev_ebit_fwd": 12.5,
        "pe_ltm": 27.4,
    },
    "peers": [
        {
            "ticker": "ACN",
            "market_cap_mm": 200_000.0,
            "tev_mm": 205_000.0,
            "revenue_ltm_mm": 64_000.0,
            "ebitda_ltm_mm": 12_000.0,
            "ebit_ltm_mm": 10_000.0,
            "eps_ltm": 11.0,
            "tev_ebitda_ltm": 17.0,
            "tev_ebitda_fwd": 15.5,
            "tev_ebit_ltm": 20.5,
            "tev_ebit_fwd": 18.0,
            "pe_ltm": 28.0,
        },
    ],
    "medians": {
        "tev_ebitda_ltm": 17.0,
        "tev_ebitda_fwd": 15.5,
        "tev_ebit_ltm": 20.5,
        "tev_ebit_fwd": 18.0,
        "pe_ltm": 28.0,
    },
}

MINIMAL_COMPS_ANALYSIS: dict = {
    "available": True,
    "primary_metric": "tev_ebitda_ltm",
    "peer_counts": {"raw": 5, "clean": 4},
    "similarity_method": "embedding_cosine",
    "similarity_model": "all-MiniLM-L6-v2",
    "weighting_formula": "0.60*description_similarity + 0.40*market_cap_proximity",
    "valuation_range": {
        "bear": 190.0,
        "base": 202.0,
        "bull": 216.0,
        "blended_base": 205.0,
    },
    "valuation_by_metric_rows": [
        {
            "metric": "tev_ebitda_ltm",
            "label": "TEV / EBITDA LTM",
            "target_multiple": 3.0,
            "peer_median_multiple": 17.0,
            "bear_multiple": 15.0,
            "base_multiple": 17.0,
            "bull_multiple": 19.0,
            "bear_iv": 190.0,
            "base_iv": 202.0,
            "bull_iv": 216.0,
            "n_raw": 5,
            "n_clean": 4,
            "outliers_removed": ["MSFT"],
            "is_primary": True,
        }
    ],
    "comparison_summary": [
        {
            "metric": "revenue_growth",
            "label": "Revenue Growth",
            "target": 0.03,
            "peer_median": 0.05,
            "delta": -0.02,
        }
    ],
    "peer_table": [
        {
            "ticker": "ACN",
            "display_name": "Accenture",
            "market_cap_mm": 200_000.0,
            "tev_ebitda_ltm": 17.0,
            "similarity_score": 0.81,
            "model_weight": 0.42,
        }
    ],
    "metric_status_rows": [
        {
            "ticker": "ACN",
            "metric": "tev_ebitda_ltm",
            "label": "TEV / EBITDA LTM",
            "raw_multiple": 17.0,
            "status": "included",
        }
    ],
    "football_field": {
        "ranges": [{"label": "TEV / EBITDA LTM", "bear": 190.0, "base": 202.0, "bull": 216.0}],
        "markers": [{"label": "Current Price", "value": 260.0, "type": "spot"}],
        "range_min": 190.0,
        "range_max": 260.0,
    },
    "historical_multiples_summary": {
        "available": True,
        "metrics": {
            "pe_trailing": {
                "current": 20.0,
                "summary": {"median": 18.0, "p25": 16.0, "p75": 22.0, "current_percentile": 0.65},
            }
        },
    },
    "audit_flags": ["Outliers removed from tev_ebitda_ltm: MSFT"],
    "notes": "primary=tev_ebitda_ltm",
    "source_lineage": {"as_of_date": "2026-03-01", "source_file": "IBM_comps.xlsx"},
}


# ── Tests: build_nested_structure ─────────────────────────────────────────────

class TestBuildNestedStructure:
    def test_top_level_keys(self):
        out = build_nested_structure(MINIMAL_RESULT)
        for key in ("$schema_version", "generated_at", "ticker", "company_name",
                    "sector", "market", "assumptions", "wacc", "valuation",
                    "method_availability", "scenarios", "scenario_policy", "context_scenarios",
                    "driver_consensus", "terminal", "health_flags", "forecast_bridge",
                    "historical_financials", "source_lineage", "ciq_lineage",
                    "story_profile", "story_adjustments", "default_resolution",
                    "drivers_raw", "excel_flat"):
            assert key in out, f"Missing top-level key: {key}"
        assert "assumption_register" in out
        assert "assumption_register_summary" in out

    def test_legacy_fcfe_shortcut_is_fail_closed(self):
        out = build_nested_structure(MINIMAL_RESULT)

        assert out["valuation"]["fcfe_iv_base"] is None
        assert out["method_availability"]["fcfe"] == {
            "status": "unavailable",
            "reason_code": "integrated_debt_interest_schedule_required",
            "detail": (
                "FCFE is withheld until after-tax interest and net borrowing "
                "tie to an integrated debt and cash schedule."
            ),
            "legacy_candidate_omitted": True,
        }
        flat = {row["key"]: row["value"] for row in out["excel_flat"]["valuation"]}
        assert flat["fcfe_iv_base"] is None

    def test_source_preflight_is_preserved(self):
        source_preflight = {
            "status": "blocked",
            "blockers": ["formula_reference_errors:24"],
            "source": {"sha256": "abc123", "run_id": 77},
            "workbook": {"formula_error_count": 24},
        }

        out = build_nested_structure(MINIMAL_RESULT, source_preflight=source_preflight)

        assert out["source_preflight"] == source_preflight

    def test_ticker_uppercase(self):
        r = dict(MINIMAL_RESULT)
        r["ticker"] = "ibm"
        out = build_nested_structure(r)
        assert out["ticker"] == "IBM"

    def test_forecast_bridge_is_array(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert isinstance(out["forecast_bridge"], list)
        assert len(out["forecast_bridge"]) == 10
        assert out["excel_flat"]["forecast"][0]["revenue_mm"] == pytest.approx(66_030.0)
        assert out["excel_flat"]["forecast"][0]["fcff_mm"] == pytest.approx(5_000.0)

    def test_excel_flat_forecast_money_units_reconcile_fcff(self):
        r = dict(MINIMAL_RESULT)
        r["forecast_bridge_json"] = json.dumps([
            {
                "year": 1,
                "nopat": 900_000_000,
                "da": 150_000_000,
                "capex": 100_000_000,
                "delta_nwc": 750_000,
                "fcff": 949_250_000,
            },
            {
                "year": 2,
                "nopat": 920_000_000,
                "da": 155_000_000,
                "capex": 110_000_000,
                "delta_nwc": -450_000,
                "fcff": 965_450_000,
            },
        ])

        forecast = build_nested_structure(r)["excel_flat"]["forecast"]

        assert forecast[0]["delta_nwc_mm"] == pytest.approx(0.75)
        assert forecast[1]["delta_nwc_mm"] == pytest.approx(-0.45)
        for row in forecast:
            assert (
                row["nopat_mm"]
                + row["da_mm"]
                - row["capex_mm"]
                - row["delta_nwc_mm"]
            ) == pytest.approx(row["fcff_mm"], abs=0.001)

    def test_forecast_bridge_empty_on_bad_json(self):
        r = dict(MINIMAL_RESULT)
        r["forecast_bridge_json"] = "not_json"
        out = build_nested_structure(r)
        assert out["forecast_bridge"] == []

    def test_drivers_raw_deserialised(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert isinstance(out["drivers_raw"], dict)
        assert "revenue_base" in out["drivers_raw"]

    def test_story_profile_deserialised(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert isinstance(out["story_profile"], dict)
        assert out["story_profile"].get("narrative") == "AI transition"

    def test_default_resolution_deserialised(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert out["default_resolution"]["status"] == "review_required_high"
        assert out["default_resolution"]["fields"][0]["field"] == "exit_multiple"

    def test_market_section(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert out["market"]["price"] == 260.0
        assert out["market"]["market_cap_mm"] == 24_000.0

    def test_wacc_section(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert out["wacc"]["wacc"] == pytest.approx(0.085, rel=1e-3)
        assert out["wacc"]["size_premium"] == pytest.approx(0.005)
        assert out["wacc"]["peers_used"] == ["ACN", "ORCL", "MSFT"]
        assert out["wacc"]["debt_weight"] == pytest.approx(0.30, rel=1e-2)

    def test_scenarios_keys(self):
        out = build_nested_structure(MINIMAL_RESULT)
        for sc in ("bear", "base", "bull"):
            assert sc in out["scenarios"]
            assert "probability" in out["scenarios"][sc]
            assert "iv" in out["scenarios"][sc]
            assert "upside_pct" in out["scenarios"][sc]

    def test_context_policy_is_advisory_and_separate(self):
        out = build_nested_structure(MINIMAL_RESULT)

        assert out["valuation"]["expected_iv"] == 250.0
        assert out["valuation"]["context_expected_iv"] == 248.0
        assert out["scenarios"]["bear"]["probability"] == pytest.approx(0.20)
        assert out["scenario_policy"]["official_policy"] == "fixed_default"
        assert out["context_scenarios"]["bear"]["growth_multiplier"] == pytest.approx(0.82)
        assert out["driver_consensus"][0]["official_action"] == "none"

    def test_health_flags_section(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert "tv_high_flag" in out["health_flags"]
        assert "tv_extreme_flag" in out["health_flags"]

    def test_no_comps_detail_by_default(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert "comps_detail" not in out

    def test_comps_detail_attached(self):
        out = build_nested_structure(MINIMAL_RESULT, comps_detail=MINIMAL_COMPS)
        assert "comps_detail" in out
        assert out["comps_detail"]["target"]["ticker"] == "IBM"
        assert len(out["comps_detail"]["peers"]) == 1

    def test_comps_analysis_attached(self):
        out = build_nested_structure(MINIMAL_RESULT, comps_analysis=MINIMAL_COMPS_ANALYSIS)
        assert "comps_analysis" in out
        assert out["comps_analysis"]["primary_metric"] == "tev_ebitda_ltm"
        assert out["comps_analysis"]["peer_table"][0]["ticker"] == "ACN"

    def test_qoe_not_present_by_default(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert "qoe" not in out

    def test_qoe_attached(self):
        out = build_nested_structure(MINIMAL_RESULT, qoe=MINIMAL_QOE)
        assert "qoe" in out
        assert out["qoe"]["qoe_score"] == 3
        assert out["qoe"]["qoe_flag"] == "amber"

    def test_ciq_lineage_section(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert out["ciq_lineage"]["snapshot_used"] is True
        assert out["ciq_lineage"]["peer_count"] == 4
        assert out["ciq_lineage"]["public_comps_fallback_used"] is False

    def test_historical_financials_attached_and_flattened(self):
        history = [
            {
                "period": "2025-12-31",
                "fiscal_year": "FY25",
                "revenue_mm": 100.0,
                "ebit_mm": 12.0,
                "ebit_margin_pct": 0.12,
            }
        ]
        out = build_nested_structure(MINIMAL_RESULT, historical_financials=history)
        assert out["historical_financials"] == history
        assert out["excel_flat"]["historical_financials"] == history

    def test_excel_flat_has_simple_power_query_tables(self):
        out = build_nested_structure(MINIMAL_RESULT, comps_detail=MINIMAL_COMPS, comps_analysis=MINIMAL_COMPS_ANALYSIS)
        flat = build_excel_flat_tables(out)
        assert flat["assumptions"][0] == {"key": "revenue_mm", "value": 62000.0}
        assert flat["scenarios"][0]["scenario"] == "bear"
        assert flat["market"][0] == {"key": "price", "value": 260.0}
        assert flat["comps_peers"][0]["ticker"] == "ACN"

    def test_assumptions_growth_terminal(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert out["assumptions"]["growth_terminal_pct"] == pytest.approx(0.025)

    def test_schema_version(self):
        out = build_nested_structure(MINIMAL_RESULT)
        assert out["$schema_version"] == "1.0"

    def test_empty_peers_used(self):
        r = dict(MINIMAL_RESULT)
        r["peers_used"] = ""
        out = build_nested_structure(r)
        assert out["wacc"]["peers_used"] == []

    def test_null_drivers_json(self):
        r = dict(MINIMAL_RESULT)
        r["drivers_json"] = ""
        out = build_nested_structure(r)
        assert out["drivers_raw"] == {}
        assert out["assumptions"]["growth_terminal_pct"] is None

    def test_assumption_register_is_top_level_full_payload(self):
        r = dict(MINIMAL_RESULT)
        r["assumption_register_json"] = json.dumps({
            "ticker": "IBM",
            "entries": [{"assumption_name": "wacc", "current_value": 0.085, "flag_level": "none"}],
        })
        r["assumption_register_summary_json"] = json.dumps({
            "model_trust_state": "clean",
            "flag_counts": {"none": 1},
            "max_flag_level": "none",
            "flagged_entries": [],
        })

        out = build_nested_structure(r)

        assert out["assumption_register"]["ticker"] == "IBM"
        assert out["assumption_register"]["entries"][0]["assumption_name"] == "wacc"
        assert out["assumption_register_summary"]["model_trust_state"] == "clean"


# ── Tests: export_ticker_json ─────────────────────────────────────────────────

@pytest.fixture()
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


class TestExportTickerJson:
    def test_creates_dated_and_latest(self, tmp_dir):
        dated = export_ticker_json(
            MINIMAL_RESULT, output_dir=tmp_dir, date_str="2026-03-09"
        )
        assert dated.exists()
        assert dated.name == "IBM_2026-03-09.json"
        latest = tmp_dir / "IBM_latest.json"
        assert latest.exists()

    def test_json_parseable(self, tmp_dir):
        dated = export_ticker_json(MINIMAL_RESULT, output_dir=tmp_dir, date_str="2026-01-01")
        content = json.loads(dated.read_text())
        assert content["ticker"] == "IBM"

    def test_json_contains_all_sections(self, tmp_dir):
        dated = export_ticker_json(
            MINIMAL_RESULT,
            qoe=MINIMAL_QOE,
            comps_detail=MINIMAL_COMPS,
            comps_analysis=MINIMAL_COMPS_ANALYSIS,
            historical_financials=[{"period": "2025-12-31", "revenue_mm": 100.0}],
            output_dir=tmp_dir, date_str="2026-01-01",
        )
        content = json.loads(dated.read_text())
        for sec in ("market", "assumptions", "wacc", "valuation", "scenarios",
                    "terminal", "health_flags", "forecast_bridge",
                    "historical_financials", "source_lineage", "ciq_lineage",
                    "comps_detail", "comps_analysis", "qoe", "excel_flat"):
            assert sec in content, f"Section {sec!r} missing from JSON"
        assert content["excel_flat"]["historical_financials"][0]["period"] == "2025-12-31"

    def test_ciq_workbook_history_extracts_annual_actuals(self, tmp_dir):
        from openpyxl import Workbook

        workbook_path = tmp_dir / "IBM_Standard.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Statements"
        ws["A1"] = "Period"
        ws["B1"] = "FY24"
        ws["C1"] = "FY25"
        ws["D1"] = "LTM"
        ws["A2"] = "Period Date"
        ws["B2"] = "2024-12-31"
        ws["C2"] = "2025-12-31"
        ws["D2"] = "2025-12-31"
        ws["A3"] = "Revenues"
        ws["B3"] = 100.0
        ws["C3"] = 110.0
        ws["D3"] = 110.0
        ws["A4"] = "Operating Income"
        ws["B4"] = 10.0
        ws["C4"] = 12.1
        ws["D4"] = 12.1
        ws["A5"] = "Capital Expenditure"
        ws["B5"] = -5.0
        ws["C5"] = -6.0
        ws["D5"] = -6.0
        ws["A6"] = "Depreciation & Amort."
        ws["B6"] = 3.0
        ws["C6"] = 3.3
        ws["D6"] = 3.3
        ws["A7"] = "Income Tax Expense"
        ws["B7"] = -2.0
        ws["C7"] = -2.5
        ws["D7"] = -2.5
        ws["A8"] = "EBT Excl Unusual Items"
        ws["B8"] = 8.0
        ws["C8"] = 9.6
        ws["D8"] = 9.6
        ws["A9"] = "Total Debt"
        ws["B9"] = 40.0
        ws["C9"] = 42.0
        ws["D9"] = 42.0
        ws["A10"] = "Cash and Equivalents"
        ws["B10"] = 5.0
        ws["C10"] = 7.0
        ws["D10"] = 7.0
        common = wb.create_sheet("Common Size")
        common["A1"] = "Period"
        common["B1"] = "FY24"
        common["A2"] = "Period Date"
        common["B2"] = "2024-12-31"
        common["A3"] = "Revenues"
        common["B3"] = 1.0
        comps = wb.create_sheet("Detailed Comps")
        comps["A1"] = "Ticker"
        comps["B1"] = "Name"
        comps["A2"] = "ACN"
        comps["B2"] = "Accenture"
        input_ws = wb.create_sheet("Input")
        input_ws["B2"] = "IBM"
        wb.save(workbook_path)

        rows = build_historical_financials_from_ciq_workbook(workbook_path)

        assert rows[-1]["period"] == "2025-12-31"
        assert rows[-1]["revenue_mm"] == 110.0
        assert rows[-1]["revenue_growth_pct"] == pytest.approx(0.10)
        assert rows[-1]["ebit_margin_pct"] == pytest.approx(0.11)
        assert rows[-1]["capex_mm"] == 6.0
        assert rows[-1]["tax_rate_pct"] == pytest.approx(2.5 / 9.6)
        assert rows[-1]["net_debt_mm"] == 35.0

        with pytest.raises(
            ValueError,
            match=r"CIQ workbook ticker mismatch: expected MSFT, found IBM",
        ):
            build_historical_financials_from_ciq_workbook(
                workbook_path,
                expected_ticker="MSFT",
            )

    def test_dated_and_latest_identical(self, tmp_dir):
        dated = export_ticker_json(MINIMAL_RESULT, output_dir=tmp_dir, date_str="2026-01-01")
        latest = tmp_dir / "IBM_latest.json"
        assert dated.read_text() == latest.read_text()

    def test_output_dir_created_if_missing(self, tmp_dir):
        nested_dir = tmp_dir / "a" / "b" / "c"
        export_ticker_json(MINIMAL_RESULT, output_dir=nested_dir, date_str="2026-01-01")
        assert nested_dir.exists()

    def test_returns_path_object(self, tmp_dir):
        result_path = export_ticker_json(MINIMAL_RESULT, output_dir=tmp_dir, date_str="2026-01-01")
        assert isinstance(result_path, Path)


# ── Tests: JSON serialisation helper ─────────────────────────────────────────

class TestJsonDefault:
    def test_handles_nan(self):
        assert _json_default(float("nan")) is None

    def test_handles_inf(self):
        assert _json_default(float("inf")) is None

    def test_handles_date(self):
        from datetime import date
        assert _json_default(date(2026, 3, 9)) == "2026-03-09"

    def test_handles_datetime(self):
        from datetime import datetime
        result = _json_default(datetime(2026, 3, 9, 12, 0, 0))
        assert result == "2026-03-09T12:00:00"

    def test_raises_on_unknown(self):
        with pytest.raises(TypeError):
            _json_default(object())


def test_wacc_debt_cost_keeps_pre_and_after_tax_fields_distinct():
    result = dict(MINIMAL_RESULT)
    result["wacc_cost_of_debt_after_tax_raw"] = 0.0316
    nested = build_nested_structure(result)

    assert nested["wacc"]["cost_of_debt"] == pytest.approx(0.04)
    assert nested["wacc"]["cost_of_debt_after_tax"] == pytest.approx(0.0316)
