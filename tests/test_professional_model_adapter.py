from __future__ import annotations

from datetime import date
import json
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest

from src.contracts.professional_financial_model import (
    AvailabilityState,
    AvailabilityStatus,
    CheckResult,
    CheckStatus,
    LineSeries,
    ModelPeriod,
    PeriodAxis,
    PeriodType,
    PeriodValue,
    PROFESSIONAL_WORKBOOK_SHEETS,
)
from src.stage_02_valuation.integrated_financial_model import HistoricalSourceReference
from src.stage_02_valuation.model_line_items import professional_line_item_registry
from src.stage_04_pipeline.professional_model_adapter import (
    BRIDGE_VALUATION_LINES,
    CORE_RENDERER_ALIASES,
    LEGACY_DIAGNOSTIC_VALUATION_FIELDS,
    ProfessionalModelAdapterError,
    adapt_professional_workbook_payload,
    build_diagnostic_scenario_forecasts,
    load_frozen_valuation_document,
    render_professional_model_v2_payload,
)


PERIODS = ("FY23", "FY24", "FY25", "LTM")
FORECAST_PERIODS = ("FY26E", "FY27E", "FY28E", "FY29E", "FY30E")
SOURCE_HASH = "a" * 64


def _available() -> AvailabilityState:
    return AvailabilityState(status=AvailabilityStatus.AVAILABLE)


def _unavailable(status: AvailabilityStatus, code: str) -> AvailabilityState:
    return AvailabilityState(status=status, reason_code=code, message=code.replace("_", " "))


class _FakeHistory:
    ticker = "MSFT"
    period_keys = PERIODS
    registry = professional_line_item_registry()

    def __init__(self) -> None:
        self.result = SimpleNamespace(
            state=_unavailable(AvailabilityStatus.PM_REQUIRED, "segment_source_required"),
            blockers=("source_dependent_modules_unavailable",),
            period_axis=PeriodAxis(
                periods=(
                    ModelPeriod(index=1, key="FY23", end_date=date(2023, 6, 30), period_type=PeriodType.FISCAL_YEAR),
                    ModelPeriod(index=2, key="FY24", end_date=date(2024, 6, 30), period_type=PeriodType.FISCAL_YEAR),
                    ModelPeriod(index=3, key="FY25", end_date=date(2025, 6, 30), period_type=PeriodType.FISCAL_YEAR),
                    ModelPeriod(index=4, key="LTM", end_date=date(2026, 3, 31), period_type=PeriodType.LTM),
                )
            ),
            check_results=(
                CheckResult(check_id="balance_sheet:LTM", status=CheckStatus.PASS, difference=0.0, tolerance=0.1),
                CheckResult(check_id="cash_flow:LTM", status=CheckStatus.PASS, difference=0.0, tolerance=0.1),
            ),
        )
        self._specs = {spec.canonical_key: spec for spec in self.registry}

    def _number(self, key: str, period: str) -> float:
        overrides = {
            "bs.cash": 32_105.0,
            "bs.short_term_investments": 40_000.0,
            "bs.long_term_investments": 6_123.0,
            "bs.total_debt": 125_432.0,
            "debt.total_debt": 100_000.0,
            "bs.short_term_borrowings": 5_000.0,
            "bs.current_long_term_debt": 5_000.0,
            "bs.long_term_debt": 90_000.0,
            "bs.current_lease_liabilities": 5_000.0,
            "bs.long_term_leases": 20_432.0,
            "debt.lease_liabilities": 25_432.0,
            "bs.minority_interest": 0.0,
            "bs.pension_liability": 12_345.0,
            "bs.net_debt": 47_204.0,
            "shares.diluted_weighted_average": 7_458.25,
            "is.revenue": 318_273.0,
            "is.ebit": 140_082.0,
            "is.net_income_common": 100_000.0,
            "is.net_income_company": 100_000.0,
            "is.minority_earnings": 0.0,
            "cf.dividends_paid": -20_000.0,
            "cf.da": 35_500.0,
            "cf.capex": -97_225.0,
            "wc.change_nwc": -5_000.0,
            "cf.intangible_amortization": 6_000.0,
            "tax.effective_rate": 20.0,
            "tax.cash_taxes": -20_000.0,
            "tax.pretax_income": 100_000.0,
            "bs.total_assets": 700_000.0,
            "bs.total_liabilities_equity": 700_000.0,
        }
        return overrides.get(key, float(self._specs[key].presentation_order * 10 + PERIODS.index(period)))

    def value(self, key: str, period: str) -> PeriodValue:
        spec = self._specs[key]
        if spec.statement_or_schedule == "segment_build":
            return PeriodValue(
                period_key=period,
                state=_unavailable(AvailabilityStatus.PM_REQUIRED, "segment_source_required"),
            )
        if spec.statement_or_schedule == "consensus_bridge":
            return PeriodValue(
                period_key=period,
                state=_unavailable(AvailabilityStatus.UNAVAILABLE, "consensus_source_required"),
            )
        return PeriodValue(period_key=period, value=self._number(key, period), state=_available())

    def period_lineage(self, key: str, period: str):
        spec = self._specs[key]
        state = self.value(key, period).state
        if spec.statement_or_schedule in {"segment_build", "consensus_bridge"}:
            return SimpleNamespace(
                canonical_key=key,
                period_key=period,
                method_id="historical:unavailable",
                formula_id=None,
                state=state,
                source_refs=(),
                normalized_value=None,
            )
        row = spec.presentation_order + 2
        column = PERIODS.index(period) + 2
        cell = f"{chr(64 + column)}{row}"
        ref = HistoricalSourceReference(
            source_ref=f"fact:{key}:{period}",
            ticker="MSFT",
            ciq_run_id=7,
            source_file="MSFT_Standard.xlsx",
            sheet_name="Financial Statements",
            row_index=row,
            column_index=column,
            a1_locator=cell,
            cell_locator=f"Financial Statements!{cell}",
            row_label=spec.display_label,
            formula_text=f"=CIQ({cell})",
            formula_status="formula_cached",
            source_value=self._number(key, period),
        )
        method = "historical:direct" if spec.source_mappings else "historical:derived"
        return SimpleNamespace(
            canonical_key=key,
            period_key=period,
            method_id=method,
            formula_id=None if method.endswith("direct") else f"historical:{key}",
            state=state,
            source_refs=(ref,),
            normalized_value=self._number(key, period),
        )


class _FakeForecastScenario:
    def __init__(self, scenario_key: str, offset: float) -> None:
        self.scenario_key = scenario_key
        periods = tuple(
            ModelPeriod(
                index=index,
                key=key,
                end_date=date(2025 + index, 6, 30),
                period_type=PeriodType.FISCAL_YEAR,
            )
            for index, key in enumerate(FORECAST_PERIODS, start=1)
        )
        self.model = SimpleNamespace(
            scenario_key=scenario_key,
            period_axis=PeriodAxis(periods=periods),
            state=_unavailable(AvailabilityStatus.PM_REQUIRED, "driver_approval_required"),
            blockers=(f"pm_approval_required:{scenario_key}:revenue_growth:queue:{scenario_key}",),
            check_results=(
                CheckResult(
                    check_id=f"balance_sheet:{FORECAST_PERIODS[-1]}",
                    status=CheckStatus.PASS,
                    difference=0.0,
                    tolerance=0.1,
                ),
            ),
        )
        self.line_index = {}
        for spec in professional_line_item_registry():
            values = []
            for index, period in enumerate(FORECAST_PERIODS, start=1):
                if spec.statement_or_schedule == "segment_build":
                    values.append(
                        PeriodValue(
                            period_key=period,
                            state=_unavailable(
                                AvailabilityStatus.PM_REQUIRED,
                                "segment_source_or_approval_required",
                            ),
                        )
                    )
                elif spec.statement_or_schedule == "consensus_bridge":
                    values.append(
                        PeriodValue(
                            period_key=period,
                            state=_unavailable(AvailabilityStatus.UNAVAILABLE, "consensus_unavailable"),
                        )
                    )
                else:
                    values.append(
                        PeriodValue(
                            period_key=period,
                            value=float(spec.presentation_order * 100 + index + offset),
                            state=_available(),
                        )
                    )
            self.line_index[spec.canonical_key] = LineSeries(
                line_key=spec.canonical_key,
                method_id="forecast:fixture",
                values=tuple(values),
            )


def _forecast_bundle():
    return SimpleNamespace(
        scenarios=(
            _FakeForecastScenario("Base", 0.0),
            _FakeForecastScenario("Upside", 10.0),
            _FakeForecastScenario("Downside", -10.0),
        )
    )


def _preflight(*, run_id: int = 7) -> dict:
    formula_errors = [
        {
            "sheet": "Detailed Comps",
            "cell": f"{column}{row}",
            "value": f"=CIQ(A{row},#REF!)",
            "kind": "formula_reference_error",
        }
        for row in range(3, 11)
        for column in ("AC", "AG", "AK")
    ]
    return {
        "ticker": "MSFT",
        "status": "blocked",
        "blockers": ["formula_reference_errors:24"],
        "warnings": [],
        "source": {
            "path": str(Path("data/exports/MSFT_Standard.xlsx").resolve()),
            "source_file": "MSFT_Standard.xlsx",
            "sha256": SOURCE_HASH,
            "run_id": run_id,
            "parser_version": "ibm_standard_v4",
            "ingest_status": "matched",
            "status": "completed",
            "workbook_as_of_date": "2026-03-31",
            "currency": "USD",
        },
        "parser": {"parser_version": "ibm_standard_v4", "rows_parsed": 8_601},
        "workbook": {
            "formula_error_count": 24,
            "cached_error_count": 0,
            "errors": formula_errors,
            "errors_truncated": False,
        },
    }

def _valuation() -> dict:
    return {
        "ticker": "MSFT",
        "company_name": "Microsoft Corporation",
        "market": {
            "price": 385.10,
            "market_cap_mm": 2_860_690.0,
            "current_fully_diluted_shares_mm": 7_450.0,
            "current_fully_diluted_shares_as_of": "2026-03-31",
        },
        "assumptions": {
            "growth_near_pct": 0.10,
            "growth_terminal_pct": 0.03,
            "tax_rate_target_pct": 0.204094,
            "capex_pct": 0.18,
            "da_pct": 0.08,
            "dso_start": 64.0,
            "dio_start": 4.0,
            "dpo_start": 115.0,
            "nopat_tax_rate_pct": 0.19,
            "shares_outstanding_mm": 7_428.435,
            "net_debt_mm": 47_204.0,
        },
        "wacc": {
            "asset_cost_disposals_mm": 1_000.0,
            "asset_disposal_accumulated_depreciation_mm": 600.0,
            "risk_free_rate": 0.045,
            "equity_risk_premium": 0.05,
            "beta_raw": None,
            "beta_relevered": 1.013,
            "cost_of_debt_after_tax": 0.0474,
            "quality_status": "degraded_fallback",
            "missing_inputs": ["beta"],
        },
        "valuation": {
            "iv_base": 266.69,
            "iv_gordon": 194.41,
            "iv_exit": 375.12,
            "iv_blended": 266.69,
            "ep_iv_base": 217.69,
            "comps_iv_ev_ebitda": 518.384,
            "comps_iv_ev_ebit": None,
            "comps_iv_pe": 475.0856,
            "comps_iv_base": 496.7348,
            "implied_growth_pct": 16.9,
            "fcfe_iv_base": 199.76,
            "expected_iv": 276.32,
        },
        "method_availability": {
            "fcfe": {
                "status": "unavailable",
                "reason_code": "integrated_debt_interest_schedule_required",
                "detail": (
                    "Legacy flat FCFE omits a supportable integrated debt and "
                    "after-tax interest schedule."
                ),
                "legacy_value_omitted": False,
            }
        },
        "forecast_bridge": [
            {
                "year": 1,
                "bridge_basis": "explicit_ytd_plus_q4_stub",
                "period_start": "2025-07-01",
                "period_end": "2026-06-30",
                "ytd_period_end": "2026-03-31",
                "stub_start": "2026-04-01",
                "fcff_stub_bridge": {
                    "base": {
                        "annual_fcff": 100.0,
                        "ytd_fcff": 75.0,
                        "stub_fcff": 25.0,
                    },
                    "upside": {
                        "annual_fcff": 110.0,
                        "ytd_fcff": 82.0,
                        "stub_fcff": 28.0,
                    },
                    "downside": {
                        "annual_fcff": 90.0,
                        "ytd_fcff": 70.0,
                        "stub_fcff": 20.0,
                    },
                },
                "stub_source_refs": ["fixture:fy26_ytd", "fixture:fy26_q4_stub"],
                "revenue": 309_896.4,
                "revenue_scale_to_model": 1.0,
                "growth_rate": 0.10,
                "ebit_margin": 0.40,
                "tax_rate": 0.20,
                "capex_pct": 0.18,
                "da_pct": 0.080,
                "dso": 64.0,
                "dio": 4.0,
                "dpo": 115.0,
            },
            {
                "year": 2,
                "growth_rate": 0.09,
                "ebit_margin": 0.39,
                "tax_rate": 0.205,
                "capex_pct": 0.175,
                "da_pct": 0.079,
                "dso": 62.0,
                "dio": 6.0,
                "dpo": 110.0,
            },
            {
                "year": 3,
                "growth_rate": 0.08,
                "ebit_margin": 0.38,
                "tax_rate": 0.21,
                "capex_pct": 0.17,
                "da_pct": 0.078,
                "dso": 60.0,
                "dio": 8.0,
                "dpo": 105.0,
            },
            {
                "year": 4,
                "growth_rate": 0.07,
                "ebit_margin": 0.37,
                "tax_rate": 0.215,
                "capex_pct": 0.165,
                "da_pct": 0.077,
                "dso": 58.0,
                "dio": 10.0,
                "dpo": 100.0,
            },
            {
                "year": 5,
                "growth_rate": 0.06,
                "ebit_margin": 0.36,
                "tax_rate": 0.22,
                "capex_pct": 0.16,
                "da_pct": 0.076,
                "dso": 56.0,
                "dio": 12.0,
                "dpo": 95.0,
            },
        ],
        "context_scenarios": {
            "bear": {
                "probability": 0.2,
                "growth_multiplier": 0.8,
                "margin_shift": -0.02,
            },
            "base": {
                "probability": 0.6,
                "growth_multiplier": 1.0,
                "margin_shift": 0.0,
            },
            "bull": {
                "probability": 0.2,
                "growth_multiplier": 1.2,
                "margin_shift": 0.02,
            },
        },
        "scenarios": {
            "base": {"probability": 0.6, "iv": 266.69},
            "bear": {"probability": 0.2, "iv": 180.53},
            "bull": {"probability": 0.2, "iv": 401.02},
        },
        "source_preflight": _preflight(),
        "ciq_lineage": {
        "dcf_scenario_governance": {
            "base": {
                "wacc": 0.08,
                "terminal_growth": 0.02,
                "nopat_tax_path_fingerprint": "1" * 64,
                "cash_tax_path_fingerprint": "2" * 64,
            },
            "upside": {
                "wacc": 0.075,
                "terminal_growth": 0.025,
                "nopat_tax_path_fingerprint": "3" * 64,
                "cash_tax_path_fingerprint": "4" * 64,
            },
            "downside": {
                "wacc": 0.09,
                "terminal_growth": 0.01,
                "nopat_tax_path_fingerprint": "5" * 64,
                "cash_tax_path_fingerprint": "6" * 64,
            },
        },
        "scenario_control": {
            "formula_first_gate": "PASS",
            "policy_gate": "PASS",
            "approval_ref": "fixture:scenario-policy",
            "current_input_hash": "7" * 64,
        },
            "snapshot_run_id": 7,
            "comps_run_id": 7,
            "comps_as_of_date": "2026-03-31",
        },
    }


def _payload():
    return adapt_professional_workbook_payload(
        ticker="MSFT",
        requested_run_id=7,
        preflight=_preflight(),
        historical=_FakeHistory(),
        forecast_bundle=_forecast_bundle(),
        valuation_document=_valuation(),
        comparables=(),
        valuation_source_path=Path("frozen-valuation.json"),
    )


def _capture_diagnostic_build(
    monkeypatch: pytest.MonkeyPatch,
    historical: _FakeHistory,
    valuation: dict,
) -> dict:
    from src.stage_02_valuation import integrated_financial_forecast as forecast_engine

    captured: dict = {}

    def fake_build(
        historical_seed,
        driver_sets,
        *,
        last_historical_period_end,
    ):
        captured["historical_seed"] = dict(historical_seed)
        captured["driver_sets"] = tuple(driver_sets)
        captured["last_historical_period_end"] = last_historical_period_end
        return SimpleNamespace(captured=True)

    monkeypatch.setattr(
        forecast_engine,
        "build_complete_scenario_forecasts",
        fake_build,
    )
    result = build_diagnostic_scenario_forecasts(historical, valuation)
    assert result.captured is True
    return captured


def test_diagnostic_paths_use_bridge_fy25_seed_and_context_deltas(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    historical = _FakeHistory()
    original_number = historical._number
    historical._number = (
        lambda key, period: 281_724.0
        if key == "is.revenue" and period == "FY25"
        else 318_273.0
        if key == "is.revenue" and period == "LTM"
        else original_number(key, period)
    )
    observed_periods: list[str] = []
    original_value = historical.value

    def tracked_value(key: str, period: str) -> PeriodValue:
        observed_periods.append(period)
        return original_value(key, period)

    historical.value = tracked_value
    valuation = _valuation()
    valuation["context_scenarios"]["base"]["probability"] = {
        "ignored": "must_not_be_parsed"
    }
    captured = _capture_diagnostic_build(monkeypatch, historical, valuation)

    assert captured["historical_seed"]["is.revenue"] == 281_724.0
    assert captured["last_historical_period_end"] == date(2025, 6, 30)
    assert "LTM" not in observed_periods

    driver_sets = {
        driver_set.scenario_key: driver_set
        for driver_set in captured["driver_sets"]
    }
    base = driver_sets["Base"].path_map
    upside = driver_sets["Upside"].path_map
    downside = driver_sets["Downside"].path_map
    bridge = valuation["forecast_bridge"]

    expected_bridge_paths = {
        "revenue_growth": "growth_rate",
        "da_percent_revenue": "da_pct",
        "effective_tax_rate": "tax_rate",
        "capex_percent_revenue": "capex_pct",
        "dso": "dso",
        "dio": "dio",
        "dpo": "dpo",
    }
    for driver_key, source_field in expected_bridge_paths.items():
        expected = tuple(row[source_field] for row in bridge)
        assert base[driver_key].values == pytest.approx(expected)
        assert len(set(base[driver_key].values)) > 1
    assert base["cash_tax_rate"].values == pytest.approx((0.2,) * 5)
    assert base["nopat_tax_rate"].values == pytest.approx((0.19,) * 5)
    assert "tax.cash_taxes" in base["cash_tax_rate"].source_ref
    assert "nopat_tax_rate_pct" in base["nopat_tax_rate"].source_ref
    assert base["effective_tax_rate"].source_ref != base["cash_tax_rate"].source_ref

    visible_keys = (
        "sga_percent_revenue",
        "rd_percent_revenue",
        "other_opex_percent_revenue",
    )
    visible_opex_ratio = sum(base[key].values[0] for key in visible_keys)
    for key in visible_keys:
        assert len(set(base[key].values)) == 1
        assert upside[key].values == base[key].values
        assert downside[key].values == base[key].values
    expected_base_gross_margin = tuple(
        row["ebit_margin"] + visible_opex_ratio for row in bridge
    )
    assert base["gross_margin"].values == pytest.approx(
        expected_base_gross_margin
    )

    expected_growth = tuple(row["growth_rate"] for row in bridge)
    assert upside["revenue_growth"].values == pytest.approx(
        tuple(value * 1.2 for value in expected_growth)
    )
    assert downside["revenue_growth"].values == pytest.approx(
        tuple(value * 0.8 for value in expected_growth)
    )
    assert upside["gross_margin"].values == pytest.approx(
        tuple(value + 0.02 for value in expected_base_gross_margin)
    )
    assert downside["gross_margin"].values == pytest.approx(
        tuple(value - 0.02 for value in expected_base_gross_margin)
    )
    assert 281_724.0 * (1.0 + base["revenue_growth"].values[0]) == pytest.approx(
        309_896.4
    )

    for driver_set in driver_sets.values():
        assert all(not path.approved for path in driver_set.paths)
        assert all(path.queue_item_id for path in driver_set.paths)
        assert all(len(path.current_driver_fingerprint) == 64 for path in driver_set.paths)
        assert all(path.driver_group.value == "finance_semantic" for path in driver_set.paths)
        assert all(
            "probability" not in f"{path.key}|{path.source_ref}|{path.method}".lower()
            for path in driver_set.paths
        )

    for key in (
        "minimum_cash",
        "scheduled_debt_issuance",
        "scheduled_debt_repayment",
    ):
        assert "common_fixed_debt_policy" in base[key].method
        assert base[key].values == upside[key].values == downside[key].values
    expected_pre_tax_cost = 0.0474 / (1.0 - 0.204094)
    assert base["cost_of_debt"].values == pytest.approx(
        (expected_pre_tax_cost,) * len(FORECAST_PERIODS)
    )
    assert "pretax_cost_of_debt" in base["cost_of_debt"].method
    assert "cost_of_debt_after_tax" in base["cost_of_debt"].source_ref
    liquid_keys = (
        "bs.cash",
        "bs.short_term_investments",
        "bs.long_term_investments",
    )
    average_liquid_assets = (
        sum(historical._number(key, "FY24") for key in liquid_keys)
        + sum(historical._number(key, "FY25") for key in liquid_keys)
    ) / 2.0
    expected_cash_yield = historical._number("is.interest_income", "FY25") / average_liquid_assets
    assert base["cash_yield"].values == pytest.approx(
        (expected_cash_yield,) * len(FORECAST_PERIODS)
    )
    assert "average_liquid_assets" in base["cash_yield"].method
    assert "cash_st_lt_investments" in base["cash_yield"].source_ref
    for key in ("cost_of_debt", "cash_yield"):
        assert base[key].values == upside[key].values == downside[key].values
    for key in (
        "dividend_payout",
        "buyback_amount",
        "common_stock_issuance",
        "average_share_price",
        "incremental_diluted_shares",
        "preferred_dividends",
    ):
        assert "common_fixed_capital_allocation_policy" in base[key].method
        assert base[key].values == upside[key].values == downside[key].values
    for key in (
        "other_operating_cash_flow",
        "other_investing_cash_flow",
        "other_financing_cash_flow",
        "fx_cash_adjustment",
        "misc_cash_adjustment",
    ):
        assert "common_fixed_zero_nonrecurring_cash_policy" in base[key].method
        assert base[key].values == (0.0,) * 5
        assert base[key].values == upside[key].values == downside[key].values
    assert base["scheduled_debt_issuance"].values == (0.0,) * 5
    assert base["scheduled_debt_repayment"].values == (0.0,) * 5


def test_diagnostic_bridge_paths_preserve_legitimate_zero_values(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    valuation = _valuation()
    first = valuation["forecast_bridge"][0]
    for key in (
        "growth_rate",
        "ebit_margin",
        "tax_rate",
        "capex_pct",
        "da_pct",
        "dso",
        "dio",
        "dpo",
    ):
        first[key] = 0.0
    first["revenue"] = _FakeHistory()._number("is.revenue", "FY25")
    valuation["context_scenarios"]["bull"].update(
        {"growth_multiplier": 0.0, "margin_shift": 0.0}
    )
    captured = _capture_diagnostic_build(
        monkeypatch,
        _FakeHistory(),
        valuation,
    )
    base = next(
        item for item in captured["driver_sets"] if item.scenario_key == "Base"
    ).path_map

    for key in (
        "revenue_growth",
        "effective_tax_rate",
        "capex_percent_revenue",
        "da_percent_revenue",
        "dso",
        "dio",
        "dpo",
    ):
        assert base[key].values[0] == 0.0
    assert base["cash_tax_rate"].values[0] == pytest.approx(0.2)
    assert base["nopat_tax_rate"].values[0] == pytest.approx(0.19)
    visible_opex_ratio = sum(
        base[key].values[0]
        for key in (
            "sga_percent_revenue",
            "rd_percent_revenue",
            "other_opex_percent_revenue",
        )
    )
    assert base["gross_margin"].values[0] == pytest.approx(
        visible_opex_ratio
    )
    upside = next(
        item for item in captured["driver_sets"] if item.scenario_key == "Upside"
    ).path_map
    assert upside["revenue_growth"].values == (0.0,) * 5
    assert upside["gross_margin"].values == pytest.approx(
        base["gross_margin"].values
    )


def test_adapter_preserves_registry_lines_lineage_and_fail_closed_states() -> None:
    payload = _payload()
    source_dependent = {
        spec.canonical_key
        for spec in professional_line_item_registry()
        if spec.statement_or_schedule in {"segment_build", "consensus_bridge"}
    }
    expected_line_count = (
        len(professional_line_item_registry()) - len(source_dependent) + len(CORE_RENDERER_ALIASES)
    )
    keys = {line.canonical_key for line in payload.lines}

    assert len(payload.lines) == expected_line_count
    assert {"is.revenue", "revenue", "cf.unlevered_fcf", "cash", "ending_cash"} <= keys
    assert not source_dependent & keys
    assert payload.source.status == "blocked"
    assert payload.source.formula_error_count == 24
    assert payload.availability["segments"].status == "pm_required"
    assert payload.availability["consensus"].status == "unavailable"
    assert payload.availability["sotp"].status == "unavailable"
    assert "source_formula_errors:24" in payload.blockers
    assert "wacc_degraded:degraded_fallback" in payload.blockers
    assert any(item.startswith("pm_approval_required:Base") for item in payload.blockers)
    assert "fcfe_unavailable" in payload.warnings
    assert "legacy_valuation_diagnostics_non_approved" in payload.warnings
    assert "legacy_valuation_diagnostics_model_blocked" in payload.warnings
    assert not any(
        "probability" in key or "expected" in key
        for key in payload.valuation_inputs
    )
    assert "comps_implied_price" not in payload.valuation_inputs
    assert not any(
        token in key
        for key in payload.valuation_inputs
        for token in ("target", "blend", "recommendation")
    )

    expected_error_cells = [
        f"Detailed Comps!{column}{row}"
        for row in range(3, 11)
        for column in ("AC", "AG", "AK")
    ]
    assert payload.backend_checks["source.formula_error_count"] == 24
    assert payload.backend_checks["source.formula_error_records_exposed"] == 24
    assert payload.backend_checks["source.formula_error_cells"].split(",") == expected_error_cells
    assert payload.backend_checks["source.formula_error.001.cell"] == "Detailed Comps!AC3"
    assert payload.backend_checks["source.formula_error.024.cell"] == "Detailed Comps!AK10"
    assert payload.backend_checks["fcfe.state"] == "UNAVAILABLE"
    assert (
        payload.backend_checks["fcfe.reason_code"]
        == "integrated_debt_interest_schedule_required"
    )
    assert "integrated debt" in payload.backend_checks["fcfe.detail"]
    assert payload.backend_checks["fcfe.legacy_value_state"] == "OMITTED_NON_APPROVED"
    assert payload.backend_checks["wacc.quality_status"] == "degraded_fallback"
    assert payload.backend_checks["wacc.missing_inputs"] == "beta"
    assert payload.backend_checks["wacc.missing_input_count"] == 1
    assert payload.backend_checks["wacc.approval_state"] == "BLOCKED_DEGRADED_INPUTS"
    assert payload.backend_checks["valuation.legacy_diagnostics.approval_state"] == "NON_APPROVED"
    assert payload.backend_checks["valuation.legacy_diagnostics.model_state"] == "BLOCKED_BY_MODEL"

    direct = next(line for line in payload.lines if line.canonical_key == "is.revenue")
    derived = next(line for line in payload.lines if line.canonical_key == "wc.receivables")
    alias = next(line for line in payload.lines if line.canonical_key == "revenue")
    assert direct.historical[-1].source_cell == "E3"
    assert direct.historical[-1].source_row_id.startswith("fact:is.revenue:LTM")
    assert derived.historical[-1].source_formula.startswith("DERIVED historical:wc.receivables")
    assert "fact:wc.receivables:LTM" in derived.historical[-1].source_formula
    assert alias.historical[-1].source_formula.startswith("RENDERER ALIAS is.revenue")


def test_adapter_carries_separate_diagnostics_and_exact_bridge_components() -> None:
    payload = _payload()

    expected_diagnostics = {
        "diagnostic_comps_ev_ebitda_per_share": 518.384,
        "diagnostic_comps_ev_ebit_per_share": None,
        "diagnostic_comps_pe_per_share": 475.0856,
        "diagnostic_economic_profit_per_share": 217.69,
        "diagnostic_v1_gordon_per_share": 194.41,
        "diagnostic_v1_exit_per_share": 375.12,
        "diagnostic_reverse_dcf_implied_growth_pct_points": 16.9,
    }
    assert {
        key: payload.valuation_inputs[key] for key in expected_diagnostics
    } == expected_diagnostics

    expected_bridge = {
        "bridge_cash": 32_105.0,
        "bridge_short_term_investments": 40_000.0,
        "bridge_long_term_investments": 6_123.0,
        "bridge_gross_debt": 125_432.0,
        "bridge_total_borrowings": 100_000.0,
        "bridge_short_term_borrowings": 5_000.0,
        "bridge_current_long_term_debt": 5_000.0,
        "bridge_long_term_debt": 90_000.0,
        "bridge_current_lease_liabilities": 5_000.0,
        "bridge_long_term_lease_liabilities": 20_432.0,
        "bridge_lease_liabilities": 25_432.0,
        "bridge_minority_interest": 0.0,
        "bridge_pension_liability": 12_345.0,
    }
    assert {key: payload.valuation_inputs[key] for key in expected_bridge} == expected_bridge
    for input_key, source_line in BRIDGE_VALUATION_LINES.items():
        component = input_key.removeprefix("bridge_")
        prefix = f"valuation_bridge.{component}"
        assert payload.backend_checks[f"{prefix}.source_line"] == source_line
        assert payload.backend_checks[f"{prefix}.period"] == "LTM"
        assert payload.backend_checks[f"{prefix}.state"] == "AVAILABLE"

    assert (
        payload.backend_checks["valuation_bridge.total_borrowings.source_lines"]
        == "bs.short_term_borrowings,bs.current_long_term_debt,bs.long_term_debt"
    )
    assert (
        payload.backend_checks["valuation_bridge.total_borrowings.derivation"]
        == "SUM_EXACT_CURRENT_LINES"
    )
    assert payload.backend_checks["valuation_bridge.claims_tie.status"] == "PASS"
    assert payload.backend_checks["valuation_bridge.claims_tie.difference"] == 0.0
    assert (
        payload.backend_checks["valuation_bridge.lease_liabilities_tie.status"]
        == "PASS"
    )
    assert (
        payload.backend_checks["valuation_bridge.lease_liabilities_tie.difference"]
        == 0.0
    )

    for input_key, source_field in LEGACY_DIAGNOSTIC_VALUATION_FIELDS.items():
        diagnostic = input_key.removeprefix("diagnostic_")
        prefix = f"valuation.legacy_diagnostic.{diagnostic}"
        assert payload.backend_checks[f"{prefix}.source_field"] == source_field
        expected_unit = (
            "PERCENTAGE_POINTS"
            if source_field == "implied_growth_pct"
            else "CURRENCY_PER_SHARE"
        )
        assert payload.backend_checks[f"{prefix}.unit"] == expected_unit
        assert payload.backend_checks[f"{prefix}.approval_state"] == "NON_APPROVED"
        assert payload.backend_checks[f"{prefix}.model_state"] == "BLOCKED_BY_MODEL"


def test_legitimate_zero_valuation_inputs_are_not_replaced_by_fallbacks() -> None:
    historical = _FakeHistory()
    original_number = historical._number
    zero_history_keys = set(BRIDGE_VALUATION_LINES.values()) | {"bs.net_debt"}
    historical._number = (
        lambda key, period: 0.0
        if key in zero_history_keys
        else original_number(key, period)
    )
    valuation = _valuation()
    valuation["market"].update({"price": 0.0, "market_cap_mm": 0.0, "current_fully_diluted_shares_mm": 0.0})
    valuation["assumptions"].update(
        {
            "growth_terminal_pct": 0.0,
            "tax_rate_target_pct": 0.0,
            "shares_outstanding_mm": 0.0,
        }
    )
    valuation["wacc"].update(
        {
            "risk_free_rate": 0.0,
            "equity_risk_premium": 0.0,
            "beta_raw": 0.0,
            "beta_relevered": 1.013,
            "cost_of_debt_after_tax": 0.0,
        }
    )
    for source_field in LEGACY_DIAGNOSTIC_VALUATION_FIELDS.values():
        valuation["valuation"][source_field] = 0.0

    payload = adapt_professional_workbook_payload(
        ticker="MSFT",
        requested_run_id=7,
        preflight=_preflight(),
        historical=historical,
        forecast_bundle=_forecast_bundle(),
        valuation_document=valuation,
        comparables=(),
        valuation_source_path=Path("frozen-zero-valuation.json"),
    )

    zero_keys = {
        "risk_free_rate",
        "equity_risk_premium",
        "beta",
        "pre_tax_cost_of_debt",
        "tax_rate",
        "debt_value",
        "terminal_growth",
        "net_debt",
        "current_basic_shares",
        "equity_value",
        *BRIDGE_VALUATION_LINES,
        "bridge_total_borrowings",
        *LEGACY_DIAGNOSTIC_VALUATION_FIELDS,
    }
    for key in zero_keys:
        assert payload.valuation_inputs[key] == 0.0
    assert payload.valuation_inputs["current_fully_diluted_shares"] is None
    assert payload.valuation_inputs["current_price"] is None
    price_evidence = payload.source_presentations[0]
    assert price_evidence.raw_value == 0.0
    assert price_evidence.normalized_value is None
    assert price_evidence.state.reason_code == "current_price_nonpositive_or_evidence_incomplete"
    assert "comps_implied_price" not in payload.valuation_inputs

def test_adapter_rejects_run_identity_mismatch() -> None:
    valuation = _valuation()
    valuation["source_preflight"] = _preflight(run_id=8)

    with pytest.raises(ProfessionalModelAdapterError, match="run identity"):
        adapt_professional_workbook_payload(
            ticker="MSFT",
            requested_run_id=7,
            preflight=_preflight(),
            historical=_FakeHistory(),
            forecast_bundle=_forecast_bundle(),
            valuation_document=valuation,
            comparables=(),
            valuation_source_path=Path("frozen-valuation.json"),
        )


def test_explicit_frozen_loader_rejects_latest_alias(tmp_path: Path) -> None:
    latest = tmp_path / "MSFT_latest.json"
    latest.write_text(json.dumps(_valuation()), encoding="utf-8")
    with pytest.raises(ProfessionalModelAdapterError, match="latest"):
        load_frozen_valuation_document(latest, ticker="MSFT", requested_run_id=7)


def test_render_writes_deterministic_run_directory_workbook_and_manifest(tmp_path: Path) -> None:
    artifacts = render_professional_model_v2_payload(_payload(), output_root=tmp_path)

    assert artifacts.output_dir == tmp_path / "MSFT" / "7"
    assert artifacts.workbook_path.exists()
    assert artifacts.manifest_path.exists()
    manifest_json = json.loads(artifacts.manifest_path.read_text(encoding="utf-8"))
    assert manifest_json["source_hash"] == SOURCE_HASH
    assert manifest_json["ticker"] == "MSFT"
    assert "source_preflight_blocked" in manifest_json["blockers"]
    assert "recalculation_not_run" in manifest_json["blockers"]

    workbook = load_workbook(artifacts.workbook_path, data_only=False)
    assert tuple(workbook.sheetnames) == PROFESSIONAL_WORKBOOK_SHEETS
    assert workbook["Cover"]["B10"].value == 7
    assert workbook["Historical_Data"]["C5"].comment is not None
    assert "Run 7" in workbook["Historical_Data"]["C5"].comment.text
    formula_count = sum(
        1
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    assert formula_count >= 1_100



def test_adapter_exposes_current_price_presentation_workflow_and_driver_approval_api() -> None:
    valuation = _valuation()
    valuation["driver_approvals"] = [
        {
            "driver_key": "revenue_growth",
            "scenario_key": "Base",
            "driver_group": "finance_semantic",
            "current_driver_fingerprint": "b" * 64,
            "approval_state": "UNAPPROVED",
        }
    ]
    valuation["ciq_lineage"]["comps_as_of_date"] = "2026-03-31"
    payload = adapt_professional_workbook_payload(
        ticker="MSFT",
        requested_run_id=7,
        preflight=_preflight(),
        historical=_FakeHistory(),
        forecast_bundle=_forecast_bundle(),
        valuation_document=valuation,
        comparables=(),
        valuation_source_path=Path("frozen-with-approval.json"),
    )

    assert payload.current_price_source.endswith("run 7 | stock_price")
    assert payload.current_price_as_of == date(2026, 3, 31)
    assert len(payload.source_presentations) == 1
    presentation = payload.source_presentations[0]
    assert presentation.canonical_key == "market.current_price"
    assert presentation.raw_value == pytest.approx(385.10)
    assert presentation.normalized_value == pytest.approx(385.10)
    assert presentation.as_of_date == date(2026, 3, 31)
    assert "dcf.current_price" in presentation.downstream_dependencies
    assert len(payload.driver_approvals) == 1
    approval = payload.driver_approvals[0]
    assert approval.driver_key == "revenue_growth"
    assert approval.current_driver_fingerprint == "b" * 64
    assert approval.approval_state.value == "UNAPPROVED"
    assert payload.backend_checks["workflow.package.state"] == "BLOCKED"
    assert (
        payload.backend_checks["workflow.package.global_blockers"]
        == "source_preflight_unproven_scope"
    )
    assert payload.backend_checks["workflow.module.forecast.state"] == (
        "NEEDS_PM_REVIEW"
    )
    assert payload.backend_checks["workflow.module.calculation.state"] == (
        "UNVERIFIED"
    )


def test_diagnostic_forecast_requires_explicit_ytd_plus_q4_stub() -> None:
    valuation = _valuation()
    valuation["forecast_bridge"][0].pop("bridge_basis")

    with pytest.raises(
        ProfessionalModelAdapterError,
        match="explicit FY26 YTD\\+Q4 stub",
    ):
        build_diagnostic_scenario_forecasts(_FakeHistory(), valuation)


def test_adapter_retains_raw_price_but_nulls_undated_normalized_price() -> None:
    valuation = _valuation()
    valuation["ciq_lineage"].pop("comps_as_of_date")
    payload = adapt_professional_workbook_payload(
        ticker="MSFT",
        requested_run_id=7,
        preflight=_preflight(),
        historical=_FakeHistory(),
        forecast_bundle=_forecast_bundle(),
        valuation_document=valuation,
        comparables=(),
        valuation_source_path=Path("frozen-undated-price.json"),
    )

    presentation = payload.source_presentations[0]
    assert presentation.raw_value == pytest.approx(385.10)
    assert presentation.normalized_value is None
    assert presentation.state.status is AvailabilityStatus.UNAVAILABLE
    assert payload.valuation_inputs["current_price"] is None
    assert payload.current_price_as_of is None


def test_render_refuses_to_overwrite_existing_run_artifacts(tmp_path: Path) -> None:
    payload = _payload()
    artifacts = render_professional_model_v2_payload(payload, output_root=tmp_path)
    workbook_bytes = artifacts.workbook_path.read_bytes()
    manifest_bytes = artifacts.manifest_path.read_bytes()

    with pytest.raises(ProfessionalModelAdapterError, match="refusing to overwrite"):
        render_professional_model_v2_payload(payload, output_root=tmp_path)

    assert artifacts.workbook_path.read_bytes() == workbook_bytes
    assert artifacts.manifest_path.read_bytes() == manifest_bytes


def test_optional_evidence_is_required_for_package_full_without_scope_proof() -> None:
    valuation = _valuation()
    payload = adapt_professional_workbook_payload(
        ticker="MSFT",
        requested_run_id=7,
        preflight=_preflight(),
        historical=_FakeHistory(),
        forecast_bundle=_forecast_bundle(),
        valuation_document=valuation,
        comparables=(),
        valuation_source_path=Path("frozen-valuation.json"),
    )

    assert payload.backend_checks["workflow.module.optional_evidence.state"] == "PARTIAL"
    assert payload.backend_checks["workflow.package.state"] != "FULL"


def test_source_backed_segment_line_is_preserved_and_rendered(tmp_path: Path) -> None:
    segment_key = next(
        spec.canonical_key
        for spec in professional_line_item_registry()
        if spec.statement_or_schedule == "segment_build"
    )
    historical = _FakeHistory()
    original_value = historical.value
    original_lineage = historical.period_lineage

    def segment_value(key: str, period: str) -> PeriodValue:
        if key == segment_key:
            return PeriodValue(
                period_key=period,
                value=42_000.0 + PERIODS.index(period),
                state=_available(),
            )
        return original_value(key, period)

    def segment_lineage(key: str, period: str):
        if key == segment_key:
            revenue_lineage = original_lineage("is.revenue", period)
            return SimpleNamespace(
                canonical_key=key,
                period_key=period,
                method_id="historical:direct",
                formula_id=None,
                state=_available(),
                source_refs=revenue_lineage.source_refs,
                normalized_value=42_000.0 + PERIODS.index(period),
            )
        return original_lineage(key, period)

    historical.value = segment_value
    historical.period_lineage = segment_lineage
    forecast_bundle = _forecast_bundle()
    for scenario in forecast_bundle.scenarios:
        scenario.line_index[segment_key] = LineSeries(
            line_key=segment_key,
            method_id="forecast:source_backed_segment_fixture",
            values=tuple(
                PeriodValue(
                    period_key=period,
                    value=45_000.0 + index,
                    state=_available(),
                )
                for index, period in enumerate(FORECAST_PERIODS)
            ),
        )
    payload = adapt_professional_workbook_payload(
        ticker="MSFT",
        requested_run_id=7,
        preflight=_preflight(),
        historical=historical,
        forecast_bundle=forecast_bundle,
        valuation_document=_valuation(),
        comparables=(),
        valuation_source_path=Path("frozen-segment.json"),
    )

    assert payload.availability["segments"].status == "available"
    assert any(line.canonical_key == segment_key for line in payload.lines)
    artifacts = render_professional_model_v2_payload(payload, output_root=tmp_path)
    workbook = load_workbook(artifacts.workbook_path, data_only=False)
    assert workbook["Segment_Build"]["B4"].value == "AVAILABLE"
    assert segment_key in {
        workbook["Segment_Build"][f"A{row}"].value
        for row in range(1, workbook["Segment_Build"].max_row + 1)
    }


def test_cy1_ltm_consensus_chain_cannot_drive_fy1_forecast() -> None:
    historical = _FakeHistory()
    original_number = historical._number
    historical._number = (
        lambda key, period: 281_724.0
        if key == "is.revenue" and period == "FY25"
        else 318_273.0
        if key == "is.revenue" and period == "LTM"
        else original_number(key, period)
    )
    valuation = _valuation()
    cy1_revenue = 354_517.9778
    ltm_revenue = 318_273.0
    cy1_growth = cy1_revenue / ltm_revenue - 1.0
    valuation["assumptions"]["revenue_mm"] = ltm_revenue
    valuation["assumptions"]["growth_near_pct"] = cy1_growth
    first = valuation["forecast_bridge"][0]
    first["revenue"] = cy1_revenue
    first["revenue_scale_to_model"] = 1.0
    first["growth_rate"] = cy1_growth

    with pytest.raises(
        ProfessionalModelAdapterError,
        match="unqualified_consensus_lineage",
    ):
        build_diagnostic_scenario_forecasts(historical, valuation)
